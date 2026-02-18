from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from CoreFisica.models import Cliente, Instalacion

HEADER_MAP = {
    'RUC': 'ruc',
    'RAZON SOCIAL': 'razon_social',
    'RAZÓN SOCIAL': 'razon_social',
    'NOMBRE COMERCIAL': 'nombre_comercial',
    'CLASIFICACION': 'clasificacion',
    'CLASIFICACIÓN': 'clasificacion',
    'INSTALACION': 'instalacion',
    'INSTALACIÓN': 'instalacion',
    'PROVINCIA': 'provincia',
    'CIUDAD': 'ciudad',
}

CLASSIF_MAP = {
    'PEQUENO': 'PEQUENO',
    'PEQUEÑO': 'PEQUENO',
    'PEQUENA': 'PEQUENO',
    'PEQUEÑA': 'PEQUENO',
    'MEDIANO': 'MEDIANO',
    'MEDIANA': 'MEDIANO',
    'GRANDE': 'GRANDE',
    'GRAN': 'GRANDE',
}

def normalize_text(val):
    if val is None:
        return ''
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val.is_integer():
            return str(int(val))
        return str(val)
    return str(val).strip()

def normalize_classif(val):
    return CLASSIF_MAP.get(normalize_text(val).upper(), '')

class Command(BaseCommand):
    help = "Importa clientes e instalaciones desde un Excel (.xlsx)"

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Ruta del archivo .xlsx a importar')

    def handle(self, *args, **options):
        path = options['file_path']
        try:
            wb = load_workbook(filename=path, read_only=True)
        except Exception as exc:
            raise CommandError(f"No se pudo abrir el archivo: {exc}")

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError("El archivo está vacío")

        headers_raw = [normalize_text(h).upper() for h in rows[0]]
        header_idx = {}
        for idx, name in enumerate(headers_raw):
            if name in HEADER_MAP:
                header_idx[HEADER_MAP[name]] = idx

        if 'nombre_comercial' not in header_idx:
            raise CommandError("Falta la columna obligatoria: NOMBRE COMERCIAL")

        created_clientes = updated_clientes = 0
        created_inst = updated_inst = 0
        errors = []

        with transaction.atomic():
            for i, row in enumerate(rows[1:], start=2):
                def col(key):
                    idx = header_idx.get(key)
                    return normalize_text(row[idx]) if idx is not None and idx < len(row) else ''

                ruc = col('ruc')
                razon_social = col('razon_social')
                nombre_comercial = col('nombre_comercial') or razon_social
                clasif = normalize_classif(col('clasificacion'))
                inst_nombre = col('instalacion') or 'SIN NOMBRE'
                provincia = col('provincia')
                ciudad = col('ciudad')

                if not nombre_comercial:
                    errors.append(f"Fila {i}: sin nombre_comercial")
                    continue

                # Cliente por RUC si hay, si no por nombre_comercial
                if ruc:
                    cliente, created = Cliente.objects.get_or_create(
                        ruc=ruc,
                        defaults={
                            'razon_social': razon_social or nombre_comercial,
                            'nombre_comercial': nombre_comercial,
                            'size': clasif or 'MEDIANO'
                        }
                    )
                else:
                    cliente, created = Cliente.objects.get_or_create(
                        nombre_comercial=nombre_comercial,
                        defaults={
                            'razon_social': razon_social or nombre_comercial,
                            'ruc': None,
                            'size': clasif or 'MEDIANO'
                        }
                    )
                if created:
                    created_clientes += 1
                else:
                    if clasif and cliente.size != clasif:
                        cliente.size = clasif
                        cliente.save(update_fields=['size'])
                        updated_clientes += 1

                # Instalación por cliente + nombre
                inst_defaults = {}
                if provincia:
                    inst_defaults['provincia'] = provincia
                if ciudad:
                    inst_defaults['ciudad'] = ciudad
                instalacion, inst_created = Instalacion.objects.get_or_create(
                    cliente=cliente,
                    nombre=inst_nombre,
                    defaults=inst_defaults
                )
                if not inst_created and inst_defaults:
                    changed = False
                    if provincia and not instalacion.provincia:
                        instalacion.provincia = provincia
                        changed = True
                    if ciudad and not instalacion.ciudad:
                        instalacion.ciudad = ciudad
                        changed = True
                    if changed:
                        instalacion.save(update_fields=['provincia', 'ciudad'])
                        updated_inst += 1
                elif inst_created:
                    created_inst += 1

        summary = (
            f"Clientes creados: {created_clientes}, actualizados: {updated_clientes}. "
            f"Instalaciones creadas: {created_inst}, actualizadas: {updated_inst}."
        )
        if errors:
            summary += f" Errores: {len(errors)}. Ejemplo: {errors[:3]}"
        self.stdout.write(self.style.SUCCESS(summary))