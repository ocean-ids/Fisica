import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from ..models import ReporteVacaciones
from ..serializers import ReporteVacacionesSerializer


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def listar_reporte_vacaciones(request):
    """Lista los registros del reporte de vacaciones (orden: más recientes primero)."""
    qs = ReporteVacaciones.objects.select_related('persona_sale_ref', 'sacavacaciones_ref').all()
    return Response(ReporteVacacionesSerializer(qs, many=True).data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def crear_reporte_vacaciones(request):
    s = ReporteVacacionesSerializer(data=request.data)
    s.is_valid(raise_exception=True)
    s.save()
    return Response(s.data, status=status.HTTP_201_CREATED)


@api_view(['PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
def actualizar_reporte_vacaciones(request, id):
    fila = get_object_or_404(ReporteVacaciones, id=id)
    s = ReporteVacacionesSerializer(fila, data=request.data, partial=True)
    s.is_valid(raise_exception=True)
    s.save()
    return Response(s.data)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def eliminar_reporte_vacaciones(request, id):
    fila = get_object_or_404(ReporteVacaciones, id=id)
    fila.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


def _dma(fecha):
    if not fecha:
        return ''
    return fecha.strftime('%d/%m/%Y')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_reporte_vacaciones_excel(request):
    """Genera el REPORTE DE VACACIONES DEL PERSONAL en el formato por bloques."""
    qs = ReporteVacaciones.objects.select_related('persona_sale_ref', 'sacavacaciones_ref').all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Vacaciones'

    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center')

    # Anchos de columna (A..G)
    widths = {'A': 14, 'B': 26, 'C': 14, 'D': 16, 'E': 26, 'F': 12, 'G': 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    def bloque(r, item):
        # Fila 1: título + código
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        c = ws.cell(row=r, column=2, value='REPORTE DE VACACIONES DEL PERSONAL')
        c.font = bold
        c.alignment = center
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        ws.cell(row=r, column=6, value='000-001').alignment = center

        # Fila 2: revisión / página / código formato
        ws.merge_cells(start_row=r + 1, start_column=2, end_row=r + 1, end_column=3)
        ws.cell(row=r + 1, column=2, value='REVISION: 01').alignment = left
        ws.merge_cells(start_row=r + 1, start_column=4, end_row=r + 1, end_column=5)
        ws.cell(row=r + 1, column=4, value='PAGINA: 1 DE 1').alignment = left
        ws.merge_cells(start_row=r + 1, start_column=6, end_row=r + 1, end_column=7)
        ws.cell(row=r + 1, column=6, value='OS-REG-OPE-030').alignment = center

        # Fila 3: DESDE / HASTA / DIAS
        ws.cell(row=r + 2, column=1, value='DESDE:').font = bold
        ws.cell(row=r + 2, column=2, value=_dma(item.fecha_desde)).alignment = center
        ws.cell(row=r + 2, column=3, value='HASTA:').font = bold
        ws.merge_cells(start_row=r + 2, start_column=4, end_row=r + 2, end_column=5)
        ws.cell(row=r + 2, column=4, value=_dma(item.fecha_hasta)).alignment = center
        ws.merge_cells(start_row=r + 2, start_column=6, end_row=r + 2, end_column=7)
        dias_txt = f"{item.dias} DIAS" if item.dias else ''
        d = ws.cell(row=r + 2, column=6, value=dias_txt)
        d.font = bold
        d.alignment = center

        # Fila 4: encabezados
        headers = ['CLIENTE', 'SALE DE VACACIONES', '', 'PERIODO', 'SACAVACACIONES']
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=r + 3, column=i, value=h)
            cell.font = bold
            cell.alignment = left
        ws.merge_cells(start_row=r + 3, start_column=5, end_row=r + 3, end_column=7)

        # Fila 5: datos
        ws.cell(row=r + 3, column=2)
        ws.merge_cells(start_row=r + 3, start_column=2, end_row=r + 3, end_column=3)
        ws.cell(row=r + 4, column=1, value=item.cliente or '').alignment = center
        ws.merge_cells(start_row=r + 4, start_column=2, end_row=r + 4, end_column=3)
        ws.cell(row=r + 4, column=2, value=item.persona_sale or '').alignment = left
        ws.cell(row=r + 4, column=4, value=item.periodo or '').alignment = center
        ws.merge_cells(start_row=r + 4, start_column=5, end_row=r + 4, end_column=7)
        ws.cell(row=r + 4, column=5, value=(item.sacavacaciones or 'N/A')).alignment = center

        # Bordes de todo el bloque
        for rr in range(r, r + 5):
            for cc in range(1, 8):
                ws.cell(row=rr, column=cc).border = border

    row = 1
    for item in qs:
        bloque(row, item)
        row += 6  # 5 filas del bloque + 1 en blanco

    if not qs:
        ws.cell(row=1, column=1, value='Sin registros de vacaciones.')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_vacaciones.xlsx"'
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    return response
