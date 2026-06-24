"""Vistas de Novedades de Puesto (apertura/modificación/cierre, formato FR): CRUD, cierre que inhabilita el puesto y export Excel por día."""
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Q
from io import BytesIO
import datetime
import json
import re

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated

import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.drawing.image import Image as XLImage

from ..models import NovedadPuesto, Puesto, Instalacion, Asignacion, PuestoHorario
from ..utils import parse_input
from .reporte_asistencia_views import _find_logo_path


TITULO_FR = 'FR REPORTE DE APERTURA, MODIFICACION O CIERRE DE PUESTO'
VERSION_FR = '.04'
FECHA_APROBACION_FR = '16-may-22'

MESES_ES = [
    '', 'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
    'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre'
]
DIAS_ES = ['lunes', 'martes', 'miércoles', 'jueves', 'viernes', 'sábado', 'domingo']


def _parse_fecha(value):
    if value:
        try:
            return datetime.date.fromisoformat(str(value))
        except (ValueError, TypeError):
            pass
    return timezone.localdate()


def _fecha_larga_es(fecha_obj):
    return f"{DIAS_ES[fecha_obj.weekday()]}, {fecha_obj.day} de {MESES_ES[fecha_obj.month]} de {fecha_obj.year}"


def _serialize(n):
    return {
        'id': n.id,
        'puesto': n.puesto_id,
        'instalacion': n.instalacion_id,
        'fecha': n.fecha.isoformat() if n.fecha else None,
        'turno': n.turno,
        'cliente_denominativo': n.cliente_denominativo,
        'sector': n.sector,
        'novedad': n.novedad,
        'tipo': n.tipo,
        'horario': n.horario,
        'solicitado_por': n.solicitado_por,
        'observacion': n.observacion,
        'creado_en': n.creado_en.isoformat() if n.creado_en else None,
    }


def _payload(request):
    if request.data:
        return request.data
    try:
        return json.loads(request.body or '{}')
    except (ValueError, TypeError):
        return {}


def _cerrar_asignaciones_futuras(puesto, fecha_cierre=None):
    """Cierra el puesto con corte EXACTO AL DÍA del cierre.

    Objetivo:
      - Reporte/Consolidado: la persona se ve hasta el día ANTERIOR al cierre
        (conserva los días ya trabajados del mes en curso) y desaparece desde el
        día del cierre en adelante.
      - Asignación: el puesto sigue saliendo (mes del cierre con la persona hasta
        el corte; meses siguientes vacantes). El front lo pinta gris porque
        puesto.activo=False.
      - Meses pasados quedan intactos.

    Mecánica:
      - Mes del cierre: a las asignaciones con persona se les fija
        end_date = fecha_cierre - 1 día (el Reporte respeta end_date). Si el cierre
        es el día 1, el mes entero queda vacante (persona=None).
      - Meses posteriores: persona=None (vacante), se mantienen ACTIVO/visibles.
    """
    hoy = timezone.localdate()
    fc = fecha_cierre or hoy
    cy, cm = fc.year, fc.month
    month_start = datetime.date(cy, cm, 1)
    cutoff = fc - datetime.timedelta(days=1)  # último día que conserva persona

    acts = Asignacion.objects.filter(puesto=puesto, estado='ACTIVO')

    # 1) Meses ANTERIORES al del cierre: las filas-mes recurrentes "arrastran"
    #    (end_date=None) hacia adelante. Se cortan en cutoff conservando la persona,
    #    para que dejen de proyectarse desde el día del cierre. Las que ya terminaron
    #    antes del cutoff no se tocan.
    acts.filter(persona__isnull=False).filter(
        Q(anio__lt=cy) | Q(anio=cy, mes__lt=cm)
    ).filter(Q(end_date__isnull=True) | Q(end_date__gt=cutoff)).update(end_date=cutoff)

    # 2) Mes del cierre:
    mes_cierre = acts.filter(persona__isnull=False, anio=cy, mes=cm)
    if fc > month_start:
        # Corte a mitad de mes: conserva la persona hasta el día anterior al cierre.
        mes_cierre.filter(
            Q(end_date__isnull=True) | Q(end_date__gt=cutoff)
        ).update(end_date=cutoff)
    else:
        # Cierre el día 1: el mes entero queda sin persona (vacante).
        mes_cierre.update(persona=None, end_date=None)

    # 3) Meses POSTERIORES al del cierre: vacante (persona=None), se mantienen ACTIVO
    #    para que el puesto siga visible (gris) en Asignación.
    acts.filter(Q(anio__gt=cy) | Q(anio=cy, mes__gt=cm)).update(persona=None)


def _abrir_asignaciones(puesto, fecha_apertura=None):
    """Reabre el puesto como VACANTE con corte EXACTO AL DÍA de la apertura.

    No restaura a la persona anterior (se reasigna manualmente). Como la novedad
    APERTURA solo se permite sobre puestos CERRADOS, las filas vacantes existentes
    provienen del cierre previo, por lo que es seguro reordenarlas:

      - Meses ANTERIORES al de la apertura: las filas vacantes pasan a INACTIVO
        (el periodo cerrado no debe notificar vacante ni reaparecer).
      - Mes de la apertura: se garantiza UNA fila vacante que arranca en la fecha
        exacta de apertura (start_date = fecha_apertura), disponible por cubrir.
      - Meses POSTERIORES: quedan vacantes (persona=None) y sin tope (end_date=None).
    """
    hoy = timezone.localdate()
    fa = fecha_apertura or hoy
    ay, am = fa.year, fa.month

    acts = Asignacion.objects.filter(puesto=puesto, estado='ACTIVO')

    # 1) Periodo cerrado anterior a la apertura: vacantes -> INACTIVO (no notifica).
    acts.filter(persona__isnull=True).filter(
        Q(anio__lt=ay) | Q(anio=ay, mes__lt=am)
    ).update(estado='INACTIVO')

    # 2) Mes de la apertura: una fila vacante que arranca en la fecha exacta.
    vacante_mes = acts.filter(persona__isnull=True, anio=ay, mes=am).first()
    if vacante_mes:
        vacante_mes.start_date = fa
        vacante_mes.end_date = None
        vacante_mes.recurring = True
        vacante_mes.save(update_fields=['start_date', 'end_date', 'recurring'])
    else:
        ref = Asignacion.objects.filter(puesto=puesto).order_by('-anio', '-mes').first()
        if ref:
            Asignacion.objects.create(
                cliente=ref.cliente,
                instalacion=ref.instalacion,
                puesto=puesto,
                horario_id=ref.horario_id,
                persona=None,
                mes=am,
                anio=ay,
                estado='ACTIVO',
                recurring=True,
                start_date=fa,
                end_date=None,
            )

    # 3) Meses posteriores: vacantes sin tope (siguen visibles/por cubrir).
    acts.filter(Q(anio__gt=ay) | Q(anio=ay, mes__gt=am)).update(persona=None, end_date=None)


def _aplicar_estado_puesto(puesto, novedad_tipo, fecha_evento=None):
    """Habilita/inhabilita el puesto según el tipo de novedad."""
    if not puesto:
        return
    tipo = (novedad_tipo or '').strip().upper()
    if tipo == 'CIERRE':
        if puesto.activo:
            puesto.activo = False
            puesto.save(update_fields=['activo'])
        _cerrar_asignaciones_futuras(puesto, fecha_evento)
    elif tipo == 'APERTURA':
        if not puesto.activo:
            puesto.activo = True
            puesto.save(update_fields=['activo'])
        _abrir_asignaciones(puesto, fecha_evento)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def obtener_novedades(request):
    if not request.user.has_perm('CoreFisica.view_novedadpuesto'):
        return JsonResponse({'detail': 'No tiene permiso para ver novedades.'}, status=403)

    qs = NovedadPuesto.objects.all()
    fecha = request.GET.get('fecha')
    if fecha:
        f = _parse_fecha(fecha)
        qs = qs.filter(fecha=f)
    cliente_id = request.GET.get('cliente_id')
    if cliente_id:
        qs = qs.filter(instalacion__cliente_id=cliente_id)
    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')
    if desde:
        qs = qs.filter(fecha__gte=_parse_fecha(desde))
    if hasta:
        qs = qs.filter(fecha__lte=_parse_fecha(hasta))

    return JsonResponse([_serialize(n) for n in qs], safe=False)


def _normalizar_tipo(text):
    """Convierte el resumen compacto ('1 24HDLD') al formato que entiende parse_input
    ('1 24H L-D Diurno'). Si ya viene en formato legible/parseable, lo deja igual.

    El resumen compacto guarda turno como una letra (D/N) o vacío (24h/Ambos) y los días
    como inicio+fin del rango (ej. 'LD' = L a D). Por eso los días se interpretan como rango.
    """
    s = (text or '').strip()
    m = re.match(r'^\s*(\d+)\s+(\d+(?:\.\d+)?)\s*H\s*([DNM]?)\s*([LMXJVSD]{1,2})\s*$', s, re.I)
    if not m:
        return s
    qty, horas, tl, dias = m.group(1), m.group(2), (m.group(3) or '').upper(), m.group(4).upper()
    turno = {'D': 'Diurno', 'N': 'Nocturno'}.get(tl, 'Ambos')  # '' o 'M' -> 24h/Ambos
    dias_txt = f'{dias[0]}-{dias[1]}' if len(dias) == 2 else dias
    return f'{qty} {horas}H {dias_txt} {turno}'


def _aplicar_tipo_a_puesto(puesto, tipo_text):
    """Aplica el 'Tipo' (resumen de horario) al horario real del puesto.

    Reconstruye los PuestoHorario desde el texto (parse_input) y conserva las horas de
    ingreso/salida actuales del puesto. Recalcula el resumen. Si el texto no es válido,
    no toca nada.
    """
    texto = _normalizar_tipo(tipo_text)
    try:
        reglas = parse_input(texto)
    except (ValueError, TypeError):
        return False
    if not reglas:
        return False
    # Conservar ingreso/salida actuales (representativos) del puesto.
    ingreso = salida = None
    for h in puesto.horarios.all():
        if h.hora_ingreso:
            ingreso, salida = h.hora_ingreso, h.hora_salida
            break
    puesto.horarios.all().delete()
    for r in reglas:
        PuestoHorario.objects.create(
            puesto=puesto,
            dia=r['dia'],
            horas=r['horas'],
            turno=r.get('turno') or 'Diurno',
            hora_ingreso=ingreso,
            hora_salida=salida,
        )
    try:
        puesto.sync_from_horarios()
        puesto.save()
    except Exception:
        pass
    return True


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def crear_novedad(request):
    if not request.user.has_perm('CoreFisica.add_novedadpuesto'):
        return JsonResponse({'detail': 'No tiene permiso para crear novedades.'}, status=403)

    data = _payload(request)
    fecha_raw = data.get('fecha')
    if not fecha_raw:
        return JsonResponse({'detail': 'La fecha es obligatoria.'}, status=400)
    novedad_tipo = (data.get('novedad') or '').strip().upper()
    if not novedad_tipo:
        return JsonResponse({'detail': 'El tipo de novedad es obligatorio.'}, status=400)

    puesto_id = data.get('puesto')
    instalacion_id = data.get('instalacion')
    puesto = None
    instalacion = None
    if puesto_id:
        puesto = Puesto.objects.filter(id=puesto_id).select_related('instalacion').first()
        if puesto:
            instalacion = puesto.instalacion
    if instalacion_id:
        instalacion = Instalacion.objects.filter(id=instalacion_id).first()

    # En Modificación, el "Tipo" editado se aplica al horario real del puesto.
    if puesto and novedad_tipo in ('MODIFICACION', 'MODIFICACION INCREMENTO'):
        sent_tipo = (data.get('tipo') or '').strip()
        if sent_tipo and sent_tipo != (puesto.resumen or '').strip():
            _aplicar_tipo_a_puesto(puesto, sent_tipo)
            puesto.refresh_from_db()

    # El "tipo" del reporte FR es el RESUMEN de horario del puesto (no el texto del puesto).
    tipo_val = (data.get('tipo') or '').strip()
    if puesto and (puesto.resumen or '').strip():
        tipo_val = puesto.resumen.strip()

    n = NovedadPuesto.objects.create(
        puesto=puesto,
        instalacion=instalacion,
        fecha=_parse_fecha(fecha_raw),
        turno=data.get('turno') or 'Diurno',
        cliente_denominativo=(data.get('cliente_denominativo') or '').strip(),
        sector=(data.get('sector') or '').strip(),
        novedad=novedad_tipo,
        tipo=tipo_val,
        horario=(data.get('horario') or '').strip(),
        solicitado_por=(data.get('solicitado_por') or '').strip(),
        observacion=(data.get('observacion') or '').strip(),
    )

    # CIERRE -> inhabilita el puesto y corta asignaciones al día del cierre.
    # APERTURA -> reactiva. Se usa la fecha de la novedad como fecha de corte.
    _aplicar_estado_puesto(puesto, novedad_tipo, n.fecha)

    return JsonResponse(_serialize(n), status=201)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def actualizar_novedad(request, id):
    if not request.user.has_perm('CoreFisica.change_novedadpuesto'):
        return JsonResponse({'detail': 'No tiene permiso para editar novedades.'}, status=403)

    n = NovedadPuesto.objects.filter(id=id).first()
    if not n:
        return JsonResponse({'detail': 'Novedad no encontrada.'}, status=404)

    data = _payload(request)
    if data.get('fecha'):
        n.fecha = _parse_fecha(data.get('fecha'))
    if data.get('turno'):
        n.turno = data.get('turno')
    if data.get('novedad'):
        n.novedad = (data.get('novedad') or '').strip().upper()
    for field in ['cliente_denominativo', 'sector', 'tipo', 'horario', 'solicitado_por', 'observacion']:
        if field in data:
            setattr(n, field, (data.get(field) or '').strip())
    if 'puesto' in data:
        n.puesto = Puesto.objects.filter(id=data.get('puesto')).first() if data.get('puesto') else None
    if 'instalacion' in data:
        n.instalacion = Instalacion.objects.filter(id=data.get('instalacion')).first() if data.get('instalacion') else None
    n.save()
    return JsonResponse(_serialize(n))


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def eliminar_novedad(request, id):
    if not request.user.has_perm('CoreFisica.delete_novedadpuesto'):
        return JsonResponse({'detail': 'No tiene permiso para eliminar novedades.'}, status=403)

    n = NovedadPuesto.objects.filter(id=id).first()
    if not n:
        return JsonResponse({'detail': 'Novedad no encontrada.'}, status=404)
    n.delete()
    return JsonResponse({'detail': 'Novedad eliminada.'})


# ---------------------------------------------------------------------------
# Export Excel formato FR
# ---------------------------------------------------------------------------

COLUMNS = ['CLIENTE\n(DENOMINATIVO)', 'SECTOR', 'NOVEDAD', 'TIPO', 'HORARIO', 'SOLICITADO POR', 'OBSERVACION']
COL_WIDTHS = [22, 18, 16, 16, 12, 18, 30]
N_COLS = len(COLUMNS)  # 7 -> A..G


def _draw_fr_header(ws, fecha_obj, thin_border):
    ws.merge_cells('A1:B2')           # logo
    ws.merge_cells('C1:E2')           # titulo
    for row in range(1, 3):
        for col in range(1, N_COLS + 1):
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws['C1'] = TITULO_FR
    ws['C1'].font = Font(bold=True, size=12)
    ws['F1'] = 'Versión:'
    ws['G1'] = VERSION_FR
    ws['F2'] = 'Fecha de aprobación:'
    ws['G2'] = FECHA_APROBACION_FR
    for ref in ['F1', 'G1', 'F2', 'G2']:
        ws[ref].font = Font(bold=True, size=9)

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 26

    # Fila FECHA (label izquierda, fecha larga centrada)
    ws.cell(row=3, column=1, value='FECHA').font = Font(bold=True, size=10)
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=N_COLS)
    fc = ws.cell(row=3, column=2, value=_fecha_larga_es(fecha_obj))
    fc.font = Font(bold=True, size=10)
    fc.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, N_COLS + 1):
        ws.cell(row=3, column=col).border = thin_border

    if _find_logo_path():
        try:
            img = XLImage(str(_find_logo_path()))
            img.width = 150
            img.height = 60
            ws.add_image(img, 'A1')
        except Exception:
            pass


def _write_turno_block(ws, start_row, turno_label, registros, thin_border):
    header_fill = PatternFill('solid', fgColor='D9D9D9')
    bold = Font(bold=True, size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Fila titulo del turno
    row = start_row
    tcell = ws.cell(row=row, column=1, value=f'TURNO {turno_label.upper()}')
    tcell.font = bold
    for col in range(1, N_COLS + 1):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

    # Cabecera de columnas
    for idx, title in enumerate(COLUMNS):
        c = ws.cell(row=row, column=idx + 1, value=title)
        c.font = bold
        c.alignment = center
        c.fill = header_fill
        c.border = thin_border
    ws.row_dimensions[row].height = 28
    row += 1

    # Filas de datos
    for n in registros:
        valores = [
            n.cliente_denominativo or '',
            n.sector or '',
            n.novedad or '',
            n.tipo or '',
            n.horario or '',
            n.solicitado_por or '',
            n.observacion or '',
        ]
        for idx, val in enumerate(valores):
            c = ws.cell(row=row, column=idx + 1, value=val)
            c.border = thin_border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row += 1

    # Filas vacías de relleno (como el formato impreso)
    for _ in range(max(0, 6 - len(registros))):
        for col in range(1, N_COLS + 1):
            ws.cell(row=row, column=col).border = thin_border
        row += 1

    return row


def _build_day_sheet(ws, fecha_obj, registros, border):
    """Dibuja en `ws` el reporte FR de un día (header + Diurno + Nocturno)."""
    for idx, width in enumerate(COL_WIDTHS):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = width

    diurnos = [n for n in registros if (n.turno or '').lower().startswith('d')]
    nocturnos = [n for n in registros if (n.turno or '').lower().startswith('n')]

    _draw_fr_header(ws, fecha_obj, border)
    next_row = _write_turno_block(ws, 5, 'Diurno', diurnos, border)
    next_row += 1
    _write_turno_block(ws, next_row, 'Nocturno', nocturnos, border)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_novedades_excel(request):
    if not request.user.has_perm('CoreFisica.export_novedadpuesto'):
        return JsonResponse({'detail': 'No tiene permiso para exportar.'}, status=403)

    # La fecha recibida define el MES a exportar. Se genera una hoja (pestaña)
    # por cada día del mes que tenga registros, nombrada con la fecha (DD-MM-YYYY).
    fecha_obj = _parse_fecha(request.GET.get('fecha'))
    anio, mes = fecha_obj.year, fecha_obj.month

    qs = (NovedadPuesto.objects
          .filter(fecha__year=anio, fecha__month=mes)
          .select_related('instalacion')
          .order_by('fecha'))

    # Agrupar por día
    por_dia = {}
    for n in qs:
        por_dia.setdefault(n.fecha, []).append(n)

    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    primera = wb.active

    if not por_dia:
        # Sin registros en el mes: una sola hoja del día seleccionado.
        primera.title = fecha_obj.strftime('%d-%m-%Y')
        _build_day_sheet(primera, fecha_obj, [], border)
    else:
        dias_ordenados = sorted(por_dia.keys())
        for i, dia in enumerate(dias_ordenados):
            ws = primera if i == 0 else wb.create_sheet()
            ws.title = dia.strftime('%d-%m-%Y')  # pestaña por día
            _build_day_sheet(ws, dia, por_dia[dia], border)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="reporte_novedades_{mes:02d}_{anio}.xlsx"'
    )
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    return response
