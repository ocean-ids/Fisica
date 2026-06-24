"""Vistas de Personas: CRUD, alta/baja, import/export Excel y gestión de sacafrancos."""
from django.http import JsonResponse, HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.db import IntegrityError, transaction
from django.db.models import Q
from ..models import Persona, AsignacionSemanal, Puesto, Asignacion, Horario, Provincia, Canton, CoberturaSacafranco
from ..utils import _strip_accents
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
import csv
import io
import re
import logging
import datetime

logger = logging.getLogger(__name__)

DAY_KEYS = ('mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun')


def _normalize_day(tok):
    t = str(tok or '').strip().lower()
    mapping = {
        'l': 'mon', 'lu': 'mon', 'lun': 'mon', 'lunes': 'mon',
        'm': 'tue', 'ma': 'tue', 'mar': 'tue', 'martes': 'tue',
        'mi': 'wed', 'mie': 'wed', 'mié': 'wed', 'mier': 'wed', 'miercoles': 'wed', 'miércoles': 'wed',
        'j': 'thu', 'ju': 'thu', 'jue': 'thu', 'jueves': 'thu',
        'v': 'fri', 'vi': 'fri', 'vie': 'fri', 'viernes': 'fri',
        's': 'sat', 'sa': 'sat', 'sab': 'sat', 'sabado': 'sat', 'sábado': 'sat',
        'd': 'sun', 'do': 'sun', 'dom': 'sun', 'domingo': 'sun'
    }
    if t in DAY_KEYS:
        return t
    return mapping.get(t, '')


def _iter_future_week_starts(start_date, years=3):
    weeks = []
    year_cursor = start_date.year
    month_cursor = start_date.month
    end_year = start_date.year + years
    while (year_cursor < end_year) or (year_cursor == end_year and month_cursor <= 12):
        base = datetime.date(year_cursor, month_cursor, 1)
        cursor = base
        while cursor.month == month_cursor:
            if cursor >= start_date:
                weeks.append(cursor)
            cursor += datetime.timedelta(days=7)
        month_cursor += 1
        if month_cursor > 12:
            month_cursor = 1
            year_cursor += 1
    return weeks


def _get_or_create_sacafranco_marker_row(puesto, week_start):
    semanal = AsignacionSemanal.objects.filter(
        puesto=puesto,
        week_start=week_start,
        asignacion__isnull=True,
    ).order_by('id').first()
    if semanal:
        return semanal
    return AsignacionSemanal.objects.create(puesto=puesto, week_start=week_start)


def _marker_value_allows_sacafranco(semanal, day):
    value = str(getattr(semanal, day, '') or '').strip()
    return value == '' or value.upper().startswith('F'), value


def _assign_sacafranco_without_asignacion(persona, puesto, week_start_date, day, replace=False):
    prop_start = week_start_date

    def _format_conflict_message(cobertura):
        conflict_puesto = cobertura.puesto
        conflict_instalacion = getattr(conflict_puesto, 'instalacion', None)
        conflict_cliente = getattr(conflict_instalacion, 'cliente', None) if conflict_instalacion else None
        cliente_nombre = getattr(conflict_cliente, 'nombre_comercial', None) or 'cliente'
        instalacion_nombre = getattr(conflict_instalacion, 'nombre', None) or 'instalacion'
        puesto_nombre = getattr(conflict_puesto, 'nombre', None) or 'puesto'
        return (
            f"La persona ya esta asignada como sacafranco en {cliente_nombre} "
            f"({instalacion_nombre} - {puesto_nombre}) para esa fecha"
        )

    existing_slot = CoberturaSacafranco.objects.filter(
        puesto=puesto,
        week_start=week_start_date,
        day=day,
    ).first()
    if existing_slot and existing_slot.persona_id != persona.id and not replace:
        return JsonResponse({'error': 'Ese puesto ya tiene un sacafranco asignado para esa fecha'}, status=400)

    existing_person = CoberturaSacafranco.objects.filter(
        persona=persona,
        week_start=week_start_date,
        day=day,
    ).exclude(puesto=puesto).first()
    if existing_person:
        return JsonResponse({'error': _format_conflict_message(existing_person)}, status=400)

    with transaction.atomic():
        weeks = _iter_future_week_starts(prop_start)
        selected_semanal = AsignacionSemanal.objects.filter(
            puesto=puesto,
            week_start=week_start_date,
        ).order_by('id').first()

        for ws in weeks:
            slot_qs = CoberturaSacafranco.objects.filter(
                puesto=puesto,
                week_start=ws,
                day=day,
            )
            slot_conflict = slot_qs.exclude(persona=persona).exists()
            if slot_conflict:
                if replace:
                    slot_qs.exclude(persona=persona).delete()
                else:
                    continue

            person_conflict = CoberturaSacafranco.objects.filter(
                persona=persona,
                week_start=ws,
                day=day,
            ).exclude(puesto=puesto).exists()
            if person_conflict:
                continue

            CoberturaSacafranco.objects.get_or_create(
                persona=persona,
                puesto=puesto,
                week_start=ws,
                day=day,
            )

            if ws == week_start_date:
                selected_semanal = AsignacionSemanal.objects.filter(
                    puesto=puesto,
                    week_start=ws,
                ).order_by('id').first() or selected_semanal

        if selected_semanal is None:
            CoberturaSacafranco.objects.get_or_create(
                persona=persona,
                puesto=puesto,
                week_start=week_start_date,
                day=day,
            )
            selected_semanal = AsignacionSemanal.objects.filter(
                puesto=puesto,
                week_start=week_start_date,
            ).order_by('id').first()

    return JsonResponse({'status': 'assigned', 'semanal_id': selected_semanal.id})


def _clear_sacafranco_marker_if_unused(puesto, week_start, day):
    if CoberturaSacafranco.objects.filter(puesto=puesto, week_start=week_start, day=day).exists():
        return None

    semanal = AsignacionSemanal.objects.filter(
        puesto=puesto,
        week_start=week_start,
        asignacion__isnull=True,
        **{f'{day}__istartswith': 'F'}
    ).order_by('id').first()
    if not semanal:
        return None

    setattr(semanal, day, '')
    semanal.save()
    return semanal


def _desasignar_sacafranco_without_asignacion(persona, puesto, week_start_date, day):
    day_offsets = {
        'mon': 0,
        'tue': 1,
        'wed': 2,
        'thu': 3,
        'fri': 4,
        'sat': 5,
        'sun': 6,
    }
    day_offset = day_offsets.get(day, 0)
    today = datetime.date.today()
    selected_cell_date = week_start_date + datetime.timedelta(days=day_offset)
    cutoff_date = today if today > selected_cell_date else selected_cell_date
    prop_end = datetime.date(week_start_date.year, 12, 31)

    candidates = CoberturaSacafranco.objects.filter(
        persona=persona,
        puesto=puesto,
        day=day,
        week_start__gte=week_start_date,
        week_start__lte=prop_end,
    ).order_by('week_start')

    delete_ids = []
    affected_weeks = []
    for cobertura in candidates:
        row_day_date = cobertura.week_start + datetime.timedelta(days=day_offset)
        if row_day_date < cutoff_date:
            continue
        delete_ids.append(cobertura.id)
        affected_weeks.append(cobertura.week_start)

    if not delete_ids:
        return None

    with transaction.atomic():
        CoberturaSacafranco.objects.filter(id__in=delete_ids).delete()
        semanal = None
        for ws in affected_weeks:
            cleared = _clear_sacafranco_marker_if_unused(puesto, ws, day)
            if ws == week_start_date:
                semanal = cleared

    return JsonResponse({'status': 'unassigned', 'semanal_id': getattr(semanal, 'id', None)})


def _resolve_provincia_id(token):
    if token is None or token == '':
        return None
    try:
        from ..utils import buscar_o_crear_provincia
        prov = buscar_o_crear_provincia(token, Provincia)
        return prov.id if prov else None
    except Exception:
        return None


def _resolve_canton_id(token, provincia_id=None):
    if token is None or token == '':
        return None
    try:
        return int(token)
    except (TypeError, ValueError):
        pass
    try:
        qs = Canton.objects.all()
        if provincia_id:
            qs = qs.filter(provincia_id=provincia_id)
        name = str(token).strip()
        canton = qs.filter(nombre__iexact=name).first()
        if canton:
            return canton.id
        if not provincia_id:
            return None
        canton = Canton.objects.create(nombre=name.upper(), provincia_id=provincia_id)
        return canton.id
    except Exception:
        return None


def _foto_url(request, foto):
    """URL absoluta de la foto (cae a la relativa si build_absolute_uri falla)."""
    if not foto:
        return None
    try:
        return request.build_absolute_uri(foto.url)
    except Exception:
        try:
            return foto.url
        except Exception:
            return None


def _fecha_persona(v):
    """Parsea 'YYYY-MM-DD' (o vacio/None) a date o None."""
    if not v:
        return None
    try:
        import datetime as _dt
        return _dt.date.fromisoformat(str(v)[:10])
    except (TypeError, ValueError):
        return None


def _aplicar_campos_persona(persona, data):
    """Asigna a la persona los campos editables del ERP que vengan en data.

    Solo toca los campos presentes (para no borrar lo que no se envia).
    """
    CHAR = [
        'sexo', 'estado_civil', 'lugar_nacimiento', 'direccion', 'telefono',
        'conyuge', 'nacionalidad', 'parroquia', 'cargo', 'departamento', 'seccion',
        'correo_personal', 'codigo_erp', 'centro_costo', 'unidad_negocio',
        'tipo_empleado', 'forma_pago', 'numero_afiliacion', 'numero_contrato',
        'actividad', 'perfil', 'motivo_salida', 'region',
    ]
    for f in CHAR:
        if f in data:
            setattr(persona, f, str(data.get(f) or '').strip())

    for f in ['fecha_nacimiento', 'fecha_ingreso', 'fecha_salida', 'fecha_pago_liquidacion']:
        if f in data:
            setattr(persona, f, _fecha_persona(data.get(f)))

    if 'estatura' in data:
        try:
            persona.estatura = data.get('estatura') if data.get('estatura') not in ('', None) else None
        except Exception:
            persona.estatura = None

    for f in ['gypaseg', 'affis', 'pbip']:
        if f in data:
            setattr(persona, f, bool(data.get(f)))

    # Activo/inactivo desde el checkbox del formulario (refleja el toggle de la lista).
    if 'is_active' in data:
        persona.is_active = bool(data.get('is_active'))

    if 'cliente' in data or 'cliente_id' in data:
        cid = data.get('cliente') if 'cliente' in data else data.get('cliente_id')
        try:
            persona.cliente_id = int(cid) if cid not in ('', None) else None
        except (TypeError, ValueError):
            persona.cliente_id = None


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def crear_persona(request):
    if not request.user.has_perm('CoreFisica.add_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    # variable data para recibir los datos de request.data, se asigna a data para evitar modificar request.data directamente
    data = request.data

    #cedula es obligatoria, se valida que exista, no esté vacía, tenga solo dígitos y máximo 10 caracteres
    cedula = (data.get('cedula') or '').strip()
    if not cedula:
        return JsonResponse({'error': 'Cédula es obligatoria'}, status=400)
    if not re.match(r'^\d{1,10}$', cedula):
        return JsonResponse({'error': 'Cédula inválida: sólo dígitos, máximo 10'}, status=400)

    #nombres y apellidos se normalizan a mayuscula y se quitan espacios al inicio y final
    nombres = str(data.get('nombres') or '').strip().upper()
    apellidos = str(data.get('apellidos') or '').strip().upper()

    
    provincia_token = data.get('provincia') or data.get('provincia_id')
    provincia_id = _resolve_provincia_id(provincia_token)
    canton_token = data.get('canton') or data.get('canton_id')
    canton_id = _resolve_canton_id(canton_token, provincia_id)

    try:
        persona = Persona(
            nombres=nombres,
            apellidos=apellidos,
            cedula=cedula,
            tipo=data.get('tipo'),
            provincia_id=provincia_id,
            canton_id=canton_id,
        )
        _aplicar_campos_persona(persona, data)
        persona.save()
        return JsonResponse({'message': 'Persona creada correctamente', 'id': persona.id}, status=201)
    except IntegrityError:
        return JsonResponse({'error': 'Cédula ya registrada'}, status=400)
    except Exception:
        logger.exception('Error creando persona')
        return JsonResponse({'error': 'No se pudo crear persona'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def obtener_personas(request):
    # si el ususario tiene permiso de ver personas, se obtiene el parametro de busqueda q y tipo para filtrar por tipo de persona (empleado, cliente, etc)
    if not request.user.has_perm('CoreFisica.view_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    try:
        q = (request.GET.get('q') or '').strip()
        tipo = (request.GET.get('tipo') or '').strip()

        # la variable persona se asigna a la consulta de todas las personas, luego se filtra por q si existe, buscando coincidencias en nombres, apellidos o cedula, y por tipo si se especifica. Finalmente se ordena por apellidos
        personas = Persona.objects.all()
        if q:
            qn = _strip_accents(q)
            personas = personas.filter(
                Q(nombres__unaccent__icontains=qn) |
                Q(apellidos__unaccent__icontains=qn) |
                Q(cedula__icontains=q) |
                Q(provincia__nombre__unaccent__icontains=qn) |
                Q(canton__nombre__unaccent__icontains=qn)
            )

        if tipo:
            personas = personas.filter(tipo=tipo)

        personas = personas.select_related('provincia', 'canton', 'cliente').order_by('apellidos')

        data = []
        for p in personas:
            data.append({
                'id': p.id,
                'nombres': p.nombres,
                'apellidos': p.apellidos,
                'cedula': p.cedula,
                'tipo': p.tipo,
                'is_active': p.is_active,
                'provincia': p.provincia_id,
                'canton': p.canton_id,
                'provincia_nombre': getattr(p.provincia, 'nombre', None),
                'canton_nombre': getattr(p.canton, 'nombre', None),
                'foto': _foto_url(request, p.foto),
                'codigo_erp': p.codigo_erp,
                'cargo': p.cargo,
                'sexo': p.sexo,
                'estado_civil': p.estado_civil,
                'fecha_nacimiento': p.fecha_nacimiento,
                'lugar_nacimiento': p.lugar_nacimiento,
                'nacionalidad': p.nacionalidad,
                'telefono': p.telefono,
                'correo_personal': p.correo_personal,
                'direccion': p.direccion,
                'parroquia': p.parroquia,
                'fecha_ingreso': p.fecha_ingreso,
                'fecha_salida': p.fecha_salida,
                'seccion': p.seccion,
                'departamento': p.departamento,
                'unidad_negocio': p.unidad_negocio,
                'tipo_empleado': p.tipo_empleado,
                'cliente': p.cliente_id,
                'cliente_nombre': getattr(p.cliente, 'razon_social', None),
            })

        return JsonResponse(data, safe=False)
    except Exception:
        logger.exception('Error obteniendo personas')
        return JsonResponse({'error': 'No se encontraron personas'}, status=404)
        

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def actualizar_persona(request, id):
    if not request.user.has_perm('CoreFisica.change_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    data = request.data

    try:
        persona = Persona.objects.get(id=id)
    except Persona.DoesNotExist:
        return JsonResponse({'error': 'Persona no encontrada'}, status=404)

    # si el request.data contiene los campos nombres o apellidos, se actualizan normalizando a mayuscula y quitando espacios al inicio y final. Si contiene cedula, se valida que no esté vacía, tenga solo dígitos y máximo 10 caracteres antes de actualizarla. El campo tipo se actualiza si está presente en el request.data, de lo contrario se mantiene el valor actual
    if 'nombres' in data:
        nombres = data.get('nombres')
        persona.nombres = str(nombres or '').strip().upper()
    if 'apellidos' in data:
        apellidos = data.get('apellidos')
        persona.apellidos = str(apellidos or '').strip().upper()

    # si se proporciona una nueva cedula, se valida que no este vacia, tenga solo dígitos y máximo 10 caracteres antes de actualizarla. Si la cedula es inválida, se retorna un error. Si es válida, se actualiza la cedula de la persona
    cedula_in = data.get('cedula')
    if cedula_in is not None:
        cedula_in = cedula_in.strip()
        if not cedula_in:
            return JsonResponse({'error': 'Cédula es obligatoria'}, status=400)
        if not re.match(r'^\d{1,10}$', cedula_in):
            return JsonResponse({'error': 'Cédula inválida: sólo dígitos, máximo 10'}, status=400)
        persona.cedula = cedula_in

    # persona.tipo se actualiza si el campo tipo está presente en el request.data, de lo contrario se mantiene el valor actual. Esto permite actualizar el tipo de persona (empleado, cliente, etc) si se proporciona en la solicitud, sin requerir que siempre esté presente
    persona.tipo = data.get('tipo', persona.tipo)

    if 'provincia' in data or 'provincia_id' in data:
        provincia_token = data.get('provincia') if 'provincia' in data else data.get('provincia_id')
        persona.provincia_id = _resolve_provincia_id(provincia_token)

    if 'canton' in data or 'canton_id' in data:
        canton_token = data.get('canton') if 'canton' in data else data.get('canton_id')
        persona.canton_id = _resolve_canton_id(canton_token, persona.provincia_id)

    # Campos del ERP editables (los que vengan en el request)
    _aplicar_campos_persona(persona, data)

    try:
        persona.save()
        return JsonResponse({'message': 'Persona actualizada correctamente', 'id': persona.id})
    except IntegrityError:
        return JsonResponse({'error': 'Cédula ya registrada'}, status=400)
    except Exception:
        logger.exception('Error actualizando persona id=%s', id)
        return JsonResponse({'error': 'No se pudo actualizar persona'}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def subir_foto_persona(request, id):
    """Sube/reemplaza la foto de una persona (multipart, campo 'foto')."""
    if not request.user.has_perm('CoreFisica.change_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        persona = Persona.objects.get(id=id)
    except Persona.DoesNotExist:
        return JsonResponse({'error': 'Persona no encontrada'}, status=404)

    archivo = request.FILES.get('foto')
    if not archivo:
        return JsonResponse({'error': 'No se envió ninguna foto'}, status=400)
    try:
        persona.foto = archivo
        persona.save(update_fields=['foto'])
        return JsonResponse({'message': 'Foto actualizada', 'foto': _foto_url(request, persona.foto)})
    except Exception:
        logger.exception('Error subiendo foto persona id=%s', id)
        return JsonResponse({'error': 'No se pudo subir la foto'}, status=500)


# --- Nómina del empleado (Ingresos / Descuentos) ---
_NOMINA_DECIMALES = [
    'sueldo', 'desc_genesis', 'bonificacion', 'transporte', 'compensacion',
    'horas_25', 'horas_50', 'horas_100',
    'decimo_tercer', 'decimo_cuarto', 'vacaciones', 'fondo_reserva',
    'moviliza', 'lunch', 'anticipo_22', 'viaticos', 'descuento', 'ingreso_extra',
    'subsidio_enfermedad_pct', 'subsidio_accidente_pct', 'subsidio_maternidad_pct',
]
_NOMINA_BOOLEANOS = [
    'pagar_fondo_reserva', 'pagar_rol_10mo_3ero', 'pagar_rol_10mo_4to', 'pagar_rol_vacaciones',
    'desc_aporte_conyuge', 'giro_contable_liquidacion',
    'subsidio_enfermedad', 'subsidio_accidente', 'subsidio_maternidad',
]
_NOMINA_TEXTO = ['observaciones', 'numero_liquidacion_ministerio', 'concepto']


def _serialize_nomina(n):
    data = {f: str(getattr(n, f)) for f in _NOMINA_DECIMALES}
    data.update({f: bool(getattr(n, f)) for f in _NOMINA_BOOLEANOS})
    data.update({f: getattr(n, f) or '' for f in _NOMINA_TEXTO})
    data['persona'] = n.persona_id
    return data


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def obtener_nomina(request, id):
    """Devuelve la nómina (ingresos/descuentos) del empleado. La crea vacía si no existe."""
    if not request.user.has_perm('CoreFisica.view_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        persona = Persona.objects.get(id=id)
    except Persona.DoesNotExist:
        return JsonResponse({'error': 'Persona no encontrada'}, status=404)
    from ..models import EmpleadoNomina
    nomina, _ = EmpleadoNomina.objects.get_or_create(persona=persona)
    return JsonResponse(_serialize_nomina(nomina))


@api_view(['PUT', 'POST'])
@permission_classes([IsAuthenticated])
def guardar_nomina(request, id):
    """Crea/actualiza la nómina (ingresos/descuentos) del empleado."""
    if not request.user.has_perm('CoreFisica.change_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        persona = Persona.objects.get(id=id)
    except Persona.DoesNotExist:
        return JsonResponse({'error': 'Persona no encontrada'}, status=404)

    from decimal import Decimal, InvalidOperation
    from ..models import EmpleadoNomina
    data = request.data if request.data else {}
    nomina, _ = EmpleadoNomina.objects.get_or_create(persona=persona)

    def to_dec(v):
        try:
            return Decimal(str(v if v not in (None, '') else 0))
        except (InvalidOperation, ValueError):
            return Decimal('0')

    for f in _NOMINA_DECIMALES:
        if f in data:
            setattr(nomina, f, to_dec(data.get(f)))
    for f in _NOMINA_BOOLEANOS:
        if f in data:
            setattr(nomina, f, bool(data.get(f)))
    for f in _NOMINA_TEXTO:
        if f in data:
            setattr(nomina, f, (data.get(f) or '').strip())
    nomina.save()
    return JsonResponse({'message': 'Nómina guardada', 'nomina': _serialize_nomina(nomina)})


# --- Otros Datos del empleado (bancario / contable / vacaciones / cargas / gastos) ---
_OTROS_DECIMALES = [
    'gasto_salud', 'gasto_vestimenta', 'gasto_educacion', 'gasto_vivienda',
    'gasto_alimentacion', 'gasto_arte_cultura', 'gasto_turismo',
]
_OTROS_ENTEROS = ['dias_vacaciones', 'numero_cargas']
_OTROS_BOOLEANOS = ['incluir_en_rol', 'acreditar']
_OTROS_FECHAS = ['ultima_liquidacion', 'fecha_ini_vacaciones', 'fecha_fin_vacaciones']
_OTROS_TEXTO = [
    'grupo_sanguineo', 'banco', 'cuenta_ahorros', 'cuenta_corriente',
    'codigo_cuenta', 'cuenta_departamento', 'cuenta_auxiliar',
]


def _serialize_otros(o):
    data = {f: str(getattr(o, f)) for f in _OTROS_DECIMALES}
    data.update({f: int(getattr(o, f) or 0) for f in _OTROS_ENTEROS})
    data.update({f: bool(getattr(o, f)) for f in _OTROS_BOOLEANOS})
    data.update({f: (getattr(o, f).isoformat() if getattr(o, f) else None) for f in _OTROS_FECHAS})
    data.update({f: getattr(o, f) or '' for f in _OTROS_TEXTO})
    data['persona'] = o.persona_id
    return data


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def obtener_otros_datos(request, id):
    """Devuelve los 'Otros Datos' del empleado. Los crea vacíos si no existen."""
    if not request.user.has_perm('CoreFisica.view_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        persona = Persona.objects.get(id=id)
    except Persona.DoesNotExist:
        return JsonResponse({'error': 'Persona no encontrada'}, status=404)
    from ..models import EmpleadoOtrosDatos
    otros, _ = EmpleadoOtrosDatos.objects.get_or_create(persona=persona)
    return JsonResponse(_serialize_otros(otros))


@api_view(['PUT', 'POST'])
@permission_classes([IsAuthenticated])
def guardar_otros_datos(request, id):
    """Crea/actualiza los 'Otros Datos' del empleado."""
    if not request.user.has_perm('CoreFisica.change_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        persona = Persona.objects.get(id=id)
    except Persona.DoesNotExist:
        return JsonResponse({'error': 'Persona no encontrada'}, status=404)

    from decimal import Decimal, InvalidOperation
    from ..models import EmpleadoOtrosDatos
    data = request.data if request.data else {}
    otros, _ = EmpleadoOtrosDatos.objects.get_or_create(persona=persona)

    def to_dec(v):
        try:
            return Decimal(str(v if v not in (None, '') else 0))
        except (InvalidOperation, ValueError):
            return Decimal('0')

    def to_int(v):
        try:
            return int(v) if v not in (None, '') else 0
        except (ValueError, TypeError):
            return 0

    for f in _OTROS_DECIMALES:
        if f in data:
            setattr(otros, f, to_dec(data.get(f)))
    for f in _OTROS_ENTEROS:
        if f in data:
            setattr(otros, f, to_int(data.get(f)))
    for f in _OTROS_BOOLEANOS:
        if f in data:
            setattr(otros, f, bool(data.get(f)))
    for f in _OTROS_FECHAS:
        if f in data:
            setattr(otros, f, _fecha_persona(data.get(f)))
    for f in _OTROS_TEXTO:
        if f in data:
            setattr(otros, f, (data.get(f) or '').strip())
    otros.save()
    return JsonResponse({'message': 'Otros datos guardados', 'otros_datos': _serialize_otros(otros)})


# --- Referencias del empleado (datos referenciales / estudios / servicios) ---
_REF_ENTEROS = ['edad', 'anios_estudio']
_REF_BOOLEANOS = [
    'primaria', 'secundaria', 'universidad',
    'miembro_fuerza_publica', 'realizo_servicio_militar',
]
_REF_TEXTO = [
    'cedula_militar', 'observacion', 'maniobras', 'carnet_conadis',
    'numero_certificado_votacion', 'licencia_conducir', 'codigo_iess',
    'certificado_violencia_intrafamiliar', 'titulo', 'contrato_inspectoria',
]


def _serialize_referencias(r):
    data = {f: getattr(r, f) for f in _REF_ENTEROS}
    data.update({f: bool(getattr(r, f)) for f in _REF_BOOLEANOS})
    data.update({f: getattr(r, f) or '' for f in _REF_TEXTO})
    data['persona'] = r.persona_id
    return data


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def obtener_referencias(request, id):
    """Devuelve las 'Referencias' del empleado. Las crea vacías si no existen."""
    if not request.user.has_perm('CoreFisica.view_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        persona = Persona.objects.get(id=id)
    except Persona.DoesNotExist:
        return JsonResponse({'error': 'Persona no encontrada'}, status=404)
    from ..models import EmpleadoReferencias
    ref, _ = EmpleadoReferencias.objects.get_or_create(persona=persona)
    return JsonResponse(_serialize_referencias(ref))


@api_view(['PUT', 'POST'])
@permission_classes([IsAuthenticated])
def guardar_referencias(request, id):
    """Crea/actualiza las 'Referencias' del empleado."""
    if not request.user.has_perm('CoreFisica.change_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        persona = Persona.objects.get(id=id)
    except Persona.DoesNotExist:
        return JsonResponse({'error': 'Persona no encontrada'}, status=404)

    from ..models import EmpleadoReferencias
    data = request.data if request.data else {}
    ref, _ = EmpleadoReferencias.objects.get_or_create(persona=persona)

    def to_int(v):
        try:
            return int(v) if v not in (None, '') else None
        except (ValueError, TypeError):
            return None

    for f in _REF_ENTEROS:
        if f in data:
            setattr(ref, f, to_int(data.get(f)))
    for f in _REF_BOOLEANOS:
        if f in data:
            setattr(ref, f, bool(data.get(f)))
    for f in _REF_TEXTO:
        if f in data:
            setattr(ref, f, (data.get(f) or '').strip())
    ref.save()
    return JsonResponse({'message': 'Referencias guardadas', 'referencias': _serialize_referencias(ref)})


# --- Documentos del empleado (rutas compartidas de red) ---
def _serialize_documento(d):
    return {
        'id': d.id,
        'tipo': d.tipo or '',
        'nombre_archivo': d.nombre_archivo or '',
        'ruta_archivo': d.ruta_archivo or '',
        'extension': d.extension or '',
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def obtener_documentos(request, id):
    """Lista los documentos (referencias a rutas) del empleado."""
    if not request.user.has_perm('CoreFisica.view_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    if not Persona.objects.filter(id=id).exists():
        return JsonResponse({'error': 'Persona no encontrada'}, status=404)
    from ..models import EmpleadoDocumento
    docs = EmpleadoDocumento.objects.filter(persona_id=id)
    return JsonResponse([_serialize_documento(d) for d in docs], safe=False)


@api_view(['PUT', 'POST'])
@permission_classes([IsAuthenticated])
def guardar_documentos(request, id):
    """Reemplaza la lista completa de documentos del empleado con la enviada."""
    if not request.user.has_perm('CoreFisica.change_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        persona = Persona.objects.get(id=id)
    except Persona.DoesNotExist:
        return JsonResponse({'error': 'Persona no encontrada'}, status=404)

    from ..models import EmpleadoDocumento
    data = request.data if request.data else {}
    lista = data.get('documentos') if isinstance(data, dict) else None
    if lista is None:
        lista = data if isinstance(data, list) else []

    with transaction.atomic():
        EmpleadoDocumento.objects.filter(persona=persona).delete()
        creados = []
        for item in lista:
            if not isinstance(item, dict):
                continue
            ruta = (item.get('ruta_archivo') or '').strip()
            nombre = (item.get('nombre_archivo') or '').strip()
            if not ruta and not nombre:
                continue  # fila vacía: se descarta
            creados.append(EmpleadoDocumento.objects.create(
                persona=persona,
                tipo=(item.get('tipo') or 'PDF GENERAL').strip(),
                nombre_archivo=nombre,
                ruta_archivo=ruta,
                extension=(item.get('extension') or '').strip(),
            ))
    return JsonResponse({'message': 'Documentos guardados',
                         'documentos': [_serialize_documento(d) for d in creados]})


# --- Más Referencias: experiencia, referencias personales, nivel de estudio, formación ---
def _ser_experiencia(e):
    return {'empresa': e.empresa or '', 'puesto_cargo': e.puesto_cargo or '',
            'tiempo': e.tiempo or '', 'motivo_salida': e.motivo_salida or ''}


def _ser_ref_personal(r):
    return {'persona_contactar': r.persona_contactar or '', 'relacion': r.relacion or '',
            'telefonos': r.telefonos or '', 'comentario': r.comentario or ''}


def _ser_nivel(n):
    return {'nivel_estudio': n.nivel_estudio or '', 'completa': bool(n.completa),
            'centro_capacitacion': n.centro_capacitacion or ''}


def _ser_formacion(f):
    return {'centro_capacitacion': f.centro_capacitacion or '', 'curso': f.curso or '',
            'area': f.area or '', 'horas': int(f.horas or 0)}


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def obtener_mas_referencias(request, id):
    """Devuelve experiencia, referencias personales, nivel de estudio y formación."""
    if not request.user.has_perm('CoreFisica.view_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    if not Persona.objects.filter(id=id).exists():
        return JsonResponse({'error': 'Persona no encontrada'}, status=404)
    from ..models import (EmpleadoExperiencia, EmpleadoReferenciaPersonal,
                          EmpleadoNivelEstudio, EmpleadoFormacion)
    return JsonResponse({
        'experiencias': [_ser_experiencia(x) for x in EmpleadoExperiencia.objects.filter(persona_id=id)],
        'referencias_personales': [_ser_ref_personal(x) for x in EmpleadoReferenciaPersonal.objects.filter(persona_id=id)],
        'niveles_estudio': [_ser_nivel(x) for x in EmpleadoNivelEstudio.objects.filter(persona_id=id)],
        'formaciones': [_ser_formacion(x) for x in EmpleadoFormacion.objects.filter(persona_id=id)],
    })


@api_view(['PUT', 'POST'])
@permission_classes([IsAuthenticated])
def guardar_mas_referencias(request, id):
    """Reemplaza las 4 listas de 'Más Referencias' con las enviadas."""
    if not request.user.has_perm('CoreFisica.change_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        persona = Persona.objects.get(id=id)
    except Persona.DoesNotExist:
        return JsonResponse({'error': 'Persona no encontrada'}, status=404)

    from ..models import (EmpleadoExperiencia, EmpleadoReferenciaPersonal,
                          EmpleadoNivelEstudio, EmpleadoFormacion)
    data = request.data if request.data else {}

    def txt(item, key):
        return (item.get(key) or '').strip()

    def rows(key):
        v = data.get(key)
        return v if isinstance(v, list) else []

    with transaction.atomic():
        EmpleadoExperiencia.objects.filter(persona=persona).delete()
        for it in rows('experiencias'):
            if not isinstance(it, dict):
                continue
            if any([txt(it, 'empresa'), txt(it, 'puesto_cargo'), txt(it, 'tiempo'), txt(it, 'motivo_salida')]):
                EmpleadoExperiencia.objects.create(
                    persona=persona, empresa=txt(it, 'empresa'), puesto_cargo=txt(it, 'puesto_cargo'),
                    tiempo=txt(it, 'tiempo'), motivo_salida=txt(it, 'motivo_salida'))

        EmpleadoReferenciaPersonal.objects.filter(persona=persona).delete()
        for it in rows('referencias_personales'):
            if not isinstance(it, dict):
                continue
            if any([txt(it, 'persona_contactar'), txt(it, 'relacion'), txt(it, 'telefonos'), txt(it, 'comentario')]):
                EmpleadoReferenciaPersonal.objects.create(
                    persona=persona, persona_contactar=txt(it, 'persona_contactar'), relacion=txt(it, 'relacion'),
                    telefonos=txt(it, 'telefonos'), comentario=txt(it, 'comentario'))

        EmpleadoNivelEstudio.objects.filter(persona=persona).delete()
        for it in rows('niveles_estudio'):
            if not isinstance(it, dict):
                continue
            if any([txt(it, 'nivel_estudio'), txt(it, 'centro_capacitacion')]):
                EmpleadoNivelEstudio.objects.create(
                    persona=persona, nivel_estudio=txt(it, 'nivel_estudio'),
                    completa=bool(it.get('completa')), centro_capacitacion=txt(it, 'centro_capacitacion'))

        EmpleadoFormacion.objects.filter(persona=persona).delete()
        for it in rows('formaciones'):
            if not isinstance(it, dict):
                continue
            if any([txt(it, 'centro_capacitacion'), txt(it, 'curso'), txt(it, 'area')]):
                try:
                    horas = int(it.get('horas') or 0)
                except (ValueError, TypeError):
                    horas = 0
                EmpleadoFormacion.objects.create(
                    persona=persona, centro_capacitacion=txt(it, 'centro_capacitacion'),
                    curso=txt(it, 'curso'), area=txt(it, 'area'), horas=horas)

    return JsonResponse({'message': 'Más referencias guardadas'})


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def eliminar_persona(request, id):
    # si solo el usuario tiene permiso de eliminar persona, se intenta obtener la persona por id, si no existe se retorna un error 404, si existe se elimina y se retorna un mensaje de éxito. Si ocurre cualquier otro error, se registra en el log y se retorna un error 500
    if not request.user.has_perm('CoreFisica.delete_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    # se intenta obtener la persona por id, si no existe se retorna un error 404, si existe se elimina y se retorna un mensaje de éxito. Si ocurre cualquier otro error, se registra en el log y se retorna un error 500
    try:
        persona = Persona.objects.get(id=id)
        persona.delete()
        return JsonResponse({'message': 'Persona eliminada correctamente'}, status=200)
    except Persona.DoesNotExist:
        return JsonResponse({'error': 'Persona no encontrada'}, status=404)
    except Exception:
        logger.exception('Error eliminando persona id=%s', id)
        return JsonResponse({'error': 'No se pudo eliminar persona'}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def disable_persona(request, id):
    # si el usuario no tiene permiso de cambiar persona, se retorna un error 403. Si tiene permiso, se intenta obtener la persona por id, si no existe se retorna un error 404. Si la persona ya está inactiva, se registra un intento de deshabilitar una persona ya inactiva en el log y se retorna un estado indicando que ya estaba deshabilitada. Si la persona está activa, se llama al método disable() de la persona, pasando el usuario que realiza la acción para registrar quién hizo el cambio. Si la deshabilitación es exitosa, se registra en el log y se retorna un estado indicando que fue deshabilitada. Si ocurre cualquier error durante el proceso, se registra en el log y se retorna un error 500
    if not request.user.has_perm('CoreFisica.change_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        persona = Persona.objects.get(id=id)
    except Persona.DoesNotExist:
        return JsonResponse({'error': 'Persona no encontrada'}, status=404)

    # si la persona no está activa, se registra un intento de deshabilitar una persona ya inactiva en el log y se retorna un estado indicando que ya estaba deshabilitada. Esto evita realizar operaciones innecesarias y proporciona información útil para auditoría
    if not persona.is_active:
        logger.info('Intento de deshabilitar persona ya inactiva id=%s', id)
        
        return JsonResponse({'status': 'already_disabled'}, status=200)

    try:
        persona.disable(by_user=request.user if request.user.is_authenticated else None)
        logger.info('Persona deshabilitada id=%s by=%s', id, getattr(request.user, 'username', None))
        return JsonResponse({'status': 'disabled'}, status=200)
    except Exception:
        logger.exception('Error deshabilitando persona id=%s', id)
        return JsonResponse({'error': 'No se pudo deshabilitar persona'}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def enable_persona(request, id):
    # si el usuario no tiene permiso de cambiar persona, se retorna un error 403. Si tiene permiso, se intenta obtener la persona por id, si no existe se retorna un error 404. Si la persona ya está activa, se registra un intento de habilitar una persona ya activa en el log y se retorna un estado indicando que ya estaba habilitada. Si la persona está inactiva, se llama al método enable() de la persona, pasando el usuario que realiza la acción para registrar quién hizo el cambio. Si la habilitación es exitosa, se registra en el log y se retorna un estado indicando que fue habilitada. Si ocurre cualquier error durante el proceso, se registra en el log y se retorna un error 500
    if not  request.user.has_perm('CoreFisica.change_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    
    try:
        persona = Persona.objects.get(id=id)
    except Persona.DoesNotExist:
        return JsonResponse({'error': 'Persona no encontrada'}, status=404)

    # si la persona ya esta activa, se re|gistra un intento de habilitar una persona ya activa en el log y se retorna un estado indicando que ya estaba habilitada. Esto evita realizar operaciones innecesarias y proporciona información útil para auditoría
    if persona.is_active:
        logger.info('Intento de habilitar persona ya activa id=%s', id)
        return JsonResponse({'status': 'already_enabled'}, status=200)

    try:
        persona.enable(by_user=request.user if request.user.is_authenticated else None)
        logger.info('Persona habilitada id=%s by=%s', id, getattr(request.user, 'username', None))
        return JsonResponse({'status': 'enabled'}, status=200)
    except Exception:
        logger.exception('Error habilitando persona id=%s', id)
        return JsonResponse({'error': 'No se pudo habilitar persona'}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def importar_personas(request):

    #Importa personas desde CSV o XLSX.
    # Requiere columnas: CEDULA, APELLIDOS, NOMBRES. Opcionales: TIPO, IS_ACTIVE.
    if not request.user.has_perm('CoreFisica.change_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    upload = request.FILES.get('file')
    if not upload:
        return JsonResponse({'error': 'Falta el archivo (campo file)'}, status=400)

    # El parámetro dry_run permite validar el archivo sin realizar cambios en la base de datos. Si dry_run es true, se procesará el archivo y se devolverá un resumen de validación sin crear ni actualizar registros. Esto es útil para verificar que el formato y los datos del archivo son correctos antes de hacer la importación real.
    dry_run = str(request.GET.get('dry_run', 'false')).lower() in ['1', 'true', 'yes']
    fullname_headers = ['APELLIDOS Y NOMBRES', 'APELLIDOS Y NOMBRE', 'NOMBRES Y APELLIDOS', 'APELLIDOS/NOMBRES']
    allowed_tipos = {choice[0] for choice in Persona.TIPO_CHOICES}

    # Función para normalizar encabezados, convirtiendo a mayúsculas, quitando espacios y acentos para facilitar la detección de columnas relevantes sin importar variaciones comunes en los nombres de las columnas.
    def normalize_header(value):
        if value is None:
            return ''
        import unicodedata
        text = str(value).strip().upper()
        # quitar acentos/diacríticos
        text = ''.join(c for c in unicodedata.normalize('NFKD', text) if not unicodedata.combining(c))
        return text

    # Función para normalizar cédula, eliminando espacios y guiones, validando que solo contenga dígitos y ajustando a formato de 10 dígitos si es necesario. Retorna la cédula normalizada o None si es inválida.
    def normalize_cedula(value):
        if value is None:
            return ''
        if isinstance(value, (int, float)):
            if isinstance(value, float) and not value.is_integer():
                return None
            value = int(value)
        raw = str(value).strip()
        if re.match(r'^\d+\.0+$', raw):
            raw = raw.split('.', 1)[0]
        compact = re.sub(r'[\s\-]', '', raw)
        if not compact:
            return ''
        if not re.match(r'^\d+$', compact):
            return None
        if len(compact) == 9:
            return f"0{compact}"
        return compact

    # funcion para parsear valores booleanos, interpretando varias formas comunes de representar falso (0, false, no, n) y considerando vacío o None como verdadero por defecto. Esto permite flexibilidad en cómo se indican los valores booleanos en el archivo de importación.
    def parse_bool(value):
        if value is None or value == '':
            return True
        return str(value).strip().lower() not in ['0', 'false', 'no', 'n']

    # funcion para dividir un nombre completo en apellidos y nombres, utilizando heurísticas para manejar casos comunes. Si el nombre completo tiene una sola palabra, se asume que son nombres sin apellidos. Si tiene dos palabras, se asigna la primera a apellidos y la segunda a nombres. Si tiene más de dos palabras, se asume que las dos primeras son apellidos y el resto son nombres. Esto permite procesar columnas combinadas de nombre completo sin requerir un formato específico.
    def split_full_name(raw: str):
        parts = [p for p in str(raw or '').strip().split() if p]
        if not parts:
            return '', ''
        if len(parts) == 1:
            return '', parts[0]
        if len(parts) == 2:
            return parts[0], parts[1]
        # Heurística: dos primeras palabras como apellidos, resto como nombres
        return ' '.join(parts[:2]), ' '.join(parts[2:])

    filas_raw = []
    has_fullname_header = False
    ext = upload.name.lower().rsplit('.', 1)[-1] if '.' in upload.name else ''

    # Procesar CSV o XLSX, detectando la fila de encabezado que contenga las columnas requeridas (CEDULA y APELLIDOS/NOMBRES o columna combinada), y extrayendo los datos en un formato intermedio sin validar aún. Esto permite manejar archivos con formatos variados y detectar correctamente las columnas relevantes para la importación.
    if ext in ['csv', 'txt']:
        try:
            raw_data = upload.read().decode('utf-8-sig', errors='ignore')
            if not raw_data.strip():
                return JsonResponse({'error': 'Archivo vacío'}, status=400)
            try:
                dialect = csv.Sniffer().sniff(raw_data[:1024], delimiters=',;\t|')
            except csv.Error:
                dialect = csv.excel
            reader_rows = list(csv.reader(io.StringIO(raw_data), dialect=dialect))
            header_idx = None
            header_norm = []
            for i, row in enumerate(reader_rows[:15]):  # busca cabecera en primeras filas
                norm = [normalize_header(c) for c in row]
                has_cedula = 'CEDULA' in norm
                has_fullname_header = any(h in fullname_headers for h in norm)
                has_apellidos = 'APELLIDOS' in norm
                has_nombres = 'NOMBRES' in norm
                if has_cedula and ((has_apellidos and has_nombres) or has_fullname_header):
                    header_idx = i
                    header_norm = norm
                    break
            if header_idx is None:
                return JsonResponse({'error': 'Faltan columnas: CEDULA y APELLIDOS/NOMBRES'}, status=400)
            header_map_idx = {h: idx for idx, h in enumerate(header_norm)}
            for idx, row in enumerate(reader_rows[header_idx+1:], start=header_idx+2):
                raw_row = {}
                for key in header_map_idx:
                    pos = header_map_idx[key]
                    raw_row[key] = row[pos] if pos < len(row) else None
                if all((val is None or str(val).strip() == '') for val in raw_row.values()):
                    continue
                filas_raw.append({
                    'fila': idx,
                    'raw': raw_row,
                    'has_fullname_header': has_fullname_header,
                    'hoja': None,
                })
        except Exception:
            logger.exception('Error procesando CSV de personas')
            return JsonResponse({'error': 'No se pudo leer el CSV'}, status=400)

    elif ext in ['xlsx']:
        try:
            wb = load_workbook(upload, read_only=True, data_only=True)
            any_sheet = False
            for ws in wb.worksheets:
                header_row = None
                header_idx = None
                sheet_has_fullname_header = False
                for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)), start=1):
                    norm = [normalize_header(cell.value) for cell in row]
                    has_cedula = 'CEDULA' in norm
                    sheet_has_fullname_header = any(h in fullname_headers for h in norm)
                    has_apellidos = 'APELLIDOS' in norm
                    has_nombres = 'NOMBRES' in norm
                    if has_cedula and ((has_apellidos and has_nombres) or sheet_has_fullname_header):
                        header_row = norm
                        header_idx = r_idx
                        break
                if header_row is None:
                    continue
                any_sheet = True
                header_map_idx = {h: idx for idx, h in enumerate(header_row)}
                for row_idx, row in enumerate(ws.iter_rows(min_row=header_idx + 1, max_row=ws.max_row), start=header_idx + 1):
                    raw_row = {}
                    for key in header_map_idx:
                        pos = header_map_idx[key]
                        raw_row[key] = row[pos].value if pos < len(row) else None
                    if all((val is None or str(val).strip() == '') for val in raw_row.values()):
                        continue
                    filas_raw.append({
                        'fila': row_idx,
                        'raw': raw_row,
                        'has_fullname_header': sheet_has_fullname_header,
                        'hoja': ws.title,
                    })
            if not any_sheet:
                return JsonResponse({'error': 'No se encontraron hojas con columnas: CEDULA y APELLIDOS/NOMBRES'}, status=400)
        except Exception:
            logger.exception('Error procesando XLSX de personas')
            return JsonResponse({'error': 'No se pudo leer el XLSX'}, status=400)
    else:
        return JsonResponse({'error': 'Formato no soportado. Use CSV o XLSX'}, status=400)

    errores = []
    filas_limpias = []

    for item in filas_raw:
        fila_num = item['fila']
        raw = item['raw']
        fila_hoja = item.get('hoja')
        row_has_fullname = item.get('has_fullname_header', False)
        cedula = normalize_cedula(raw.get('CEDULA'))
        tipo = str(raw.get('TIPO') or '').strip().upper()
        if tipo == 'EVENTUALES':
            tipo = 'EVENTUAL'
        if not tipo:
            tipo = None
        is_active = parse_bool(raw.get('IS_ACTIVE'))
        provincia_token = raw.get('PROVINCIA')
        canton_token = raw.get('CANTON')

        # obtener apellidos/nombres, permitiendo columna combinada
        apellidos = str(raw.get('APELLIDOS') or '').strip()
        nombres = str(raw.get('NOMBRES') or '').strip()
        if not (apellidos and nombres) and row_has_fullname:
            # buscar la primera columna fullname presente
            fullname_val = None
            for key in raw.keys():
                if normalize_header(key) in fullname_headers:
                    fullname_val = raw.get(key)
                    break
            a_split, n_split = split_full_name(fullname_val)
            if not apellidos:
                apellidos = a_split
            if not nombres:
                nombres = n_split

        #Validaciones: cedula no vacia, solo dgitos, maximo 10 caracteres, apellidos y nombres no vacios, tipo valido si se proporciona. Se acumulan errores para cada fila sin detener el proceso, permitiendo reportar múltiples problemas en un solo intento de importación.
        if cedula is None:
            errores.append({'fila': fila_num, 'hoja': fila_hoja, 'error': 'CEDULA invalida: solo digitos'})
            continue
        if not cedula:
            errores.append({'fila': fila_num, 'hoja': fila_hoja, 'error': 'CEDULA vacia'})
            continue
        if not re.match(r'^\d{1,10}$', cedula):
            errores.append({'fila': fila_num, 'hoja': fila_hoja, 'error': 'CEDULA invalida: solo digitos, maximo 10'})
            continue
        if not apellidos:
            errores.append({'fila': fila_num, 'hoja': fila_hoja, 'error': 'APELLIDOS vacios'})
            continue
        if not nombres:
            errores.append({'fila': fila_num, 'hoja': fila_hoja, 'error': 'NOMBRES vacios'})
            continue
        if tipo and tipo not in allowed_tipos:
            errores.append({'fila': fila_num, 'hoja': fila_hoja, 'error': f'TIPO invalido: {tipo}'})
            continue
        
        # Si la fila es válida, se agrega a filas_limpias para su posterior procesamiento. Esto permite separar claramente las filas que tienen problemas de formato o datos inválidos de aquellas que están listas para ser importadas, facilitando la gestión de errores y la importación efectiva.
        filas_limpias.append({
            'fila': fila_num,
            'hoja': fila_hoja,
            'cedula': cedula,
            'apellidos': apellidos,
            'nombres': nombres,
            'tipo': tipo or None,
            'is_active': bool(is_active),
            'provincia': provincia_token,
            'canton': canton_token,
        })
    # Se construye un resumen del proceso de validación, incluyendo el total de filas procesadas, cuántas son válidas, cuántas se crearían o actualizarían (inicialmente 0 en este caso) y los errores encontrados. Este resumen se devuelve al cliente para proporcionar retroalimentación sobre el resultado de la validación, especialmente útil cuando se utiliza el modo dry_run para verificar el archivo sin realizar cambios en la base de datos.
    resumen = {
        'total_filas': len(filas_raw),
        'filas_validas': len(filas_limpias),
        'creadas': 0,
        'actualizadas': 0,
        'omitidas': 0,
        'errores': errores,
    }

    # Si dry_run es true, se devuelve el resumen de validación sin realizar la importación real. Esto permite al usuario verificar que el archivo tiene el formato correcto y que los datos son válidos antes de proceder con la importación, evitando cambios no deseados en la base de datos.
    if dry_run:
        resumen['mensaje'] = 'Validación realizada (dry_run)'
        return JsonResponse(resumen, status=200)
    
    # si no hay filas válidas para importar, se devuelve un error con el resumen de errores encontrados. Esto evita intentar realizar una importación cuando no hay datos correctos, proporcionando retroalimentación clara sobre los problemas que deben corregirse en el archivo antes de intentar importar nuevamente.
    if not filas_limpias:
        return JsonResponse({'error': 'No hay filas válidas para importar', 'detalles': errores}, status=400)
    # Importar filas válidas, evitando duplicados tanto en el archivo como con registros existentes en la base de datos
    try:
        with transaction.atomic():
            cedulas = [f['cedula'] for f in filas_limpias]
            cedula_candidates = set(cedulas)
            for c in cedulas:
                if len(c) == 10 and c.startswith('0'):
                    cedula_candidates.add(c.lstrip('0') or c)
                elif len(c) == 9:
                    cedula_candidates.add(f"0{c}")
            existentes = {}
            for p in Persona.objects.filter(cedula__in=list(cedula_candidates)):
                norm = normalize_cedula(p.cedula)
                if norm:
                    existentes[norm] = p
            procesadas = set()

            for fila in filas_limpias:
                cedula = fila['cedula']
                if cedula in procesadas:
                    # Evitar duplicados en el mismo archivo y existentes en BD.
                    continue
                persona = existentes.get(cedula)
                if not persona:
                    provincia_id = _resolve_provincia_id(fila.get('provincia'))
                    canton_id = _resolve_canton_id(fila.get('canton'), provincia_id)
                    Persona.objects.create(
                        nombres=str(fila.get('nombres') or '').strip().upper(),
                        apellidos=str(fila.get('apellidos') or '').strip().upper(),
                        cedula=cedula,
                        tipo=fila.get('tipo') or None,
                        is_active=bool(fila.get('is_active')),
                        provincia_id=provincia_id,
                        canton_id=canton_id,
                    )
                    resumen['creadas'] += 1
                    procesadas.add(cedula)
                    continue
                provincia_id = _resolve_provincia_id(fila.get('provincia'))
                canton_id = _resolve_canton_id(fila.get('canton'), provincia_id)
                update_fields = []
                if provincia_id and persona.provincia_id != provincia_id:
                    persona.provincia_id = provincia_id
                    update_fields.append('provincia')
                if canton_id and persona.canton_id != canton_id:
                    persona.canton_id = canton_id
                    update_fields.append('canton')
                if update_fields:
                    persona.save(update_fields=update_fields)
                    resumen['actualizadas'] += 1
                procesadas.add(cedula)
        resumen['mensaje'] = 'Importación completada'
        return JsonResponse(resumen, status=200)
    except IntegrityError as exc:
        logger.exception('Error de integridad importando personas')
        return JsonResponse({'error': 'Conflicto de integridad', 'detalle': str(exc)}, status=400)
    except Exception:
        logger.exception('Error importando personas')
        return JsonResponse({'error': 'No se pudo importar personas'}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_personas_excel(request):
    # si el usuario no tiene permiso de ver personas, se retorna un error 403. Si tiene permiso, se obtienen los parámetros de búsqueda q y tipo para filtrar las personas. Solo se exportan personas activas para evitar confusión, pero se pueden ajustar filtros según necesidad. Se crea un archivo Excel con los datos de las personas filtradas, aplicando formato a las celdas para mejorar la legibilidad. Finalmente, se devuelve el archivo Excel como una respuesta de descarga al cliente.
    if not request.user.has_perm('CoreFisica.view_persona'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    
    # Obtener parámetros de búsqueda y strip para eliminar espacios al inicio y final. Esto permite filtrar las personas por cédula, nombres o apellidos (q) y por tipo de persona (tipo) antes de generar el archivo Excel, proporcionando una exportación más relevante según los criterios especificados por el usuario.
    q = (request.GET.get('q') or '').strip()
    tipo = (request.GET.get('tipo') or '').strip()

    # Solo exportamos personas activas para evitar confusión, pero se pueden ajustar filtros según necesidad
    personas = Persona.objects.filter(is_active=True)

    # Si se proporciona un parámetro de búsqueda q, se filtran las personas buscando coincidencias en los campos nombres, apellidos o cedula utilizando una consulta Q para combinar las condiciones. Si se especifica un tipo, se filtran las personas por ese tipo. Finalmente, se ordenan las personas por apellidos y nombres para una presentación más organizada en el archivo Excel.
    if q:
        qn = _strip_accents(q)
        personas = personas.filter(
            Q(nombres__unaccent__icontains=qn) |
            Q(apellidos__unaccent__icontains=qn) |
            Q(cedula__icontains=q)
        )
    
    # si tipo no esta vacio se filtra por tipo
    if tipo:
        personas = personas.filter(tipo=tipo)

    # variable personas se ordena por apellidos y nombres en el excel
    personas = personas.select_related('provincia', 'canton').order_by('apellidos', 'nombres')
    
    # wb es un objeto Workbook de openpyxl que representa el archivo Excel que se va a generar
    wb = Workbook()
    # wb.active obtiene la hoja activa del libro de Excel, que es donde se escribirán los datos de las personas. Se asigna a la variable ws para facilitar su manipulación. Luego se establece el título de la hoja como "Personas" para identificar claramente el contenido del archivo.
    ws = wb.active
    # ws.title establece el título de la hoja activa a "Personas", lo que ayuda a identificar el contenido del archivo Excel. Luego, se agregan los encabezados de las columnas (CEDULA, APELLIDOS, NOMBRES, TIPO) como la primera fila de la hoja utilizando ws.append(). Esto proporciona una estructura clara para los datos que se agregarán a continuación.
    ws.title = "Personas"
    #ws.append agrega los encabezados de las columnas al archivo Excel, definiendo claramente qué información se encuentra en cada columna. Esto es importante para la legibilidad del archivo y para que los usuarios puedan entender fácilmente los datos que se presentan. Los encabezados son: CEDULA, APELLIDOS, NOMBRES y TIPO, que corresponden a los campos principales de la entidad Persona.
    ws.append(['CEDULA', 'APELLIDOS', 'NOMBRES', 'TIPO', 'PROVINCIA', 'CANTON'])

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="F1F3F5", end_color="F1F3F5", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for p in personas:
        provincia = getattr(getattr(p, 'provincia', None), 'nombre', '') or ''
        canton = getattr(getattr(p, 'canton', None), 'nombre', '') or ''
        ws.append([p.cedula, p.apellidos, p.nombres, p.tipo, provincia, canton])

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    resp = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    resp['Content-Disposition'] = 'attachment; filename=personal.xlsx'
    return resp

class SacafrancoListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.has_perm('CoreFisica.view_persona'):
            return JsonResponse({'error': 'No autorizado'}, status=403)

        week_start = request.query_params.get('week_start')
        day = request.query_params.get('day')
        puesto_id = request.query_params.get('puesto_id')
        week_start_date = None

        if day:
            day_norm = _normalize_day(day)
            if not day_norm:
                return Response({'error': 'Día inválido'}, status=400)
            day = day_norm

        if week_start:
            try:
                week_start_date = datetime.date.fromisoformat(str(week_start))
            except Exception:
                return Response({'error': 'week_start inválida, formato YYYY-MM-DD'}, status=400)

        qs = Persona.objects.filter(tipo='SACAFRANCO', is_active=True).order_by('nombres', 'apellidos')
        results = []
        day_offsets = {
            'mon': 0,
            'tue': 1,
            'wed': 2,
            'thu': 3,
            'fri': 4,
            'sat': 5,
            'sun': 6,
        }
        today = datetime.date.today()
        slot_occupied_by = None
        if puesto_id and week_start_date and day:
            slot = CoberturaSacafranco.objects.filter(
                puesto_id=puesto_id,
                week_start=week_start_date,
                day=day,
            ).select_related('persona').first()
            if slot and slot.persona:
                slot_occupied_by = slot.persona

        for p in qs:
            occupied = False
            if week_start_date and day:
                selected_day_date = None
                if day in day_offsets:
                    selected_day_date = week_start_date + datetime.timedelta(days=day_offsets[day])

                allowed_days = []
                week_end_date = week_start_date + datetime.timedelta(days=6)
                if week_start_date >= today:
                    allowed_days = list(DAY_KEYS)
                elif week_start_date <= today <= week_end_date:
                    for key, offset in day_offsets.items():
                        if week_start_date + datetime.timedelta(days=offset) >= today:
                            allowed_days.append(key)

                if allowed_days and CoberturaSacafranco.objects.filter(
                    persona=p,
                    week_start=week_start_date,
                    day__in=allowed_days,
                ).exists():
                    occupied = True
                else:
                    min_week_start_for_day = week_start_date
                    if selected_day_date and selected_day_date < today:
                        min_week_start_for_day = week_start_date + datetime.timedelta(days=7)

                    if CoberturaSacafranco.objects.filter(
                        persona=p,
                        week_start__gte=min_week_start_for_day,
                        day=day,
                    ).exists():
                        occupied = True
            assigned_for_puesto = None
            if puesto_id and week_start_date and day:
                assigned = CoberturaSacafranco.objects.filter(
                    puesto_id=puesto_id,
                    week_start=week_start_date,
                    day=day,
                    persona=p,
                ).first()
                if assigned:
                    assigned_for_puesto = p.id

            results.append({
                'id': p.id,
                'nombres': p.nombres,
                'apellidos': p.apellidos,
                'cedula': p.cedula,
                'status': 'asignado' if (assigned_for_puesto or occupied) else 'available',
                'assigned_for_puesto': assigned_for_puesto,
                'slot_occupied_by_id': getattr(slot_occupied_by, 'id', None),
                'slot_occupied_by_name': f"{slot_occupied_by.nombres} {slot_occupied_by.apellidos}".strip() if slot_occupied_by else None,
                'slot_occupied_by_cedula': getattr(slot_occupied_by, 'cedula', None) if slot_occupied_by else None,
            })

        return Response(results)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def asignar_sacafranco(request):
    # si el ususario no tiene permiso de cambiar asignacionsemanal, se retorna un error 403. Si tiene permiso, se obtienen los parámetros necesarios del request.data para realizar la asignación de sacafranco a una persona en un puesto específico para una semana y día determinados. Se valida que todos los parámetros requeridos estén presentes y que el día sea válido. Luego se intenta obtener la persona y el puesto por sus IDs, retornando errores 404 si no se encuentran. Se determina el mes y año de referencia a partir de week_start para mantener la unicidad de asignación por persona/mes/año. Si no existe una asignación para esa persona/mes/año, se crea una nueva asignación con el contexto del puesto actual. Si ya existe una asignación, se sincroniza su contexto con el puesto actual si es necesario. Finalmente, se propaga la asignación a semanas futuras a partir de la semana indicada, sin tocar semanas pasadas ni sobrescribir códigos existentes.
    if not request.user.has_perm('CoreFisica.change_asignacionsemanal'):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    data = request.data
    persona_id = data.get('persona_id')
    puesto_id = data.get('puesto_id')
    week_start = data.get('week_start')
    day = data.get('day')
    replace = bool(data.get('replace'))
   

    if not all([persona_id, puesto_id, week_start, day]):
        return JsonResponse({'error': 'Faltan parámetros requeridos'}, status=400)

    # Normalizar día a las claves del modelo (mon..sun)
    day_norm = _normalize_day(day)
    if not day_norm:
        return JsonResponse({'error': 'Día inválido'}, status=400)
    day = day_norm

    try:
        persona = Persona.objects.get(id=persona_id)
    except Persona.DoesNotExist:
        return JsonResponse({'error': 'Persona no encontrada'}, status=404)

    try:
        puesto = Puesto.objects.get(id=puesto_id)
    except Puesto.DoesNotExist:
        return JsonResponse({'error': 'Puesto no encontrado'}, status=404)

    # Determinar mes/año desde week_start (respeta unicidad persona/mes/anio)
    try:
        if isinstance(week_start, str):
            week_start_date = datetime.date.fromisoformat(week_start)
        else:
            week_start_date = week_start if isinstance(week_start, datetime.date) else datetime.date.today()
    except Exception:
        week_start_date = datetime.date.today()

    mes_ref = week_start_date.month
    anio_ref = week_start_date.year

    if persona.tipo == 'SACAFRANCO':
        try:
            return _assign_sacafranco_without_asignacion(persona, puesto, week_start_date, day, replace=replace)
        except Exception:
            logger.exception('Error asignando sacafranco sin crear asignación')
            return JsonResponse({'error': 'No se pudo asignar sacafranco'}, status=500)

    asignacion = Asignacion.objects.filter(persona=persona, mes=mes_ref, anio=anio_ref).first()
    try:
        if not asignacion:
            cliente = getattr(puesto.instalacion, 'cliente', None)
            instalacion = getattr(puesto, 'instalacion', None)
            horario = Horario.objects.first()
            if horario is None:
                horario = Horario.objects.create(hora_ingreso='08:00', hora_salida='20:00')
            try:
                asignacion = Asignacion.objects.create(
                    persona=persona,
                    cliente=cliente,
                    instalacion=instalacion,
                    puesto=puesto,
                    horario=horario,
                    mes=mes_ref,
                    anio=anio_ref,
                    recurring=False
                )
            except IntegrityError:
                asignacion = Asignacion.objects.filter(persona=persona, mes=mes_ref, anio=anio_ref).first()
                if not asignacion:
                    return JsonResponse({'error': 'Conflicto de asignación existente'}, status=400)
        else:
            # Reusar la asignación del mes/año y sincronizar contexto del puesto actual.
            changed = False
            if asignacion.puesto_id != puesto.id:
                asignacion.puesto = puesto
                changed = True
            try:
                cliente = getattr(puesto.instalacion, 'cliente', None)
                instalacion = getattr(puesto, 'instalacion', None)
                if asignacion.instalacion_id != getattr(instalacion, 'id', None):
                    asignacion.instalacion = instalacion
                    changed = True
                if asignacion.cliente_id != getattr(cliente, 'id', None):
                    asignacion.cliente = cliente
                    changed = True
            except Exception:
                pass
            if not asignacion.horario_id:
                horario = Horario.objects.first()
                if horario is None:
                    horario = Horario.objects.create(hora_ingreso='08:00', hora_salida='20:00')
                asignacion.horario = horario
                changed = True
            if changed:
                asignacion.save()

        # Propagar la asignación a semanas futuras a partir de la semana indicada,
        # pero sin tocar semanas pasadas ni sobrescribir códigos existentes.
        try:
            # parsear week_start a date
            try:
                if isinstance(week_start, str):
                    week_start_date = datetime.date.fromisoformat(week_start)
                else:
                    week_start_date = week_start
            except Exception:
                week_start_date = week_start

            # empezamos a propagar desde hoy (nunca hacia atras)
            today = datetime.date.today()
            prop_start = week_start_date if isinstance(week_start_date, datetime.date) else today
            if prop_start < today:
                prop_start = today
            # asegurar filas semanales alineadas con el front (semanas por mes: día 1 y saltos de 7)
            # y propagar varios años hacia adelante para que quede "de largo".
            weeks = []
            year_cursor = prop_start.year
            month_cursor = prop_start.month
            end_year = prop_start.year + 3
            while (year_cursor < end_year) or (year_cursor == end_year and month_cursor <= 12):
                base = datetime.date(year_cursor, month_cursor, 1)
                cursor = base
                while cursor.month == month_cursor:
                    if cursor >= prop_start:
                        weeks.append(cursor)
                    cursor += datetime.timedelta(days=7)
                month_cursor += 1
                if month_cursor > 12:
                    month_cursor = 1
                    year_cursor += 1

            for ws in weeks:
                semanal_obj, _ = AsignacionSemanal.objects.get_or_create(
                    asignacion_id=asignacion.id,
                    week_start=ws,
                    defaults={'puesto': puesto}
                )
                try:
                    cur_val = getattr(semanal_obj, day, '') or ''
                except Exception:
                    cur_val = ''
                cur_str = str(cur_val).strip()
                
                if cur_str == '':
                    setattr(semanal_obj, day, 'F')
                    semanal_obj.asignacion_id = asignacion.id
                    semanal_obj.save()
                elif cur_str.upper().startswith('F'):
                    if semanal_obj.asignacion_id != asignacion.id:
                        semanal_obj.asignacion_id = asignacion.id
                        semanal_obj.save()
                
        except Exception:
            logger.exception('Error propagando sacafranco a semanas futuras')

        semanal, created = AsignacionSemanal.objects.get_or_create(
            asignacion_id=asignacion.id,
            week_start=week_start_date if isinstance(week_start_date, datetime.date) else week_start,
            defaults={'asignacion': asignacion, 'puesto': puesto}
        )
        
        current_val = None
        try:
            current_val = getattr(semanal, day, None)
        except Exception:
            current_val = None

        # Vinculamos la asignación para la semana seleccionada si la celda está vacía
        # o contiene 'F' (no sobrescribimos el marcador 'F').
        if hasattr(semanal, day):
            cur = (current_val or '')
            cur_str = str(cur).strip() if cur is not None else ''
            # permitimos enlace si la celda está vacía o ya es 'F'
            if cur_str == '' or cur_str.upper().startswith('F'):
                if semanal.asignacion_id != asignacion.id:
                    semanal.asignacion = asignacion
                # si está vacía escribimos 'F', si ya es 'F' no la tocamos
                if cur_str == '':
                    setattr(semanal, day, 'F')
                semanal.save()
                return JsonResponse({'status': 'assigned', 'semanal_id': semanal.id})
            else:
                return JsonResponse({'status': 'preserved', 'semanal_id': semanal.id, 'value': current_val})
        else:
            return JsonResponse({'error': 'Día inválido'}, status=400)
    except Exception:
        logger.exception('Error asignando sacafranco')
        return JsonResponse({'error': 'No se pudo asignar sacafranco'}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def desasignar_sacafranco(request):
    if not request.user.has_perm('CoreFisica.change_asignacionsemanal'):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    data = request.data
    persona_id = data.get('persona_id')
    puesto_id = data.get('puesto_id')
    week_start = data.get('week_start')
    day = data.get('day')

    if not all([persona_id, puesto_id, week_start, day]):
        return JsonResponse({'error': 'Faltan parámetros requeridos'}, status=400)

    try:
        # Obtener la persona por ID, retornando un error 404 si no se encuentra. Esto es esencial para asegurarse de que la persona a la que se le va a desasignar el sacafranco existe en la base de datos antes de intentar realizar cualquier operación relacionada con ella.
        persona = Persona.objects.get(id=persona_id)
    except Persona.DoesNotExist:
        return JsonResponse({'error': 'Persona no encontrada'}, status=404)

    try:
        # Obtener el puesto por ID, retornando un error 404 si no se encuentra. Esto es esencial para asegurarse de que el puesto al que se le va a desasignar el sacafranco existe en la base de datos antes de intentar realizar cualquier operación relacionada con él.
        puesto = Puesto.objects.get(id=puesto_id)
    except Puesto.DoesNotExist:
        return JsonResponse({'error': 'Puesto no encontrado'}, status=404)

    # Normalizar día a las claves del modelo (mon..sun)
    day_norm = _normalize_day(day)
    if not day_norm:
        return JsonResponse({'error': 'Día inválido'}, status=400)
    day = day_norm

    try:
        try:
            if isinstance(week_start, str):
                week_start_date = datetime.date.fromisoformat(week_start)
            else:
                week_start_date = week_start if isinstance(week_start, datetime.date) else datetime.date.today()
        except Exception:
            week_start_date = datetime.date.today()

        if persona.tipo == 'SACAFRANCO':
            unassigned = _desasignar_sacafranco_without_asignacion(persona, puesto, week_start_date, day)
            if unassigned is not None:
                return unassigned

        mes_ref = week_start_date.month
        anio_ref = week_start_date.year
        asignacion = Asignacion.objects.filter(persona=persona, mes=mes_ref, anio=anio_ref).first()

        today = datetime.date.today()
        prop_start = week_start_date if isinstance(week_start_date, datetime.date) else today
        if prop_start < today:
            prop_start = today
        try:
            year_for_end = week_start_date.year if isinstance(week_start_date, datetime.date) else datetime.date.today().year
            prop_end = datetime.date(year_for_end, 12, 31)
        except Exception:
            today = datetime.date.today()
            prop_end = datetime.date(today.year, 12, 31)

        semanal = None
        if asignacion:
            semanal = AsignacionSemanal.objects.filter(
                puesto=puesto,
                week_start=week_start_date if isinstance(week_start_date, datetime.date) else week_start,
                asignacion_id=asignacion.id
            ).first()

        # Fallback: en datos históricos puede existir la marca F sin vínculo consistente
        # con asignación/persona; en ese caso buscamos la fila por puesto/semana/día.
        if not semanal:
            semanal = AsignacionSemanal.objects.filter(
                puesto=puesto,
                week_start=week_start_date if isinstance(week_start_date, datetime.date) else week_start,
                **{f"{day}__istartswith": 'F'}
            ).filter(
                Q(asignacion__persona=persona) | Q(asignacion__isnull=True)
            ).first() or AsignacionSemanal.objects.filter(
                puesto=puesto,
                week_start=week_start_date if isinstance(week_start_date, datetime.date) else week_start,
                **{f"{day}__istartswith": 'F'}
            ).first()

        if not semanal:
            return JsonResponse({'error': 'No hay programación semanal para ese puesto/semana'}, status=404)

        if not asignacion and semanal.asignacion_id:
            asignacion = semanal.asignacion

        if not hasattr(semanal, day):
            return JsonResponse({'error': 'Día inválido'}, status=400)

        day_offsets = {
            'mon': 0,
            'tue': 1,
            'wed': 2,
            'thu': 3,
            'fri': 4,
            'sat': 5,
            'sun': 6,
        }
        day_offset = day_offsets.get(day, 0)
        # Corta la propagación desde hoy hacia adelante (nunca hacia atras)
        selected_cell_date = week_start_date + datetime.timedelta(days=day_offset)
        cutoff_date = today if today > selected_cell_date else selected_cell_date

        # Propagar la desasignacion solo a fechas vigentes/futuras,
        # sin borrar la marca 'F'.
        try:
            if asignacion:
                candidate_qs = AsignacionSemanal.objects.filter(
                    puesto=puesto,
                    week_start__gte=prop_start,
                    week_start__lte=prop_end,
                    asignacion_id=asignacion.id
                )
            else:
                candidate_qs = AsignacionSemanal.objects.filter(
                    puesto=puesto,
                    week_start__gte=prop_start,
                    week_start__lte=prop_end,
                    asignacion_id=semanal.asignacion_id
                ) if semanal.asignacion_id else AsignacionSemanal.objects.filter(
                    puesto=puesto,
                    week_start__gte=prop_start,
                    week_start__lte=prop_end
                )

            for fila in candidate_qs:
                row_day_date = fila.week_start + datetime.timedelta(days=day_offset)
                if row_day_date < cutoff_date:
                    continue

                cell_value = str(getattr(fila, day, '') or '').strip().upper()
                if not cell_value.startswith('F'):
                    continue

                if fila.asignacion_id is not None:
                    fila.asignacion = None
                    fila.save()
        except Exception:
            logger.exception('Error propagando desasignación a semanas futuras')

        return JsonResponse({'status': 'unassigned', 'semanal_id': semanal.id})
    except Exception:
        logger.exception('Error desasignando sacafranco')
        return JsonResponse({'error': 'No se pudo desasignar sacafranco'}, status=500)


