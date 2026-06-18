"""Vistas de Asignaciones: listado paginado por cantón/vista, creación/edición/reasignación, filas sacafranco, orden y export Excel."""
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from django.http import JsonResponse
from rest_framework import status
from ..models import Asignacion, AsignacionSemanal, Persona, Puesto, ReporteAsistencia, SacafrancoFila, SacafrancoFilaSemanal, Provincia, Canton
from django.db.models import Q, Max, Value
from django.db.models.functions import Coalesce
from django.db import transaction
from django.utils import timezone
from ..serializers import AsignacionSerializer, AsignacionLiteSerializer, SacafrancoFilaSerializer
from ..utils import _strip_accents
import openpyxl
import datetime
from io import BytesIO
from pathlib import Path
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU


def _find_asignaciones_logo_path():
    this_file = Path(__file__).resolve()
    root_candidates = [
        this_file.parents[i] for i in [4, 3] if i < len(this_file.parents)
    ] + [Path.cwd()]
    relative_candidates = [
        Path('frontendf/public/logodescargable.jpg'),
        Path('frontendf/public/favicon.png'),
        Path('frontendf/src/assets/images/logo.png'),
    ]

    for root in root_candidates:
        for rel in relative_candidates:
            candidate = root / rel
            if candidate.exists():
                return candidate
    return None


def _rebuild_asignacion_semanal(asignacion, force_all: bool = False):
    # mes es igual a un entero entre 1 y 12
    mes = int(asignacion.mes)
    #anio es un entero de 4 digitos
    anio = int(asignacion.anio)
    # cacular primer y último día del mes
    first_day = datetime.date(anio, mes, 1)
    # si mes es igual a 12, el siguiente mes es enero del año siguiente
    if mes == 12:
        next_month_first = datetime.date(anio + 1, 1, 1)
    # De no ser diciembre, el siguiente mes es el mismo año y mes + 1
    else:
        next_month_first = datetime.date(anio, mes + 1, 1)
    last_day = next_month_first - datetime.timedelta(days=1)
    current = first_day

    puesto_obj = None
    try:
        puesto_obj = getattr(asignacion, 'puesto')
        if isinstance(puesto_obj, int) or isinstance(puesto_obj, str):
            puesto_obj = Puesto.objects.get(id=int(puesto_obj))
    except Exception:
        try:
            puesto_obj = Puesto.objects.get(id=asignacion.puesto_id)
        except Exception:
            puesto_obj = None
    # dias_puesto es una lista de días de la semana asociados al puesto, obtenidos directamente del campo 'dias' del puesto o inferidos a partir de los horarios asociados al puesto. Se normalizan a nombres de días en español para facilitar comparación con las fechas del calendario.
    dias_puesto = []
    if puesto_obj:
        try:
            if hasattr(puesto_obj, 'dias') and getattr(puesto_obj, 'dias'):
                dias_puesto = puesto_obj.dias or []
            else:
                horarios_qs = getattr(puesto_obj, 'horarios', None)
                if horarios_qs is not None:
                    dias_nums = list(horarios_qs.values_list('dia', flat=True))
                    day_map = {1: 'lunes', 2: 'martes', 3: 'miercoles', 4: 'jueves', 5: 'viernes', 6: 'sabado', 7: 'domingo'}
                    dias_puesto = [day_map.get(n, '') for n in dias_nums if n]
        except Exception:
            dias_puesto = []

    def normalize_day_token(tok: str) -> str:
        t = str(tok).strip().lower()
        if not t:
            return ''
        map_short = {
            'l': 'lunes', 'lu': 'lunes', 'lun': 'lunes', 'lunes': 'lunes',
            'm': 'martes', 'ma': 'martes', 'mar': 'martes', 'martes': 'martes',
            'mi': 'miercoles', 'mie': 'miercoles', 'miercoles': 'miercoles', 'miércoles':'miercoles',
            'j': 'jueves', 'ju': 'jueves', 'jue': 'jueves', 'jueves': 'jueves',
            'v': 'viernes', 'vi': 'viernes', 'vie': 'viernes', 'viernes': 'viernes',
            's': 'sabado', 'sa': 'sabado', 'sab': 'sabado', 'sabado': 'sabado', 'sábado': 'sabado',
            'd': 'domingo', 'do': 'domingo', 'dom': 'domingo', 'domingo': 'domingo'
        }
        return map_short.get(t, t)

    dias_norm = [normalize_day_token(d) for d in dias_puesto if d]
    turno = (getattr(puesto_obj, 'turno', '') or '').strip().lower() if puesto_obj else ''
    default_code = 'N' if turno.startswith('n') else 'D'

    dias_nums = []
    try:
        horarios_qs = getattr(puesto_obj, 'horarios', None)
        if horarios_qs is not None:
            try:
                dias_nums = [int(n) for n in horarios_qs.values_list('dia', flat=True) if n is not None]
            except Exception:
                dias_nums = list(horarios_qs.values_list('dia', flat=True))
    except Exception:
        dias_nums = []
    #weekday_names es una lista de los nombres de los dias de la semana usada para iterar sobre las fechas del mes y determinar el nombre del día correspondiente a cada fecha. Se utiliza para comparar con los días asociados al puesto y decidir si se asigna un código específico o se deja vacío.
    weekday_names = ['lunes','martes','miercoles','jueves','viernes','sabado','domingo']
    # hoy es la fecha actual, se utiliza para evitar modificar dias pasados 
    hoy = timezone.localdate()

    def resolve_day_code(day_date):
        try:
            horarios_qs = getattr(puesto_obj, 'horarios', None)
            if horarios_qs is None:
                return default_code
            dia_num = day_date.isoweekday()
            turnos = list(horarios_qs.filter(dia=dia_num).values_list('turno', flat=True))
            if not turnos:
                turnos = list(horarios_qs.values_list('turno', flat=True))
            tokens = [str(t).strip().lower() for t in turnos if t]
            if any(t.startswith('n') for t in tokens):
                return 'N'
            if any(t.startswith('d') for t in tokens):
                return 'D'
            if any(t.startswith('a') for t in tokens):
                return 'D'
        except Exception:
            return default_code
        return default_code
    #mientras la fecha actual este dentro del mes de la asignacion
    while current <= last_day:
        existing = AsignacionSemanal.objects.filter(
            asignacion_id=asignacion.id,
            week_start=current
        ).first()
        defaults = {}
        # iterar sobre los 7 dias de la semana comenzando por la fecha actual (current) y determinar el código a asignar para cada día según la lógica de secuencia, días del puesto, días de la semana y fechas de la asignación. Si force_all es False, no se modificarán los días pasados (antes de hoy) y se mantendrán los valores existentes en esos días.
        for idx in range(7):
            day_date = current + datetime.timedelta(days=idx)
            name = weekday_names[day_date.weekday()]
            weekday_keys = ['mon','tue','wed','thu','fri','sat','sun']
            key = weekday_keys[day_date.weekday()]

            if day_date < hoy and not force_all:
                if existing:
                    defaults[key] = getattr(existing, key, '') or ''
                continue

            seq = None
            patron = None
            try:
                patron = getattr(asignacion, 'patronAsignacion', None)
                if patron and not hasattr(patron, 'secuencia'):
                    from ..models import PatronAsignacion as _PA
                    try:
                        patron = _PA.objects.get(id=int(patron))
                    except Exception:
                        patron = None
                if patron and getattr(patron, 'secuencia', None):
                    seq = [str(x).strip().upper() for x in patron.secuencia if x]
            except Exception:
                seq = None

            if seq:
                applies_by_puesto = True
            elif dias_nums:
                applies_by_puesto = (day_date.isoweekday() in dias_nums)
            else:
                applies_by_puesto = any(name == d or d in name or name in d for d in dias_norm)

            value = ''
            if seq:
                effective_start = asignacion.start_date
                ref_date = None
                if effective_start:
                    ref_date = effective_start
                elif getattr(asignacion, 'fecha', None):
                    ref_date = asignacion.fecha
                else:
                    try:
                        ref_date = datetime.date(int(asignacion.anio), int(asignacion.mes), 1)
                    except Exception:
                        ref_date = current

                active = True
                if asignacion.recurring:
                    if effective_start and day_date < effective_start:
                        active = False
                    if asignacion.end_date and asignacion.end_date and day_date > asignacion.end_date:
                        active = False
                else:
                    if getattr(asignacion, 'fecha', None):
                        active = (day_date == asignacion.fecha)
                    else:
                        active = (day_date.month == asignacion.mes and day_date.year == asignacion.anio)

                if active and applies_by_puesto:
                    try:
                        days_diff = (day_date - ref_date).days
                        offset = 0
                        try:
                            is_24h = False
                            if getattr(asignacion, 'horario', None):
                                hi = asignacion.horario.hora_ingreso
                                ho = asignacion.horario.hora_salida
                                dt1 = datetime.datetime.combine(datetime.date(1,1,1), hi)
                                dt2 = datetime.datetime.combine(datetime.date(1,1,1), ho)
                                if dt2 <= dt1:
                                    dt2 += datetime.timedelta(days=1)
                                dur = (dt2 - dt1).total_seconds() / 3600.0
                                is_24h = dur >= 23.5
                            if not is_24h and puesto_obj is not None:
                                horarios_qs = getattr(puesto_obj, 'horarios', None)
                                if horarios_qs is not None:
                                    try:
                                        dia_num = day_date.isoweekday()
                                        horas_por_dia = list(horarios_qs.filter(dia=dia_num).values_list('horas', flat=True))
                                        if horas_por_dia:
                                            is_24h = any((int(h) if h is not None else 0) == 24 for h in horas_por_dia)
                                        else:
                                            horas_list = list(horarios_qs.values_list('horas', flat=True))
                                            is_24h = any((int(h) if h is not None else 0) == 24 for h in horas_list)
                                    except Exception:
                                        horas_list = list(horarios_qs.values_list('horas', flat=True))
                                        is_24h = any((int(h) if h is not None else 0) == 24 for h in horas_list)
                            if is_24h and seq:
                                first = seq[0]
                                cnt = 0
                                for s in seq:
                                    if s == first:
                                        cnt += 1
                                    else:
                                        break
                                offset = cnt
                        except Exception:
                            offset = 0

                        idx_seq = (days_diff + offset) % len(seq)
                        value = seq[idx_seq]
                    except Exception:
                        value = ''
            else:
                if applies_by_puesto:
                    value = resolve_day_code(day_date)

            defaults[key] = value

        pid = puesto_obj.id if puesto_obj and hasattr(puesto_obj, 'id') else getattr(asignacion, 'puesto_id', None)
        obj, created = AsignacionSemanal.objects.update_or_create(
            asignacion_id=asignacion.id,
            week_start=current,
            defaults={**defaults, 'asignacion': asignacion, 'puesto_id': pid}
        )
        if not created:
            if getattr(obj, 'asignacion_id', None) != getattr(asignacion, 'id', None):
                obj.asignacion = asignacion
            for k, v in defaults.items():
                setattr(obj, k, v)
            obj.save()

        current += datetime.timedelta(days=7)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def obtener_asignaciones(request, mes=None, anio=None):
    # si el usuario no tiene permiso para ver asignaciones, devolver error 403 antes de procesar parámetros
    if not request.user.has_perm('CoreFisica.view_asignacion'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    # obtener parametros de filtro: mes, año, instalacion_id, cliente_id, q (texto libre para buscar en varios campos)
    instalacion_id = request.GET.get('instalacion_id')
    #cliente_id se puede recibir como query param o como parte de la ruta (en este caso se prioriza el query param para mantener consistencia con otros filtros)
    cliente_id = request.GET.get('cliente_id')
    provincia_id = request.GET.get('provincia_id')
    canton_id = request.GET.get('canton_id')
    canton_ids_raw = (request.GET.get('canton_ids') or '').strip()
    canton_ids: list[int] = []
    if canton_ids_raw:
        try:
            canton_ids = [int(x) for x in canton_ids_raw.split(',') if str(x).strip()]
        except (TypeError, ValueError):
            return Response({'error': 'Canton IDs invalidos'}, status=status.HTTP_400_BAD_REQUEST)
    # Vista por empresa: filtra por uno o varios clientes (todas sus instalaciones).
    cliente_ids_raw = (request.GET.get('cliente_ids') or '').strip()
    cliente_ids: list[int] = []
    if cliente_ids_raw:
        try:
            cliente_ids = [int(x) for x in cliente_ids_raw.split(',') if str(x).strip()]
        except (TypeError, ValueError):
            return Response({'error': 'Cliente IDs invalidos'}, status=status.HTTP_400_BAD_REQUEST)
    lite = str(request.GET.get('lite', 'false')).lower() in ['true', '1', 'yes']
    # Vista por tipo de persona (aditiva): filtra por persona__tipo, lista plana.
    tipos_raw = (request.GET.get('tipos') or '').strip()
    tipos: list[str] = [t.strip().upper() for t in tipos_raw.split(',') if t.strip()] if tipos_raw else []
    #q es un texto libre q se busca
    q = (request.GET.get('q') or '').strip()
    # si se proporcionan mes y año, filtrar asignaciones activas que correspondan al mes/año o que sean recurrentes y tengan rango de fechas que incluya el mes/año. Si no se proporcionan mes/año, devolver todas las asignaciones activas. En ambos casos, excluir personas de tipo SACAFRANCO y ordenar por orden y id para mantener un orden consistente.   
    base_qs = Asignacion.objects.filter(
        estado='ACTIVO'
    ).exclude(persona__tipo='SACAFRANCO')

    if mes and anio:
        month_start = datetime.date(int(anio), int(mes), 1)
        if int(mes) == 12:
            month_end = datetime.date(int(anio) + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            month_end = datetime.date(int(anio), int(mes) + 1, 1) - datetime.timedelta(days=1)

        # Evitar duplicados por persona: si ya existe asignacion del mes solicitado,
        # no incluir para esa misma persona la asignacion recurrente heredada de meses previos.
        personas_con_mes = base_qs.filter(mes=mes, anio=anio).values('persona_id')

        base_qs = base_qs.filter(
            Q(mes=mes, anio=anio) |
            (
                Q(recurring=True) &
                Q(start_date__lte=month_end) &
                (Q(end_date__isnull=True) | Q(end_date__gte=month_start)) &
                ~Q(persona_id__in=personas_con_mes)
            )
        )

    asignaciones = base_qs.select_related(
        'persona',
        'cliente',
        'instalacion',
        'instalacion__canton',
        'instalacion__canton__provincia',
        'puesto',
        'puesto__zona',
        'horario'
    ).prefetch_related(
        'puesto__horarios'
    ).order_by(
        Coalesce('instalacion__canton__provincia_id', Value(999999)),
        'orden',
        'id'
    )
    # si se proporciona cliente_id, filtrar por cliente_id despues de filtrar por mes/año para optimizar la consulta y evitar crear filas semanales innecesarias para clientes no relacionados. Si se proporciona instalacion_id, filtrar por instalacion_id después de filtrar por mes/año para optimizar la consulta y evitar crear filas semanales innecesarias para instalaciones no relacionadas. Si se proporciona un término de búsqueda q, aplicarlo a campos relevantes de asignación, persona y puesto para facilitar búsqueda rápida. Esto se hace después de filtrar por mes/año para optimizar la consulta y evitar aplicar filtros de texto a asignaciones que no corresponden al periodo seleccionado.
    if cliente_id:
        asignaciones = asignaciones.filter(cliente_id=cliente_id)
    if cliente_ids:
        asignaciones = asignaciones.filter(cliente_id__in=cliente_ids)
    if tipos:
        asignaciones = asignaciones.filter(persona__tipo__in=tipos)
    if instalacion_id:
        asignaciones = asignaciones.filter(instalacion_id=instalacion_id)
    if canton_ids:
        asignaciones = asignaciones.filter(instalacion__canton_id__in=canton_ids)
    elif canton_id:
        try:
            canton_val = int(canton_id)
            asignaciones = asignaciones.filter(instalacion__canton_id=canton_val)
        except (TypeError, ValueError):
            return Response({'error': 'Canton invalido'}, status=status.HTTP_400_BAD_REQUEST)
    if provincia_id:
        try:
            provincia_val = int(provincia_id)
            asignaciones = asignaciones.filter(instalacion__canton__provincia_id=provincia_val)
        except (TypeError, ValueError):
            return Response({'error': 'Provincia invalida'}, status=status.HTTP_400_BAD_REQUEST)
    if q:
        # Búsqueda insensible a acentos: unaccent en la columna (__unaccent) y
        # se quitan acentos al texto escrito, así 'JOSE' encuentra 'JOSÉ' y viceversa.
        tokens = [t for t in q.split() if t]
        filtros = Q()
        for token in tokens:
            tok = _strip_accents(token)
            token_filter = (
                Q(cliente__nombre_comercial__unaccent__icontains=tok) |
                Q(cliente__razon_social__unaccent__icontains=tok) |
                Q(persona__cedula__icontains=tok) |
                Q(persona__nombres__unaccent__icontains=tok) |
                Q(persona__apellidos__unaccent__icontains=tok) |
                Q(puesto__nombre__unaccent__icontains=tok) |
                Q(instalacion__codigo__unaccent__icontains=tok) |
                Q(instalacion__nombre__unaccent__icontains=tok)
            )
            if token.isdigit():
                token_filter = token_filter | Q(semanales__id=int(token))
            filtros &= token_filter
        asignaciones = asignaciones.filter(filtros).distinct()

    # Al BUSCAR por texto sin una vista específica: devolver TODAS las coincidencias en
    # lista plana (sin paginar por cantón y sin excluir clientes de vista de empresa).
    # Así el usuario ve todo lo que coincide, venga de un cantón o de una vista de empresa,
    # y nunca cae en una página vacía.
    if q and not cliente_ids and not canton_ids:
        serializer = (AsignacionLiteSerializer if lite else AsignacionSerializer)(asignaciones, many=True)
        return Response({
            'results': serializer.data,
            'canton_page': 1,
            'canton_total': 1,
            'canton_id': None,
            'canton_options': [],
            'provincia_page': 1,
            'provincia_total': 1,
            'provincia_id': None
        })

    # Vista por TIPO DE PERSONA: lista plana con todas las asignaciones de esos tipos.
    # Es ADITIVA (no excluye empresas ni los saca de sus cantones): se devuelve antes
    # de la exclusión de empresas para que aparezcan todos los del tipo.
    if tipos:
        serializer = (AsignacionLiteSerializer if lite else AsignacionSerializer)(asignaciones, many=True)
        return Response({
            'results': serializer.data,
            'canton_page': 1,
            'canton_total': 1,
            'canton_id': None,
            'canton_options': [],
            'provincia_page': 1,
            'provincia_total': 1,
            'provincia_id': None
        })

    # Un cliente que ya tiene VISTA DE EMPRESA se muestra SOLO en esa vista: se
    # excluye de las vistas por cantón / generales para que no salga duplicado,
    # INCLUSO al buscar por texto (debe encontrarse desde su propia vista de empresa,
    # no aparecer en los cantones a los que pertenece geográficamente).
    if not cliente_ids:
        from ..models import VistaCanton
        clientes_empresa = set()
        for v in VistaCanton.objects.filter(tipo='cliente'):
            clientes_empresa.update(int(c) for c in (v.clientes or []) if str(c).strip())
        if clientes_empresa:
            asignaciones = asignaciones.exclude(cliente_id__in=clientes_empresa)

    # Vista por empresa: devolver TODAS las asignaciones de esos clientes (lista plana,
    # sin paginar por cantón). El frontend la trata como una vista (igual que canton_ids).
    if cliente_ids:
        serializer = (AsignacionLiteSerializer if lite else AsignacionSerializer)(asignaciones, many=True)
        return Response({
            'results': serializer.data,
            'canton_page': 1,
            'canton_total': 1,
            'canton_id': None,
            'canton_options': [],
            'provincia_page': 1,
            'provincia_total': 1,
            'provincia_id': None
        })

    canton_page = request.GET.get('canton_page')
    restore_canton_id = request.GET.get('restore_canton_id')
    if canton_page and not canton_ids:
        try:
            cant_page = int(canton_page)
            if cant_page < 1:
                raise ValueError
        except ValueError:
            return Response({'error': 'Pagina de canton invalida'}, status=status.HTTP_400_BAD_REQUEST)

        canton_ids = list(
            asignaciones
            .order_by('instalacion__canton_id')
            .values_list('instalacion__canton_id', flat=True)
            .distinct()
        )
        sac_qs = SacafrancoFila.objects.all()
        if mes and anio:
            try:
                mes_val = int(mes)
                anio_val = int(anio)
                sac_qs = sac_qs.filter(Q(anio__lt=anio_val) | Q(anio=anio_val, mes__lte=mes_val))
            except (TypeError, ValueError):
                pass
        canton_ids += list(
            sac_qs.order_by('persona__canton_id').values_list('persona__canton_id', flat=True).distinct()
        )

        seen = set()
        ordered: list = []
        for cid in canton_ids:
            if cid in seen:
                continue
            seen.add(cid)
            ordered.append(cid)
        ordered.sort(key=lambda v: (v is None, v if v is not None else 999999))

        # Si se pide restaurar un cantón específico, encontrar su página
        if restore_canton_id:
            try:
                restore_id = int(restore_canton_id)
                if restore_id in ordered:
                    cant_page = ordered.index(restore_id) + 1
            except (TypeError, ValueError):
                pass

        total_cantones = len(ordered)
        if total_cantones == 0:
            serializer = AsignacionSerializer(asignaciones.none(), many=True)
            return Response({
                'results': serializer.data,
                'canton_page': 1,
                'canton_total': 0,
                'canton_id': None,
                'canton_options': [],
                'provincia_page': 1,
                'provincia_total': 0,
                'provincia_id': None
            })

        if cant_page > total_cantones:
            cant_page = total_cantones
        canton_id = ordered[cant_page - 1]

        canton_ids = [cid for cid in ordered if cid is not None]
        canton_map = {}
        if canton_ids:
            canton_map = {
                c.id: c.nombre for c in Canton.objects.filter(id__in=canton_ids)
            }
        canton_options = []
        for cid in ordered:
            if cid is None:
                canton_options.append({'id': None, 'nombre': 'SIN CANTON'})
                continue
            nombre = (canton_map.get(cid) or '').strip()
            canton_options.append({'id': cid, 'nombre': nombre or f'CANTON {cid}'})

        if canton_id is None:
            asignaciones = asignaciones.filter(instalacion__canton_id__isnull=True)
        else:
            asignaciones = asignaciones.filter(instalacion__canton_id=canton_id)

        serializer = (AsignacionLiteSerializer if lite else AsignacionSerializer)(asignaciones, many=True)
        return Response({
            'results': serializer.data,
            'canton_page': cant_page,
            'canton_total': total_cantones,
            'canton_id': canton_id,
            'canton_options': canton_options,
            'provincia_page': cant_page,
            'provincia_total': total_cantones,
            'provincia_id': canton_id
        })

    # Vista mixta: devolver resultados filtrados por canton_ids + lista COMPLETA de cantones
    if canton_ids:
        full_canton_ids = list(
            base_qs
            .order_by('instalacion__canton_id')
            .values_list('instalacion__canton_id', flat=True)
            .distinct()
        )
        sac_full_qs = SacafrancoFila.objects.all()
        if mes and anio:
            try:
                sac_full_qs = sac_full_qs.filter(Q(anio__lt=int(anio)) | Q(anio=int(anio), mes__lte=int(mes)))
            except (TypeError, ValueError):
                pass
        full_canton_ids += list(
            sac_full_qs.order_by('persona__canton_id').values_list('persona__canton_id', flat=True).distinct()
        )
        seen_full = set()
        ordered_full = []
        for cid in full_canton_ids:
            if cid in seen_full:
                continue
            seen_full.add(cid)
            ordered_full.append(cid)
        ordered_full.sort(key=lambda v: (v is None, v if v is not None else 999999))

        real_full_ids = [cid for cid in ordered_full if cid is not None]
        canton_map_full = {}
        if real_full_ids:
            canton_map_full = {c.id: c.nombre for c in Canton.objects.filter(id__in=real_full_ids)}
        canton_options_full = []
        for cid in ordered_full:
            if cid is None:
                canton_options_full.append({'id': None, 'nombre': 'SIN CANTON'})
                continue
            nombre = (canton_map_full.get(cid) or '').strip()
            canton_options_full.append({'id': cid, 'nombre': nombre or f'CANTON {cid}'})

        serializer = (AsignacionLiteSerializer if lite else AsignacionSerializer)(asignaciones, many=True)
        return Response({
            'results': serializer.data,
            'canton_page': 1,
            'canton_total': len(ordered_full),
            'canton_id': None,
            'canton_options': canton_options_full,
            'provincia_page': 1,
            'provincia_total': len(ordered_full),
            'provincia_id': None
        })

    page_param = request.GET.get('page')
    size_param = request.GET.get('size')
    if page_param is not None or size_param is not None:
        try:
            page = int(page_param or 1)
            size = int(size_param or 100)
            if page < 1 or size < 1:
                raise ValueError
        except ValueError:
            return Response({'error': 'Parametros de paginacion invalidos'}, status=status.HTTP_400_BAD_REQUEST)

        total = asignaciones.count()
        start = (page - 1) * size
        end = start + size
        asignaciones = asignaciones[start:end]
        serializer = (AsignacionLiteSerializer if lite else AsignacionSerializer)(asignaciones, many=True)
        return Response({
            'total': total,
            'page': page,
            'size': size,
            'results': serializer.data
        })

    serializer = (AsignacionLiteSerializer if lite else AsignacionSerializer)(asignaciones, many=True)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def asignar_servicio(request):
    #si el usuario no tiene permiso para agregar asignaciones, devolver error 403 antes de procesar la solicitud
    if not request.user.has_perm('CoreFisica.add_asignacion'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    # global datetime se declara para asegurar que se utiliza el módulo datetime importado en lugar de cualquier variable local con el mismo nombre, ya que se realizan operaciones con fechas dentro de esta función y es crucial que se utilice el módulo correcto para evitar errores.
    global datetime

    data = request.data.copy()
    data['recurring'] = True
    data['end_date'] = None
    # El horario es opcional (proviene del puesto). Evitar enviar pk inválido (0/'').
    try:
        if str(data.get('horario', '')).strip() in ('', '0', 'None', 'null'):
            data.pop('horario', None)
    except Exception:
        pass

    reasignar = str(data.get('reasignar', '')).strip().lower() in ('1', 'true', 'yes', 'si', 'sí')
    if 'reasignar' in data:
        data.pop('reasignar')

    # Reasignación: mover persona a un puesto, liberando al ocupante anterior
    # y dejando su puesto previo como vacante (conservando su calendario).
    if reasignar:
        try:
            persona_id = int(data.get('persona'))
            puesto_id = int(data.get('puesto') or data.get('puesto_id'))
            mes_val = int(data.get('mes'))
            anio_val = int(data.get('anio'))
        except (TypeError, ValueError):
            return Response({'error': 'Datos insuficientes para reasignar'}, status=status.HTTP_400_BAD_REQUEST)
        horario_id = data.get('horario')

        with transaction.atomic():
            # 1. Vacar la asignación previa de la persona (en otro puesto este mes)
            old_p = Asignacion.objects.filter(
                persona_id=persona_id, mes=mes_val, anio=anio_val, estado='ACTIVO'
            ).exclude(puesto_id=puesto_id).first()
            if old_p:
                old_p.persona = None
                old_p.save(update_fields=['persona'])
                ReporteAsistencia.objects.filter(asignacion=old_p).update(
                    persona=None, estado='TURNO', estado_asistencia='',
                    reemplazo=None, descripcion=None, row_color=None
                )

            # 2. Tomar el puesto destino si ya está ocupado (el ocupante queda libre)
            target = Asignacion.objects.filter(
                puesto_id=puesto_id, mes=mes_val, anio=anio_val, estado='ACTIVO'
            ).first()
            if target:
                target.persona_id = persona_id
                if horario_id:
                    try:
                        target.horario_id = int(horario_id)
                    except (TypeError, ValueError):
                        pass
                target.save()
                ReporteAsistencia.objects.filter(asignacion=target).update(
                    persona_id=persona_id, estado='TURNO', estado_asistencia='',
                    reemplazo=None, descripcion=None, row_color=None
                )
                return Response(AsignacionSerializer(target).data, status=status.HTTP_200_OK)
            # Si el puesto destino está vacío, continúa al flujo normal de creación.

    try:
        puesto_id = data.get('puesto') or data.get('puesto_id')
        mes_val = data.get('mes')
        anio_val = data.get('anio')
        if not reasignar and puesto_id and mes_val and anio_val:
            existentes = Asignacion.objects.filter(
                puesto_id=int(puesto_id),
                mes=int(mes_val),
                anio=int(anio_val),
                estado='ACTIVO'
            )
            # Si hay una asignación VACANTE (sin persona) para este puesto, llenarla en vez de bloquear.
            vacante = existentes.filter(persona__isnull=True).first()
            if vacante:
                vacante.persona_id = int(data.get('persona'))
                horario_id = data.get('horario')
                if horario_id:
                    try:
                        vacante.horario_id = int(horario_id)
                    except (TypeError, ValueError):
                        pass
                vacante.save()
                ReporteAsistencia.objects.filter(asignacion=vacante).update(
                    persona_id=vacante.persona_id, estado='TURNO', estado_asistencia='',
                    reemplazo=None, descripcion=None, row_color=None
                )
                # Reflejar horario en el puesto si no tiene
                try:
                    if vacante.puesto and not vacante.puesto.horario_id and vacante.horario_id:
                        vacante.puesto.horario_id = vacante.horario_id
                        vacante.puesto.save(update_fields=['horario'])
                except Exception:
                    pass
                return Response(AsignacionSerializer(vacante).data, status=status.HTTP_200_OK)
            # Capacidad del puesto: cuántos cupos (asignaciones) admite en el mes.
            try:
                cupo = Puesto.objects.filter(id=int(puesto_id)).values_list('cantidad_puestos', flat=True).first() or 1
            except Exception:
                cupo = 1
            ocupadas = existentes.filter(persona__isnull=False).count()
            # Solo se bloquea cuando ya se llenaron todos los cupos del puesto.
            if ocupadas >= cupo:
                return Response(
                    {'error': 'Este puesto ya alcanzó su cantidad máxima de asignaciones para el mes seleccionado.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
    except Exception:
        pass
    serializer = AsignacionSerializer(data=data)
    if serializer.is_valid():
        asignacion = serializer.save()
        # Si la asignación es recurrente y no tiene start_date, fijar start_date al primer día del mes de la asignación
        try:
            if getattr(asignacion, 'recurring', False) and not getattr(asignacion, 'start_date', None):
                asignacion.start_date = datetime.date(int(asignacion.anio), int(asignacion.mes), 1)
                asignacion.save()
        except Exception:
            pass
        # Si el puesto no tiene horario, asignarle el de esta asignación (se refleja en el puesto)
        try:
            puesto_obj = asignacion.puesto
            if puesto_obj and not puesto_obj.horario_id and asignacion.horario_id:
                puesto_obj.horario_id = asignacion.horario_id
                puesto_obj.save(update_fields=['horario'])
        except Exception:
            pass
        # Crear filas de AsignacionSemanal para el puesto en las semanas del mes/año de la asignación
        # Forzamos creación de calendario siempre (evita depender del flag del front y cubre el mes actual)
        create_calendar = True

        # Asegurar un registro base en ReporteAsistencia para la asignación
        try:
            reporte, _ = ReporteAsistencia.objects.get_or_create(asignacion=asignacion)
            reporte.persona = asignacion.persona
            reporte.cliente = asignacion.cliente
            reporte.instalacion = asignacion.instalacion
            reporte.puesto = asignacion.puesto
            reporte.horario = asignacion.horario
            reporte.puesto_tipo = getattr(asignacion.puesto, 'tipo', None) if asignacion.puesto else None
            reporte.save()
        except Exception:
            pass

        if create_calendar:
            try:
                mes = int(asignacion.mes)
                anio = int(asignacion.anio)
                first_day = datetime.date(anio, mes, 1)
                if mes == 12:
                    next_month_first = datetime.date(anio + 1, 1, 1)
                else:
                    next_month_first = datetime.date(anio, mes + 1, 1)
                last_day = next_month_first - datetime.timedelta(days=1)

                # semanas del mes iniciando el día 1 y sumando bloques de 7 días (1,8,15,22,29)
                current = first_day

                weekday_names = ['lunes','martes','miercoles','jueves','viernes','sabado','domingo']

                # Obtener instancia de Puesto (por si asignacion.puesto es id)
                puesto_obj = None
                try:
                    puesto_obj = getattr(asignacion, 'puesto')
                    if isinstance(puesto_obj, int) or isinstance(puesto_obj, str):
                        puesto_obj = Puesto.objects.get(id=int(puesto_obj))
                except Exception:
                    try:
                        puesto_obj = Puesto.objects.get(id=asignacion.puesto_id)
                    except Exception:
                        puesto_obj = None

                dias_puesto = []
                if puesto_obj:
                    try:
                        if hasattr(puesto_obj, 'dias') and getattr(puesto_obj, 'dias'):
                            dias_puesto = puesto_obj.dias or []
                        else:
                            horarios_qs = getattr(puesto_obj, 'horarios', None)
                            if horarios_qs is not None:
                                dias_nums = list(horarios_qs.values_list('dia', flat=True))
                                day_map = {1: 'lunes', 2: 'martes', 3: 'miercoles', 4: 'jueves', 5: 'viernes', 6: 'sabado', 7: 'domingo'}
                                dias_puesto = [day_map.get(n, '') for n in dias_nums if n]
                    except Exception:
                        dias_puesto = []

                def normalize_day_token(tok: str) -> str:
                    # Normaliza un token de día a un nombre de día completo en español. Se utiliza para comparar los días asociados al puesto con los días del calendario. El mapeo incluye formas abreviadas y completas de los días de la semana, y se ignoran mayúsculas, espacios y acentos para facilitar la comparación.
                    t = str(tok).strip().lower()
                    if not t:
                        return ''
                    map_short = {
                        'l': 'lunes', 'lu': 'lunes', 'lun': 'lunes', 'lunes': 'lunes',
                        'm': 'martes', 'ma': 'martes', 'mar': 'martes', 'martes': 'martes',
                        'mi': 'miercoles', 'mie': 'miercoles', 'miercoles': 'miercoles', 'miércoles':'miercoles',
                        'j': 'jueves', 'ju': 'jueves', 'jue': 'jueves', 'jueves': 'jueves',
                        'v': 'viernes', 'vi': 'viernes', 'vie': 'viernes', 'viernes': 'viernes',
                        's': 'sabado', 'sa': 'sabado', 'sab': 'sabado', 'sabado': 'sabado', 'sábado': 'sabado',
                        'd': 'domingo', 'do': 'domingo', 'dom': 'domingo', 'domingo': 'domingo'
                    }
                    return map_short.get(t, t)

                dias_norm = [normalize_day_token(d) for d in dias_puesto if d]
                # también obtener números de día (1..7) desde los horarios si existen (normalizados a int)
                dias_nums = []
                try:
                    if puesto_obj is not None:
                        horarios_qs = getattr(puesto_obj, 'horarios', None)
                        if horarios_qs is not None:
                            try:
                                dias_nums = [int(n) for n in horarios_qs.values_list('dia', flat=True) if n is not None]
                            except Exception:
                                dias_nums = list(horarios_qs.values_list('dia', flat=True))
                except Exception:
                    dias_nums = []
                # iterar sobre las semanas del mes y crear o actualzar filas de AsignacionSemanal según corresponda, aplicando la lógica de secuencia, días del puesto, días de la semana y fechas de la asignación para determinar el código a asignar en cada día. Se utiliza update_or_create para garantizar que haya una fila 1:1 por asignación/semana, y se actualizan los campos relacionados (persona, cliente, instalacion, horario) para mantener consistencia con la asignación principal.
                while current <= last_day:
                    defaults = {}
                    for idx in range(7):
                        day_date = current + datetime.timedelta(days=idx)
                        name = weekday_names[day_date.weekday()]
                        weekday_keys = ['mon','tue','wed','thu','fri','sat','sun']
                        key = weekday_keys[day_date.weekday()]

                        # Obtener secuencia del patron antes de evaluar applies_by_puesto
                        seq = None
                        patron = None
                        try:
                            patron = getattr(asignacion, 'patronAsignacion', None)
                            # si es id, buscar objeto
                            if patron and not hasattr(patron, 'secuencia'):
                                from ..models import PatronAsignacion as _PA
                                try:
                                    patron = _PA.objects.get(id=int(patron))
                                except Exception:
                                    patron = None
                            if patron and getattr(patron, 'secuencia', None):
                                seq = [str(x).strip().upper() for x in patron.secuencia if x]
                        except Exception:
                            seq = None

                        # Si hay patrón, el ciclo se aplica de forma continua por día.
                        if seq:
                            applies_by_puesto = True
                        # Sin patrón: respetar días/horarios del puesto.
                        elif dias_nums:
                            applies_by_puesto = (day_date.isoweekday() in dias_nums)
                        else:
                            applies_by_puesto = any(name == d or d in name or name in d for d in dias_norm)

                        value = ''
                        if seq:
                            effective_start = asignacion.start_date
                            # definir fecha referencia
                            ref_date = None
                            if effective_start:
                                ref_date = effective_start
                            elif getattr(asignacion, 'fecha', None):
                                ref_date = asignacion.fecha
                            else:
                                try:
                                    ref_date = datetime.date(int(asignacion.anio), int(asignacion.mes), 1)
                                except Exception:
                                    ref_date = current

                            # determinar si la asignación aplica ese día
                            active = True
                            if asignacion.recurring:
                                if effective_start and day_date < effective_start:
                                    active = False
                                if asignacion.end_date and asignacion.end_date and day_date > asignacion.end_date:
                                    active = False
                            else:
                                # non-recurring: match by fecha or month/year
                                if getattr(asignacion, 'fecha', None):
                                    active = (day_date == asignacion.fecha)
                                else:
                                    active = (day_date.month == asignacion.mes and day_date.year == asignacion.anio)

                            if active and applies_by_puesto:
                                try:
                                    days_diff = (day_date - ref_date).days
                                    # detectar si corresponde aplicar offset por 24h
                                    offset = 0
                                    try:
                                        is_24h = False
                                        # preferir la duración del horario de la asignación
                                        if getattr(asignacion, 'horario', None):
                                            hi = asignacion.horario.hora_ingreso
                                            ho = asignacion.horario.hora_salida
                                            dt1 = datetime.datetime.combine(datetime.date(1,1,1), hi)
                                            dt2 = datetime.datetime.combine(datetime.date(1,1,1), ho)
                                            if dt2 <= dt1:
                                                dt2 += datetime.timedelta(days=1)
                                            dur = (dt2 - dt1).total_seconds() / 3600.0
                                            is_24h = dur >= 23.5
                                        # si no se detectó por asignacion, caer en puesto horarios
                                        if not is_24h and puesto_obj is not None:
                                            horarios_qs = getattr(puesto_obj, 'horarios', None)
                                            if horarios_qs is not None:
                                                try:
                                                    dia_num = day_date.isoweekday()
                                                    horas_por_dia = list(horarios_qs.filter(dia=dia_num).values_list('horas', flat=True))
                                                    if horas_por_dia:
                                                        is_24h = any((int(h) if h is not None else 0) == 24 for h in horas_por_dia)
                                                    else:
                                                        horas_list = list(horarios_qs.values_list('horas', flat=True))
                                                        is_24h = any((int(h) if h is not None else 0) == 24 for h in horas_list)
                                                except Exception:
                                                    horas_list = list(horarios_qs.values_list('horas', flat=True))
                                                    is_24h = any((int(h) if h is not None else 0) == 24 for h in horas_list)
                                        if is_24h and seq:
                                            # offset = longitud del primer bloque consecutivo igual al primer símbolo
                                            first = seq[0]
                                            cnt = 0
                                            for s in seq:
                                                if s == first:
                                                    cnt += 1
                                                else:
                                                    break
                                            offset = cnt
                                    except Exception:
                                        offset = 0

                                    idx_seq = (days_diff + offset) % len(seq)
                                    value = seq[idx_seq]
                                except Exception:
                                    value = ''
                        else:
                            # Sin patrón: dejar el día vacío para llenarlo manualmente.
                            value = ''

                        defaults[key] = value

                    try:
                        pid = puesto_obj.id if puesto_obj and hasattr(puesto_obj, 'id') else getattr(asignacion, 'puesto_id', None)
                        # update_or_create garantiza fila 1:1 por asignacion/semana
                        obj, created = AsignacionSemanal.objects.update_or_create(
                            asignacion_id=asignacion.id,
                            week_start=current,
                            defaults={**defaults, 'asignacion': asignacion, 'puesto_id': pid}
                        )
                        if not created:
                            # Asegurar asignacion vinculada y días actualizados
                            if getattr(obj, 'asignacion_id', None) != getattr(asignacion, 'id', None):
                                obj.asignacion = asignacion
                            for k, v in defaults.items():
                                setattr(obj, k, v)
                            obj.save()
                    except Exception as e:
                        raise Exception(f"Error creando/asegurando AsignacionSemanal en asignar_servicio (puesto {getattr(puesto_obj,'id',None)} week_start {current}): {e}")
                    current += datetime.timedelta(days=7)
            except Exception as e:
                try:
                    asignacion.delete()
                except Exception:
                    pass
                return Response({'error': f'No se pudo crear el calendario semanal: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.data, status=status.HTTP_201_CREATED)
    # Si hay error de unicidad (persona, mes, anio) intentamos actualizar la asignación existente
    try:
        errs = serializer.errors or {}
        nonf = errs.get('non_field_errors') if isinstance(errs, dict) else None
        is_unique = False
        if nonf:
            for e in nonf:
                if 'unique' in str(e).lower() or 'únic' in str(e).lower():
                    is_unique = True
                    break
        if is_unique:
            persona = request.data.get('persona')
            mes = request.data.get('mes')
            anio = request.data.get('anio')
            if persona and mes and anio:
                existing = Asignacion.objects.filter(persona_id=persona, mes=mes, anio=anio).first()
                if existing:
                    # actualizar la asignación existente con los nuevos datos
                    serializer2 = AsignacionSerializer(existing, data=request.data, partial=True)
                    if serializer2.is_valid():
                        asignacion = serializer2.save()
                        # crear/actualizar filas semanales siempre
                        create_calendar = True
                        if create_calendar:
                            try:
                                mes = int(asignacion.mes)
                                anio = int(asignacion.anio)
                                first_day = datetime.date(anio, mes, 1)
                                if mes == 12:
                                    next_month_first = datetime.date(anio + 1, 1, 1)
                                else:
                                    next_month_first = datetime.date(anio, mes + 1, 1)
                                last_day = next_month_first - datetime.timedelta(days=1)
                                # semanas del mes iniciando el día 1 y sumando bloques de 7 días (1,8,15,22,29)
                                current = first_day

                                # Obtener puesto_obj
                                puesto_obj = None
                                try:
                                    puesto_obj = getattr(asignacion, 'puesto')
                                    if isinstance(puesto_obj, int) or isinstance(puesto_obj, str):
                                        puesto_obj = Puesto.objects.get(id=int(puesto_obj))
                                except Exception:
                                    try:
                                        puesto_obj = Puesto.objects.get(id=asignacion.puesto_id)
                                    except Exception:
                                        puesto_obj = None

                                dias_puesto = []
                                if puesto_obj:
                                    try:
                                        if hasattr(puesto_obj, 'dias') and getattr(puesto_obj, 'dias'):
                                            dias_puesto = puesto_obj.dias or []
                                        else:
                                            horarios_qs = getattr(puesto_obj, 'horarios', None)
                                            if horarios_qs is not None:
                                                dias_nums = list(horarios_qs.values_list('dia', flat=True))
                                                day_map = {1: 'lunes', 2: 'martes', 3: 'miercoles', 4: 'jueves', 5: 'viernes', 6: 'sabado', 7: 'domingo'}
                                                dias_puesto = [day_map.get(n, '') for n in dias_nums if n]
                                    except Exception:
                                        dias_puesto = []

                                def normalize_day_token(tok: str) -> str:
                                    t = str(tok).strip().lower()
                                    if not t:
                                        return ''
                                    map_short = {
                                        'l': 'lunes', 'lu': 'lunes', 'lun': 'lunes', 'lunes': 'lunes',
                                        'm': 'martes', 'ma': 'martes', 'mar': 'martes', 'martes': 'martes',
                                        'mi': 'miercoles', 'mie': 'miercoles', 'miercoles': 'miercoles', 'miércoles':'miercoles',
                                        'j': 'jueves', 'ju': 'jueves', 'jue': 'jueves', 'jueves': 'jueves',
                                        'v': 'viernes', 'vi': 'viernes', 'vie': 'viernes', 'viernes': 'viernes',
                                        's': 'sabado', 'sa': 'sabado', 'sab': 'sabado', 'sabado': 'sabado', 'sábado': 'sabado',
                                        'd': 'domingo', 'do': 'domingo', 'dom': 'domingo', 'domingo': 'domingo'
                                    }
                                    return map_short.get(t, t)

                                dias_norm = [normalize_day_token(d) for d in dias_puesto if d]

                                # también obtener números de día (1..7) desde los horarios si existen (normalizados a int)
                                dias_nums = []
                                try:
                                    horarios_qs = getattr(puesto_obj, 'horarios', None)
                                    if horarios_qs is not None:
                                        try:
                                            dias_nums = [int(n) for n in horarios_qs.values_list('dia', flat=True) if n is not None]
                                        except Exception:
                                            dias_nums = list(horarios_qs.values_list('dia', flat=True))
                                except Exception:
                                    dias_nums = []

                                weekday_names = ['lunes','martes','miercoles','jueves','viernes','sabado','domingo']
                                hoy = timezone.localdate()
                                while current <= last_day:
                                    existing = AsignacionSemanal.objects.filter(
                                        asignacion_id=asignacion.id,
                                        week_start=current
                                    ).first()
                                    defaults = {}
                                    for idx in range(7):
                                        day_date = current + datetime.timedelta(days=idx)
                                        name = weekday_names[day_date.weekday()]
                                        weekday_keys = ['mon','tue','wed','thu','fri','sat','sun']
                                        key = weekday_keys[day_date.weekday()]
                                        # si la fecha de la semana es anterior a hoy, no actualizar el código para no afectar reportes históricos, pero sí asegurar que se mantiene el mismo código existente (en caso de que se haya modificado la asignación para un mes/año pasado, o se esté creando una asignación con fecha pasada). Esto permite corregir asignaciones pasadas sin perder la consistencia de los reportes históricos, y evita que cambios en asignaciones futuras afecten datos de semanas anteriores.
                                        if day_date < hoy:
                                            if existing:
                                                defaults[key] = getattr(existing, key, '') or ''
                                            continue

                                        # Obtener secuencia del patron antes de evaluar applies_by_puesto
                                        seq = None
                                        patron = None
                                        try:
                                            #Obtener patronAsignacion relacionado a la asignación, ya sea por relación directa o por id, para luego obtener su secuencia y aplicarla si corresponde. Esto permite que la lógica de asignación semanal respete el patrón definido en la asignación, aplicando la secuencia de códigos de forma continua
                                            patron = getattr(asignacion, 'patronAsignacion', None)
                                            if patron and not hasattr(patron, 'secuencia'):
                                                from ..models import PatronAsignacion as _PA
                                                try:
                                                    patron = _PA.objects.get(id=int(patron))
                                                except Exception:
                                                    patron = None
                                            #si se obtuvo un patron con secuencia, normalizar la secuencia a una lista de códigos en mayúscula sin espacios para facilitar su aplicación continua por día
                                            if patron and getattr(patron, 'secuencia', None):
                                                seq = [str(x).strip().upper() for x in patron.secuencia if x]
                                        except Exception:
                                            seq = None

                                        # Si hay patrón, el ciclo se aplica de forma continua por día.
                                        if seq:
                                            applies_by_puesto = True
                                        # Sin patrón: respetar días/horarios del puesto.
                                        elif dias_nums:
                                            applies_by_puesto = (day_date.isoweekday() in dias_nums)
                                        else:
                                            applies_by_puesto = any(name == d or d in name or name in d for d in dias_norm)

                                        value = ''
                                        if seq:
                                            ref_date = None
                                            if asignacion.start_date:
                                                ref_date = asignacion.start_date
                                            elif getattr(asignacion, 'fecha', None):
                                                ref_date = asignacion.fecha
                                            else:
                                                try:
                                                    ref_date = datetime.date(int(asignacion.anio), int(asignacion.mes), 1)
                                                except Exception:
                                                    ref_date = current

                                            active = True
                                            if asignacion.recurring:
                                                if asignacion.start_date and day_date < asignacion.start_date:
                                                    active = False
                                                if asignacion.end_date and asignacion.end_date and day_date > asignacion.end_date:
                                                    active = False
                                            else:
                                                if getattr(asignacion, 'fecha', None):
                                                    active = (day_date == asignacion.fecha)
                                                else:
                                                    active = (day_date.month == asignacion.mes and day_date.year == asignacion.anio)

                                            if active and applies_by_puesto:
                                                try:
                                                    days_diff = (day_date - ref_date).days
                                                    # detectar si corresponde aplicar offset por 24h
                                                    offset = 0
                                                    try:
                                                        is_24h = False
                                                        # preferir la duración del horario de la asignación
                                                        if getattr(asignacion, 'horario', None):
                                                            hi = asignacion.horario.hora_ingreso
                                                            ho = asignacion.horario.hora_salida
                                                            dt1 = datetime.datetime.combine(datetime.date(1,1,1), hi)
                                                            dt2 = datetime.datetime.combine(datetime.date(1,1,1), ho)
                                                            if dt2 <= dt1:
                                                                dt2 += datetime.timedelta(days=1)
                                                            dur = (dt2 - dt1).total_seconds() / 3600.0
                                                            is_24h = dur >= 23.5
                                                        # si no se detectó por asignacion, caer en puesto horarios
                                                        if not is_24h and puesto_obj is not None:
                                                            horarios_qs = getattr(puesto_obj, 'horarios', None)
                                                            if horarios_qs is not None:
                                                                # Preferir horario del día específico (dia de la semana)
                                                                try:
                                                                    dia_num = day_date.isoweekday()
                                                                    horas_por_dia = list(horarios_qs.filter(dia=dia_num).values_list('horas', flat=True))
                                                                    if horas_por_dia:
                                                                        is_24h = any((int(h) if h is not None else 0) == 24 for h in horas_por_dia)
                                                                    else:
                                                                        # fallback: si no hay fila para ese día, revisar cualquier horario 24h
                                                                        horas_list = list(horarios_qs.values_list('horas', flat=True))
                                                                        is_24h = any((int(h) if h is not None else 0) == 24 for h in horas_list)
                                                                except Exception:
                                                                    horas_list = list(horarios_qs.values_list('horas', flat=True))
                                                                    is_24h = any((int(h) if h is not None else 0) == 24 for h in horas_list)
                                                        # si corresponde aplicar secuencia con offset por 24h, calcular el índice de la secuencia sumando el offset a la diferencia de días y aplicando módulo por la longitud de la secuencia para que sea continua
                                                        if is_24h and seq:
                                                            # offset = longitud del primer bloque consecutivo igual al primer símbolo
                                                            first = seq[0]
                                                            cnt = 0
                                                            for s in seq:
                                                                if s == first:
                                                                    cnt += 1
                                                                else:
                                                                    break
                                                            offset = cnt
                                                    except Exception:
                                                        offset = 0

                                                    idx_seq = (days_diff + offset) % len(seq)
                                                    value = seq[idx_seq]
                                                except Exception:
                                                    value = ''
                                        else:
                                            # Sin patrón: dejar el día vacío para llenarlo manualmente.
                                            value = ''

                                        defaults[key] = value
                                    try:
                                        pid = puesto_obj.id if puesto_obj and hasattr(puesto_obj, 'id') else getattr(asignacion, 'puesto_id', None)
                                        obj, created = AsignacionSemanal.objects.update_or_create(
                                            asignacion_id=asignacion.id,
                                            week_start=current,
                                            defaults={**defaults, 'asignacion': asignacion, 'puesto_id': pid}
                                        )
                                        if not created:
                                            if getattr(obj, 'asignacion_id', None) != getattr(asignacion, 'id', None):
                                                obj.asignacion = asignacion
                                            for k, v in defaults.items():
                                                setattr(obj, k, v)
                                            obj.save()
                                    except Exception as e:
                                        raise Exception(f"Error creando/asegurando AsignacionSemanal al actualizar: {e}")
                                    current += datetime.timedelta(days=7)
                            except Exception as e:
                                return Response({'error': f'No se pudo actualizar el calendario semanal: {e}'}, status=status.HTTP_400_BAD_REQUEST)
                        return Response(serializer2.data, status=status.HTTP_200_OK)
                    else:
                        return Response(serializer2.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception:
        pass

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def editar_servicio(request, id):
    # si el ususario no tiene permiso de cambio de asignacion, retornar error 403
    if not request.user.has_perm('CoreFisica.change_asignacion'):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    try:
        # Obtener la asignación a editar por id. Si no existe, retornar error 404.
        asignacion = Asignacion.objects.get(id=id)
    except Asignacion.DoesNotExist:
        return Response({'error': 'Asignación no encontrada'}, status=status.HTTP_404_NOT_FOUND)

    old_persona_id = asignacion.persona_id
    old_patron_id = getattr(asignacion, 'patronAsignacion_id', None)
    data = request.data.copy()
    reset_calendar_raw = data.get('reset_calendar', None)
    if 'reset_calendar' in data:
        data.pop('reset_calendar')
    reset_calendar = str(reset_calendar_raw).strip().lower() in ('1', 'true', 'yes', 'y', 'si', 'sí')

    # Detectar cambio de persona
    new_persona_raw = data.get('persona')
    try:
        new_persona_id = int(new_persona_raw) if new_persona_raw not in [None, '', 'null'] else None
    except (TypeError, ValueError):
        new_persona_id = None

    persona_cambio = new_persona_id and new_persona_id != old_persona_id

    if persona_cambio:
        try:
            with transaction.atomic():
                # Si la nueva persona ya tiene asignación activa este mes en OTRO puesto,
                # liberar ese puesto (la persona se mueve, su puesto anterior queda vacante).
                otras = Asignacion.objects.filter(
                    persona_id=new_persona_id,
                    mes=asignacion.mes,
                    anio=asignacion.anio,
                    estado='ACTIVO'
                ).exclude(id=asignacion.id)
                for otra in otras:
                    otra.persona = None
                    otra.save(update_fields=['persona'])
                    # El puesto queda vacante: limpiar estado del reporte
                    ReporteAsistencia.objects.filter(asignacion=otra).update(
                        persona=None, estado='TURNO', estado_asistencia='',
                        reemplazo=None, descripcion=None, row_color=None
                    )

                # Cambiar la persona EN LA MISMA fila: el puesto conserva su calendario.
                asignacion.persona_id = new_persona_id
                nuevo_horario = data.get('horario')
                if nuevo_horario:
                    try:
                        asignacion.horario_id = int(nuevo_horario)
                    except (TypeError, ValueError):
                        pass
                asignacion.save()
                # Nueva persona: resetear estado del reporte (no heredar el de la anterior)
                ReporteAsistencia.objects.filter(asignacion=asignacion).update(
                    persona_id=new_persona_id, estado='TURNO', estado_asistencia='',
                    reemplazo=None, descripcion=None, row_color=None
                )
                return Response(AsignacionSerializer(asignacion).data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    data['recurring'] = True
    data['end_date'] = None
    serializer = AsignacionSerializer(asignacion, data=data, partial=True)
    if serializer.is_valid():
        asignacion = serializer.save()
        patron_changed = 'patronAsignacion' in request.data and old_patron_id != getattr(asignacion, 'patronAsignacion_id', None)
        if reset_calendar or patron_changed:
            try:
                if reset_calendar:
                    AsignacionSemanal.objects.filter(asignacion_id=asignacion.id).delete()
                _rebuild_asignacion_semanal(asignacion, force_all=reset_calendar)
            except Exception as e:
                return Response({'error': f'No se pudo actualizar el calendario semanal: {e}'}, status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def guardar_orden_asignacion(request):
    # si el usuario no tiene permiso de cambio de asignacion, retornar error 403
    if not request.user.has_perm('CoreFisica.change_asignacion'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    # ordenes es igual a la lista de objetos con id y orden que viene en el body de la petición. Por ejemplo: [{"id": 1, "orden": 2}, {"id": 2, "orden": 1}]
    ordenes = request.data.get('ordenes', [])

    #itera sobre la lista de ordenes y para cada una intenta obtener la asignación por id, si existe actualiza su campo orden con el valor que viene en la petición y guarda la asignación. Si no existe, continúa con la siguiente orden sin hacer nada. Al final retorna un mensaje de éxito indicando que el orden se actualizó correctamente.
    for item in ordenes:
        try:
            asignacion = Asignacion.objects.get(id=item['id'])
            asignacion.orden = item['orden']
            asignacion.save()
        except Asignacion.DoesNotExist:
            continue

    # Propagar el nuevo orden de PUESTOS a los meses SIGUIENTES (no afecta meses pasados).
    # Se permutan los puestos reordenados dentro de los mismos "slots" de orden que ya
    # ocupan en los meses futuros, para no alterar el orden de otros clientes.
    try:
        ref_mes = int(request.data.get('mes')) if request.data.get('mes') else None
        ref_anio = int(request.data.get('anio')) if request.data.get('anio') else None
    except (TypeError, ValueError):
        ref_mes = ref_anio = None
    if ref_mes and ref_anio and ordenes:
        _propagar_orden_puestos_a_futuros(ordenes, ref_mes, ref_anio)

    return Response({'mensaje': 'Orden actualizado correctamente'})


def _propagar_orden_puestos_a_futuros(ordenes, ref_mes, ref_anio):
    """Propaga el orden de puestos del mes reordenado hacia los meses siguientes.

    - Toma la secuencia de puestos según el nuevo orden del mes actual.
    - En los meses futuros, esos mismos puestos se reubican dentro de los slots de
      `orden` que ya ocupaban (permutación), sin tocar otros puestos/clientes ni
      los meses pasados.
    """
    items = [it for it in ordenes if it.get('id') is not None]
    ids = [it['id'] for it in items]
    if not ids:
        return
    asigs = Asignacion.objects.in_bulk(ids)
    # Mapas del mes de referencia:
    #  - orden EXACTO por (puesto, persona): preserva el orden fila por fila.
    #  - orden mínimo por puesto: respaldo cuando en el futuro hay otra persona
    #    en ese puesto (p. ej. un reemplazo) que no estaba en el mes de referencia.
    orden_by_pp = {}
    orden_by_puesto = {}
    for it in items:
        a = asigs.get(it['id'])
        pid = getattr(a, 'puesto_id', None) if a else None
        if pid is None:
            continue
        o = it.get('orden', 0)
        orden_by_pp[(pid, getattr(a, 'persona_id', None))] = o
        if pid not in orden_by_puesto or o < orden_by_puesto[pid]:
            orden_by_puesto[pid] = o
    if not orden_by_puesto:
        return

    # Solo los puestos reordenados, solo meses futuros. Cada fila futura toma el
    # orden EXACTO de su (puesto, persona) en el mes de referencia.
    fut = Asignacion.objects.filter(
        estado='ACTIVO', puesto_id__in=list(orden_by_puesto.keys())
    ).filter(Q(anio__gt=ref_anio) | Q(anio=ref_anio, mes__gt=ref_mes))
    cambios = []
    for a in fut.only('id', 'orden', 'puesto_id', 'persona_id'):
        nv = orden_by_pp.get((a.puesto_id, a.persona_id))
        if nv is None:
            nv = orden_by_puesto.get(a.puesto_id)
        if nv is not None and a.orden != nv:
            a.orden = nv
            cambios.append(a)
    if cambios:
        Asignacion.objects.bulk_update(cambios, ['orden'], batch_size=1000)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def guardar_orden_sacafranco(request):
    # si el usuario no tiene permiso de cambio de asignacion, retornar error 403
    if not request.user.has_perm('CoreFisica.change_asignacion'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    #ordenes es igual a la lista de objetos con id y orden que viene en el body de la petición. Por ejemplo: [{"id": 1, "orden": 2}, {"id": 2, "orden": 1}]
    ordenes = request.data.get('ordenes', [])
    # itera sobre la lista de ordenes y para cada una intenta obtener la fila de sacafranco por id, si existe actualiza su campo orden con el valor que viene en la petición y guarda la fila. Si no existe, continúa con la siguiente orden sin hacer nada. Al final retorna un mensaje de éxito indicando que el orden de sacafranco se actualizó correctamente.
    for item in ordenes:
        
        try:
            fila = SacafrancoFila.objects.get(id=item['id'])
            fila.orden = item['orden']
            fila.save(update_fields=['orden'])
        except SacafrancoFila.DoesNotExist:
            continue
    return Response({'mensaje': 'Orden sacafranco actualizado correctamente'})

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def eliminar_asignacion(request, id):
    # si el usuario no tiene permiso de eliminar asignacion, retornar error 403
    if not request.user.has_perm('CoreFisica.delete_asignacion'):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    try:
        asignar = Asignacion.objects.get(id=id)
        # Guardar datos antes de eliminar para poder limpiar calendario y patron si corresponde
        # No recuperamos/eliminamos el PatronAsignacion aquí: preservamos los
        # patrones creados por los usuarios (se mantienen en el select).

        # Calcular una ventana de fechas para limpiar filas semanales huérfanas del mismo puesto.
        # Preferimos start_date/end_date si existen; si no, usamos mes/año y extendemos
        # hasta la última semana existente para ese puesto.
        puesto_id = getattr(asignar, 'puesto_id', None)
        window_start = None
        window_end = None
        try:
            start_ref = getattr(asignar, 'start_date', None) or getattr(asignar, 'fecha', None)
            if not start_ref and getattr(asignar, 'mes', None) and getattr(asignar, 'anio', None):
                start_ref = datetime.date(int(asignar.anio), int(asignar.mes), 1)

            end_ref = getattr(asignar, 'end_date', None)
            if not end_ref and start_ref:
                # Fin de mes del start_ref
                next_month = (start_ref.replace(day=28) + datetime.timedelta(days=4)).replace(day=1)
                end_ref = next_month - datetime.timedelta(days=1)

            if start_ref and end_ref:
                window_start = start_ref - datetime.timedelta(days=6)
                window_end = end_ref
        except Exception:
            window_start = None
            window_end = None

        # Ampliar la ventana usando las filas existentes del mismo puesto (cubrir recurrentes sin end_date).
        try:
            if puesto_id:
                max_week = AsignacionSemanal.objects.filter(puesto_id=puesto_id).aggregate(Max('week_start')).get('week_start__max')
                if max_week:
                    window_end = window_end or max_week
                if start_ref and not window_start:
                    window_start = start_ref - datetime.timedelta(days=6)
        except Exception:
            pass

        # Borrar filas semanales y luego la asignación dentro de una transacción
        try:
            with transaction.atomic():
                # Eliminar filas ligadas explícitamente a la asignación
                hoy = timezone.localdate()
                week_start_today = hoy - datetime.timedelta(days=hoy.weekday())

                AsignacionSemanal.objects.filter(
                    asignacion_id=id,
                    week_start__gte=week_start_today
                ).delete()

                # Limpiar filas del mismo puesto sin asignación pero con códigos (evita celdas huérfanas)
                try:
                    if puesto_id:
                        qs_clean = AsignacionSemanal.objects.filter(
                            puesto_id=puesto_id,
                            week_start__gte=week_start_today
                        )
                        if window_start:
                            qs_clean = qs_clean.filter(week_start__gte=window_start)
                        if window_end:
                            qs_clean = qs_clean.filter(week_start__lte=window_end)
                        # filas sin asignación, ligadas a esta asignación, o relinkeadas a SACAFRANCO
                        qs_clean = qs_clean.filter(
                            Q(asignacion__isnull=True) |
                            Q(asignacion_id=id) |
                            Q(asignacion__persona__tipo='SACAFRANCO')
                        )
                        for row in qs_clean:
                            dirty = False
                            for d in ['mon','tue','wed','thu','fri','sat','sun']:
                                val = getattr(row, d, '') or ''
                                if val != '':
                                    setattr(row, d, '')
                                    dirty = True
                            if dirty or row.asignacion_id:
                                row.asignacion = None
                                row.save()
                except Exception:
                    pass

                # Limpiar celdas 'F' del mismo puesto dentro de la ventana para evitar sacafranco huérfano
                try:
                    if puesto_id:
                        from django.db.models import Q as _Q
                        qs_f = AsignacionSemanal.objects.filter(
                            puesto_id=puesto_id,
                            week_start__gte=week_start_today
                        )
                        if window_start:
                            qs_f = qs_f.filter(week_start__gte=window_start)
                        if window_end:
                            qs_f = qs_f.filter(week_start__lte=window_end)
                        # Celdas 'F' y asignacion nula o la misma que se elimina
                        f_filter = (
                            _Q(mon__istartswith='F') | _Q(tue__istartswith='F') | _Q(wed__istartswith='F') |
                            _Q(thu__istartswith='F') | _Q(fri__istartswith='F') | _Q(sat__istartswith='F') |
                            _Q(sun__istartswith='F')
                        )
                        qs_f = qs_f.filter(f_filter).filter(
                            _Q(asignacion__isnull=True) |
                            _Q(asignacion_id=id) |
                            _Q(asignacion__persona__tipo='SACAFRANCO')
                        )
                        for row in qs_f:
                            try:
                                for d in ['mon','tue','wed','thu','fri','sat','sun']:
                                    val = getattr(row, d, '') or ''
                                    if str(val).upper().startswith('F'):
                                        setattr(row, d, '')
                                row.asignacion = None
                                row.save()
                            except Exception:
                                pass
                except Exception:
                    pass

                # Además, intentar eliminar posibles filas huérfanas creadas
                # para el mismo puesto dentro de la ventana calculada.
                try:
                    if puesto_id:
                        qs = AsignacionSemanal.objects.filter(
                            puesto_id=puesto_id,
                            asignacion__isnull=True,
                            week_start__gte=week_start_today
                        )
                        if window_start:
                            qs = qs.filter(week_start__gte=window_start)
                        if window_end:
                            qs = qs.filter(week_start__lte=window_end)
                        orphan_count = qs.count()
                        if orphan_count:
                            qs.delete()
                except Exception:
                    # no crítico: continuar con la eliminación principal
                    pass

                asignar.delete()
        except Exception:
            return Response({'error': 'No se pudo eliminar la asignación'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({'mensaje': 'Asignación eliminada correctamente'}, status=status.HTTP_204_NO_CONTENT)
    except Asignacion.DoesNotExist:
        return Response({'error': 'Asignacion no encontrada'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def sacafranco_filas(request):
    # si el usuario no tiene permiso de ver asignacion, retornar error 403
    if request.method == 'GET':
        if not request.user.has_perm('CoreFisica.view_asignacion'):
            return JsonResponse({'error': 'No autorizado'}, status=403)
        # Obtener mes y año desde los parámetros de la solicitud
        mes = request.GET.get('mes')
        anio = request.GET.get('anio')
        provincia_id = request.GET.get('provincia_id')
        canton_id = request.GET.get('canton_id')
        canton_ids_raw = (request.GET.get('canton_ids') or '').strip()
        canton_ids: list[int] = []
        if canton_ids_raw:
            try:
                canton_ids = [int(x) for x in canton_ids_raw.split(',') if str(x).strip()]
            except (TypeError, ValueError):
                return Response({'error': 'Canton IDs invalidos'}, status=status.HTTP_400_BAD_REQUEST)
        # Vista por empresa: scope por clientes (filas creadas en esa vista). Para
        # filas heredadas (sin scope) se derivan los cantones de las instalaciones.
        cliente_ids = []
        cliente_ids_raw = (request.GET.get('cliente_ids') or '').strip()
        if cliente_ids_raw:
            try:
                cliente_ids = [int(x) for x in cliente_ids_raw.split(',') if str(x).strip()]
            except (TypeError, ValueError):
                return Response({'error': 'Cliente IDs invalidos'}, status=status.HTTP_400_BAD_REQUEST)

        # se obtiene las filas de sacafranco
        qs = SacafrancoFila.objects.all()
        if mes and anio:
            try:
                mes_val = int(mes)
                anio_val = int(anio)
                qs = qs.filter(Q(anio__lt=anio_val) | Q(anio=anio_val, mes__lte=mes_val))
                # Mostrar todas las filas del mes (no filtrar por semanales).
            except (TypeError, ValueError):
                pass

        # Una fila con scope propio (cantones/clientes) se muestra SOLO donde coincide.
        # Una fila sin scope ("heredada") se muestra por el cantón de su persona.
        sin_scope = Q(cantones__len=0) & Q(clientes__len=0)
        if cliente_ids and not canton_ids:
            from ..models import Instalacion
            derived_cantones = list(
                Instalacion.objects.filter(cliente_id__in=cliente_ids)
                .exclude(canton_id__isnull=True)
                .values_list('canton_id', flat=True).distinct()
            )
            qs = qs.filter(
                Q(clientes__overlap=cliente_ids)
                | (sin_scope & Q(persona__canton_id__in=derived_cantones))
            )
        elif canton_ids:
            qs = qs.filter(
                Q(cantones__overlap=canton_ids)
                | (sin_scope & Q(persona__canton_id__in=canton_ids))
            )
        elif canton_id:
            try:
                canton_val = int(canton_id)
            except (TypeError, ValueError):
                return Response({'error': 'Canton invalido'}, status=status.HTTP_400_BAD_REQUEST)
            qs = qs.filter(
                Q(cantones__contains=[canton_val])
                | (sin_scope & Q(persona__canton_id=canton_val))
            )
        elif provincia_id:
            try:
                provincia_val = int(provincia_id)
            except (TypeError, ValueError):
                return Response({'error': 'Provincia invalida'}, status=status.HTTP_400_BAD_REQUEST)
            qs = qs.filter(Q(provincia_id=provincia_val) | Q(persona__provincia_id=provincia_val))
        # Búsqueda de texto: filtrar sacafranco por su persona (nombre/apellido/cédula),
        # insensible a acentos. Así no aparecen siempre al buscar un cliente/instalación.
        q = (request.GET.get('q') or '').strip()
        if q:
            qn = _strip_accents(q)
            saca_filtros = (
                Q(persona__nombres__unaccent__icontains=qn) |
                Q(persona__apellidos__unaccent__icontains=qn) |
                Q(persona__cedula__icontains=q)
            )
            if q.isdigit():
                saca_filtros = saca_filtros | Q(id=int(q))
            qs = qs.filter(saca_filtros)
        qs = qs.order_by(Coalesce('provincia_id', Value(999999)), 'orden', 'id')

        page_param = request.GET.get('page')
        size_param = request.GET.get('size')
        if page_param is not None or size_param is not None:
            try:
                page = int(page_param or 1)
                size = int(size_param or 100)
                if page < 1 or size < 1:
                    raise ValueError
            except ValueError:
                return Response({'error': 'Parametros de paginacion invalidos'}, status=status.HTTP_400_BAD_REQUEST)

            total = qs.count()
            start = (page - 1) * size
            end = start + size
            qs = qs[start:end]
            serializer = SacafrancoFilaSerializer(qs, many=True)
            return Response({
                'total': total,
                'page': page,
                'size': size,
                'results': serializer.data
            })

        serializer = SacafrancoFilaSerializer(qs, many=True)
        return Response(serializer.data)

    if not request.user.has_perm('CoreFisica.add_asignacion'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    serializer = SacafrancoFilaSerializer(data=request.data)
    if serializer.is_valid():
        fila = serializer.save()
        if request.data.get('orden') is None:
            max_orden = SacafrancoFila.objects.filter(mes=fila.mes, anio=fila.anio).aggregate(Max('orden')).get('orden__max')
            fila.orden = (max_orden or 0) + 1
            fila.save()
        return Response(SacafrancoFilaSerializer(fila).data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE', 'PATCH'])
@permission_classes([IsAuthenticated])
def eliminar_sacafranco_fila(request, id):
    if request.method == 'PATCH':
        if not request.user.has_perm('CoreFisica.change_asignacion'):
            return JsonResponse({'error': 'No autorizado'}, status=403)
        try:
            fila = SacafrancoFila.objects.get(id=id)
        except SacafrancoFila.DoesNotExist:
            return Response({'error': 'Fila no encontrada'}, status=status.HTTP_404_NOT_FOUND)

        update_fields = []

        provincia_id = request.data.get('provincia', None)
        if provincia_id not in ['', None]:
            try:
                provincia_id = int(provincia_id)
            except (TypeError, ValueError):
                return Response({'error': 'Provincia invalida'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            provincia_id = None
        if fila.provincia_id != provincia_id:
            fila.provincia_id = provincia_id
            update_fields.append('provincia')

        persona_id = request.data.get('persona', None)
        if persona_id not in ['', None]:
            try:
                persona_id = int(persona_id)
            except (TypeError, ValueError):
                return Response({'error': 'Persona invalida'}, status=status.HTTP_400_BAD_REQUEST)
            try:
                persona = Persona.objects.get(id=persona_id)
            except Persona.DoesNotExist:
                return Response({'error': 'Persona no encontrada'}, status=status.HTTP_404_NOT_FOUND)
            if (persona.tipo or '').upper() != 'SACAFRANCO':
                return Response({'error': 'La persona no es SACAFRANCO'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            persona = None
        if fila.persona_id != (persona.id if persona else None):
            fila.persona = persona
            update_fields.append('persona')

        # Hora de ingreso/salida (libre, formato HH:MM o vacío para limpiar).
        for campo in ('hora_ingreso', 'hora_salida'):
            if campo in request.data:
                val = request.data.get(campo)
                val = str(val).strip() if val not in (None, '') else None
                setattr(fila, campo, val or None)
                update_fields.append(campo)

        if update_fields:
            fila.save(update_fields=update_fields)
        return Response(SacafrancoFilaSerializer(fila).data)

    if not request.user.has_perm('CoreFisica.delete_asignacion'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        fila = SacafrancoFila.objects.get(id=id)
    except SacafrancoFila.DoesNotExist:
        return Response({'error': 'Fila no encontrada'}, status=status.HTTP_404_NOT_FOUND)

    # Clean up auto-generated cobertura records before removing weekly tokens.
    from .asignacion_semanal_views import _cleanup_auto_sacafranco_from_token_map

    semanales = SacafrancoFilaSemanal.objects.filter(sacafranco_fila_id=fila.id)
    for semanal in semanales:
        token_map = {
            'mon': semanal.mon,
            'tue': semanal.tue,
            'wed': semanal.wed,
            'thu': semanal.thu,
            'fri': semanal.fri,
            'sat': semanal.sat,
            'sun': semanal.sun,
        }
        _cleanup_auto_sacafranco_from_token_map(
            sacafranco_fila_id=fila.id,
            week_start_date=semanal.week_start,
            token_map=token_map,
        )

    semanales.delete()
    fila.delete()
    return Response({'mensaje': 'Fila sacafranco eliminada', 'eliminadas': 1})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_asignaciones_excel(request):
    if not request.user.has_perm('CoreFisica.export_asignacion'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    import calendar
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asignaciones y Calendario"

    # Obtener mes/año desde query params si vienen, si no usar mes actual
    qs = request.GET
    try:
        if 'mes' in qs and 'anio' in qs:
            month = int(qs.get('mes'))
            year = int(qs.get('anio'))
        else:
            today = datetime.date.today()
            year = today.year
            month = today.month
    except Exception:
        today = datetime.date.today()
        year = today.year
        month = today.month

    first_day = datetime.date(year, month, 1)
    last_day_num = calendar.monthrange(year, month)[1]
    dates = [first_day + datetime.timedelta(days=i) for i in range(last_day_num)]

    # Columnas izquierdas (datos de asignación)
    # Se eliminó 'Dirección Instalación' según solicitud
    left_headers = ['HORARIO', 'NOMINATIVO', 'CLIENTE', 'NOMBRE PUESTO', 'RESUMEN', 'CÉDULA', 'PERSONA', 'TIPO']
    left_cols = len(left_headers)

    # Columnas de fecha comienzan después de las columnas izquierdas
    date_start_col = left_cols + 1
    num_days = len(dates)

    # Encabezado superior (formato institucional)
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    celeste_fill = PatternFill(fill_type='solid', fgColor='B1C2CC')
    sacafranco_fill = PatternFill(fill_type='solid', fgColor='FFF3CD')

    ws.merge_cells('A1:A3')
    ws.merge_cells('B1:E2')
    ws.merge_cells('B3:E3')
    ws.merge_cells('G1:H1')
    ws.merge_cells('G2:H2')
    ws.merge_cells('G3:H3')

    ws['B1'] = 'REPORTE DE HORARIOS DE PERSONAL'
    ws['B3'] = 'SEGURIDAD FÍSICA'
    ws['F1'] = 'Código:'
    ws['F2'] = 'Versión:'
    ws['F3'] = 'Fecha:'
    ws['G1'] = 'FOR-SF-001'
    ws['G2'] = '02'
    ws['G3'] = datetime.date.today().strftime('%d/%m/%Y')

    ws['B1'].font = Font(bold=True, size=18)
    ws['B3'].font = Font(bold=True, size=14)
    ws['F1'].font = ws['F2'].font = ws['F3'].font = Font(bold=True, size=11)
    ws['G1'].font = ws['G2'].font = ws['G3'].font = Font(bold=True, size=11)

    for row in range(1, 4):
        for col in range(1, 9):
            c = ws.cell(row=row, column=col)
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws['F1'].alignment = ws['F2'].alignment = ws['F3'].alignment = Alignment(horizontal='left', vertical='center')
    ws['G1'].alignment = ws['G2'].alignment = ws['G3'].alignment = Alignment(horizontal='left', vertical='center')

    ws.row_dimensions[1].height = 42
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[3].height = 34

    logo_path = _find_asignaciones_logo_path()
    if logo_path:
        try:
            img = XLImage(str(logo_path))

            # Escalado proporcional para evitar deformación
            max_w = 80
            max_h = 80
            orig_w = max(float(img.width), 1.0)
            orig_h = max(float(img.height), 1.0)
            scale = min(max_w / orig_w, max_h / orig_h)
            final_w = int(orig_w * scale)
            final_h = int(orig_h * scale)
            img.width = final_w
            img.height = final_h

            # Centrar en recuadro A1:A3
            box_w_px = 126
            box_h_px = 110
            x_offset_px = int((box_w_px - final_w) / 2)
            y_offset_px = int((box_h_px - final_h) / 2)
            y_offset_px += 16
            
            img.anchor = OneCellAnchor(
                _from=AnchorMarker(
                    col=0,
                    row=0,
                    colOff=pixels_to_EMU(max(x_offset_px, 0)),
                    rowOff=pixels_to_EMU(max(y_offset_px, 0)),
                ),
                ext=XDRPositiveSize2D(
                    cx=pixels_to_EMU(final_w),
                    cy=pixels_to_EMU(final_h),
                ),
            )
            ws.add_image(img)
        except Exception:
            pass

    # Rellenar filas: una fila por Asignacion activa para el mes solicitado
    # Aplicar la misma lógica que en obtener_asignaciones: incluir recurrentes que aplican al mes
    month_start = first_day
    if month == 12:
        month_end = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
    else:
        month_end = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)

    asignaciones = Asignacion.objects.filter(estado='ACTIVO').filter(
        Q(mes=month, anio=year) |
        (Q(recurring=True) & Q(start_date__lte=month_end) & (Q(end_date__isnull=True) | Q(end_date__gte=month_start)))
    ).select_related('horario', 'cliente', 'puesto', 'persona', 'instalacion').prefetch_related('puesto__horarios')
    asignaciones = list(asignaciones)

    # Pre-cargar AsignacionSemanal del mes en una sola consulta para evitar N x dias queries.
    semanal_cache = {}
    asignacion_ids = {getattr(a, 'id', None) for a in asignaciones if getattr(a, 'id', None)}
    puesto_ids = {getattr(a, 'puesto_id', None) for a in asignaciones if getattr(a, 'puesto_id', None)}
    week_starts = set()
    for d in dates:
        week_index = (d.day - 1) // 7
        week_starts.add(first_day + datetime.timedelta(days=week_index * 7))
        week_starts.add(d - datetime.timedelta(days=d.weekday()))

    if week_starts and (asignacion_ids or puesto_ids):
        semanales = AsignacionSemanal.objects.filter(week_start__in=week_starts).filter(
            Q(asignacion_id__in=asignacion_ids) | Q(puesto_id__in=puesto_ids)
        )
        for s in semanales:
            if getattr(s, 'asignacion_id', None):
                semanal_cache[('a', s.asignacion_id, s.week_start)] = s
            if getattr(s, 'puesto_id', None):
                semanal_cache[('p', s.puesto_id, s.week_start)] = s

    def build_resumen(puesto):
        try:
            if not puesto:
                return ''
            horarios_qs = getattr(puesto, 'horarios', None)
            horarios = list(horarios_qs.all()) if horarios_qs is not None else []
            if not horarios:
                return ''

            day_map = {1: 'L', 2: 'M', 3: 'X', 4: 'J', 5: 'V', 6: 'S', 7: 'D'}
            groups = {}

            for h in horarios:
                horas_val = float(getattr(h, 'horas', 0) or 0)
                turno_val = (getattr(h, 'turno', '') or '').strip()
                key = f"{horas_val}-{turno_val}"
                if key not in groups:
                    groups[key] = {'horas': horas_val, 'turno': turno_val, 'dias': []}
                dia_val = getattr(h, 'dia', None)
                if dia_val and dia_val not in groups[key]['dias']:
                    groups[key]['dias'].append(dia_val)

            def turno_letter(t: str) -> str:
                lower = (t or '').strip().lower()
                if lower.startswith('d'):
                    return 'D'
                if lower.startswith('n'):
                    return 'N'
                return ''  # 24h/Ambos: sin letra de turno (evita la doble H)

            parts = []
            for g in groups.values():
                ordered = sorted(g['dias']) if g['dias'] else []
                first = day_map.get(ordered[0], '') if ordered else ''
                last = day_map.get(ordered[-1], '') if ordered else ''
                dias_str = first if len(ordered) <= 1 else f"{first}{last}"
                base = f"{g['horas']:g}H{turno_letter(g['turno'])}".strip()
                parts.append(f"{base} {dias_str}".strip())

            # Ordenar priorizando horas numéricas bajas primero (coincide con la vista)
            def sort_key(p: str):
                try:
                    return int(''.join([ch for ch in p if ch.isdigit()]) or 0)
                except Exception:
                    return 0

            parts = sorted(parts, key=sort_key)
            body = ' / '.join([p for p in parts if p])
            # El resumen representa UN puesto (registro), no su capacidad de cupos.
            cant = '1'
            if cant and body:
                return f"{cant} {body}"
            if cant:
                return cant
            return body
        except Exception:
            return ''
    sacafranco_rows = SacafrancoFila.objects.filter(
        Q(anio__lt=year) | Q(anio=year, mes__lte=month)
    ).select_related('persona').order_by(Coalesce('provincia_id', Value(999999)), 'orden', 'id')

    sac_week_rows = SacafrancoFilaSemanal.objects.filter(
        sacafranco_fila_id__in=list(sacafranco_rows.values_list('id', flat=True)),
        week_start__gte=month_start,
        week_start__lte=month_end
    )
    sac_sem_cache = {(r.sacafranco_fila_id, r.week_start): r for r in sac_week_rows}

    def get_asignacion_provincia_id(item: Asignacion) -> int:
        return getattr(getattr(getattr(item, 'instalacion', None), 'canton', None), 'provincia_id', None) or 999999

    def get_sacafranco_provincia_id(item: SacafrancoFila) -> int:
        return getattr(item, 'provincia_id', None) or 999999

    combined_rows = []
    for asignacion in asignaciones:
        combined_rows.append({
            'kind': 'asignacion',
            'provincia': get_asignacion_provincia_id(asignacion),
            'orden': getattr(asignacion, 'orden', 0) or 0,
            'id': getattr(asignacion, 'id', 0) or 0,
            'item': asignacion
        })
    for fila in sacafranco_rows:
        combined_rows.append({
            'kind': 'sacafranco',
            'provincia': get_sacafranco_provincia_id(fila),
            'orden': getattr(fila, 'orden', 0) or 0,
            'id': getattr(fila, 'id', 0) or 0,
            'item': fila
        })

    combined_rows.sort(key=lambda r: (
        r['provincia'],
        r['orden'],
        0 if r['kind'] == 'asignacion' else 1,
        r['id']
    ))

    def safe_sheet_title(name: str) -> str:
        title = (name or '').strip().upper() or 'SIN PROVINCIA'
        for ch in ['\\', '/', '?', '*', '[', ']', ':']:
            title = title.replace(ch, ' ')
        title = title.strip() or 'SIN PROVINCIA'
        return title[:31]

    def unique_sheet_title(base: str) -> str:
        title = base
        idx = 2
        while title in wb.sheetnames:
            suffix = f" {idx}"
            title = f"{base[:31 - len(suffix)]}{suffix}"
            idx += 1
        return title

    def build_sheet(target_ws, title: str, rows: list) -> None:
        target_ws.title = title
        # Encabezado institucional con logo (igual a la hoja general)
        target_ws.merge_cells('A1:A3')
        target_ws.merge_cells('B1:E2')
        target_ws.merge_cells('B3:E3')
        target_ws.merge_cells('G1:H1')
        target_ws.merge_cells('G2:H2')
        target_ws.merge_cells('G3:H3')

        target_ws['B1'] = 'REPORTE DE HORARIOS DE PERSONAL'
        target_ws['B3'] = 'SEGURIDAD FÍSICA'
        target_ws['F1'] = 'Código:'
        target_ws['F2'] = 'Versión:'
        target_ws['F3'] = 'Fecha:'
        target_ws['G1'] = 'FOR-SF-001'
        target_ws['G2'] = '02'
        target_ws['G3'] = datetime.date.today().strftime('%d/%m/%Y')

        target_ws['B1'].font = Font(bold=True, size=18)
        target_ws['B3'].font = Font(bold=True, size=14)
        target_ws['F1'].font = target_ws['F2'].font = target_ws['F3'].font = Font(bold=True, size=11)
        target_ws['G1'].font = target_ws['G2'].font = target_ws['G3'].font = Font(bold=True, size=11)

        for row in range(1, 4):
            for col in range(1, 9):
                c = target_ws.cell(row=row, column=col)
                c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        target_ws['F1'].alignment = target_ws['F2'].alignment = target_ws['F3'].alignment = Alignment(horizontal='left', vertical='center')
        target_ws['G1'].alignment = target_ws['G2'].alignment = target_ws['G3'].alignment = Alignment(horizontal='left', vertical='center')

        target_ws.row_dimensions[1].height = 42
        target_ws.row_dimensions[2].height = 34
        target_ws.row_dimensions[3].height = 34

        if logo_path:
            try:
                img = XLImage(str(logo_path))

                max_w = 80
                max_h = 80
                orig_w = max(float(img.width), 1.0)
                orig_h = max(float(img.height), 1.0)
                scale = min(max_w / orig_w, max_h / orig_h)
                final_w = int(orig_w * scale)
                final_h = int(orig_h * scale)
                img.width = final_w
                img.height = final_h

                box_w_px = 126
                box_h_px = 110
                x_offset_px = int((box_w_px - final_w) / 2)
                y_offset_px = int((box_h_px - final_h) / 2)
                y_offset_px += 16

                img.anchor = OneCellAnchor(
                    _from=AnchorMarker(
                        col=0,
                        row=0,
                        colOff=pixels_to_EMU(max(x_offset_px, 0)),
                        rowOff=pixels_to_EMU(max(y_offset_px, 0)),
                    ),
                    ext=XDRPositiveSize2D(
                        cx=pixels_to_EMU(final_w),
                        cy=pixels_to_EMU(final_h),
                    ),
                )
                target_ws.add_image(img)
            except Exception:
                pass

        month_row = 5
        dow_row = 6
        header_row = 7
        data_start_row = 8

        month_name = first_day.strftime('%B').upper()
        if num_days > 0:
            target_ws.merge_cells(
                start_row=month_row,
                start_column=date_start_col,
                end_row=month_row,
                end_column=date_start_col + num_days - 1
            )
            cell = target_ws.cell(row=month_row, column=date_start_col)
            cell.value = f"{month_name} {year}"
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        dow_names = ['L', 'M', 'M', 'J', 'V', 'S', 'D']
        for i, d in enumerate(dates):
            c = target_ws.cell(row=dow_row, column=date_start_col + i)
            c.value = dow_names[d.weekday()]
            c.alignment = Alignment(horizontal='center')

        for idx, h in enumerate(left_headers, start=1):
            ch = target_ws.cell(row=header_row, column=idx)
            ch.value = h
            ch.font = Font(bold=True)
            ch.alignment = Alignment(horizontal='left')
        for i, d in enumerate(dates):
            c = target_ws.cell(row=header_row, column=date_start_col + i)
            c.value = d.day
            c.alignment = Alignment(horizontal='center')

        target_ws.freeze_panes = target_ws.cell(row=data_start_row, column=date_start_col)

        start_row = data_start_row
        for entry in rows:
            row_idx = start_row
            if entry['kind'] == 'asignacion':
                asignacion = entry['item']
                horario_txt = ''
                try:
                    horario_txt = f"{asignacion.horario.hora_ingreso} - {asignacion.horario.hora_salida}"
                except Exception:
                    horario_txt = ''
                inst_codigo = getattr(getattr(asignacion, 'instalacion', None), 'codigo', '') or ''
                puesto_obj = getattr(asignacion, 'puesto', None)
                resumen_val = build_resumen(puesto_obj) or getattr(puesto_obj, 'resumen', '')
                vals = [
                    horario_txt,
                    inst_codigo,
                    getattr(asignacion.cliente, 'nombre_comercial', ''),
                    getattr(puesto_obj, 'nombre', ''),
                    resumen_val,
                    getattr(asignacion.persona, 'cedula', '') if asignacion.persona else '',
                    f"{getattr(asignacion.persona, 'apellidos', '')} {getattr(asignacion.persona, 'nombres', '')}".strip() if asignacion.persona else '(VACANTE)',
                    getattr(asignacion.persona, 'tipo', '') if asignacion.persona else ''
                ]
                for ci, v in enumerate(vals, start=1):
                    cell = target_ws.cell(row=row_idx, column=ci)
                    cell.value = v
                    cell.border = border

                for di, d in enumerate(dates):
                    col = date_start_col + di
                    week_index = (d.day - 1) // 7
                    week_start = first_day + datetime.timedelta(days=week_index * 7)
                    legacy_week_start = d - datetime.timedelta(days=d.weekday())
                    puesto_id = getattr(asignacion.puesto, 'id', getattr(asignacion, 'puesto_id', None))
                    asign_key = ('a', getattr(asignacion, 'id', None), week_start)
                    legacy_asign_key = ('a', getattr(asignacion, 'id', None), legacy_week_start)
                    puesto_key = ('p', puesto_id, week_start)
                    legacy_puesto_key = ('p', puesto_id, legacy_week_start)

                    semanal = (
                        semanal_cache.get(asign_key)
                        or semanal_cache.get(legacy_asign_key)
                        or semanal_cache.get(puesto_key)
                        or semanal_cache.get(legacy_puesto_key)
                    )

                    val = ''
                    if semanal:
                        day_field = ['mon','tue','wed','thu','fri','sat','sun'][d.weekday()]
                        try:
                            val = getattr(semanal, day_field, '') or ''
                        except Exception:
                            val = ''

                    cell = target_ws.cell(row=row_idx, column=col)
                    cell.value = val
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border
                    if str(val).strip().upper() == 'F':
                        cell.fill = celeste_fill
            else:
                fila = entry['item']
                persona = getattr(fila, 'persona', None)
                _hi = getattr(fila, 'hora_ingreso', None)
                _ho = getattr(fila, 'hora_salida', None)
                horario_saca = ''
                if _hi or _ho:
                    horario_saca = f"{_hi.strftime('%H:%M') if _hi else ''} {_ho.strftime('%H:%M') if _ho else ''}".strip()
                vals = [
                    horario_saca,
                    '',
                    '',
                    'SACAFRANCO',
                    '',
                    getattr(persona, 'cedula', '') if persona else '',
                    f"{getattr(persona, 'apellidos', '')} {getattr(persona, 'nombres', '')}".strip(),
                    getattr(persona, 'tipo', '') if persona else ''
                ]
                for ci, v in enumerate(vals, start=1):
                    cell = target_ws.cell(row=row_idx, column=ci)
                    cell.value = v
                    cell.border = border
                    cell.fill = sacafranco_fill

                for di, d in enumerate(dates):
                    col = date_start_col + di
                    week_index = (d.day - 1) // 7
                    week_start = first_day + datetime.timedelta(days=week_index * 7)
                    sac_key = (fila.id, week_start)
                    semanal = sac_sem_cache.get(sac_key)

                    val = ''
                    if semanal:
                        day_field = ['mon','tue','wed','thu','fri','sat','sun'][d.weekday()]
                        try:
                            val = getattr(semanal, day_field, '') or ''
                        except Exception:
                            val = ''

                    cell = target_ws.cell(row=row_idx, column=col)
                    cell.value = val
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border
                    if str(val).strip().upper() == 'F':
                        cell.fill = celeste_fill

            start_row += 1

        for i in range(1, left_cols + 1):
            target_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18
        target_ws.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 28
        target_ws.column_dimensions[openpyxl.utils.get_column_letter(7)].width = 38
        for i in range(date_start_col, date_start_col + num_days):
            target_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 5

    build_sheet(ws, 'Asignaciones y Calendario', combined_rows)

    prov_ids = sorted({r['provincia'] for r in combined_rows})
    prov_lookup = {}
    prov_db_ids = [pid for pid in prov_ids if pid != 999999]
    if prov_db_ids:
        prov_lookup = {p.id: p.nombre for p in Provincia.objects.filter(id__in=prov_db_ids)}

    for prov_id in prov_ids:
        prov_name = prov_lookup.get(prov_id)
        if not prov_name:
            prov_name = 'SIN PROVINCIA' if prov_id == 999999 else f"PROVINCIA {prov_id}"
        base_title = safe_sheet_title(prov_name)
        sheet_title = unique_sheet_title(base_title)
        prov_rows = [r for r in combined_rows if r['provincia'] == prov_id]
        ws_prov = wb.create_sheet(title=sheet_title)
        build_sheet(ws_prov, sheet_title, prov_rows)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reporte_asignaciones_calendario_{year}_{month}.xlsx"'
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def asignaciones_vacantes(request, mes, anio):
    """Lista las asignaciones activas sin persona (puestos que quedaron vacantes) del mes."""
    if not request.user.has_perm('CoreFisica.view_asignacion'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        mes = int(mes)
        anio = int(anio)
    except (TypeError, ValueError):
        return Response({'error': 'mes o anio invalidos'}, status=status.HTTP_400_BAD_REQUEST)

    month_start = datetime.date(anio, mes, 1)
    if mes == 12:
        month_end = datetime.date(anio + 1, 1, 1) - datetime.timedelta(days=1)
    else:
        month_end = datetime.date(anio, mes + 1, 1) - datetime.timedelta(days=1)

    mes_filter = (
        Q(mes=mes, anio=anio) |
        (Q(recurring=True) & Q(start_date__lte=month_end) & (Q(end_date__isnull=True) | Q(end_date__gte=month_start)))
    )

    qs = Asignacion.objects.filter(
        estado='ACTIVO',
        persona__isnull=True,
        puesto__activo=True  # un puesto CERRADO no es una vacante por cubrir (no notifica)
    ).filter(mes_filter).select_related(
        'cliente', 'instalacion', 'instalacion__canton', 'puesto', 'horario'
    ).order_by('instalacion__canton__nombre', 'orden', 'id')

    resultado = []
    for a in qs:
        horario_txt = ''
        if a.horario:
            try:
                horario_txt = f"{a.horario.hora_ingreso.strftime('%H:%M')} - {a.horario.hora_salida.strftime('%H:%M')}"
            except Exception:
                horario_txt = ''
        resultado.append({
            'id': a.id,
            'codigo': getattr(a.instalacion, 'codigo', '') or '',
            'cliente': getattr(a.cliente, 'nombre_comercial', '') or '',
            'instalacion': getattr(a.instalacion, 'nombre', '') or '',
            'puesto': getattr(a.puesto, 'nombre', '') or '',
            'canton': getattr(getattr(a.instalacion, 'canton', None), 'nombre', '') or 'SIN CANTON',
            'canton_id': getattr(a.instalacion, 'canton_id', None),
            'horario': horario_txt,
        })

    return JsonResponse({'total': len(resultado), 'results': resultado}, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def personas_asignadas(request, mes, anio):
    """Devuelve los IDs de personas con asignación activa en el mes (todos los cantones)."""
    if not request.user.has_perm('CoreFisica.view_asignacion'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        mes = int(mes)
        anio = int(anio)
    except (TypeError, ValueError):
        return Response({'error': 'mes o anio invalidos'}, status=status.HTTP_400_BAD_REQUEST)

    month_start = datetime.date(anio, mes, 1)
    if mes == 12:
        month_end = datetime.date(anio + 1, 1, 1) - datetime.timedelta(days=1)
    else:
        month_end = datetime.date(anio, mes + 1, 1) - datetime.timedelta(days=1)

    ids = Asignacion.objects.filter(
        estado='ACTIVO',
        persona__isnull=False
    ).filter(
        Q(mes=mes, anio=anio) |
        (Q(recurring=True) & Q(start_date__lte=month_end) & (Q(end_date__isnull=True) | Q(end_date__gte=month_start)))
    ).values_list('persona_id', flat=True).distinct()

    return JsonResponse({'persona_ids': list(ids)}, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def puestos_ocupacion(request, mes, anio):
    """Cuántas asignaciones OCUPADAS tiene cada puesto en el mes (todos los cantones).

    Sirve para mostrar el contador de cupos (ocupadas/capacidad) en el modal de
    asignación, independientemente del cantón/vista cargado en pantalla.
    """
    if not request.user.has_perm('CoreFisica.view_asignacion'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        mes = int(mes)
        anio = int(anio)
    except (TypeError, ValueError):
        return Response({'error': 'mes o anio invalidos'}, status=status.HTTP_400_BAD_REQUEST)

    from django.db.models import Count
    filas = (Asignacion.objects
             .filter(estado='ACTIVO', persona__isnull=False, mes=mes, anio=anio)
             .values('puesto_id')
             .annotate(total=Count('id')))
    ocupacion = {str(f['puesto_id']): f['total'] for f in filas if f['puesto_id'] is not None}
    return JsonResponse({'ocupacion': ocupacion}, status=status.HTTP_200_OK)
