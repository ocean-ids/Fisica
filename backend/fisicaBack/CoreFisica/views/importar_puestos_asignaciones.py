"""Importación masiva de puestos y asignaciones desde archivo Excel."""
from datetime import date, datetime, time, timedelta
import logging
import os
import re
import tempfile
import threading

from django.db import transaction
from django.db.models import Q
from django.http import JsonResponse
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from ..models import Cliente, Instalacion, Puesto, PuestoHorario, Persona, Horario, Asignacion, PatronAsignacion, AsignacionSemanal, ReporteAsistencia

logger = logging.getLogger(__name__)

HEADER_MAP = {
    'INSTALACION': 'instalacion',
    'INSTALACION ID': 'instalacion_id',
    'PUESTO': 'puesto',
    'PUESTO NOMBRE': 'puesto',
    'NOMBRE PUESTO': 'puesto',
    'TIPO PUESTO': 'puesto_tipo',
    'PUESTO TIPO': 'puesto_tipo',
    'CEDULA': 'cedula',
    'APELLIDOS': 'apellidos',
    'NOMBRES': 'nombres',
    'TIPO': 'tipo',
    'HORARIO INGRESO': 'hora_ingreso',
    'HORA INGRESO': 'hora_ingreso',
    'INGRESO': 'hora_ingreso',
    'HORARIO SALIDA': 'hora_salida',
    'HORA SALIDA': 'hora_salida',
    'SALIDA': 'hora_salida',
    'HORAS': 'horas',
    'TURNO': 'turno',
    'DIAS': 'dias',
    'DIAS TURNO': 'dias',
    'DIA': 'dias',
    'FECHA': 'fecha',
    'CLIENTE': 'cliente',
    'CLIENTE NOMBRE': 'cliente',
    'NOMBRE COMERCIAL': 'cliente',
    'CLIENTE RUC': 'cliente_ruc',
    'RUC': 'cliente_ruc',
    'CANTIDAD PUESTOS': 'cantidad_puestos',
    'CANTIDAD': 'cantidad_puestos',
    'PUESTO CANTIDAD': 'cantidad_puestos',
    'PATRON': 'patron',
    'PATRON CODIGO': 'patron',
    'CODIGO PATRON': 'patron',
    'PATRON ID': 'patron_id',
}

for day_num in range(1, 32):
    HEADER_MAP[str(day_num)] = f'day_{day_num}'


def normalize_header(value):
    if value is None:
        return ''
    import unicodedata
    text = str(value).strip().upper()
    text = ''.join(c for c in unicodedata.normalize('NFKD', text) if not unicodedata.combining(c))
    text = text.replace('_', ' ')
    text = ' '.join(text.split())
    return text


def norm(val):
    if val is None:
        return ''
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val.is_integer():
            return str(int(val))
        return str(val)
    return re.sub(r'\s+', ' ', str(val)).strip()


def parse_excel_date(val):
    if not val:
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, (int, float)):
        try:
            return from_excel(val).date()
        except Exception:
            return None
    if isinstance(val, str):
        raw = val.strip()
        if not raw:
            return None
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y'):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
    return None


def parse_excel_time(val):
    if not val:
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, (int, float)):
        try:
            return from_excel(val).time()
        except Exception:
            return None
    if isinstance(val, str):
        raw = val.strip()
        if not raw:
            return None
        for fmt in ('%H:%M', '%H:%M:%S'):
            try:
                return datetime.strptime(raw, fmt).time()
            except ValueError:
                continue
    return None


def normalize_cedula(value):
    if value is None:
        return ''
    raw = str(value).strip()
    compact = re.sub(r'[\s\-]', '', raw)
    if not compact:
        return ''
    if not re.match(r'^\d+$', compact):
        return None
    if len(compact) == 9:
        return f"0{compact}"
    return compact


def parse_turno(value):
    if not value:
        return 'Diurno'
    token = str(value).strip().upper()
    if token in ('D', 'DIURNO', 'DIURNA'):
        return 'Diurno'
    if token in ('N', 'NOCTURNO', 'NOCTURNA'):
        return 'Nocturno'
    if token in ('A', 'AMBOS', '24H', '24'):
        return 'Ambos'
    return 'Diurno'


def parse_turno_groups(value):
    if not value:
        return []
    text = str(value).strip()
    if not text:
        return []
    parts = [p.strip() for p in text.split('/') if p.strip()]
    return [parse_turno(p) for p in parts]


def parse_hours_groups(value):
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    parts = [p.strip() for p in text.split('/') if p.strip()]
    hours = []
    for part in parts:
        try:
            hours.append(int(float(part)))
        except (TypeError, ValueError):
            continue
    return hours


def parse_dias(value):
    if not value:
        return []
    text = str(value).strip().upper()
    text = text.replace('"', '').replace("'", '')
    if not text:
        return []
    tokens = re.split(r'[\s,;/|]+', text)
    mapping = {
        'L': 1, 'LU': 1, 'LUN': 1, 'LUNES': 1,
        'M': 2, 'MA': 2, 'MAR': 2, 'MARTES': 2,
        'X': 3, 'MI': 3, 'MIE': 3, 'MIERCOLES': 3,
        'J': 4, 'JU': 4, 'JUE': 4, 'JUEVES': 4,
        'V': 5, 'VI': 5, 'VIE': 5, 'VIERNES': 5,
        'S': 6, 'SA': 6, 'SAB': 6, 'SABADO': 6,
        'D': 7, 'DO': 7, 'DOM': 7, 'DOMINGO': 7,
    }
    result = set()
    for tok in tokens:
        if not tok:
            continue
        tok = re.sub(r'[^A-Z0-9]', '', tok.strip('.').upper())
        if tok.isdigit():
            num = int(tok)
            if 1 <= num <= 7:
                result.add(num)
            continue
        if tok in mapping:
            result.add(mapping[tok])
    return sorted(result)


def parse_dias_groups(value):
    if not value:
        return []
    text = str(value).strip().upper()
    text = text.replace('"', '').replace("'", '')
    if not text:
        return []
    # Support "LUNES-JUEVES / VIERNES-SABADO" or "L,M,X,J / V,S"
    group_tokens = [g.strip() for g in text.split('/') if g.strip()]
    mapping = {
        'L': 1, 'LU': 1, 'LUN': 1, 'LUNES': 1,
        'M': 2, 'MA': 2, 'MAR': 2, 'MARTES': 2,
        'X': 3, 'MI': 3, 'MIE': 3, 'MIERCOLES': 3,
        'J': 4, 'JU': 4, 'JUE': 4, 'JUEVES': 4,
        'V': 5, 'VI': 5, 'VIE': 5, 'VIERNES': 5,
        'S': 6, 'SA': 6, 'SAB': 6, 'SABADO': 6,
        'D': 7, 'DO': 7, 'DOM': 7, 'DOMINGO': 7,
    }
    def normalize_tok(tok: str):
        tok = re.sub(r'[^A-Z0-9]', '', tok.strip().upper())
        return tok

    groups = []
    for group in group_tokens:
        group = group.replace(' A ', '-').replace(' A\t', '-').replace('\tA ', '-')
        days = []
        if '-' in group:
            parts = [p for p in re.split(r'-+', group) if p.strip()]
            if len(parts) >= 2:
                start = mapping.get(normalize_tok(parts[0]))
                end = mapping.get(normalize_tok(parts[1]))
                if start and end:
                    if start <= end:
                        days = list(range(start, end + 1))
                    else:
                        days = list(range(start, 8)) + list(range(1, end + 1))
        if not days:
            tokens = re.split(r'[\s,;|]+', group)
            for tok in tokens:
                if not tok:
                    continue
                key = normalize_tok(tok.strip('.'))
                if key.isdigit():
                    num = int(key)
                    if 1 <= num <= 7:
                        days.append(num)
                    continue
                if key in mapping:
                    days.append(mapping[key])
        if days:
            groups.append(sorted(set(days)))
    return groups


def _expand_compact_dias(token: str):
    if not token:
        return []
    token = re.sub(r'[^A-Z]', '', token.upper())
    mapping = {'L': 1, 'M': 2, 'X': 3, 'J': 4, 'V': 5, 'S': 6, 'D': 7}
    if len(token) == 2 and token[0] in mapping and token[1] in mapping and token[0] != token[1]:
        start = mapping[token[0]]
        end = mapping[token[1]]
        if start <= end:
            return list(range(start, end + 1))
        return list(range(start, 8)) + list(range(1, end + 1))
    days = []
    for ch in token:
        if ch in mapping:
            days.append(mapping[ch])
    return sorted(set(days))


def parse_compact_horas_turno_dias(value):
    """Parsea el resumen compacto de un puesto. Acepta ambos formatos:
      - Con H: '24HLD', '12HDJM', '9HDLV', '5HDLU' (numero + H + turno? + dias)
      - Con turno y espacio: '24D LD', '12D LV', '12N LD' (numero + turno + dias)
      - Decimales y multi-grupo: '5HDLU - 4.5HDV - 5HDSD'
    """
    if value is None:
        return []
    text = str(value).strip().upper()
    if not text:
        return []
    # separar grupos por '/', ' - ', ' + ' o ' Y ' (texto ya en mayúsculas)
    raw_parts = re.split(r'\s*/\s*|\s+-\s+|\s*\+\s*|\s+Y\s+', text)
    groups = []
    turno_heredado = None  # el turno se declara una vez y se hereda a los grupos siguientes
    for part in raw_parts:
        p = re.sub(r'\s+', '', part)
        if not p:
            continue
        # numero(decimal) + H opcional + turno opcional (D/N) + dias (1-2 letras de LMXJVSD)
        m = re.match(r'^(\d+(?:[.,]\d+)?)H?([DN])?([LMXJVSD]{1,2})?$', p)
        if not m:
            continue
        try:
            hours_val = int(round(float(m.group(1).replace(',', '.'))))
        except ValueError:
            hours_val = 12
        turno_token = m.group(2) or ''
        dias_token = m.group(3) or ''
        dias = _expand_compact_dias(dias_token)
        turno = parse_turno(turno_token) if turno_token else None
        # Ambiguedad de la letra D: en un grupo tipo "5 H D" (D suelta, sin mas dias),
        # la D es el DIA Domingo, no el turno Diurno (ej. "14HDLV + 9HS + 5HD" =
        # L-V + Sabado + Domingo). Solo aplica cuando el grupo quedaria SIN dias.
        if not dias and turno_token == 'D':
            dias = [7]  # Domingo
            turno = None
        # Herencia de turno: si el grupo no trae turno propio, usa el del grupo anterior
        # (ej. "12HNLV y 24HSD" -> el 24HSD tambien es Nocturno). Si lo trae, actualiza.
        if turno is None:
            turno = turno_heredado
        else:
            turno_heredado = turno
        groups.append({
            'hours': hours_val,
            'turno': turno,
            'dias': dias,
        })
    return groups


def parse_calendar_value(val):
    if val is None:
        return ''
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val.is_integer():
            return str(int(val)).strip().upper()
        return str(val).strip().upper()
    return str(val).strip().upper()


def _cal_valor_ok(v, es_saca):
    """True si el valor del calendario es RECONOCIDO. Solo lo establecido:
      - Normal: D, N, F (y vacio).
      - Sacafranco: ademas cualquier token valido (D/N+codigo, DB, NB, F...).
    Todo lo demas (T, numeros, #REF!, M, V, S, L, J...) => NO reconocido (se ignora)."""
    v = (v or '').strip().upper()
    if not v:
        return True
    if v in ('D', 'N', 'F'):
        return True
    if es_saca:
        try:
            from .asignacion_semanal_views import _parse_sacafranco_token
            return _parse_sacafranco_token(v)[0] != 'invalid'
        except Exception:
            return False
    return False


def _detect_month_year_from_sheet(rows):
    month_map = {
        'ENERO': 1,
        'FEBRERO': 2,
        'MARZO': 3,
        'ABRIL': 4,
        'MAYO': 5,
        'JUNIO': 6,
        'JULIO': 7,
        'AGOSTO': 8,
        'SEPTIEMBRE': 9,
        'SETIEMBRE': 9,
        'OCTUBRE': 10,
        'NOVIEMBRE': 11,
        'DICIEMBRE': 12,
    }
    for row in rows[:30]:
        for cell in row:
            if not cell:
                continue
            text = str(cell).strip().upper()
            match = re.search(r'(ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|SETIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE)\s+(\d{4})', text)
            if match:
                month = month_map.get(match.group(1))
                year = int(match.group(2))
                if month:
                    return year, month
    return None, None


def compute_horas(hora_ingreso, hora_salida):
    if not hora_ingreso or not hora_salida:
        return 12
    base = datetime(2000, 1, 1)
    dt_in = datetime.combine(base.date(), hora_ingreso)
    dt_out = datetime.combine(base.date(), hora_salida)
    if dt_out <= dt_in:
        dt_out += timedelta(days=1)
    hours = (dt_out - dt_in).total_seconds() / 3600.0
    if hours <= 0:
        return 12
    return int(round(hours))


def _get_or_create_horario(hora_ingreso, hora_salida):
    """get_or_create seguro para Horario: si hay duplicados (misma hora
    ingreso/salida), usa el primero en vez de reventar con MultipleObjectsReturned."""
    h = Horario.objects.filter(
        hora_ingreso=hora_ingreso, hora_salida=hora_salida
    ).order_by('id').first()
    if h:
        return h, False
    return Horario.objects.create(hora_ingreso=hora_ingreso, hora_salida=hora_salida), True


def _norm_puesto(s):
    """Normaliza el nombre de un puesto para comparar sin importar
    mayus/minus, acentos ni espacios extra."""
    import unicodedata
    t = str(s or '').strip().upper()
    t = ''.join(c for c in unicodedata.normalize('NFKD', t) if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', t)


_CLI_STOPWORDS = {'DE', 'DEL', 'LA', 'LAS', 'LOS', 'EL', 'Y', 'SA', 'CIA', 'SAC', 'EP', 'CIA.'}


def _cli_norm_alnum(s):
    """Cliente normalizado a solo alfanumerico (sin espacios/acentos/simbolos)."""
    import unicodedata
    t = str(s or '').strip().upper()
    t = ''.join(c for c in unicodedata.normalize('NFKD', t) if not unicodedata.combining(c))
    return re.sub(r'[^A-Z0-9]', '', t)


def _cli_tokens(s):
    import unicodedata
    t = str(s or '').strip().upper()
    t = ''.join(c for c in unicodedata.normalize('NFKD', t) if not unicodedata.combining(c))
    return [w for w in re.split(r'[^A-Z0-9]+', t) if w and w not in _CLI_STOPWORDS]


def _cliente_coincide(excel_cli, db_cli):
    """Compara el cliente del Excel contra el de la instalacion de forma TOLERANTE
    para no dar falsos positivos por variaciones de nombre. Coincide si:
    uno contiene al otro, o son muy similares (typos), o comparten un token fuerte,
    o uno es la SIGLA del otro (ej. ATD = Agencia Transito Duran). Solo devuelve
    False cuando son clientes realmente distintos."""
    import difflib
    a, b = _cli_norm_alnum(excel_cli), _cli_norm_alnum(db_cli)
    if not a or not b:
        return True  # sin dato para comparar -> no bloquea
    if a in b or b in a:
        return True
    if difflib.SequenceMatcher(None, a, b).ratio() >= 0.62:
        return True
    ta, tb = _cli_tokens(excel_cli), _cli_tokens(db_cli)
    for x in ta:
        for y in tb:
            if x == y or (len(x) >= 4 and len(y) >= 4
                          and difflib.SequenceMatcher(None, x, y).ratio() >= 0.8):
                return True
    # sigla: una cadena = iniciales de las palabras de la otra (ej. UPS = Universidad
    # Politecnica Salesiana). Se prueba con el texto completo Y con cada token (por si
    # viene la sigla junto al nombre de la instalacion, ej. "UPS CENTENARIO").
    ini_a = ''.join(w[0] for w in ta if w)
    ini_b = ''.join(w[0] for w in tb if w)
    cands_a = [_cli_norm_alnum(excel_cli)] + ta
    cands_b = [_cli_norm_alnum(db_cli)] + tb
    if ini_b and any(c == ini_b for c in cands_a):
        return True
    if ini_a and any(c == ini_a for c in cands_b):
        return True
    return False


def _get_or_create_puesto(instalacion, nombre, defaults=None):
    """get_or_create seguro para Puesto: empareja por nombre normalizado
    (ignora mayus/espacios/acentos) para NO crear duplicados, y tolera duplicados
    existentes usando el primero en vez de reventar con MultipleObjectsReturned."""
    nombre = (nombre or '').strip()
    p = Puesto.objects.filter(instalacion=instalacion, nombre=nombre).order_by('id').first()
    if p:
        return p, False
    target = _norm_puesto(nombre)
    for cand in Puesto.objects.filter(instalacion=instalacion).order_by('id'):
        if _norm_puesto(cand.nombre) == target:
            return cand, False
    return Puesto.objects.create(instalacion=instalacion, nombre=nombre, **(defaults or {})), True


def _get_puesto(instalacion, nombre):
    """Busca un Puesto por nombre (normalizado, ignora mayus/espacios/acentos).
    NO crea. Devuelve el puesto o None. (Regla: el import no crea puestos.)"""
    nombre = (nombre or '').strip()
    p = Puesto.objects.filter(instalacion=instalacion, nombre=nombre).order_by('id').first()
    if p:
        return p
    target = _norm_puesto(nombre)
    for cand in Puesto.objects.filter(instalacion=instalacion).order_by('id'):
        if _norm_puesto(cand.nombre) == target:
            return cand
    return None


def _get_or_create_sacafranco_fila(persona, mes, anio, orden):
    """get_or_create seguro para SacafrancoFila. Si ya existen duplicados de
    imports viejos (misma persona/mes/anio), usa el PRIMERO y BORRA los sobrantes
    (con su calendario), evitando MultipleObjectsReturned y limpiando lo duplicado."""
    from ..models import SacafrancoFila, SacafrancoFilaSemanal
    filas = list(SacafrancoFila.objects.filter(persona=persona, mes=mes, anio=anio).order_by('id'))
    if filas:
        principal = filas[0]
        if len(filas) > 1:
            extra_ids = [f.id for f in filas[1:]]
            SacafrancoFilaSemanal.objects.filter(sacafranco_fila_id__in=extra_ids).delete()
            SacafrancoFila.objects.filter(id__in=extra_ids).delete()
        return principal, False
    return SacafrancoFila.objects.create(persona=persona, mes=mes, anio=anio, orden=orden), True


def _meses_proyeccion(request):
    """Cuantos meses se proyecta el patron hacia adelante. Por defecto 36.
    Se ajusta con ?meses=N (0 = solo el mes importado, sin proyeccion). Tope 60."""
    try:
        val = request.GET.get('meses') or request.POST.get('meses')
    except Exception:
        val = None
    try:
        n = int(val)
    except (TypeError, ValueError):
        n = 36
    return max(0, min(n, 60))


def _meses_sync_saca(request):
    """Cuantos meses proyectados del SACAFRANCO se reflejan en Reporte de
    Asistencia y Consolidado. El CALENDARIO del sacafranco si se proyecta a
    todos los meses (barato), pero sincronizar Reporte/Consolidado por cada
    semana de 36 meses es carisimo y operativamente inutil tan a futuro.
    El mes base SIEMPRE se sincroniza; esto controla cuantos meses MAS.
    Por defecto 2 (mes actual + proximos). Se ajusta con ?meses_sync=N."""
    try:
        val = request.GET.get('meses_sync') or request.POST.get('meses_sync')
    except Exception:
        val = None
    try:
        n = int(val)
    except (TypeError, ValueError):
        n = 2
    return max(0, min(n, 60))


def _quiere_desactivar_sobrantes(request):
    """Por defecto el import NO desactiva a nadie: solo AGREGA/ACTUALIZA.
    La desactivacion de quien ya no aparece en el archivo solo ocurre si se pide
    explicitamente con ?desactivar_sobrantes=1 (protege los datos existentes)."""
    try:
        val = request.GET.get('desactivar_sobrantes') or request.POST.get('desactivar_sobrantes') or ''
    except Exception:
        val = ''
    return str(val).strip().lower() in ('1', 'true', 'si', 'on')


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def importar_puestos_asignaciones(request):
    # si el usuario no tiene permiso de importacion, retorna un jsonreponse con error 403
    if not request.user.has_perm('CoreFisica.import_puestos_asignaciones'):
        return JsonResponse({'error': 'No Autorizado'}, status=403)
    # se espera  un respuesta con un archivo excel
    upload = request.FILES.get('file')
    # si no se recibe un archivo, retorna un jsonresponse con error 400
    if not upload:
        return JsonResponse({'error': 'Falta el archivo (campo file)'}, status=400)

    cliente_id = request.GET.get('cliente_id')
    # si se recibe cliente_id, intenta convertirlo a entero, si no es posible, lo ignora y continua sin filtro de cliente
    if cliente_id:
        try:
            cliente_id = int(cliente_id)
        except (TypeError, ValueError):
            cliente_id = None

    try:
        # intenta abrir el archivo excel usando openpyxl, si no es posible, retorna un jsonresponse con error 400
        wb = load_workbook(upload, read_only=False, data_only=True)
    except Exception as exc:
        return JsonResponse({'error': f'No se pudo abrir el archivo: {exc}'}, status=400)

    # Si es el FORMATO REPORTE (el que genera Descargar), usar el importador dedicado
    try:
        if es_formato_reporte(wb):
            resumen = importar_formato_reporte(request, wb, cliente_id)
            return JsonResponse(resumen, status=200)
    except Exception:
        logger.exception('Error importando formato reporte')
        return JsonResponse({'error': 'No se pudo importar el formato reporte'}, status=500)

    # ws es la hoja activa del libro excel
    ws = wb.active
    # rows es una lista de filas de la hoja excel, cada fila es una tupla de valores de celdas, usando values_only=True para obtener solo los valores sin formato
    rows = list(ws.iter_rows(values_only=True))
    # si no se encuentran filas en el archivo, retorna un jsonresponse con error 400
    if not rows:
        return JsonResponse({'error': 'El archivo esta vacio'}, status=400)
    # se buscan las columnas obligatorias en las primeras 15 filas del archivo, normalizando los encabezados y comparandolos con el HEADER_MAP para encontrar los indices de las columnas necesarias, si no se encuentran todas las columnas obligatorias, retorna un jsonresponse con error 400 indicando las columnas faltantes y los encabezados detectados
    header_idx = None
    header_row_num = None
    headers_raw = []
    for ridx, row in enumerate(rows[:15]):
        candidate_headers = [normalize_header(h) for h in row]
        tmp_idx = {HEADER_MAP[h]: i for i, h in enumerate(candidate_headers) if h in HEADER_MAP}
        headers_raw = candidate_headers
        required = {'puesto', 'cedula', 'hora_ingreso', 'hora_salida'}
        alt_required = {'dias', 'horas'}
        if required.issubset(set(tmp_idx.keys())):
            if not alt_required.intersection(set(tmp_idx.keys())):
                continue
            header_idx = tmp_idx
            header_row_num = ridx
            break
    # si no se encuentran todas las columnas obligatorias, retorna un jsonresponse con error 400 indicando las columnas faltantes y los encabezados detectados
    if header_idx is None:
        return JsonResponse({
            'error': 'Faltan columnas obligatorias: PUESTO, CEDULA, HORA INGRESO, HORA SALIDA y (DIAS o HORAS)',
            'headers_detectados': headers_raw
        }, status=400)
    # detectar si la fila siguiente contiene los numeros de dias (1..31) para el calendario
    day_header_row_num = None
    if header_row_num is not None and (header_row_num + 1) < len(rows):
        candidate = rows[header_row_num + 1]
        candidate_headers = [normalize_header(h) for h in candidate]
        day_idx = {HEADER_MAP[h]: i for i, h in enumerate(candidate_headers) if h in HEADER_MAP and h.startswith(tuple(str(d) for d in range(1, 32)))}
        # si hay suficientes columnas de dias, usar esta fila como encabezado de calendario
        if len(day_idx) >= 10:
            day_header_row_num = header_row_num + 1
            for key, idx in day_idx.items():
                if key not in header_idx:
                    header_idx[key] = idx
    sheet_year, sheet_month = _detect_month_year_from_sheet(rows)
    req_month = request.GET.get('mes') or request.POST.get('mes')
    req_year = request.GET.get('anio') or request.POST.get('anio')
    try:
        req_month = int(req_month) if req_month is not None else None
    except (TypeError, ValueError):
        req_month = None
    try:
        req_year = int(req_year) if req_year is not None else None
    except (TypeError, ValueError):
        req_year = None
    if req_month is not None and not (1 <= req_month <= 12):
        req_month = None
    if req_year is not None and req_year < 1:
        req_year = None
    meses_proy = _meses_proyeccion(request)
    # se inicializa un diccionario de resumen para llevar conteo de filas procesadas, personas/puestos/horarios/asignaciones creadas o actualizadas, y errores encontrados durante el proceso de importacion
    resumen = {
        'total_filas': 0,
        'filas_validas': 0,
        'personas_creadas': 0,
        'puestos_creados': 0,
        'horarios_creados': 0,
        'asignaciones_creadas': 0,
        'asignaciones_actualizadas': 0,
        'errores': [],
    }
    # start_row se establece como la fila siguiente a la fila de encabezado encontrada, para comenzar a procesar los datos desde esa fila en adelante
    start_row = (header_row_num or 0) + 1
    if day_header_row_num is not None and day_header_row_num >= start_row:
        start_row = day_header_row_num + 1

    try:
        touched_asig_ids = set()
        touched_dates = set()
        regenerated_future_signatures = {}
        # Personas importadas por (puesto, mes, anio) -> para reemplazar sin duplicar
        from collections import defaultdict as _dd
        puesto_personas = _dd(set)
        with transaction.atomic():
            for i, row in enumerate(rows[start_row:], start=start_row + 1):
                if not row or all(v is None or str(v).strip() == '' for v in row):
                    continue
                resumen['total_filas'] += 1

                def col(key):
                    idx = header_idx.get(key)
                    return norm(row[idx]) if idx is not None and idx < len(row) else ''

                def col_raw(key):
                    idx = header_idx.get(key)
                    return row[idx] if idx is not None and idx < len(row) else None

                instalacion_id = col('instalacion_id')
                instalacion_nombre = col('instalacion')
                cliente_ruc = col('cliente_ruc')
                cliente_nombre = col('cliente')
                puesto_nombre = col('puesto')
                puesto_tipo = col('puesto_tipo')
                cantidad_puestos = col('cantidad_puestos')
                patron_codigo = col('patron')
                patron_id = col('patron_id')

                cedula = normalize_cedula(col('cedula'))
                apellidos = col('apellidos')
                nombres = col('nombres')
                if not apellidos and not nombres:
                    full_name = col('apellidos_nombres')
                    if full_name:
                        tokens = [t for t in re.split(r'\s+', full_name.strip()) if t]
                        if len(tokens) >= 3:
                            apellidos = ' '.join(tokens[:2])
                            nombres = ' '.join(tokens[2:])
                        elif len(tokens) == 2:
                            apellidos = tokens[0]
                            nombres = tokens[1]
                        elif len(tokens) == 1:
                            apellidos = tokens[0]
                            nombres = tokens[0]
                persona_tipo = col('tipo')

                hora_ingreso = parse_excel_time(col_raw('hora_ingreso'))
                hora_salida = parse_excel_time(col_raw('hora_salida'))
                horas_raw = col('horas')
                turno = parse_turno(col('turno'))
                dias_groups = parse_dias_groups(col('dias'))
                compact_groups = []
                if horas_raw and re.search(r'[A-Z]', str(horas_raw).upper()):
                    compact_groups = parse_compact_horas_turno_dias(horas_raw)
                    if compact_groups and not dias_groups:
                        dias_groups = [g['dias'] for g in compact_groups if g.get('dias')]
                fecha = parse_excel_date(col_raw('fecha')) if header_idx.get('fecha') is not None else None
                # si el puesto_nombre se agrega un mensaje de error al resumen indicando que el puesto esta vacio y se continua con la siguiente fila sin procesar la fila actual, ya que el puesto es un campo obligatorio para crear o actualizar una asignacion
                if not puesto_nombre:
                    resumen['errores'].append(f'Fila {i}: puesto vacio')
                    continue
                # si la cedula se agrega un mensaje de error al resumen indicando que la cedula esta vacia o invalida y se continua con la siguiente fila sin procesar la fila actual, ya que la cedula es un campo obligatorio para crear o actualizar una persona y asignacion
                if not cedula:
                    resumen['errores'].append(f'Fila {i}: cedula vacia o invalida')
                    continue
                # si la hora_ingreso o hora_salida es invalida se agrega un mensaje de error al resumen indicando que la hora de ingreso o salida es invalida y se continua con la siguiente fila sin procesar la fila actual, ya que ambos campos son obligatorios para crear o actualizar un horario y asignacion
                if (not hora_ingreso or not hora_salida) and not horas_raw:
                    resumen['errores'].append(f'Fila {i}: hora ingreso/salida invalida')
                    continue
                # si la lista de dias esta vacia se agrega un mensaje de error al resumen indicando que los dias estan vacios y se continua con la siguiente fila sin procesar la fila actual, ya que los dias son un campo obligatorio para crear o actualizar un puesto horario y asignacion
                if not dias_groups:
                    resumen['errores'].append(f'Fila {i}: dias vacio')
                    continue
                # se intenta obtener la instalacion usando el instalacion_id si se proporciona, si no se encuentra la instalacion con ese id, se agrega un mensaje de error al resumen indicando que la instalacion no fue encontrada y se continua con la siguiente fila sin procesar la fila actual, ya que la instalacion es un campo obligatorio para crear o actualizar un puesto y asignacion. Si no se proporciona el instalacion_id, se intenta buscar la instalacion usando el nombre de la instalacion y opcionalmente filtrando por cliente usando el cliente_id, cliente_ruc o cliente_nombre si se proporcionan, si se encuentran varias instalaciones que coinciden con el nombre, se agrega un mensaje de error al resumen indicando que hay instalaciones duplicadas y se continua con la siguiente fila sin procesar la fila actual, ya que no se puede determinar a cual instalacion asociar el puesto y asignacion. Si no se encuentra ninguna instalacion que coincida con el nombre (y filtro de cliente), se agrega un mensaje de error al resumen indicando que la instalacion no fue encontrada y se continua con la siguiente fila sin procesar la fila actual, ya que la instalacion es un campo obligatorio para crear o actualizar un puesto y asignacion
                instalacion = None
                # si la instalacion_id se proporciona
                if instalacion_id:
                    try:
                        # se intenta obtener la instalacion usando el instalacion_id
                        instalacion = Instalacion.objects.get(id=int(instalacion_id))
                    except (ValueError, Instalacion.DoesNotExist):
                        resumen['errores'].append(f'Fila {i}: instalacion_id no encontrada')
                        continue
                # si la instalacion_id no se proporciona, se intenta buscar la instalacion usando el nombre de la instalacion y opcionalmente filtrando por cliente usando el cliente_id, cliente_ruc o cliente_nombre si se proporcionan
                else:
                    # si el nombre de la instalcion est avacio se agrega un mensaje de error el resumen indicando que la instalacion esta vacia y se continua con la siguiente fila sin procesar la fila actual, ya que la instalacion es un campo obligatorio para crear o actualizar un puesto y asignacion
                    if not instalacion_nombre:
                        # intentar inferir instalacion desde un puesto existente
                        existing_puestos = Puesto.objects.filter(nombre__iexact=puesto_nombre)
                        if existing_puestos.count() == 1:
                            instalacion = existing_puestos.first().instalacion
                        else:
                            resumen['errores'].append(f'Fila {i}: instalacion vacia, use INSTALACION o instale puesto existente')
                            continue
                    # se busca la instalacion usando el nombre de la instalacion
                    inst_qs = Instalacion.objects.filter(nombre__iexact=instalacion_nombre)
                    # si el cliente_id se proporciona, se filtra la instalacion por cliente_id
                    if cliente_id:
                        inst_qs = inst_qs.filter(cliente_id=cliente_id)
                    # si el cliente_ruc se proporciona, se busca el cliente por ruc
                    elif cliente_ruc:
                        cliente = Cliente.objects.filter(ruc=cliente_ruc).first()
                        # si se encuentra el cliente, se filtra la instalacion por cliente
                        if cliente:
                            inst_qs = inst_qs.filter(cliente=cliente)
                    # si el cliente_nombre se proporciona, se busca el cliente por nombre comercial
                    elif cliente_nombre:
                        cliente = Cliente.objects.filter(nombre_comercial__iexact=cliente_nombre).first()
                        if cliente:
                            inst_qs = inst_qs.filter(cliente=cliente)
                    # si la cantidad de instalaciones que coinciden con el nombre (y filtro de cliente) es mayor a 1, se agrega un mensaje de error al resumen indicando que hay instalaciones duplicadas y se continua con la siguiente fila sin procesar la fila actual, ya que no se puede determinar a cual instalacion asociar el puesto y asignacion
                    if inst_qs.count() > 1:
                        resumen['errores'].append(f'Fila {i}: instalacion duplicada, use instalacion_id o cliente')
                        continue
                    instalacion = inst_qs.first()
                    if not instalacion:
                        resumen['errores'].append(f'Fila {i}: instalacion no encontrada')
                        continue
                # REGLA: el import NO crea personas. Solo referencia por cedula.
                persona = Persona.objects.filter(cedula=cedula).first()
                if not persona:
                    resumen['errores'].append(f'Fila {i}: persona con cedula {cedula} no esta registrada — no se importa')
                    continue
                # se obtiene o crea un horario usando la hora_ingreso y hora_salida
                horario, created_horario = _get_or_create_horario(hora_ingreso, hora_salida)
                # si se creo un nuevo horario, se incrementa el contador de horarios_creados en el resumen
                if created_horario:
                    resumen['horarios_creados'] += 1

                try:
                    #se intenta convertir la cantidad de puestos a entero, si no es posible, se establece en 1 por defecto, ya que la cantidad de puestos es un campo opcional para crear o actualizar un puesto y asignacion, pero si se proporciona debe ser un numero entero positivo
                    cantidad_int = int(cantidad_puestos) if cantidad_puestos else 1
                except (TypeError, ValueError):
                    cantidad_int = 1
                if cantidad_int < 1:
                    cantidad_int = 1
                # REGLA: el import NO crea puestos. Debe existir (configurado antes con
                # su cantidad de guardias). Si no existe -> avisa y salta.
                puesto = _get_puesto(instalacion, puesto_nombre)
                if not puesto:
                    resumen['errores'].append(
                        f"Fila {i}: el puesto '{puesto_nombre}' no existe en la instalacion — no se importa"
                    )
                    continue
                # se actualizan o crean los puestos horarios para el puesto usando la lista de dias, hora_ingreso, hora_salida y turno proporcionados, ya que los puestos horarios son necesarios para crear o actualizar una asignacion y se asume que el horario y turno proporcionados aplican para todos los dias indicados
                horas_groups = parse_hours_groups(horas_raw)
                turnos_groups = parse_turno_groups(col('turno'))
                if compact_groups:
                    horas_groups = [g['hours'] for g in compact_groups]
                    turnos_groups = [g['turno'] for g in compact_groups]
                default_horas = compute_horas(hora_ingreso, hora_salida)
                for idx, dias in enumerate(dias_groups):
                    horas = horas_groups[idx] if idx < len(horas_groups) else default_horas
                    turno_val = turnos_groups[idx] if idx < len(turnos_groups) else turno
                    if not turno_val:
                        turno_val = turno
                    for dia in dias:
                        PuestoHorario.objects.update_or_create(
                            puesto=puesto,
                            dia=dia,
                            defaults={'horas': min(max(horas, 0), 24), 'turno': turno_val}
                        )

                try:
                    # se sincroniza el puesto con los horarios usando el metodo sync_from_horarios, para asegurar que el puesto tenga el horario correcto basado en los puestos horarios asociados, ya que el horario es un campo necesario para crear o actualizar una asignacion y se asume que el horario del puesto debe reflejar los horarios definidos en los puestos horarios asociados
                    puesto.sync_from_horarios()
                    puesto.save()
                except Exception:
                    pass
                
                # se actualizan o crean las asignaciones para la persona, puesto, instalacion, horario
                if fecha:
                    ref_date = fecha
                elif req_year and req_month:
                    ref_date = date(req_year, req_month, 1)
                elif sheet_year and sheet_month:
                    ref_date = date(sheet_year, sheet_month, 1)
                else:
                    ref_date = date.today()
                mes = ref_date.month
                anio = ref_date.year
                patron_obj = None
                if patron_id:
                    try:
                        patron_obj = PatronAsignacion.objects.filter(id=int(patron_id)).first()
                    except (TypeError, ValueError):
                        patron_obj = None
                if not patron_obj and patron_codigo:
                    token = str(patron_codigo).strip()
                    patron_obj = PatronAsignacion.objects.filter(codigo=token).first()
                    if patron_obj is None and token.isdigit():
                        patron_obj = PatronAsignacion.objects.filter(id=int(token)).first()
                # asig se actualiza o crea una asignacion usando persona + puesto + mes + anio como claves
                defaults = {
                    'cliente': instalacion.cliente,
                    'instalacion': instalacion,
                    'puesto': puesto,
                    'horario': horario,
                    'fecha': None,
                    'patronAsignacion': patron_obj,
                    'estado': 'ACTIVO',
                    'publicada_calendario': True,
                    'recurring': True,
                    'start_date': date(ref_date.year, ref_date.month, 1),
                    'end_date': None,
                }
                if not patron_obj:
                    defaults['patronAsignacion'] = None

                asig, created_asig = Asignacion.objects.update_or_create(
                    persona=persona,
                    mes=mes,
                    anio=anio,
                    defaults=defaults,
                )
                touched_asig_ids.add(asig.id)
                touched_dates.add(ref_date)
                puesto_personas[(puesto.id, mes, anio)].add(persona.id)
                # si se creo una nueva asignacion
                if created_asig:
                    resumen['asignaciones_creadas'] += 1
                else:
                    resumen['asignaciones_actualizadas'] += 1

                # aplicar calendario manual desde columnas 1..31 si existen
                calendar_updates = {}
                raw_month_values = []
                days_in_month = (date(ref_date.year, ref_date.month + 1, 1) - timedelta(days=1)).day if ref_date.month < 12 else 31
                has_calendar_values = False
                for day_num in range(1, 32):
                    if day_num > days_in_month:
                        continue
                    key = f'day_{day_num}'
                    idx = header_idx.get(key)
                    raw_val = row[idx] if (idx is not None and idx < len(row)) else None
                    val = parse_calendar_value(raw_val)
                    raw_month_values.append(val)
                    if not val:
                        continue
                    has_calendar_values = True

                if has_calendar_values:
                    # Tomar solo valores no vacios como ciclo base y completar todo el primer mes.
                    seq_values_future = [v for v in raw_month_values if str(v or '').strip()]
                    cycle_len = len(seq_values_future)
                    if cycle_len <= 0:
                        raise ValueError('Secuencia vacia')

                    filled_month_values = []
                    sequence_index = 0
                    for day_num in range(1, days_in_month + 1):
                        explicit = raw_month_values[day_num - 1] if (day_num - 1) < len(raw_month_values) else ''
                        if str(explicit or '').strip():
                            val = explicit
                        else:
                            val = seq_values_future[sequence_index % cycle_len]
                        filled_month_values.append(val)
                        sequence_index += 1

                    for day_num, val in enumerate(filled_month_values, start=1):
                        day_date = date(ref_date.year, ref_date.month, day_num)
                        week_index = (day_num - 1) // 7
                        week_start = date(ref_date.year, ref_date.month, 1) + timedelta(days=week_index * 7)
                        day_field = ['mon','tue','wed','thu','fri','sat','sun'][day_date.weekday()]
                        calendar_updates.setdefault(week_start, {})[day_field] = val

                    week_count = ((days_in_month - 1) // 7) + 1
                    month_start = date(ref_date.year, ref_date.month, 1)
                    target_week_starts = [month_start + timedelta(days=7 * i) for i in range(week_count)]
                    if target_week_starts:
                        AsignacionSemanal.objects.filter(
                            asignacion_id=asig.id,
                            week_start__in=target_week_starts,
                        ).delete()
                    for ws, day_map in calendar_updates.items():
                        defaults = {**day_map, 'puesto_id': puesto.id}
                        obj, created = AsignacionSemanal.objects.get_or_create(
                            asignacion_id=asig.id,
                            week_start=ws,
                            defaults=defaults,
                        )
                        if not created:
                            changed = False
                            for d_key, d_val in day_map.items():
                                setattr(obj, d_key, d_val)
                                changed = True
                            if getattr(obj, 'puesto_id', None) is None:
                                obj.puesto_id = puesto.id
                                changed = True
                            if changed:
                                obj.save()
                    # Regenerar los proximos 24 meses para continuar la secuencia.
                    regen_key = (persona.id, puesto.id, ref_date.year, ref_date.month)
                    seq_signature = tuple(filled_month_values)
                    if regenerated_future_signatures.get(regen_key) == seq_signature:
                        resumen['filas_validas'] += 1
                        continue
                    regenerated_future_signatures[regen_key] = seq_signature
                    try:
                        def add_months(d, months):
                            year = d.year + (d.month - 1 + months) // 12
                            month = (d.month - 1 + months) % 12 + 1
                            return date(year, month, 1)

                        # Continuar en el mes siguiente justo donde terminó el primer mes completado.
                        base_month_start = date(ref_date.year, ref_date.month, 1)
                        for offset in range(1, meses_proy + 1):
                            target_start = add_months(base_month_start, offset)
                            target_year = target_start.year
                            target_month = target_start.month
                            if target_month == 12:
                                target_end = date(target_year + 1, 1, 1) - timedelta(days=1)
                            else:
                                target_end = date(target_year, target_month + 1, 1) - timedelta(days=1)
                            target_asig, _ = Asignacion.objects.update_or_create(
                                persona=persona,
                                mes=target_month,
                                anio=target_year,
                                defaults={
                                    'cliente': instalacion.cliente,
                                    'instalacion': instalacion,
                                    'puesto': puesto,
                                    'horario': horario,
                                    'fecha': None,
                                    'patronAsignacion': patron_obj,
                                    'estado': 'ACTIVO',
                                    'publicada_calendario': True,
                                    'recurring': True,
                                    'start_date': date(target_year, target_month, 1),
                                    'end_date': None,
                                }
                            )
                            days_in_target = target_end.day
                            weekly_payload = {}
                            for day_num in range(1, days_in_target + 1):
                                day_date = date(target_year, target_month, day_num)
                                week_index = (day_num - 1) // 7
                                week_start = date(target_year, target_month, 1) + timedelta(days=week_index * 7)
                                day_field = ['mon','tue','wed','thu','fri','sat','sun'][day_date.weekday()]
                                val = seq_values_future[sequence_index % cycle_len]
                                sequence_index += 1
                                if week_start not in weekly_payload:
                                    weekly_payload[week_start] = {}
                                weekly_payload[week_start][day_field] = val

                            if weekly_payload:
                                target_week_starts = list(weekly_payload.keys())
                                AsignacionSemanal.objects.filter(
                                    asignacion_id=target_asig.id,
                                    week_start__in=target_week_starts
                                ).delete()
                                bulk_rows = []
                                for ws_key, day_map in weekly_payload.items():
                                    row_data = {
                                        'asignacion_id': target_asig.id,
                                        'week_start': ws_key,
                                        'puesto_id': puesto.id,
                                        'mon': day_map.get('mon', ''),
                                        'tue': day_map.get('tue', ''),
                                        'wed': day_map.get('wed', ''),
                                        'thu': day_map.get('thu', ''),
                                        'fri': day_map.get('fri', ''),
                                        'sat': day_map.get('sat', ''),
                                        'sun': day_map.get('sun', ''),
                                    }
                                    bulk_rows.append(AsignacionSemanal(**row_data))
                                AsignacionSemanal.objects.bulk_create(bulk_rows)
                    except Exception:
                        logger.exception('Error regenerando secuencia de los proximos meses')
                else:
                    try:
                        from .asignacion_views import _rebuild_asignacion_semanal
                        _rebuild_asignacion_semanal(asig, force_all=created_asig)
                    except Exception:
                        logger.exception('Error reconstruyendo asignacion semanal')

                resumen['filas_validas'] += 1

            # Por defecto el import NO desactiva a nadie (solo agrega/actualiza).
            # Solo con ?desactivar_sobrantes=1 se desactiva a quien ya no viene.
            if _quiere_desactivar_sobrantes(request):
                for (pid, pmes, panio), persona_ids in puesto_personas.items():
                    sobrantes = Asignacion.objects.filter(
                        puesto_id=pid, mes=pmes, anio=panio, estado='ACTIVO'
                    ).exclude(persona_id__in=persona_ids)
                    for o in sobrantes:
                        o.estado = 'INACTIVO'
                        o.save(update_fields=['estado'])
                        ReporteAsistencia.objects.filter(asignacion=o).update(
                            estado='TURNO', estado_asistencia='', reemplazo=None,
                            descripcion=None, row_color=None
                        )

            if touched_asig_ids:
                # Asegurar ReporteAsistencia base para las asignaciones importadas
                asig_qs = Asignacion.objects.select_related(
                    'persona', 'cliente', 'instalacion', 'puesto', 'horario'
                ).filter(id__in=touched_asig_ids)
                for asig in asig_qs:
                    try:
                        reporte, _ = ReporteAsistencia.objects.get_or_create(asignacion=asig)
                        reporte.persona = asig.persona
                        reporte.cliente = asig.cliente
                        reporte.instalacion = asig.instalacion
                        reporte.puesto = asig.puesto
                        reporte.horario = asig.horario
                        reporte.puesto_tipo = getattr(asig.puesto, 'tipo', None) if asig.puesto else None
                        reporte.save()
                    except Exception:
                        pass

                # Actualizar resumen de consolidado para fechas importadas (turnos diurno/nocturno)
                try:
                    from .reporte_asistencia_views import _build_reporte_asistencia_data
                    from .consolidado_views import _build_resumen_manual
                    for ref_date in touched_dates:
                        for turno_val in ('Diurno', 'Nocturno'):
                            rows_data = _build_reporte_asistencia_data(
                                fecha=ref_date.isoformat(),
                                turno=turno_val
                            )
                            _build_resumen_manual(ref_date, turno_val, rows_data)
                except Exception:
                    pass

        return JsonResponse(resumen, status=200)
    except Exception:
        logger.exception('Error importando puestos/asignaciones')
        return JsonResponse({'error': 'No se pudo importar puestos/asignaciones'}, status=500)


# ============================================================================
# Importación del FORMATO REPORTE (el que genera el botón Descargar de Asignaciones)
# ============================================================================
def _rep_norm(v):
    import unicodedata
    if v is None:
        return ''
    t = str(v).strip().upper().replace('.', ' ').replace('_', ' ')
    t = ''.join(c for c in unicodedata.normalize('NFKD', t) if not unicodedata.combining(c))
    return ' '.join(t.split())


def _periodo_minimo(seq):
    """Detecta el período mínimo EXACTO de una secuencia (ej. DDDNNNF -> 7).
    Si no hay período exacto más corto, devuelve la secuencia completa."""
    n = len(seq)
    if n == 0:
        return seq
    for p in range(1, n):
        if all(seq[i] == seq[i % p] for i in range(n)):
            return seq[:p]
    return list(seq)


def _ciclo_para_continuar(seq):
    """Devuelve el ciclo a usar para continuar el patrón en meses futuros.
    1) Período exacto si existe. 2) Si la fila es irregular, prueba longitudes
    típicas (7,6,5,8,14,10,...) y usa la que mejor encaje (>=85%). 3) Si nada
    encaja, repite el mes completo."""
    n = len(seq)
    if n == 0:
        return list(seq)
    # 1) período exacto
    exacto = _periodo_minimo(seq)
    if len(exacto) < n:
        return exacto
    # 2) mejor aproximado entre longitudes típicas de rotación
    candidatos = [7, 6, 5, 8, 14, 10, 12, 4, 3, 2, 21, 28]
    mejor_p, mejor_ratio = None, 0.0
    for p in candidatos:
        if p >= n:
            continue
        aciertos = sum(1 for i in range(n) if seq[i] == seq[i % p])
        ratio = aciertos / n
        if ratio > mejor_ratio:
            mejor_ratio, mejor_p = ratio, p
    if mejor_p and mejor_ratio >= 0.85:
        return seq[:mejor_p]
    # 3) sin patrón claro -> repetir el mes completo
    return list(seq)


def _rep_detectar_columnas(rows):
    for ri, row in enumerate(rows[:20]):
        H = {_rep_norm(c): j for j, c in enumerate(row) if c is not None and str(c).strip()}
        if 'CLIENTE' in H and 'PUESTO' in H and 'CEDULA' in H and ('H INGRESO' in H or 'H SALIDA' in H):
            dias = []
            for j, c in enumerate(row):
                s = str(c).strip() if c is not None else ''
                if s.isdigit() and 1 <= int(s) <= 31:
                    dias.append(j)
            col = {
                'ing': H.get('H INGRESO'), 'sal': H.get('H SALIDA'),
                'cli': H['CLIENTE'], 'pue': H['PUESTO'], 'resumen': H.get('TIPO'),
                'ced': H['CEDULA'], 'nombre': H.get('APELLIDOS Y NOMBRES'),
                'nominativo': H['CLIENTE'] - 1, 'dias': dias,
            }
            # Si no hay encabezado "TIPO", inferir la columna del resumen del puesto:
            # entre PUESTO y CEDULA, la columna cuyos datos tengan patrón de horas
            # (24H, 4.5HDV, 5HDLJ...). Evita confundirla con el '#'/orden (solo dígitos).
            if col['resumen'] is None and col['pue'] is not None and col['ced'] is not None:
                lo, hi = col['pue'] + 1, col['ced']
                hpat = re.compile(r'\d+(?:[.,]\d+)?\s*H', re.IGNORECASE)
                conteo = {}
                for r2 in rows[ri + 1:ri + 60]:
                    if not r2:
                        continue
                    for j in range(lo, hi):
                        if j < len(r2) and r2[j] is not None and hpat.search(str(r2[j])):
                            conteo[j] = conteo.get(j, 0) + 1
                if conteo:
                    col['resumen'] = max(conteo, key=conteo.get)

            # Deteccion de BLOQUES de mes: algunas hojas (p.ej. PUESTOS GENERALES)
            # traen DOS meses pegados lado a lado (JULIO en una columna, AGOSTO en
            # otra), con 62 columnas de dias. Sin esto el importador tomaria las
            # primeras 31 columnas (mes equivocado) y desalinearia el calendario.
            # Se buscan los nombres de mes en las filas encima del encabezado de
            # numeros y se parte 'dias' en bloques por columna de inicio.
            _MESES_BLOQUE = {
                'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4, 'MAYO': 5,
                'JUNIO': 6, 'JULIO': 7, 'AGOSTO': 8, 'SEPTIEMBRE': 9,
                'SETIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12,
            }
            by_col = {}
            for rr in range(max(0, ri - 4), ri):
                for j, c in enumerate(rows[rr]):
                    if c is None:
                        continue
                    txt = str(c).strip().upper()
                    for name, mnum in _MESES_BLOQUE.items():
                        if name in txt:
                            by_col.setdefault(j, mnum)
                            break
            bloques = []
            mcols = sorted(by_col.items())
            for idx, (start_col, mnum) in enumerate(mcols):
                end_col = mcols[idx + 1][0] if idx + 1 < len(mcols) else 10 ** 9
                bcols = [d for d in dias if start_col <= d < end_col]
                if bcols:
                    bloques.append({'mes': mnum, 'dias': bcols})
            col['bloques'] = bloques
            return ri, col
    return None, None


def es_formato_reporte(wb):
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        ri, _ = _rep_detectar_columnas(rows)
        if ri is not None:
            return True
    return False


def importar_formato_reporte(request, wb, cliente_id_filter=None):
    from .asignacion_semanal_views import _sync_sacafranco_to_reporte_y_consolidado, _validate_sacafranco_tokens, _purge_auto_sacafranco_persona
    from ..models import SacafrancoFilaSemanal, SacafrancoFila
    from ..audit import suppress_audit

    resumen = {
        'total_filas': 0, 'filas_validas': 0, 'personas_creadas': 0,
        'puestos_creados': 0, 'horarios_creados': 0, 'asignaciones_creadas': 0,
        'asignaciones_actualizadas': 0, 'sacafranco_creados': 0, 'errores': [],
    }

    req_month = request.GET.get('mes') or request.POST.get('mes')
    req_year = request.GET.get('anio') or request.POST.get('anio')
    try:
        req_month = int(req_month) if req_month else None
    except (TypeError, ValueError):
        req_month = None
    try:
        req_year = int(req_year) if req_year else None
    except (TypeError, ValueError):
        req_year = None
    meses_proy = _meses_proyeccion(request)
    meses_sync_saca = _meses_sync_saca(request)

    touched_asig_ids = set()
    touched_dates = set()
    puesto_personas = {}
    touched_saca_ids = set()   # SacafrancoFila que SI vinieron en este Excel
    touched_periodos = set()   # (mes, anio) que trajo el Excel
    orden_counter = 0  # orden de presentación según el orden del Excel (igual en todos los meses)
    WEEK_KEYS = ['mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun']

    objetivo_mes = req_month or date.today().month

    with suppress_audit(), transaction.atomic():
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            ri, col = _rep_detectar_columnas(rows)
            if ri is None:
                continue
            # Si la hoja proviene del DESCARGABLE (hoja 'DATOS'), solo se actualiza
            # el mes exportado: NO se proyectan 36 meses (seria lentisimo y pisaria
            # meses futuros). Las plantillas normales si proyectan el patron.
            proyectar = not str(ws.title or '').strip().upper().startswith('DATOS')

            # Elegir el BLOQUE de mes a importar. Si la hoja trae varios meses
            # pegados (AGOSTO..DICIEMBRE), se usa: el mes que pida la UI (req_month);
            # si no, el MES ACTUAL si esta entre los bloques; y como ultimo recurso,
            # el ultimo bloque. Asi por defecto se importa el mes en curso, no diciembre.
            bloques = col.get('bloques') or []
            mes_bloque = None
            dias_cols = col['dias']
            if bloques:
                elegido = next((b for b in bloques if b['mes'] == objetivo_mes), None)
                if elegido is None:
                    elegido = bloques[-1]
                dias_cols = elegido['dias']
                mes_bloque = elegido['mes']
                if len(bloques) > 1:
                    resumen['errores'].append(
                        f"Hoja {ws.title}: contiene {len(bloques)} meses "
                        f"({', '.join(str(b['mes']) for b in bloques)}); se importa el "
                        f"mes {mes_bloque} (los demas se ignoran por ser referencia)."
                    )

            sheet_year, sheet_month = _detect_month_year_from_sheet(rows)
            mes = req_month or mes_bloque or sheet_month
            anio = req_year or sheet_year
            # Fallback: detectar nombre de mes suelto (sin año) y usar año actual
            if not mes:
                meses_map = {'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4, 'MAYO': 5,
                             'JUNIO': 6, 'JULIO': 7, 'AGOSTO': 8, 'SEPTIEMBRE': 9, 'SETIEMBRE': 9,
                             'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12}
                for r in rows[:ri + 2]:
                    for c in r:
                        key = _rep_norm(c)
                        if key in meses_map:
                            mes = meses_map[key]
                            break
                    if mes:
                        break
            if not anio:
                anio = date.today().year
            if not mes:
                resumen['errores'].append(f'Hoja {ws.title}: no se detecto el mes')
                continue
            month_start = date(anio, mes, 1)
            days_in_month = (date(anio, mes + 1, 1) - timedelta(days=1)).day if mes < 12 else 31
            carry = {'nominativo': '', 'cli': '', 'pue': '', 'resumen': ''}
            carry_time = {'ing': None, 'sal': None}
            # Vista de ESTA hoja: cantones/clientes de sus asignaciones y los sacafranco
            # que aparecen en ella. Al final de la hoja, a esos sacafranco se les "sella"
            # esta vista (cantones/clientes) para que queden por vista, no globales.
            _sheet_saca_ids = []
            _sheet_cantones = set()
            _sheet_clientes = set()
            _sheet_inst_ids = set()

            # Anti-mezcla: si una persona aparece en VARIAS filas de la hoja (p.ej. GARITA
            # y RETEN), se procesa SOLO la fila con mas dias de calendario (su asignacion
            # real); las demas se saltan para no mezclar/pisar su calendario.
            _cj = col.get('ced')
            best_row_idx = {}
            for _i2, _r2 in enumerate(rows[ri + 1:], start=ri + 2):
                _ced2 = normalize_cedula(norm(_r2[_cj])) if (_cj is not None and _cj < len(_r2)) else ''
                if not _ced2:
                    continue
                _cnt = 0
                for _dj in col['dias']:
                    if _dj < len(_r2) and str(parse_calendar_value(_r2[_dj]) or '').strip():
                        _cnt += 1
                _prev = best_row_idx.get(_ced2)
                if _prev is None or _cnt > _prev[1]:
                    best_row_idx[_ced2] = (_i2, _cnt)

            for i, row in enumerate(rows[ri + 1:], start=ri + 2):
                if not row or all(v is None or str(v).strip() == '' for v in row):
                    continue
                resumen['total_filas'] += 1

                def g(key):
                    j = col.get(key)
                    return row[j] if (j is not None and j < len(row)) else None

                # Nominativo: SU propio valor por fila; si va en blanco la fila se salta
                # ("sin nominativo") y NO se hereda el de arriba (evita que una fila tome el
                # nominativo de otra y quede con el cliente equivocado).
                _raw_nom = g('nominativo')
                carry['nominativo'] = str(_raw_nom).strip() if _raw_nom is not None else ''
                for k in ('cli', 'pue', 'resumen'):
                    v = g(k)
                    if v is not None and str(v).strip():
                        carry[k] = str(v).strip()

                cedula = normalize_cedula(norm(g('ced')))
                # Anti-mezcla: si esta persona esta repetida en la hoja, solo se procesa
                # la fila con mas calendario; las demas se saltan (no mezclar calendarios).
                if cedula and best_row_idx.get(cedula, (i, 0))[0] != i:
                    resumen['errores'].append(
                        f'Hoja {ws.title}, Fila {i}: cedula {cedula} repetida en la hoja; se usa la fila con calendario mas completo'
                    )
                    continue

                # Bloque de calendario de ESTA fila. Normalmente el mes objetivo
                # (p.ej. agosto). Pero si para esta persona el bloque objetivo esta
                # VACIO y otro bloque tiene datos (caso RONDA GENERAL: agosto vacio,
                # julio lleno), se usa el ultimo bloque CON datos como mes base; asi
                # la proyeccion cubre agosto y el calendario no queda en blanco.
                if bloques:
                    def _cnt_fila(b):
                        return sum(
                            1 for d in b['dias']
                            if d < len(row) and str(parse_calendar_value(row[d]) or '').strip()
                        )
                    _obj_mes = req_month or mes_bloque
                    _target = next((b for b in bloques if b['mes'] == _obj_mes), None) if _obj_mes else None
                    if _target is None:
                        _target = bloques[-1]
                    _chosen = _target
                    if _cnt_fila(_target) == 0:
                        for b in bloques:
                            if _cnt_fila(b) > 0:
                                _chosen = b
                    dias_cols = _chosen['dias']
                    mes = _chosen['mes']
                    month_start = date(anio, mes, 1)
                    days_in_month = (date(anio, mes + 1, 1) - timedelta(days=1)).day if mes < 12 else 31

                puesto_nombre = carry['pue']
                es_saca = _rep_norm(puesto_nombre) == 'SACAFRANCO'

                cal = []
                _cal_ignorados = []
                for d_i, dj in enumerate(dias_cols):
                    if d_i >= days_in_month:
                        break
                    v = (parse_calendar_value(row[dj]) if dj < len(row) else '') or ''
                    # REGLA (opcion B): valores no reconocidos se IGNORAN (casilla vacia),
                    # el resto de la fila SI se importa. Solo D/N/F y tokens de sacafranco.
                    if v and not _cal_valor_ok(v, es_saca):
                        _cal_ignorados.append(v)
                        v = ''
                    cal.append(v)
                if _cal_ignorados:
                    from collections import Counter as _Cnt
                    _det = ', '.join(f"{k}(x{n})" for k, n in _Cnt(_cal_ignorados).items())
                    resumen['errores'].append(
                        f'Hoja {ws.title}, Fila {i}: valores ignorados en calendario (no reconocidos): {_det}'
                    )

                if es_saca:
                    if not cedula:
                        continue
                    # REGLA: no crear personas. Sacafranco tambien debe estar registrado.
                    persona = Persona.objects.filter(cedula=cedula).first()
                    if not persona:
                        resumen['errores'].append(
                            f'Hoja {ws.title}, Fila {i}: sacafranco con cedula {cedula} no esta registrado — no se importa'
                        )
                        continue
                    orden_counter += 1
                    row_orden = orden_counter
                    fila, _ = _get_or_create_sacafranco_fila(persona, mes, anio, row_orden)
                    touched_saca_ids.add(fila.id)
                    _sheet_saca_ids.append(fila.id)
                    touched_periodos.add((mes, anio))
                    if fila.orden != row_orden:
                        fila.orden = row_orden
                        fila.save(update_fields=['orden'])
                    for d_i, val in enumerate(cal):
                        day_num = d_i + 1
                        ws_start = month_start + timedelta(days=((day_num - 1) // 7) * 7)
                        day_field = WEEK_KEYS[date(anio, mes, day_num).weekday()]
                        sem, _ = SacafrancoFilaSemanal.objects.get_or_create(sacafranco_fila=fila, week_start=ws_start)
                        setattr(sem, day_field, (val or '').upper())
                        sem.save()
                    # RECONCILIAR: borrar toda la cobertura AUTO vieja de este sacafranco en
                    # el mes (incluye huerfanas de imports previos, p.ej. un nominativo que ya
                    # no cubre); el sync de abajo regenera solo la vigente.
                    try:
                        _purge_auto_sacafranco_persona(persona.id, month_start, date(anio, mes, days_in_month))
                    except Exception:
                        pass
                    _saca_errs = set()
                    for wk in range(((days_in_month - 1) // 7) + 1):
                        ws_start = month_start + timedelta(days=wk * 7)
                        sem = SacafrancoFilaSemanal.objects.filter(sacafranco_fila=fila, week_start=ws_start).first()
                        if not sem:
                            continue
                        payload = {k: getattr(sem, k, '') for k in WEEK_KEYS}
                        try:
                            err, resolved = _validate_sacafranco_tokens(payload, ws_start)
                            if not err:
                                _sync_sacafranco_to_reporte_y_consolidado(fila.id, ws_start, payload, resolved)
                            else:
                                # Token de sacafranco que NO resuelve (nominativo/turno mal):
                                # se reporta para que se revise (validacion por codigo, confiable).
                                _saca_errs.add(err)
                        except Exception:
                            pass
                    for _e in _saca_errs:
                        resumen['errores'].append(f"Hoja {ws.title}, Fila {i}: sacafranco — {_e}")

                    # Continuar el patrón del sacafranco en los próximos 36 meses
                    sf_vals = [(v or '').upper() for v in cal]
                    sf_ciclo = _ciclo_para_continuar(sf_vals)
                    if proyectar and sf_ciclo and any(str(x).strip() for x in sf_ciclo):
                        sf_len = len(sf_ciclo)
                        sf_idx = len(sf_vals)

                        def _add_m(y, mo, off):
                            return (y + (mo - 1 + off) // 12), ((mo - 1 + off) % 12 + 1)

                        for off in range(1, meses_proy + 1):
                            ty, tm = _add_m(anio, mes, off)
                            t_start = date(ty, tm, 1)
                            t_days = (date(ty, tm + 1, 1) - timedelta(days=1)).day if tm < 12 else 31
                            t_fila, _ = _get_or_create_sacafranco_fila(persona, tm, ty, row_orden)
                            _sheet_saca_ids.append(t_fila.id)
                            if t_fila.orden != row_orden:
                                t_fila.orden = row_orden
                                t_fila.save(update_fields=['orden'])
                            wp = {}
                            for dn in range(1, t_days + 1):
                                wss = t_start + timedelta(days=((dn - 1) // 7) * 7)
                                df = WEEK_KEYS[date(ty, tm, dn).weekday()]
                                wp.setdefault(wss, {})[df] = sf_ciclo[sf_idx % sf_len]
                                sf_idx += 1
                            # Guardar el CALENDARIO del sacafranco en bloque (rapido).
                            SacafrancoFilaSemanal.objects.filter(
                                sacafranco_fila=t_fila, week_start__in=list(wp.keys())
                            ).delete()
                            SacafrancoFilaSemanal.objects.bulk_create([
                                SacafrancoFilaSemanal(
                                    sacafranco_fila=t_fila, week_start=wss,
                                    **{k: dm.get(k, '') for k in WEEK_KEYS}
                                )
                                for wss, dm in wp.items()
                            ])
                            # RECONCILIAR el mes proyectado: borrar cobertura AUTO vieja
                            # (incluye huerfanas de imports previos hasta futuros lejanos,
                            # p.ej. un K16 que quedo hasta 2029). El sync de abajo (si esta
                            # dentro del horizonte) regenera solo la vigente.
                            try:
                                _purge_auto_sacafranco_persona(persona.id, t_start, date(ty, tm, t_days))
                            except Exception:
                                pass
                            # ...y sincronizar por semana para que el sacafranco SI se refleje
                            # en Reporte de Asistencia y Consolidado. Solo para los primeros
                            # meses (horizonte operativo): sincronizar los 36 meses por semana
                            # es carisimo y no sirve tan a futuro. El calendario si queda completo.
                            if off <= meses_sync_saca:
                                for wss, dm in wp.items():
                                    payload2 = {k: dm.get(k, '') for k in WEEK_KEYS}
                                    try:
                                        err2, resolved2 = _validate_sacafranco_tokens(payload2, wss)
                                        if not err2:
                                            _sync_sacafranco_to_reporte_y_consolidado(t_fila.id, wss, payload2, resolved2)
                                    except Exception:
                                        pass

                    resumen['sacafranco_creados'] += 1
                    resumen['filas_validas'] += 1
                    touched_dates.add(month_start)
                    continue

                if not cedula:
                    continue
                # Solo registros COMPLETOS: el cronograma del mes debe estar lleno.
                # Si falta ALGUN dia (hueco en blanco) o vienen menos dias que el mes,
                # la fila NO se importa (cronograma incompleto).
                _cal_mes = [(v or '').strip() for v in cal]
                _dias_faltantes = [d_i + 1 for d_i, v in enumerate(_cal_mes) if not v]
                if len(_cal_mes) < days_in_month or _dias_faltantes:
                    _faltan = _dias_faltantes if len(_cal_mes) >= days_in_month else 'todos los dias del mes'
                    resumen['errores'].append(
                        f'Hoja {ws.title}, Fila {i}: cronograma incompleto (dias sin marcar: {_faltan}) — no se importa'
                    )
                    continue
                if not carry['nominativo']:
                    resumen['errores'].append(f'Hoja {ws.title}, Fila {i}: sin nominativo (codigo de instalacion)')
                    continue
                instalacion = Instalacion.objects.filter(codigo__iexact=carry['nominativo']).first()
                if not instalacion:
                    resumen['errores'].append(f"Hoja {ws.title}, Fila {i}: instalacion con codigo '{carry['nominativo']}' no existe")
                    continue
                if cliente_id_filter and instalacion.cliente_id != cliente_id_filter:
                    continue

                # Validacion: el nominativo (codigo) debe ser del cliente/instalacion que
                # dice la fila. Comparacion TOLERANTE (typos/siglas) para no botar validos;
                # si el cliente es realmente distinto -> alerta y NO se importa.
                _cli_inst = getattr(instalacion, 'cliente', None)
                _cli_excel = carry.get('cli') or ''
                if _cli_excel and _cli_inst:
                    _ok = (_cliente_coincide(_cli_excel, getattr(_cli_inst, 'nombre_comercial', ''))
                           or _cliente_coincide(_cli_excel, getattr(_cli_inst, 'razon_social', ''))
                           or _cliente_coincide(_cli_excel, getattr(instalacion, 'nombre', '')))
                    if not _ok:
                        _nom_persona = norm(g('nombre'))
                        _quien = f"{cedula} {_nom_persona}".strip() or 'sin persona'
                        resumen['errores'].append(
                            f"Hoja {ws.title}, Fila {i}: [{_quien}] nominativo '{carry['nominativo']}' "
                            f"es del cliente '{getattr(_cli_inst, 'nombre_comercial', '')}', "
                            f"no coincide con '{_cli_excel}' del archivo — no se importa"
                        )
                        continue

                # arrastrar hora ingreso/salida de celdas combinadas (mismo puesto)
                raw_ing = g('ing')
                raw_sal = g('sal')
                if raw_ing is not None and str(raw_ing).strip():
                    carry_time['ing'] = raw_ing
                if raw_sal is not None and str(raw_sal).strip():
                    carry_time['sal'] = raw_sal
                hora_ingreso = parse_excel_time(raw_ing if (raw_ing is not None and str(raw_ing).strip()) else carry_time['ing'])
                hora_salida = parse_excel_time(raw_sal if (raw_sal is not None and str(raw_sal).strip()) else carry_time['sal'])
                if not hora_ingreso or not hora_salida:
                    resumen['errores'].append(f'Hoja {ws.title}, Fila {i}: hora ingreso/salida invalida')
                    continue

                resumen_txt = carry['resumen'] or ''
                m = re.match(r'\s*(\d+)\s+(.*)', resumen_txt)
                cantidad = int(m.group(1)) if m else 1
                resto = m.group(2) if m else resumen_txt
                grupos = parse_compact_horas_turno_dias(resto) or []

                # REGLA: el import NO crea personas. Solo referencia por cedula.
                # Si la persona no esta registrada -> avisa y salta (no se importa esa fila).
                persona = Persona.objects.filter(cedula=cedula).first()
                if not persona:
                    resumen['errores'].append(
                        f'Hoja {ws.title}, Fila {i}: persona con cedula {cedula} no esta registrada — no se importa'
                    )
                    continue

                horario, h_created = _get_or_create_horario(hora_ingreso, hora_salida)
                if h_created:
                    resumen['horarios_creados'] += 1

                # REGLA: el import NO crea puestos. Deben estar configurados antes
                # (con su cantidad de guardias). Si no existe -> avisa y salta.
                puesto = _get_puesto(instalacion, puesto_nombre)
                if not puesto:
                    resumen['errores'].append(
                        f"Hoja {ws.title}, Fila {i}: el puesto '{puesto_nombre}' no existe en la "
                        f"instalacion {carry['nominativo']} — no se importa"
                    )
                    continue
                if not puesto.horario_id:
                    puesto.horario = horario
                    puesto.save(update_fields=['horario'])
                for grp in grupos:
                    for dia in grp.get('dias', []):
                        PuestoHorario.objects.update_or_create(
                            puesto=puesto, dia=dia,
                            defaults={'horas': min(max(grp.get('hours', 12), 0), 24), 'turno': grp.get('turno') or 'Diurno'}
                        )
                try:
                    puesto.sync_from_horarios()
                    puesto.save()
                except Exception:
                    pass

                orden_counter += 1
                row_orden = orden_counter
                # RE-ORDENAR SIEMPRE segun el Excel: el orden refleja el ultimo import
                # (fila por fila, pestaña por pestaña). Asi el grid sigue el orden del Excel.
                asig, created = Asignacion.objects.update_or_create(
                    persona=persona, mes=mes, anio=anio,
                    defaults={
                        'cliente': instalacion.cliente, 'instalacion': instalacion,
                        'puesto': puesto, 'horario': horario, 'fecha': None,
                        'patronAsignacion': None, 'estado': 'ACTIVO',
                        'publicada_calendario': True, 'recurring': True,
                        'start_date': month_start, 'end_date': None,
                        'orden': row_orden,
                    }
                )
                touched_asig_ids.add(asig.id)
                _sheet_cantones.add(getattr(instalacion, 'canton_id', None))
                _sheet_clientes.add(getattr(instalacion, 'cliente_id', None))
                _sheet_inst_ids.add(instalacion.id)
                touched_dates.add(month_start)
                touched_periodos.add((mes, anio))
                puesto_personas.setdefault((puesto.id, mes, anio), set()).add(persona.id)
                resumen['asignaciones_creadas' if created else 'asignaciones_actualizadas'] += 1

                cal_by_week = {}
                for d_i, val in enumerate(cal):
                    day_num = d_i + 1
                    ws_start = month_start + timedelta(days=((day_num - 1) // 7) * 7)
                    day_field = WEEK_KEYS[date(anio, mes, day_num).weekday()]
                    cal_by_week.setdefault(ws_start, {})[day_field] = (val or '').upper()
                # Escritura en BLOQUE del mes base: borrar las semanas del mes y
                # recrearlas de una (mucho mas rapido que get_or_create + save por
                # semana; ademas garantiza que no queden datos viejos mezclados).
                AsignacionSemanal.objects.filter(
                    asignacion_id=asig.id, week_start__in=list(cal_by_week.keys())
                ).delete()
                AsignacionSemanal.objects.bulk_create([
                    AsignacionSemanal(
                        asignacion_id=asig.id, week_start=ws_start, puesto_id=puesto.id,
                        mon=dm.get('mon', ''), tue=dm.get('tue', ''), wed=dm.get('wed', ''),
                        thu=dm.get('thu', ''), fri=dm.get('fri', ''), sat=dm.get('sat', ''),
                        sun=dm.get('sun', ''),
                    )
                    for ws_start, dm in cal_by_week.items()
                ])

                # Continuar el patrón D/N/F en los próximos 36 meses (ciclo continuo)
                # Detectar el período REAL (ej. DDDNNNF=7) para no repetir el mes entero.
                cal_vals = [(v or '').upper() for v in cal]
                ciclo = _ciclo_para_continuar(cal_vals)
                if proyectar and ciclo and any(str(x).strip() for x in ciclo):
                    cycle_len = len(ciclo)
                    seq_idx = len(cal_vals)  # desfase global: continúa donde terminó el mes

                    def _add_months(y, mo, off):
                        ny = y + (mo - 1 + off) // 12
                        nm = (mo - 1 + off) % 12 + 1
                        return ny, nm

                    # 1) Calcular el calendario proyectado de TODOS los meses objetivo.
                    targets = []  # (ty, tm, t_start, wp)
                    for off in range(1, meses_proy + 1):
                        ty, tm = _add_months(anio, mes, off)
                        t_start = date(ty, tm, 1)
                        t_days = (date(ty, tm + 1, 1) - timedelta(days=1)).day if tm < 12 else 31
                        wp = {}
                        for dn in range(1, t_days + 1):
                            wss = t_start + timedelta(days=((dn - 1) // 7) * 7)
                            df = WEEK_KEYS[date(ty, tm, dn).weekday()]
                            wp.setdefault(wss, {})[df] = ciclo[seq_idx % cycle_len]
                            seq_idx += 1
                        targets.append((ty, tm, t_start, wp))
                        puesto_personas.setdefault((puesto.id, tm, ty), set()).add(persona.id)

                    # 2) Asignaciones de esos meses en BLOQUE (1 query para leer las
                    #    existentes; bulk_update + bulk_create en vez de update_or_create
                    #    por cada mes). Esto es lo que hacia lento el import de 36 meses.
                    want = {(tm, ty) for (ty, tm, _s, _w) in targets}
                    existing = {}
                    for a in Asignacion.objects.filter(persona=persona):
                        if (a.mes, a.anio) in want:
                            existing[(a.mes, a.anio)] = a
                    _upd_fields = ['cliente', 'instalacion', 'puesto', 'horario', 'fecha',
                                   'patronAsignacion', 'estado', 'publicada_calendario',
                                   'recurring', 'start_date', 'end_date', 'orden']
                    to_update, to_create = [], []
                    for (ty, tm, t_start, wp) in targets:
                        a = existing.get((tm, ty))
                        if a:
                            a.cliente = instalacion.cliente
                            a.instalacion = instalacion
                            a.puesto = puesto
                            a.horario = horario
                            a.fecha = None
                            a.patronAsignacion = None
                            a.estado = 'ACTIVO'
                            a.publicada_calendario = True
                            a.recurring = True
                            a.start_date = t_start
                            a.end_date = None
                            a.orden = row_orden
                            to_update.append(a)
                        else:
                            to_create.append(Asignacion(
                                persona=persona, mes=tm, anio=ty,
                                cliente=instalacion.cliente, instalacion=instalacion,
                                puesto=puesto, horario=horario, fecha=None,
                                patronAsignacion=None, estado='ACTIVO',
                                publicada_calendario=True, recurring=True,
                                start_date=t_start, end_date=None, orden=row_orden,
                            ))
                    if to_update:
                        Asignacion.objects.bulk_update(to_update, _upd_fields)
                    created_objs = Asignacion.objects.bulk_create(to_create) if to_create else []

                    asig_by_month = {}
                    for a in to_update:
                        asig_by_month[(a.mes, a.anio)] = a.id
                    for a in created_objs:
                        asig_by_month[(a.mes, a.anio)] = a.id
                    # Registrar TODAS las asignaciones proyectadas como "tocadas" para que
                    # la desactivacion de sobrantes (mas abajo) no las marque como sobrantes.
                    touched_asig_ids.update(asig_by_month.values())

                    # 3) Calendario semanal de todos los meses objetivo en BLOQUE:
                    #    borrar de una y recrear de una (2 queries en total).
                    all_ids = list(asig_by_month.values())
                    if all_ids:
                        AsignacionSemanal.objects.filter(asignacion_id__in=all_ids).delete()
                    sem_bulk = []
                    for (ty, tm, t_start, wp) in targets:
                        aid = asig_by_month.get((tm, ty))
                        if not aid:
                            continue
                        for wss, dm in wp.items():
                            sem_bulk.append(AsignacionSemanal(
                                asignacion_id=aid, week_start=wss, puesto_id=puesto.id,
                                mon=dm.get('mon', ''), tue=dm.get('tue', ''), wed=dm.get('wed', ''),
                                thu=dm.get('thu', ''), fri=dm.get('fri', ''), sat=dm.get('sat', ''),
                                sun=dm.get('sun', ''),
                            ))
                    if sem_bulk:
                        AsignacionSemanal.objects.bulk_create(sem_bulk)

                resumen['filas_validas'] += 1

            # Fin de la hoja: sellar la VISTA en sus sacafranco. Se elige la vista que
            # MEJOR representa la hoja (por sus datos) y se estampa segun su TIPO:
            #   - vista de CANTON  -> solo cantones (clientes vacio)
            #   - vista de CLIENTE -> solo clientes (cantones vacio)
            # Asi una hoja de canton (que abarca muchos clientes) NO se filtra a las
            # vistas de cliente, y viceversa. Sin match claro: cantones de la hoja.
            if _sheet_saca_ids:
                from ..models import VistaCanton
                cant_set = {c for c in _sheet_cantones if c is not None}
                cli_set = {c for c in _sheet_clientes if c is not None}
                inst_set = set(_sheet_inst_ids)
                best, best_score = None, 0
                for v in VistaCanton.objects.all():
                    if v.tipo == 'canton':
                        score = len(cant_set & set(v.cantones or []))
                    elif v.tipo == 'cliente':
                        vin = set(v.instalaciones or [])
                        score = len(inst_set & vin) if vin else len(cli_set & set(v.clientes or []))
                    else:
                        score = 0
                    if score > best_score:
                        best, best_score = v, score
                if best is not None and best.tipo == 'canton':
                    stamp_cant, stamp_cli = sorted(set(best.cantones or [])), []
                elif best is not None and best.tipo == 'cliente':
                    stamp_cant, stamp_cli = [], sorted(set(best.clientes or []))
                else:
                    stamp_cant, stamp_cli = sorted(cant_set), []
                SacafrancoFila.objects.filter(id__in=set(_sheet_saca_ids)).update(
                    cantones=stamp_cant, clientes=stamp_cli,
                )

        if _quiere_desactivar_sobrantes(request):
            # (a+b) Sobrantes EN BLOQUE: toda asignacion ACTIVA en los periodos tocados
            #       (mes base + proyectados) que NO fue creada/actualizada por este import
            #       se desactiva. `touched_asig_ids` ya incluye base + proyeccion, asi que
            #       "lo que no esta ahi" = lo que ya no vino en el Excel (guardias sobrantes
            #       de puestos que si vinieron + puestos/instalaciones que desaparecieron).
            #       Antes esto era un bucle por (puesto x mes) con save() fila por fila
            #       (miles de queries); ahora son 3 queries en bloque.
            #       Import COMPLETO (sin cliente) = verdad total del mes; por cliente se acota.
            #       No toca meses ANTERIORES al import (se conserva el historial).
            all_periods = {(pm, pa) for (_pid, pm, pa) in puesto_personas.keys()}
            if all_periods:
                period_q = Q()
                for (pm, pa) in all_periods:
                    period_q |= Q(mes=pm, anio=pa)
                sob = (Asignacion.objects.filter(period_q, estado='ACTIVO')
                       .exclude(id__in=touched_asig_ids)
                       .exclude(persona__tipo='SACAFRANCO'))
                if cliente_id_filter is not None:
                    sob = sob.filter(cliente_id=cliente_id_filter)
                sob_ids = list(sob.values_list('id', flat=True))
                if sob_ids:
                    Asignacion.objects.filter(id__in=sob_ids).update(estado='INACTIVO')
                    ReporteAsistencia.objects.filter(asignacion_id__in=sob_ids).update(
                        estado='TURNO', estado_asistencia='', reemplazo=None,
                        descripcion=None, row_color=None,
                    )

            # (b2) Cortar la PROYECCION de quien ya no viene en el Excel.
            #      Cada fila es recurring con end_date=None y se proyecta hacia adelante
            #      para siempre. Desactivar la fila del mes no basta: sus filas de meses
            #      PREVIOS se siguen proyectando al mes importado. A esas personas (que no
            #      vinieron) se les pone end_date al dia previo al mes MAS ANTIGUO importado:
            #      dejan de proyectarse de ahi en adelante y se CONSERVA su historial
            #      (los meses anteriores siguen visibles). Solo aplica al import COMPLETO
            #      del mes; por cliente se acota a ese cliente.
            if all_periods:
                min_anio, min_mes = min((pa, pm) for (pm, pa) in all_periods)
                cutoff_start = date(min_anio, min_mes, 1)
                prev_day = cutoff_start - timedelta(days=1)
                imported_personas = set(
                    Asignacion.objects.filter(id__in=touched_asig_ids, persona_id__isnull=False)
                    .values_list('persona_id', flat=True)
                )
                fuga_qs = Asignacion.objects.filter(
                    estado='ACTIVO', recurring=True, start_date__lt=cutoff_start
                ).filter(
                    Q(end_date__isnull=True) | Q(end_date__gte=cutoff_start)
                ).exclude(persona_id__in=imported_personas).exclude(persona__tipo='SACAFRANCO')
                if cliente_id_filter is not None:
                    fuga_qs = fuga_qs.filter(cliente_id=cliente_id_filter)
                fuga_qs.update(end_date=prev_day)

            # (c) Sacafranco sobrantes -> BORRAR. Solo en import COMPLETO (sin filtro de cliente),
            #     acotado a los periodos que trajo el Excel.
            if cliente_id_filter is None:
                for (pmes, panio) in touched_periodos:
                    ids = list(SacafrancoFila.objects.filter(mes=pmes, anio=panio)
                               .exclude(id__in=touched_saca_ids).values_list('id', flat=True))
                    if ids:
                        SacafrancoFilaSemanal.objects.filter(sacafranco_fila_id__in=ids).delete()
                        SacafrancoFila.objects.filter(id__in=ids).delete()

        for asig in Asignacion.objects.select_related('persona', 'cliente', 'instalacion', 'puesto', 'horario').filter(id__in=touched_asig_ids):
            rep, _ = ReporteAsistencia.objects.get_or_create(asignacion=asig)
            rep.persona = asig.persona
            rep.cliente = asig.cliente
            rep.instalacion = asig.instalacion
            rep.puesto = asig.puesto
            rep.horario = asig.horario
            rep.puesto_tipo = getattr(asig.puesto, 'tipo', None) if asig.puesto else None
            rep.save()

        try:
            from .reporte_asistencia_views import _build_reporte_asistencia_data
            from .consolidado_views import _build_resumen_manual
            for ref_date in touched_dates:
                for turno_val in ('Diurno', 'Nocturno'):
                    rows_data = _build_reporte_asistencia_data(fecha=ref_date.isoformat(), turno=turno_val)
                    _build_resumen_manual(ref_date, turno_val, rows_data)
        except Exception:
            pass

    return resumen


# ============================================================================
# Importacion en SEGUNDO PLANO (async) — evita el timeout 524 de Cloudflare.
# El POST responde al instante con un job_id; el proceso corre en un hilo y el
# frontend consulta el estado hasta que termina.
# ============================================================================
class _ReqLite:
    """Request minimo para reusar el importador desde el hilo (solo GET/POST)."""
    def __init__(self, get_params):
        self.GET = get_params or {}
        self.POST = {}


def _run_import_job(job_id, path, get_params, cliente_id):
    from django.db import connection
    from ..models import ImportJob
    try:
        wb = load_workbook(path, read_only=False, data_only=True)
        req = _ReqLite(get_params)
        if es_formato_reporte(wb):
            resumen = importar_formato_reporte(req, wb, cliente_id)
        else:
            resumen = {'error': 'El archivo no tiene el formato reporte reconocido.'}
        job = ImportJob.objects.get(id=job_id)
        job.resumen = resumen
        job.estado = 'error' if (resumen.get('error') and not resumen.get('filas_validas')) else 'ok'
        job.save(update_fields=['resumen', 'estado', 'actualizado_en'])
    except Exception as exc:
        logger.exception('Error en import async')
        try:
            job = ImportJob.objects.get(id=job_id)
            job.estado = 'error'
            job.error = str(exc)
            job.save(update_fields=['estado', 'error', 'actualizado_en'])
        except Exception:
            pass
    finally:
        try:
            os.remove(path)
        except Exception:
            pass
        connection.close()


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def importar_puestos_asignaciones_async(request):
    """Lanza la importacion en segundo plano y responde al instante con un job_id."""
    if not request.user.has_perm('CoreFisica.import_puestos_asignaciones'):
        return JsonResponse({'error': 'No Autorizado'}, status=403)
    upload = request.FILES.get('file')
    if not upload:
        return JsonResponse({'error': 'Falta el archivo (campo file)'}, status=400)

    cliente_id = request.GET.get('cliente_id')
    try:
        cliente_id = int(cliente_id) if cliente_id else None
    except (TypeError, ValueError):
        cliente_id = None

    fd, path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    with open(path, 'wb') as fh:
        for chunk in upload.chunks():
            fh.write(chunk)

    from ..models import ImportJob
    job = ImportJob.objects.create(tipo='puestos_asignaciones', estado='procesando')
    get_params = {k: request.GET.get(k) for k in ('mes', 'anio', 'meses', 'desactivar_sobrantes')}
    threading.Thread(
        target=_run_import_job,
        args=(str(job.id), path, get_params, cliente_id),
        daemon=True,
    ).start()
    return JsonResponse({'job_id': str(job.id), 'estado': 'procesando'}, status=202)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def estado_import(request, job_id):
    """Consulta el estado/resultado de una importacion en segundo plano."""
    from ..models import ImportJob
    try:
        job = ImportJob.objects.get(id=job_id)
    except (ImportJob.DoesNotExist, ValueError, Exception):
        return JsonResponse({'error': 'Trabajo de importacion no encontrado'}, status=404)
    return JsonResponse({
        'estado': job.estado,
        'resumen': job.resumen,
        'error': job.error,
    })
