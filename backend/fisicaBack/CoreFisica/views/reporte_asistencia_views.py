"""Vistas del Reporte de Asistencia: armado por día (rutea D/N/F del calendario), edición, historial y export Excel/PDF."""
from django.http import JsonResponse
from django.db.models import Q, Subquery
from django.http import HttpResponse
from django.utils import timezone
from io import BytesIO
import datetime
from pathlib import Path
from types import SimpleNamespace
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from ..models import Asignacion, Persona, ReporteAsistencia, ReporteAsistenciaHistorial, AsignacionSemanal, SacafrancoFilaSemanal, Instalacion, SacafrancoAsistencia
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader


TIPOS_REEMPLAZO_PERMITIDOS = set(ReporteAsistencia.TIPOS_REEMPLAZO)
HEADER_TITULO = 'FR REPORTE DE ASISTENCIA SEGURIDAD FÍSICA'
HEADER_VERSION = '.01'
HEADER_FECHA_APROBACION = '12-Aug-22'

DIAS_SEMANA_ES = [
    'LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES', 'SÁBADO', 'DOMINGO'
]

def _normalize_zona_filter(zona_param):
    """Resuelve el filtro de zona contra ZonaOperativa (por numero o por nombre) y
    devuelve el NOMBRE canonico de la zona (ej. 'ZONA 1'); '' = todas. Soporta crear
    Zona 4 y nombres personalizados sin tocar codigo."""
    import re as _re
    from ..models import ZonaOperativa
    zona_raw = str(zona_param or '').strip()
    if not zona_raw:
        return ''
    z = None
    if zona_raw.isdigit():
        z = ZonaOperativa.objects.filter(numero=int(zona_raw)).first()
    if not z:
        z = ZonaOperativa.objects.filter(nombre__iexact=zona_raw).first()
    if not z:
        # compatibilidad: 'Zona 1' -> numero 1
        m = _re.search(r'\d+', zona_raw)
        if m:
            z = ZonaOperativa.objects.filter(numero=int(m.group())).first()
    return z.nombre if z else ''


def _parse_fecha_reporte(fecha_param):
    if fecha_param:
        try:
            return datetime.date.fromisoformat(str(fecha_param))
        except ValueError:
            pass
    return timezone.localdate()


def _build_operador_consola(request):
    user = getattr(request, 'user', None)
    if not user or not user.is_authenticated:
        return ''
    full_name = f"{user.first_name} {user.last_name}".strip()
    return full_name or user.get_username()


def _format_fecha_reporte_es(fecha_obj):
    dia = DIAS_SEMANA_ES[fecha_obj.weekday()]
    return f"{dia} {fecha_obj.strftime('%d/%m/%Y')}"


def _find_logo_path():
    this_file = Path(__file__).resolve()
    root_candidates = [
        this_file.parents[i] for i in [4, 3] if i < len(this_file.parents)
    ] + [Path.cwd()]
    relative_candidates = [
        Path('frontendf/public/logodescargable.jpg'),
        Path('frontendf/public/favicon.png'),
        Path('frontendf/public/logo.png'),
        Path('frontendf/src/assets/images/logo.png'),
        Path('src/assets/images/logo.png'),
    ]

    for root in root_candidates:
        for rel in relative_candidates:
            candidate = root / rel
            if candidate.exists():
                return candidate
    return None


def _build_header_context(request, fecha, turno):
    fecha_obj = _parse_fecha_reporte(fecha)
    turno_label = turno if turno in ['Diurno', 'Nocturno', 'Tarde', 'Veinticuatro'] else 'TODOS'
    return {
        'titulo': HEADER_TITULO,
        'version': HEADER_VERSION,
        'fecha_aprobacion': HEADER_FECHA_APROBACION,
        'fecha_reporte': _format_fecha_reporte_es(fecha_obj),
        'turno': turno_label,
        'operador_consola': _build_operador_consola(request),
        'logo_path': _find_logo_path(),
    }


def _texto_hueca(item):
    """Motivo de la hueca para mostrar en los reportes (ej. 'HUECA POR UNIDAD FIJA').
    Devuelve '' cuando la fila no es hueca."""
    if item.get('hueca'):
        return (item.get('hueca_motivo') or '').strip() or 'HUECA'
    return ''


def _descripcion_con_hueca(item):
    """DESCRIPCIÓN para el descargable: si la fila es hueca, antepone el motivo
    al texto que ya tuviera la descripción (para que la hueca salga en el reporte)."""
    desc = (item.get('descripcion') or '').strip()
    h = _texto_hueca(item)
    if h:
        return f"{h} · {desc}" if desc else h
    return desc


def _draw_excel_header(ws, ctx, border):
    ws.merge_cells('A1:B2')
    ws.merge_cells('C1:F2')
    ws.merge_cells('A4:D4')
    ws.merge_cells('E4:F4')
    ws.merge_cells('G4:I4')

    for row in range(1, 5):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws['C1'] = ctx['titulo']
    ws['C1'].font = Font(bold=True, size=14)

    ws['G1'] = 'Versión:'
    ws['H1'] = ctx['version']
    ws['G2'] = 'Fecha de aprobación:'
    ws['H2'] = ctx['fecha_aprobacion']
    ws['G1'].font = ws['G2'].font = Font(bold=True, size=10)
    ws['H1'].font = ws['H2'].font = Font(bold=True, size=10)

    ws['A4'] = f"FECHA: {ctx['fecha_reporte']}"
    ws['E4'] = f"TURNO: {ctx['turno'].upper()}"
    ws['G4'] = f"OPERADOR DE CONSOLA: {ctx['operador_consola']}"
    ws['A4'].font = ws['E4'].font = ws['G4'].font = Font(bold=True, size=11)

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 10
    ws.row_dimensions[4].height = 24

    if ctx['logo_path']:
        try:
            img = XLImage(str(ctx['logo_path']))

            # Escalar proporcionalmente dentro del recuadro A1:B2 (sin deformar)
            box_w_px = 203
            box_h_px = 74
            padding_px = 6
            max_w = box_w_px - (padding_px * 2)
            max_h = box_h_px - (padding_px * 2)

            orig_w = max(float(img.width), 1.0)
            orig_h = max(float(img.height), 1.0)
            scale = min(max_w / orig_w, max_h / orig_h)

            final_w = int(orig_w * scale)
            final_h = int(orig_h * scale)
            img.width = final_w
            img.height = final_h

            x_offset_px = int((box_w_px - final_w) / 2)
            y_offset_px = int((box_h_px - final_h) / 2)

            img.anchor = OneCellAnchor(
                _from=AnchorMarker(
                    col=0,
                    row=0,
                    colOff=pixels_to_EMU(max(x_offset_px, 0)),
                    rowOff=pixels_to_EMU(max(y_offset_px, 0)),
                ),
                ext=XDRPositiveSize2D(
                    cx=pixels_to_EMU(final_w),
                    cy=pixels_to_EMU(final_h),
                ),
            )
            ws.add_image(img)
        except Exception:
            pass


def _draw_pdf_header(p, width, height, x_margin, y_margin, ctx):
    table_w = width - (2 * x_margin)
    top = height - y_margin

    top_h = 0.95 * inch
    info_h = 0.32 * inch

    logo_w = 1.8 * inch
    title_w = 4.8 * inch

    x_logo = x_margin
    x_title = x_logo + logo_w
    x_meta = x_title + title_w

    p.setStrokeColor(colors.black)
    p.rect(x_margin, top - top_h, table_w, top_h, stroke=1, fill=0)
    p.line(x_title, top - top_h, x_title, top)
    p.line(x_meta, top - top_h, x_meta, top)

    p.line(x_meta, top - (top_h / 2), x_margin + table_w, top - (top_h / 2))

    if ctx['logo_path']:
        try:
            logo = ImageReader(str(ctx['logo_path']))
            draw_w = logo_w * 0.62
            draw_h = top_h * 0.62
            draw_x = x_logo + ((logo_w - draw_w) / 2)
            draw_y = (top - top_h) + ((top_h - draw_h) / 2)
            p.drawImage(
                logo,
                draw_x,
                draw_y,
                width=draw_w,
                height=draw_h,
                preserveAspectRatio=True,
                mask='auto'
            )
        except Exception:
            pass

    title_text = str(ctx['titulo'] or '')
    title_font = 15
    max_title_w = title_w - 0.20 * inch
    while title_font > 9 and pdfmetrics.stringWidth(title_text, 'Helvetica-Bold', title_font) > max_title_w:
        title_font -= 1

    p.setFont('Helvetica-Bold', title_font)
    title_y = top - (top_h / 2) + 0.1 * inch
    p.drawCentredString(x_title + (title_w / 2), title_y, title_text)

    p.setFont('Helvetica-Bold', 10)
    p.drawString(x_meta + 8, top - 0.28 * inch, 'Versión:')
    p.drawRightString(x_margin + table_w - 8, top - 0.28 * inch, ctx['version'])

    p.drawString(x_meta + 8, top - 0.75 * inch, 'Fecha de aprobación:')
    p.drawRightString(x_margin + table_w - 8, top - 0.75 * inch, ctx['fecha_aprobacion'])

    info_top = top - top_h
    p.rect(x_margin, info_top - info_h, table_w, info_h, stroke=1, fill=0)

    info_a_w = 2.8 * inch
    info_b_w = 2.3 * inch
    x_info_b = x_margin + info_a_w
    x_info_c = x_info_b + info_b_w

    p.line(x_info_b, info_top - info_h, x_info_b, info_top)
    p.line(x_info_c, info_top - info_h, x_info_c, info_top)

    p.setFont('Helvetica-Bold', 10)
    text_y = info_top - 0.22 * inch
    p.drawString(x_margin + 8, text_y, f"FECHA: {ctx['fecha_reporte']}")
    p.drawString(x_info_b + 8, text_y, f"TURNO: {ctx['turno'].upper()}")
    p.drawString(x_info_c + 8, text_y, f"OPERADOR DE CONSOLA: {ctx['operador_consola']}")

    return info_top - info_h - 0.2 * inch


def _draw_pdf_table_headers(p, x_margin, y, headers, col_widths):
    p.setFont('Helvetica-Bold', 8)
    x = x_margin
    for i, header in enumerate(headers):
        header_w = pdfmetrics.stringWidth(header, 'Helvetica-Bold', 8)
        p.drawString(x + max((col_widths[i] - header_w) / 2, 0), y, header)
        x += col_widths[i]
    return y - 0.25 * inch


def _fit_text_to_width(text, max_width, font_name='Helvetica', font_size=7):
    s = str(text) if text is not None else ''
    if not s:
        return ''
    if pdfmetrics.stringWidth(s, font_name, font_size) <= max_width:
        return s

    ellipsis = '...'
    ellipsis_w = pdfmetrics.stringWidth(ellipsis, font_name, font_size)
    allowed = max_width - ellipsis_w
    if allowed <= 0:
        return ''

    while s and pdfmetrics.stringWidth(s, font_name, font_size) > allowed:
        s = s[:-1]
    return s + ellipsis


def _wrap_text_to_width(text, max_width, font_name='Helvetica', font_size=6):
    """Parte el texto en varias líneas que caben en max_width (sin cortar/elipsis).
    Quiebra por palabras; si una palabra sola excede el ancho, la corta por caracteres."""
    s = str(text) if text is not None else ''
    if not s.strip():
        return ['']
    lines = []
    current = ''
    for word in s.split():
        trial = (current + ' ' + word).strip()
        if pdfmetrics.stringWidth(trial, font_name, font_size) <= max_width:
            current = trial
            continue
        if current:
            lines.append(current)
            current = ''
        # palabra más larga que la columna: cortarla por caracteres
        while pdfmetrics.stringWidth(word, font_name, font_size) > max_width and len(word) > 1:
            cut = len(word)
            while cut > 1 and pdfmetrics.stringWidth(word[:cut], font_name, font_size) > max_width:
                cut -= 1
            lines.append(word[:cut])
            word = word[cut:]
        current = word
    if current:
        lines.append(current)
    return lines or ['']


def _normalize_hex_color(color_value):
    c = str(color_value or '').strip()
    if not c:
        return None
    if c.startswith('#'):
        c = c[1:]
    if len(c) != 6:
        return None
    try:
        int(c, 16)
    except ValueError:
        return None
    return c.upper()


def _normalize_estado_asistencia(value):
    v = str(value or '').strip().upper()
    if v in ('ASISTIO', 'FALTO'):
        return v
    return ''


def _is_auto_sacafranco_desc(desc):
    return str(desc or '').strip().upper().startswith('COBERTURA SACAFRANCO AUTO')


def _is_falto(item):
    return str(item.get('estado_asistencia') or '').strip().upper() == 'FALTO'


def _zona_sort_key(zona_titulo):
    """Ordena por el numero que aparece en el nombre de la zona (Zona 1, 2, 3, 4...).
    Las que no tienen numero (SIN ZONA, SIN ASIGNACION) van al final."""
    import re as _re
    z = str(zona_titulo or '').strip()
    m = _re.search(r'\d+', z)
    if m:
        return (int(m.group()), z.lower())
    return (9999, z.lower())


def _normalize_zona_label(zona_titulo):
    # Conserva el nombre real de la zona (ej. 'ZONA 1'); solo limpia espacios.
    return str(zona_titulo or '').strip()


def _format_zona_label(zona_titulo):
    label = _normalize_zona_label(zona_titulo)
    if label.lower().startswith('zona '):
        return label
    return f"Zona {label}" if label else 'Zona'


def _build_resumen_asistencia(data):
    evaluables = [item for item in data if item.get('asignacion_id')]
    faltos = sum(1 for item in evaluables if _is_falto(item))
    # Asistencias = SOLO los marcados ASISTIO (los sin marcar quedan pendientes, no cuentan).
    asistencias = sum(
        1 for item in evaluables
        if _normalize_estado_asistencia(item.get('estado_asistencia')) == 'ASISTIO'
    )
    return asistencias, faltos


def _build_resumen_asistencia_por_zona(data):
    evaluables = [item for item in data if item.get('asignacion_id')]
    zonas = {}
    for item in evaluables:
        zona = _normalize_zona_label((item.get('zona_titulo') or 'SIN ZONA').strip())
        entry = zonas.setdefault(zona, {'total': 0, 'asistencias': 0, 'faltas': 0})
        entry['total'] += 1
        if _is_falto(item):
            entry['faltas'] += 1
        elif _normalize_estado_asistencia(item.get('estado_asistencia')) == 'ASISTIO':
            entry['asistencias'] += 1

    resumen = []
    for zona in sorted(zonas.keys(), key=_zona_sort_key):
        z = zonas[zona]
        resumen.append({
            'zona': zona,
            'total': z['total'],
            'asistencias': z['asistencias'],   # solo ASISTIO, ya no total - faltas
            'faltas': z['faltas'],
        })
    return resumen



def _group_reporte_por_zona_y_provincia(data):
    zonas = {}
    for item in data:
        if item.get('asignacion_id'):
            zona = _normalize_zona_label((item.get('zona_titulo') or 'SIN ZONA').strip())
            provincia = (item.get('provincia') or 'SIN PROVINCIA').strip()
        else:
            zona = 'SIN ASIGNACION'
            provincia = 'SIN ASIGNACION'
        zonas.setdefault(zona, {}).setdefault(provincia, []).append(item)

    grouped = []
    for zona in sorted(zonas.keys(), key=_zona_sort_key):
        provincias = []
        for prov in sorted(zonas[zona].keys(), key=lambda v: str(v).lower()):
            provincias.append({
                'provincia': prov,
                'rows': zonas[zona][prov],
            })
        grouped.append({'zona': zona, 'provincias': provincias})
    return grouped


def _write_excel_resumen(ws, row_idx, asistencias, faltos, border):
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
    ws.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=9)

    left_cell = ws.cell(row=row_idx, column=1)
    left_cell.value = f"{asistencias}\nASISTENCIAS"
    left_cell.font = Font(bold=True, color='9C0006', size=12)
    left_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_cell.fill = PatternFill(start_color='E6B8BE', end_color='E6B8BE', fill_type='solid')

    right_cell = ws.cell(row=row_idx, column=5)
    right_cell.value = f"{faltos}\nFALTOS"
    right_cell.font = Font(bold=True, color='000000', size=12)
    right_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx in range(1, 10):
        ws.cell(row=row_idx, column=col_idx).border = border
    ws.row_dimensions[row_idx].height = 38
    return row_idx


def _reemplazo_esta_ocupado(persona_id, fecha_reporte=None, asignacion_id_actual=None, permitir_asignado=False):
    # Un reemplazo debe estar LIBRE ese dia. Coherente con personas_asignadas /
    # el badge DISPONIBLE: si ese dia esta en FRANCO (F) o no esta programado, esta
    # disponible aunque tenga asignacion en el mes; solo se bloquea si TRABAJA (D/N).
    # permitir_asignado=True (movimiento interno / ADICIONAL con FIJOS): se omite el
    # bloqueo por tener asignacion activa; solo se valida el doble uso como reemplazo.
    if not permitir_asignado:
        if fecha_reporte:
            asig_ids = list(
                Asignacion.objects.filter(persona_id=persona_id, estado='ACTIVO').filter(
                    Q(mes=fecha_reporte.month, anio=fecha_reporte.year) |
                    Q(fecha=fecha_reporte) |
                    (Q(recurring=True) & Q(start_date__lte=fecha_reporte) &
                     (Q(end_date__isnull=True) | Q(end_date__gte=fecha_reporte)))
                ).exclude(end_date__isnull=False, end_date__lt=fecha_reporte)
                .values_list('id', flat=True)
            )
            if asig_ids:
                dnf = _calendar_dnf_for_date(fecha_reporte)  # {asignacion_id: 'D'|'N'|'F'}
                if any(dnf.get(aid) != 'F' for aid in asig_ids):
                    return True, 'La persona seleccionada ya tiene una asignacion activa ese dia.'
        else:
            # Sin fecha del reporte no se puede evaluar el dia: cualquier asignacion bloquea.
            if Asignacion.objects.filter(persona_id=persona_id, estado='ACTIVO').exists():
                return True, 'La persona seleccionada ya tiene una asignacion activa.'

    if fecha_reporte:
        qs = ReporteAsistencia.objects.filter(
            fecha_reporte=fecha_reporte,
            reemplazo_id=persona_id,
        )
        if asignacion_id_actual:
            qs = qs.exclude(asignacion_id=asignacion_id_actual)
        if qs.exists():
            return True, 'La persona seleccionada ya esta asignada como reemplazo en este reporte.'

    return False, ''


def _resolver_reemplazo_desde_request(request, fecha_reporte=None, asignacion_id_actual=None):
    if 'reemplazo_id' not in request.data:
        return 'no-enviado', None

    reemplazo_id = request.data.get('reemplazo_id')
    if reemplazo_id in [None, '', 'null']:
        return None, None

    try:
        reemplazo_id = int(reemplazo_id)
    except (TypeError, ValueError):
        return None, JsonResponse(
            {'error': 'reemplazo_id debe ser un entero valido'},
            status=status.HTTP_400_BAD_REQUEST
        )

    persona = Persona.objects.filter(id=reemplazo_id, is_active=True).first()
    if not persona:
        return None, JsonResponse(
            {'error': 'La persona enviada en reemplazo_id no existe o esta inactiva'},
            status=status.HTTP_400_BAD_REQUEST
        )

    if persona.tipo not in TIPOS_REEMPLAZO_PERMITIDOS:
        return None, JsonResponse(
            {'error': f'El tipo {persona.tipo} no puede ser reemplazo. Permitidos: {sorted(TIPOS_REEMPLAZO_PERMITIDOS)}'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Se permite elegir como reemplazo a alguien con asignacion activa (movimiento
    # interno, para cualquier estado); no modifica su asignacion, solo registra la
    # cobertura. Solo se valida que no este ya usado como reemplazo en otro registro.
    ocupado, motivo = _reemplazo_esta_ocupado(
        persona_id=persona.id,
        fecha_reporte=fecha_reporte,
        asignacion_id_actual=asignacion_id_actual,
        permitir_asignado=True,
    )
    if ocupado:
        return None, JsonResponse(
            {'error': motivo},
            status=status.HTTP_400_BAD_REQUEST
        )

    return persona, None


def _calendar_dnf_for_date(fecha_obj):
    """Letra del calendario (AsignacionSemanal) para esa fecha, por asignación.

    Devuelve {asignacion_id: 'D' | 'N' | 'F' | 'T' | 'V'} según el valor del día:
    D=diurno, N=nocturno, F=franco, T=tarde, V=24 horas (se toma solo la primera letra).
    Considera semana estilo mensual (día 1 + saltos de 7) y estilo ISO (lunes).
    """
    if not fecha_obj:
        return {}
    day_field_map = {0: 'mon', 1: 'tue', 2: 'wed', 3: 'thu', 4: 'fri', 5: 'sat', 6: 'sun'}
    day_field = day_field_map.get(fecha_obj.weekday())
    if not day_field:
        return {}
    month_base = fecha_obj.replace(day=1)
    week_start_month = month_base + datetime.timedelta(days=((fecha_obj.day - 1) // 7) * 7)
    week_start_iso = fecha_obj - datetime.timedelta(days=fecha_obj.weekday())
    rows = AsignacionSemanal.objects.filter(
        week_start__in=[week_start_month, week_start_iso]
    ).exclude(asignacion_id__isnull=True).values_list('asignacion_id', day_field)
    out = {}
    for aid, val in rows:
        letra = (str(val).strip()[:1].upper()) if val else ''
        if letra in ('D', 'N', 'F', 'T', 'V') and aid not in out:
            out[aid] = letra
    return out


def _build_reporte_asistencia_data(
    fecha=None,
    cliente_id=None,
    turno=None,
    exclude_sacafranco=False,
    zona='',
    q='',
    page=None,
    page_size=None,
    include_total=False,
):
    fecha_obj = None
    if fecha:
        try:
            fecha_obj = datetime.date.fromisoformat(str(fecha))
        except ValueError:
            fecha_obj = None

    hoy = timezone.localdate()
    fin_anio_actual = datetime.date(hoy.year, 12, 31)

    reporte_qs = ReporteAsistencia.objects.select_related('asignacion', 'modificado_por', 'reemplazo')
    # Solo considerar overrides de asignaciones activas
    reporte_qs = reporte_qs.filter(asignacion__estado='ACTIVO')
    historial_asig_ids_qs = ReporteAsistenciaHistorial.objects.none().values('asignacion_id')
    if fecha_obj:
        reporte_qs = reporte_qs.filter(fecha_reporte=fecha_obj)
        historial_asig_ids_qs = ReporteAsistenciaHistorial.objects.filter(
            fecha_reporte=fecha_obj,
            asignacion__estado='ACTIVO'
        ).values('asignacion_id').distinct()

    asig_qs = Asignacion.objects.select_related(
        'cliente', 'instalacion', 'instalacion__canton', 'instalacion__canton__provincia',
        'puesto', 'horario', 'persona',
        'instalacion__nominativo', 'instalacion__nominativo__zona',
    ).filter(
        persona__isnull=False,
        persona__is_active=True,
        estado='ACTIVO'
    ).exclude(persona__tipo='SACAFRANCO')

    if fecha_obj:
        # El reporte de asistencia es de UN dia: solo debe traer registros del MES/AÑO de
        # esa fecha. Antes, para fechas hoy/futuras se filtraba solo por año, lo que colaba
        # las asignaciones proyectadas de los meses siguientes (sept, oct...) en la vista
        # "todos". Ahora se restringe siempre al mes de la fecha.
        if fecha_obj >= hoy:
            asig_qs = asig_qs.filter(mes=fecha_obj.month, anio=fecha_obj.year)
            asig_qs = asig_qs.filter(
                Q(fecha__gte=fecha_obj, fecha__lte=fin_anio_actual) |
                Q(fecha__isnull=True) |
                Q(id__in=reporte_qs.values('asignacion_id')) |
                Q(id__in=historial_asig_ids_qs)
            )
        else:
            # Fecha pasada: incluir asignaciones del mes/año exacto, las recurrentes
            # activas en esa fecha (para que el día pasado muestre su calendario D/N),
            # y las que tengan datos guardados/override.
            asig_qs = asig_qs.filter(
                Q(mes=fecha_obj.month, anio=fecha_obj.year) |
                Q(fecha=fecha_obj) |
                (
                    Q(recurring=True) &
                    Q(start_date__lte=fecha_obj) &
                    (Q(end_date__isnull=True) | Q(end_date__gte=fecha_obj))
                ) |
                Q(id__in=reporte_qs.values('asignacion_id')) |
                Q(id__in=historial_asig_ids_qs)
            )
    if cliente_id:
        asig_qs = asig_qs.filter(cliente_id=cliente_id)
    if zona:
        asig_qs = asig_qs.filter(instalacion__nominativo__zona__nombre__iexact=zona).distinct()

    # Corte exacto al día: una asignación con end_date NO debe mostrarse en fechas
    # POSTERIORES a su end_date. Esto hace efectivo el cierre de puesto al día
    # (la persona se ve hasta el día anterior al cierre) y evita que filas con
    # historial fuguen a meses siguientes.
    if fecha_obj:
        asig_qs = asig_qs.exclude(end_date__isnull=False, end_date__lt=fecha_obj)

    # Ruteo por el calendario del día (D/N), no por la config de turno del puesto.
    # D -> Diurno, N -> Nocturno. Los francos (F) NO aparecen en el reporte.
    # NOTA: la cobertura de sacafranco NO se enruta aqui; se agrega mas abajo como
    # fila propia (independiente del fijo) desde su calendario semanal.
    dnf = _calendar_dnf_for_date(fecha_obj)
    if fecha_obj:
        franco_ids = {aid for aid, lt in dnf.items() if lt == 'F'}
        if franco_ids:
            asig_qs = asig_qs.exclude(id__in=franco_ids)
    _TURNO_LETRA = {'Diurno': 'D', 'Nocturno': 'N', 'Tarde': 'T', 'Veinticuatro': 'V'}
    if turno in _TURNO_LETRA and fecha_obj:
        letra_turno = _TURNO_LETRA[turno]
        ids_turno = {aid for aid, lt in dnf.items() if lt == letra_turno}
        asig_qs = asig_qs.filter(id__in=ids_turno)

    # Evitar duplicado: el sacafranco se muestra como su PROPIA fila (mas abajo, desde su
    # calendario) tanto con filtro Diurno/Nocturno como sin filtro de turno. Por eso el fijo
    # que ese dia esta cubierto por un sacafranco NO debe salir tambien como fila de fijo.
    if fecha_obj and (turno in ('Diurno', 'Nocturno') or not turno):
        _cov_latest = (ReporteAsistenciaHistorial.objects
                       .filter(fecha_reporte=fecha_obj)
                       .order_by('asignacion_id', '-creado_en')
                       .distinct('asignacion_id').values('id'))
        covered_ids = {aid for aid, desc in (ReporteAsistenciaHistorial.objects
                       .filter(id__in=Subquery(_cov_latest))
                       .values_list('asignacion_id', 'descripcion'))
                       if _is_auto_sacafranco_desc(desc)}
        if covered_ids:
            asig_qs = asig_qs.exclude(id__in=covered_ids)

    if exclude_sacafranco:
        # Excluir asignaciones marcadas como cobertura de sacafranco (F) en la fecha consultada.
        fecha_ref = fecha_obj or hoy
        day_field_map = {
            0: 'mon',
            1: 'tue',
            2: 'wed',
            3: 'thu',
            4: 'fri',
            5: 'sat',
            6: 'sun',
        }
        day_field = day_field_map.get(fecha_ref.weekday())
        if day_field:
            # Semana estilo calendario mensual del sistema (dia 1 y saltos de 7)
            month_base = fecha_ref.replace(day=1)
            week_start_month = month_base + datetime.timedelta(days=((fecha_ref.day - 1) // 7) * 7)
            # Semana estilo ISO (lunes)
            week_start_iso = fecha_ref - datetime.timedelta(days=fecha_ref.weekday())

            sacafranco_asig_ids = AsignacionSemanal.objects.filter(
                week_start__in=[week_start_month, week_start_iso],
                **{f"{day_field}__istartswith": 'F'}
            ).exclude(asignacion_id__isnull=True).values_list('asignacion_id', flat=True)

            asig_qs = asig_qs.exclude(id__in=sacafranco_asig_ids)

    term = (q or '').strip().lower()
    # Keep pagination in-memory for date report so BASE rows (NB/DB) can be merged reliably.
    use_db_pagination = bool(include_total and page and page_size and not term and not fecha_obj)
    ordered_qs = asig_qs.order_by('mes', 'fecha', 'id')

    total = None
    if use_db_pagination:
        total = ordered_qs.count()
        start = (int(page) - 1) * int(page_size)
        end = start + int(page_size)
        ordered_qs = ordered_qs[start:end]

    asig_list = list(ordered_qs)
    asig_ids = [a.id for a in asig_list]

    overrides = {}
    if fecha_obj:
        latest_hist_ids = ReporteAsistenciaHistorial.objects.filter(
            fecha_reporte=fecha_obj
        )
        if use_db_pagination and asig_ids:
            latest_hist_ids = latest_hist_ids.filter(asignacion_id__in=asig_ids)
        latest_hist_ids = latest_hist_ids.order_by('asignacion_id', '-creado_en').distinct('asignacion_id').values('id')

        hist_qs = ReporteAsistenciaHistorial.objects.select_related('usuario', 'reemplazo')
        hist_qs = hist_qs.filter(id__in=Subquery(latest_hist_ids)).order_by('asignacion_id')
        for h in hist_qs:
            is_auto_sacafranco = _is_auto_sacafranco_desc(h.descripcion)
            overrides[h.asignacion_id] = SimpleNamespace(
                codigo=h.codigo,
                estado=h.estado,
                estado_asistencia=getattr(h, 'estado_asistencia', None),
                descripcion=h.descripcion,
                reemplazo=h.reemplazo,
                persona_cobertura=h.reemplazo if is_auto_sacafranco else None,
                auto_sacafranco=is_auto_sacafranco,
                modificado_por=h.usuario,
                modificado_en=h.creado_en,
                row_color=h.row_color,
                hueca=bool(getattr(h, 'hueca', False)),
                hueca_motivo=getattr(h, 'hueca_motivo', '') or '',
            )

    reporte_iter = reporte_qs
    if use_db_pagination and asig_ids:
        reporte_iter = reporte_iter.filter(asignacion_id__in=asig_ids)
    for r in reporte_iter:
        if r.asignacion_id not in overrides:
            overrides[r.asignacion_id] = r

    data = []
    personas_con_asignacion = set()

    # Horario del PUESTO para el día del reporte (manda sobre el Horario de la asignación):
    # se busca el PuestoHorario del puesto para ese día de la semana.
    horario_puesto_dia = {}
    if fecha_obj:
        from ..models import PuestoHorario
        dia_semana = fecha_obj.weekday() + 1  # 1=Lunes ... 7=Domingo
        puesto_ids = [getattr(a, 'puesto_id', None) for a in asig_list if getattr(a, 'puesto_id', None)]
        if puesto_ids:
            for ph in PuestoHorario.objects.filter(dia=dia_semana, puesto_id__in=puesto_ids):
                if ph.hora_ingreso and ph.puesto_id not in horario_puesto_dia:
                    horario_puesto_dia[ph.puesto_id] = (ph.hora_ingreso, ph.hora_salida)

    for asig in asig_list:
        p = asig.persona
        personas_con_asignacion.add(p.id)
        override = overrides.get(asig.id)

        cliente_nombre = getattr(asig.cliente, 'nombre_comercial', '') if asig else ''

        puesto_nombre = getattr(asig.puesto, 'nombre', '') if asig else ''
        if not puesto_nombre and asig:
            puesto_nombre = getattr(asig.puesto, 'tipo', '')
        horario_str = ''
        ph_horas = horario_puesto_dia.get(getattr(asig, 'puesto_id', None)) if asig else None
        if ph_horas and ph_horas[0]:
            hi, ho = ph_horas
            horario_str = f"{hi.strftime('%H:%M')} - {ho.strftime('%H:%M')}" if ho else hi.strftime('%H:%M')
        elif asig and asig.horario:
            horario_str = f"{asig.horario.hora_ingreso.strftime('%H:%M')} - {asig.horario.hora_salida.strftime('%H:%M')}"
        nombre_apellidos = f"{p.nombres} {p.apellidos}".strip()
        auto_sacafranco = _is_auto_sacafranco_desc(getattr(override, 'descripcion', '')) if override else False
        persona_cobertura = None
        if override:
            persona_cobertura = getattr(override, 'persona_cobertura', None)
            if not persona_cobertura and auto_sacafranco:
                persona_cobertura = getattr(override, 'reemplazo', None)
        if auto_sacafranco and persona_cobertura:
            nombre_apellidos = f"{persona_cobertura.nombres} {persona_cobertura.apellidos}".strip()
        zona_titulo = ''
        provincia_nombre = ''
        if asig and asig.instalacion:
            # La zona sale del NOMINATIVO de la instalacion (letra+numero -> zona 1/2/3).
            # getattr con default: si la instalacion no tiene nominativo, queda 'SIN ZONA'.
            nom_obj = getattr(asig.instalacion, 'nominativo', None)
            if nom_obj and nom_obj.zona_id:
                zona_titulo = nom_obj.zona.nombre
        if asig and asig.instalacion and asig.instalacion.canton and asig.instalacion.canton.provincia:
            provincia_nombre = asig.instalacion.canton.provincia.nombre
        reemplazo_nombre = ''
        reemplazo_id = None
        if override and override.reemplazo:
            reemplazo_id = override.reemplazo.id
            reemplazo_nombre = f"{override.reemplazo.nombres} {override.reemplazo.apellidos}".strip()
        if auto_sacafranco:
            reemplazo_id = None
            reemplazo_nombre = ''

        modificado_por_nombre = ''
        modificado_en_iso = None
        if override:
            if override.modificado_por:
                full_name = f"{override.modificado_por.first_name} {override.modificado_por.last_name}".strip()
                modificado_por_nombre = full_name or override.modificado_por.get_username()
            modificado_en_iso = override.modificado_en.isoformat() if override.modificado_en else None

        codigo_instalacion = getattr(asig.instalacion, 'codigo', '') if asig and asig.instalacion else ''
        estado_asistencia = _normalize_estado_asistencia(getattr(override, 'estado_asistencia', '') if override else '')
        estado = (getattr(override, 'estado', None) if override else None) or 'TURNO'
        descripcion = (override.descripcion or '') if override else ''
        if auto_sacafranco:
            puesto_nombre = ''
            horario_str = ''
            estado = ''
            estado_asistencia = ''
            descripcion = ''

        data.append({
            'asignacion_id': asig.id,
            'codigo': override.codigo if (override and override.codigo) else (codigo_instalacion or ''),
            'cliente': cliente_nombre,
            'instalacion_nombre': (getattr(asig.instalacion, 'nombre', '') or '') if asig else '',
            'puesto': puesto_nombre,
            'puesto_tipo': (getattr(asig.puesto, 'tipo', '') or '') if asig else '',
            'horario': horario_str,
            'nombre_apellidos': nombre_apellidos,
            'cedula': getattr(p, 'cedula', '') or '',
            'reemplazo_id': reemplazo_id,
            'reemplazo': reemplazo_nombre,
            'estado_asistencia': estado_asistencia,
            'estado': estado,
            'descripcion': descripcion,
            'modificado_por': modificado_por_nombre,
            'row_color': (override.row_color or '') if override else '',
            'hueca': bool(getattr(override, 'hueca', False)) if override else False,
            'hueca_motivo': (getattr(override, 'hueca_motivo', '') or '') if override else '',
            'modificado_en': modificado_en_iso,
            'zona_titulo': zona_titulo,
            'provincia': provincia_nombre,
        })

    # Filas de SACAFRANCO desde su calendario semanal, ruteadas por el token del dia.
    # Token = letra de turno (D/N) + nominativo (codigo de instalacion) o 'B' (base):
    #   D{nominativo}/N{nominativo} -> fila en ese turno con el CLIENTE de ese nominativo.
    #   DB/NB -> base, cliente 'SEGURIDAD FISICA', puesto 'DIA BASE'/'NOCHE BASE'.
    #   F u otro turno -> no aparece.
    if fecha_obj and (turno in ['Diurno', 'Nocturno'] or not turno):
        day_key_map = {0: 'mon', 1: 'tue', 2: 'wed', 3: 'thu', 4: 'fri', 5: 'sat', 6: 'sun'}
        day_key = day_key_map.get(fecha_obj.weekday())
        if day_key:
            # Sin filtro de turno -> se muestran TODOS los sacafranco (tokens D y N).
            # Con filtro Diurno/Nocturno -> solo los de ese turno.
            _turno_letters = ('D',) if turno == 'Diurno' else (('N',) if turno == 'Nocturno' else ('D', 'N'))
            month_base = fecha_obj.replace(day=1)
            week_start_month = month_base + datetime.timedelta(days=((fecha_obj.day - 1) // 7) * 7)
            week_start_iso = fecha_obj - datetime.timedelta(days=fecha_obj.weekday())
            zona_filter_norm = _normalize_zona_filter(zona) if zona else ''

            sac_qs = SacafrancoFilaSemanal.objects.select_related(
                'sacafranco_fila', 'sacafranco_fila__persona', 'sacafranco_fila__provincia'
            ).filter(week_start__in=[week_start_month, week_start_iso])

            # Asistencia marcada del sacafranco (ASISTIO/FALTO) para ESTA fecha, por fila.
            _saca_asist = {}
            for _sa in SacafrancoAsistencia.objects.filter(fecha=fecha_obj).select_related('modificado_por'):
                _saca_asist[_sa.sacafranco_fila_id] = _sa

            inst_cache = {}

            def _ctx_nominativo(nom):
                if nom in inst_cache:
                    return inst_cache[nom]
                inst = Instalacion.objects.select_related(
                    'cliente', 'canton', 'canton__provincia', 'nominativo', 'nominativo__zona'
                ).filter(codigo__iexact=nom).first()
                if inst:
                    _nomobj = getattr(inst, 'nominativo', None)
                    _zona = _nomobj.zona.nombre if (_nomobj and getattr(_nomobj, 'zona_id', None)) else 'SIN ZONA'
                    ctx_val = {
                        'cliente': getattr(inst.cliente, 'nombre_comercial', '') or '',
                        'puesto': inst.nombre or '',
                        'codigo': inst.codigo or nom,
                        'provincia': getattr(getattr(getattr(inst, 'canton', None), 'provincia', None), 'nombre', '') or 'SIN PROVINCIA',
                        'zona': _zona,
                    }
                else:
                    ctx_val = {'cliente': '', 'puesto': '', 'codigo': nom, 'provincia': 'SIN PROVINCIA', 'zona': 'SIN ZONA'}
                inst_cache[nom] = ctx_val
                return ctx_val

            seen_fila_ids = set()
            for srow in sac_qs:
                token_val = str(getattr(srow, day_key, '') or '').strip().upper()
                if not token_val or token_val[0] not in _turno_letters:
                    continue
                fila = getattr(srow, 'sacafranco_fila', None)
                persona = getattr(fila, 'persona', None) if fila else None
                if not fila or not persona:
                    continue
                if fila.id in seen_fila_ids:
                    continue
                seen_fila_ids.add(fila.id)

                persona_nombre = f"{persona.nombres} {persona.apellidos}".strip()
                nominativo = token_val[1:].strip()
                if nominativo in ('', 'B'):
                    codigo_val = 'BASE'
                    cliente_val = 'SEGURIDAD FISICA'
                    puesto_val = 'DIA BASE' if token_val[0] == 'D' else 'NOCHE BASE'
                    provincia_val = (getattr(getattr(fila, 'provincia', None), 'nombre', None)
                                     or getattr(getattr(persona, 'provincia', None), 'nombre', None)
                                     or 'SIN PROVINCIA')
                    zona_val = 'SIN ZONA'
                else:
                    # Cobertura: el nominativo es el código de la instalación cubierta.
                    # Toma cliente, provincia y ZONA de esa instalación para que la fila
                    # del sacafranco salga agrupada/ordenada junto a ese nominativo.
                    ctx_n = _ctx_nominativo(nominativo)
                    codigo_val = ctx_n['codigo']
                    cliente_val = ctx_n['cliente']
                    puesto_val = ctx_n['puesto']
                    provincia_val = ctx_n['provincia']
                    zona_val = ctx_n['zona']

                # Respetar filtro de zona: si hay filtro activo, solo salen los sacafranco
                # cuyo nominativo pertenece a esa zona.
                if zona_filter_norm and _normalize_zona_filter(zona_val) != zona_filter_norm:
                    continue

                _hi = getattr(fila, 'hora_ingreso', None)
                _ho = getattr(fila, 'hora_salida', None)
                horario_saca = ''
                if _hi or _ho:
                    horario_saca = f"{_hi.strftime('%H:%M') if _hi else ''} {_ho.strftime('%H:%M') if _ho else ''}".strip()

                # Asistencia/edicion marcada del sacafranco para esta fecha (si existe).
                _sa = _saca_asist.get(fila.id)
                _sa_estado = _normalize_estado_asistencia(getattr(_sa, 'estado_asistencia', '')) if _sa else ''
                _sa_modpor = ''
                _sa_moden = None
                _sa_rem = getattr(_sa, 'reemplazo', None) if _sa else None
                if _sa:
                    _u = getattr(_sa, 'modificado_por', None)
                    if _u:
                        _sa_modpor = f"{_u.first_name} {_u.last_name}".strip() or _u.get_username()
                    _sa_moden = _sa.modificado_en.isoformat() if _sa.modificado_en else None

                data.append({
                    'asignacion_id': None,
                    'sacafranco_fila_id': fila.id,
                    'codigo': codigo_val,
                    'cliente': cliente_val,
                    'puesto': puesto_val,
                    'horario': horario_saca,
                    'nombre_apellidos': persona_nombre or 'Libre en base',
                    'reemplazo_id': _sa_rem.id if _sa_rem else None,
                    'reemplazo': f"{_sa_rem.nombres} {_sa_rem.apellidos}".strip() if _sa_rem else '',
                    'estado_asistencia': _sa_estado,
                    'estado': (getattr(_sa, 'estado', '') or 'TURNO') if _sa else 'TURNO',
                    'descripcion': (getattr(_sa, 'descripcion', '') or '') if _sa else '',
                    'modificado_por': _sa_modpor,
                    'row_color': (getattr(_sa, 'row_color', '') or '') if _sa else '',
                    'hueca': bool(getattr(_sa, 'hueca', False)) if _sa else False,
                    'hueca_motivo': (getattr(_sa, 'hueca_motivo', '') or '') if _sa else '',
                    'modificado_en': _sa_moden,
                    'zona_titulo': zona_val,
                    'provincia': provincia_val,
                })

    if term:
        tokens = [t for t in term.split() if t]
        filtered = []
        for item in data:
            haystack = ' '.join([
                str(item.get('codigo') or ''),
                str(item.get('cliente') or ''),
                str(item.get('puesto') or ''),
                str(item.get('horario') or ''),
                str(item.get('nombre_apellidos') or ''),
                str(item.get('estado') or ''),
                str(item.get('estado_asistencia') or ''),
                str(item.get('reemplazo') or ''),
                str(item.get('descripcion') or ''),
                str(item.get('zona_titulo') or ''),
                str(item.get('provincia') or ''),
            ]).lower()
            if all(token in haystack for token in tokens):
                filtered.append(item)
        data = filtered

    # Ordenar dentro de cada provincia por el NOMINATIVO (codigo): agrupado por letra
    # y de menor a mayor por numero (K1, K2, K3, luego S1, S2...). Se conserva el orden
    # de aparicion de las provincias. Solo cuando esta todo en memoria (vista por fecha).
    if not use_db_pagination:
        import re as _re
        _prov_order = {}
        for _it in data:
            _p = str(_it.get('provincia') or '')
            if _p not in _prov_order:
                _prov_order[_p] = len(_prov_order)

        def _cod_key(item):
            p = str(item.get('provincia') or '')
            cod = str(item.get('codigo') or '').strip().upper()
            m = _re.match(r'([A-Z]+)\s*0*(\d+)', cod)
            if m:
                return (_prov_order.get(p, 9999), 0, m.group(1), int(m.group(2)), cod)
            # codigos sin patron letra+numero: al final de su provincia
            return (_prov_order.get(p, 9999), 1, cod, 0, cod)

        data.sort(key=_cod_key)

    if include_total:
        if use_db_pagination:
            return data, int(total or 0), True
        return data, len(data), False

    return data

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def obtener_reporte_asistencia(request):
    if not request.user.has_perm('CoreFisica.view_reporteasistencia'):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    fecha = request.GET.get('fecha')
    cliente_id = request.GET.get('cliente_id')
    turno = request.GET.get('turno')
    q = (request.GET.get('q') or '').strip()
    zona = _normalize_zona_filter(request.GET.get('zona'))
    try:
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 50))
        if page < 1:
            page = 1
        if page_size < 1:
            page_size = 50
        page_size = min(page_size, 200)
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Parámetros de paginación inválidos'}, status=400)

    exclude_sacafranco = str(request.GET.get('exclude_sacafranco', '')).strip() == '1'
    data, total, db_paginated = _build_reporte_asistencia_data(
        fecha=fecha,
        cliente_id=cliente_id,
        turno=turno,
        exclude_sacafranco=exclude_sacafranco,
        zona=zona,
        q=q,
        page=page,
        page_size=page_size,
        include_total=True,
    )

    start = (page - 1) * page_size
    end = start + page_size
    page_results = data if db_paginated else data[start:end]
    total_pages = (total + page_size - 1) // page_size if total else 1

    return JsonResponse({
        'results': page_results,
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': total_pages,
        'zona': zona or None,
    }, safe=False, status=status.HTTP_200_OK)
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def listar_descripciones_reporte(request):
    if not request.user.has_perm('CoreFisica.view_reporteasistencia'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    descripciones = (
        ReporteAsistencia.objects
        .exclude(descripcion__isnull=True)
        .exclude(descripcion__exact='')
        .values_list('descripcion', flat=True)
        .distinct()
        .order_by('descripcion')
    )

    return JsonResponse(list(descripciones), safe=False, status=status.HTTP_200_OK)


def _sync_reporte_guardia(override, asignacion, fecha_reporte):
    """Refleja el registro de asistencia en el REPORTE DE GUARDIA.
    - Titular que FALTÓ -> Faltos.
    - Reemplazo según su estado -> Dobladas/Adicionales/Adelantos (con SU tipo en 'proviene').
    Puede generar 2 filas (titular faltó + reemplazo cubrió). Idempotente por reporte+fecha."""
    from ..models import ReporteGuardia

    # Preservar el motivo escrito a mano en la HUECA antes de regenerar (se borra y recrea).
    hueca_motivo = ''
    if fecha_reporte:
        _h = ReporteGuardia.objects.filter(
            reporte_asistencia=override, fecha=fecha_reporte, seccion='HUECA', auto=True
        ).first()
        if _h:
            hueca_motivo = _h.motivo or ''

    # Preservar los campos editados a mano (overrides) por sección, para
    # re-aplicarlos tras regenerar las filas auto.
    prev_overrides = {}
    for _r in ReporteGuardia.objects.filter(reporte_asistencia=override, auto=True):
        if _r.overrides:
            prev_overrides.setdefault(_r.seccion, _r.overrides)

    # Quitar TODAS las filas auto previas de este reporte (cualquier fecha) para reflejar
    # cambios y no dejar huérfanos si cambió la fecha del reporte. El ReporteAsistencia
    # tiene una sola fecha_reporte vigente, así que solo debe existir la de la fecha actual.
    ReporteGuardia.objects.filter(reporte_asistencia=override, auto=True).delete()
    if not fecha_reporte:
        return

    # Turno desde el calendario D/N de esa fecha.
    letra = _calendar_dnf_for_date(fecha_reporte).get(asignacion.id)
    turno = 'Diurno' if letra == 'D' else ('Nocturno' if letra == 'N' else '')
    if turno not in ('Diurno', 'Nocturno'):
        return

    cliente = getattr(asignacion.cliente, 'nombre_comercial', '') or ''
    puesto = getattr(asignacion.puesto, 'nombre', '') or ''

    def _nombre(p):
        return f"{p.nombres} {p.apellidos}".strip() if p else ''

    # (seccion, persona, motivo) — se crean todas las que apliquen.
    filas = []

    # 1) Titular que faltó -> Faltos.
    if (override.estado_asistencia or '').upper() == 'FALTO':
        filas.append(('FALTOS', asignacion.persona, 'FALTO'))

    # 2) Reemplazo según su estado -> sección correspondiente (acepta DOBLA y DOBLADO).
    #    FR/TRABAJADO se refleja aparte, automático al guardar (ver _sync_frtrabajado_dobladas).
    estado = (override.estado or '').upper()
    mapa = {
        'DOBLA': 'DOBLADAS', 'DOBLADO': 'DOBLADAS',
        'ADICIONAL': 'ADICIONALES',
        'ADEL/TURNO': 'ADELANTOS',
    }
    seccion_reemplazo = mapa.get(estado)
    if seccion_reemplazo and override.reemplazo:
        filas.append((seccion_reemplazo, override.reemplazo, ''))

    for seccion, persona, motivo in filas:
        _row = ReporteGuardia.objects.create(
            fecha=fecha_reporte,
            turno=turno,
            seccion=seccion,
            cliente=cliente,
            puesto=puesto,
            persona_nombre=_nombre(persona),
            persona_ref=persona,          # 'proviene' se autocompleta con persona.tipo en el save()
            reporte_asistencia=override,
            auto=True,
            motivo=motivo,
        )
        _ov = prev_overrides.get(seccion)
        if _ov:
            for _k, _v in _ov.items():
                setattr(_row, _k, _v)
            _row.overrides = _ov
            _row.save()

    # HUECA espeja al ADICIONAL: solo Cliente/Puesto/Fecha + Motivo editable (preservado).
    # PERO si el check "Hueca" está marcado, la fila manual (auto=False) que crea
    # _sync_hueca_reporte_guardia ya la representa (con su motivo). No duplicar aquí.
    if seccion_reemplazo == 'ADICIONALES' and not getattr(override, 'hueca', False):
        _hueca = ReporteGuardia.objects.create(
            fecha=fecha_reporte,
            turno=turno,
            seccion='HUECA',
            cliente=cliente,
            puesto=puesto,
            fecha_evento=fecha_reporte,   # columna "Fecha" = el día del reporte
            reporte_asistencia=override,
            auto=True,
            motivo=hueca_motivo,          # conserva el motivo escrito a mano
        )
        _ov = prev_overrides.get('HUECA')
        if _ov:
            for _k, _v in _ov.items():
                setattr(_hueca, _k, _v)
            _hueca.overrides = _ov
            _hueca.save()


def _sync_hueca_reporte_guardia(override, asignacion, fecha_reporte):
    """Refleja el check 'Hueca' (marcado a mano en el diálogo de asistencia) en el
    REPORTE DE GUARDIA, sección HUECA. Se crea como fila MANUAL (auto=False) para que
    el botón 'Regenerar desde asistencia' (que borra las auto=True) NO la pise.

    - hueca=True  -> crea/actualiza la fila HUECA con su motivo (dropdown).
    - hueca=False -> la elimina.
    """
    from ..models import ReporteGuardia

    qs = ReporteGuardia.objects.filter(reporte_asistencia=override, seccion='HUECA', auto=False)
    # Sin check o sin fecha: no debe existir la hueca manual.
    if not getattr(override, 'hueca', False) or not fecha_reporte:
        qs.delete()
        return

    # Turno según el calendario D/N de esa fecha; si no hay, Diurno por defecto.
    letra = _calendar_dnf_for_date(fecha_reporte).get(asignacion.id)
    turno = 'Diurno' if letra == 'D' else ('Nocturno' if letra == 'N' else 'Diurno')

    cliente = getattr(asignacion.cliente, 'nombre_comercial', '') or ''
    puesto = getattr(asignacion.puesto, 'nombre', '') or ''

    # Borrar y recrear (una sola hueca manual por reporte).
    qs.delete()
    ReporteGuardia.objects.create(
        fecha=fecha_reporte,
        turno=turno,
        seccion='HUECA',
        cliente=cliente,
        puesto=puesto,
        fecha_evento=fecha_reporte,
        reporte_asistencia=override,
        auto=False,
        motivo=override.hueca_motivo or '',
    )


def _sync_frtrabajado_dobladas(override, asignacion, fecha_reporte):
    """Refleja el estado FR/TRABAJADO (franco trabajado) en REPORTE DE GUARDIA,
    sección DOBLADAS, automático al guardar. Fila MANUAL (auto=False) para que el
    botón 'Regenerar' (que borra las auto=True) NO la pise ni la duplique.

    - estado == FR/TRABAJADO y hay reemplazo -> crea/actualiza la fila DOBLADAS
      con la persona que trabajó su franco.
    - en otro caso -> la elimina.
    """
    from ..models import ReporteGuardia

    qs = ReporteGuardia.objects.filter(
        reporte_asistencia=override, seccion='DOBLADAS', auto=False
    )
    estado = (getattr(override, 'estado', '') or '').upper()
    rem = getattr(override, 'reemplazo', None)
    if estado != 'FR/TRABAJADO' or not fecha_reporte or not rem:
        qs.delete()
        return

    letra = _calendar_dnf_for_date(fecha_reporte).get(asignacion.id)
    turno = 'Diurno' if letra == 'D' else ('Nocturno' if letra == 'N' else 'Diurno')
    cliente = getattr(asignacion.cliente, 'nombre_comercial', '') or ''
    puesto = getattr(asignacion.puesto, 'nombre', '') or ''

    # PROVIENE = de dónde viene el que trabajó su franco: FT-<cliente>-<puesto> de SU
    # propio puesto (su asignación activa). Si no se encuentra, queda solo "FT".
    origen = Asignacion.objects.select_related('cliente', 'puesto').filter(
        persona=rem, estado='ACTIVO'
    ).order_by('-anio', '-mes').first()
    if origen:
        _cli = getattr(origen.cliente, 'nombre_comercial', '') or ''
        _pue = getattr(origen.puesto, 'nombre', '') or ''
        proviene = '-'.join([x for x in ['FT', _cli, _pue] if x])
    else:
        proviene = 'FT'

    qs.delete()
    ReporteGuardia.objects.create(
        fecha=fecha_reporte,
        turno=turno,
        seccion='DOBLADAS',
        cliente=cliente,
        puesto=puesto,
        persona_nombre=f"{rem.nombres} {rem.apellidos}".strip(),
        persona_ref=rem,
        proviene=proviene,               # FT-<cliente>-<puesto> de origen (no el tipo)
        reporte_asistencia=override,
        auto=False,
    )


def _sacafranco_token_for_date(sacafranco_fila_id, fecha):
    """Token (ej 'DS6') del sacafranco para esa fecha, o '' si no hay."""
    from ..models import SacafrancoFilaSemanal
    day_field_map = {0: 'mon', 1: 'tue', 2: 'wed', 3: 'thu', 4: 'fri', 5: 'sat', 6: 'sun'}
    day_field = day_field_map.get(fecha.weekday())
    if not day_field:
        return ''
    month_base = fecha.replace(day=1)
    week_start_month = month_base + datetime.timedelta(days=((fecha.day - 1) // 7) * 7)
    week_start_iso = fecha - datetime.timedelta(days=fecha.weekday())
    row = SacafrancoFilaSemanal.objects.filter(
        sacafranco_fila_id=sacafranco_fila_id,
        week_start__in=[week_start_month, week_start_iso],
    ).first()
    if not row:
        return ''
    return str(getattr(row, day_field, '') or '').strip().upper()


def _saca_guardia_ctx(fila, fecha_reporte):
    """(turno, cliente, puesto) del nominativo que cubre el sacafranco ese dia, segun su
    token. Devuelve None si ese dia no tiene cobertura Diurno/Nocturno."""
    from ..models import Instalacion
    from .asignacion_semanal_views import _parse_sacafranco_token
    token = _sacafranco_token_for_date(fila.id, fecha_reporte)
    _t, tturno, tcode, _i, _p = _parse_sacafranco_token(token)
    if tturno not in ('Diurno', 'Nocturno'):
        return None
    cliente = ''
    puesto = ''
    if tcode and tcode != 'BASE':
        inst = Instalacion.objects.select_related('cliente').filter(codigo__iexact=tcode).first()
        if inst:
            cliente = getattr(inst.cliente, 'nombre_comercial', '') or ''
            puesto = inst.nombre or ''
    else:
        cliente = 'SEGURIDAD FISICA'
    return tturno, cliente, puesto


def _sync_reporte_guardia_sacafranco(sa, fecha_reporte):
    """Refleja la asistencia de un SACAFRANCO (SacafrancoAsistencia) en el REPORTE DE
    GUARDIA: FALTO -> Faltos; reemplazo segun estado -> Dobladas/Adicionales/Adelantos;
    ADICIONAL -> Hueca auto. Idempotente por sacafranco_fila. Preserva overrides y motivo."""
    from ..models import ReporteGuardia
    fila = getattr(sa, 'sacafranco_fila', None)
    if not fila:
        return

    hueca_motivo_prev = ''
    if fecha_reporte:
        _h = ReporteGuardia.objects.filter(
            sacafranco_fila=fila, fecha=fecha_reporte, seccion='HUECA', auto=True
        ).first()
        if _h:
            hueca_motivo_prev = _h.motivo or ''
    prev_overrides = {}
    for _r in ReporteGuardia.objects.filter(sacafranco_fila=fila, auto=True):
        if _r.overrides:
            prev_overrides.setdefault(_r.seccion, _r.overrides)

    ReporteGuardia.objects.filter(sacafranco_fila=fila, auto=True).delete()
    if not fecha_reporte:
        return
    ctx = _saca_guardia_ctx(fila, fecha_reporte)
    if not ctx:
        return
    turno, cliente, puesto = ctx
    persona = getattr(fila, 'persona', None)

    def _nombre(p):
        return f"{p.nombres} {p.apellidos}".strip() if p else ''

    filas = []
    if (sa.estado_asistencia or '').upper() == 'FALTO':
        filas.append(('FALTOS', persona, 'FALTO'))
    estado = (sa.estado or '').upper()
    mapa = {'DOBLA': 'DOBLADAS', 'DOBLADO': 'DOBLADAS', 'ADICIONAL': 'ADICIONALES', 'ADEL/TURNO': 'ADELANTOS'}
    seccion_reemplazo = mapa.get(estado)
    rem = getattr(sa, 'reemplazo', None)
    if seccion_reemplazo and rem:
        filas.append((seccion_reemplazo, rem, ''))

    for seccion, per, motivo in filas:
        _row = ReporteGuardia.objects.create(
            fecha=fecha_reporte, turno=turno, seccion=seccion,
            cliente=cliente, puesto=puesto,
            persona_nombre=_nombre(per), persona_ref=per,
            sacafranco_fila=fila, auto=True, motivo=motivo,
        )
        _ov = prev_overrides.get(seccion)
        if _ov:
            for _k, _v in _ov.items():
                setattr(_row, _k, _v)
            _row.overrides = _ov
            _row.save()

    if seccion_reemplazo == 'ADICIONALES' and not getattr(sa, 'hueca', False):
        _hueca = ReporteGuardia.objects.create(
            fecha=fecha_reporte, turno=turno, seccion='HUECA',
            cliente=cliente, puesto=puesto, fecha_evento=fecha_reporte,
            sacafranco_fila=fila, auto=True, motivo=hueca_motivo_prev,
        )
        _ov = prev_overrides.get('HUECA')
        if _ov:
            for _k, _v in _ov.items():
                setattr(_hueca, _k, _v)
            _hueca.overrides = _ov
            _hueca.save()


def _sync_hueca_reporte_guardia_sacafranco(sa, fecha_reporte):
    """Refleja el check 'Hueca' del sacafranco como fila MANUAL (auto=False) en HUECA."""
    from ..models import ReporteGuardia
    fila = getattr(sa, 'sacafranco_fila', None)
    if not fila:
        return
    qs = ReporteGuardia.objects.filter(sacafranco_fila=fila, seccion='HUECA', auto=False)
    if not getattr(sa, 'hueca', False) or not fecha_reporte:
        qs.delete()
        return
    ctx = _saca_guardia_ctx(fila, fecha_reporte)
    turno, cliente, puesto = ctx if ctx else ('Diurno', '', '')
    qs.delete()
    ReporteGuardia.objects.create(
        fecha=fecha_reporte, turno=turno, seccion='HUECA',
        cliente=cliente, puesto=puesto, fecha_evento=fecha_reporte,
        sacafranco_fila=fila, auto=False, motivo=sa.hueca_motivo or '',
    )


@api_view(['PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
def insertar_reporte_asistencia(request, asignacion_id):
    if not request.user.has_perm('CoreFisica.change_reporteasistencia'):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    override, _ = ReporteAsistencia.objects.get_or_create(asignacion_id=asignacion_id)

    asignacion = Asignacion.objects.select_related(
        'persona', 'cliente', 'instalacion', 'puesto', 'horario'
    ).filter(id=asignacion_id).first()
    if not asignacion:
        return JsonResponse({'error': 'Asignacion no encontrada'}, status=status.HTTP_404_NOT_FOUND)

    override.persona = asignacion.persona
    override.cliente = asignacion.cliente
    override.instalacion = asignacion.instalacion
    override.puesto = asignacion.puesto
    override.horario = asignacion.horario
    override.puesto_tipo = getattr(asignacion.puesto, 'tipo', None) if asignacion.puesto else None

    override.codigo = getattr(asignacion.instalacion, 'codigo', None)

    for field in ['estado', 'estado_asistencia', 'descripcion', 'row_color']:
        if field in request.data:
            val = request.data.get(field) or None
            setattr(override, field, val)

    if 'hueca' in request.data:
        _hv = request.data.get('hueca')
        override.hueca = str(_hv).strip().lower() in ('1', 'true', 'si', 'on', 'yes', 'verdadero')
    if 'hueca_motivo' in request.data:
        override.hueca_motivo = (request.data.get('hueca_motivo') or '').strip()

    fecha_reporte = override.fecha_reporte
    if 'fecha' in request.data:
        fecha_param = request.data.get('fecha')
        if fecha_param in [None, '', 'null']:
            fecha_reporte = None
        else:
            try:
                fecha_reporte = datetime.date.fromisoformat(str(fecha_param))
            except ValueError:
                fecha_reporte = None

    reemplazo_result, err = _resolver_reemplazo_desde_request(
        request,
        fecha_reporte=fecha_reporte,
        asignacion_id_actual=asignacion_id,
    )
    if err:
        return err
    if reemplazo_result != 'no-enviado':
        override.reemplazo = reemplazo_result

    if request.user and request.user.is_authenticated:
        override.modificado_por = request.user
    override.fecha_reporte = fecha_reporte
    override.modificado_en = timezone.now()
    override.save()

    try:
        ReporteAsistenciaHistorial.objects.create(
            reporte=override,
            asignacion=asignacion,
            fecha_reporte=override.fecha_reporte,
            usuario=request.user if request.user and request.user.is_authenticated else None,
            codigo=override.codigo,
            estado_asistencia=override.estado_asistencia,
            estado=override.estado,
            reemplazo=override.reemplazo,
            descripcion=override.descripcion,
            row_color=override.row_color,
            hueca=bool(override.hueca),
            hueca_motivo=override.hueca_motivo or '',
        )
    except Exception:
        pass

    # El reporte de guardia YA NO se sincroniza automaticamente al guardar la
    # asistencia. Ahora se regenera bajo demanda con el boton "Regenerar desde
    # asistencia" (endpoint regenerar_reporte_guardia), para que las correcciones
    # manuales del reporte de guardia no se pisen solas.
    # EXCEPCION: el check 'Hueca' y el estado FR/TRABAJADO SI se reflejan
    # automaticamente al guardar (HUECA y DOBLADAS respectivamente).
    try:
        _sync_hueca_reporte_guardia(override, asignacion, fecha_reporte)
    except Exception:
        pass
    try:
        _sync_frtrabajado_dobladas(override, asignacion, fecha_reporte)
    except Exception:
        pass

    modificado_por_nombre = ''
    if override.modificado_por:
        full_name = f"{override.modificado_por.first_name} {override.modificado_por.last_name}".strip()
        modificado_por_nombre = full_name or override.modificado_por.get_username()

    reemplazo_nombre = ''
    if override.reemplazo:
        reemplazo_nombre = f"{override.reemplazo.nombres} {override.reemplazo.apellidos}".strip()

    return JsonResponse({
        'codigo': override.codigo or '',
        'estado_asistencia': _normalize_estado_asistencia(override.estado_asistencia),
        'estado': override.estado or 'TURNO',
        'descripcion': override.descripcion or '',
        'reemplazo_id': override.reemplazo_id,
        'reemplazo': reemplazo_nombre,
        'modificado_por': modificado_por_nombre,
        'row_color': override.row_color or '',
        'hueca': bool(override.hueca),
        'hueca_motivo': override.hueca_motivo or '',
        'modificado_en': override.modificado_en.isoformat() if override.modificado_en else None,
    }, status=status.HTTP_200_OK)


@api_view(['PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
def marcar_sacafranco_asistencia(request, sacafranco_fila_id):
    """Marca la asistencia (ASISTIO/FALTO/'') de un sacafranco para una fecha. El
    sacafranco no tiene Asignacion; se guarda por SacafrancoFila + fecha."""
    if not request.user.has_perm('CoreFisica.change_reporteasistencia'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    from ..models import SacafrancoFila
    fila = SacafrancoFila.objects.filter(id=sacafranco_fila_id).first()
    if not fila:
        return JsonResponse({'error': 'Sacafranco no encontrado'}, status=status.HTTP_404_NOT_FOUND)
    try:
        fecha = datetime.date.fromisoformat(str(request.data.get('fecha')))
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Fecha invalida'}, status=status.HTTP_400_BAD_REQUEST)

    obj, _ = SacafrancoAsistencia.objects.get_or_create(sacafranco_fila=fila, fecha=fecha)

    if 'estado_asistencia' in request.data:
        est_a = str(request.data.get('estado_asistencia') or '').strip().upper()
        if est_a not in ('ASISTIO', 'FALTO', ''):
            return JsonResponse({'error': 'estado_asistencia invalido'}, status=status.HTTP_400_BAD_REQUEST)
        obj.estado_asistencia = est_a
    if 'estado' in request.data:
        obj.estado = (request.data.get('estado') or '').strip()
    if 'descripcion' in request.data:
        obj.descripcion = (request.data.get('descripcion') or '').strip()
    if 'row_color' in request.data:
        obj.row_color = (request.data.get('row_color') or '').strip()
    if 'hueca' in request.data:
        _hv = request.data.get('hueca')
        obj.hueca = str(_hv).strip().lower() in ('1', 'true', 'si', 'on', 'yes', 'verdadero')
    if 'hueca_motivo' in request.data:
        obj.hueca_motivo = (request.data.get('hueca_motivo') or '').strip()
    if 'reemplazo_id' in request.data:
        _rid = request.data.get('reemplazo_id')
        if _rid in (None, '', 'null'):
            obj.reemplazo = None
        else:
            try:
                obj.reemplazo = Persona.objects.filter(id=int(_rid)).first()
            except (ValueError, TypeError):
                obj.reemplazo = None
    obj.modificado_por = request.user
    obj.save()

    # Reflejar en el REPORTE DE GUARDIA (Faltos/Dobladas/Adicionales/Hueca).
    try:
        _sync_reporte_guardia_sacafranco(obj, fecha)
        _sync_hueca_reporte_guardia_sacafranco(obj, fecha)
    except Exception:
        pass

    _u = obj.modificado_por
    _rem = obj.reemplazo
    return JsonResponse({
        'sacafranco_fila_id': fila.id,
        'fecha': fecha.isoformat(),
        'estado_asistencia': obj.estado_asistencia,
        'estado': obj.estado or 'TURNO',
        'reemplazo_id': _rem.id if _rem else None,
        'reemplazo': f"{_rem.nombres} {_rem.apellidos}".strip() if _rem else '',
        'descripcion': obj.descripcion or '',
        'hueca': obj.hueca,
        'hueca_motivo': obj.hueca_motivo or '',
        'row_color': obj.row_color or '',
        'modificado_por': (f"{_u.first_name} {_u.last_name}".strip() or _u.get_username()) if _u else '',
        'modificado_en': obj.modificado_en.isoformat() if obj.modificado_en else None,
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def historial_reporte_asistencia(request, asignacion_id):
    if not request.user.has_perm('CoreFisica.view_reporteasistencia'):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    fecha = request.GET.get('fecha')
    qs = ReporteAsistenciaHistorial.objects.select_related('usuario', 'reemplazo').filter(
        asignacion_id=asignacion_id
    )
    if fecha:
        try:
            fecha_obj = datetime.date.fromisoformat(str(fecha))
            qs = qs.filter(fecha_reporte=fecha_obj)
        except ValueError:
            pass

    data = []
    for h in qs:
        usuario_nombre = ''
        if h.usuario:
            full_name = f"{h.usuario.first_name} {h.usuario.last_name}".strip()
            usuario_nombre = full_name or h.usuario.get_username()
        reemplazo_nombre = ''
        if h.reemplazo:
            reemplazo_nombre = f"{h.reemplazo.nombres} {h.reemplazo.apellidos}".strip()
        data.append({
            'fecha_reporte': h.fecha_reporte.isoformat() if h.fecha_reporte else None,
            'usuario': usuario_nombre,
            'codigo': h.codigo or '',
            'estado_asistencia': _normalize_estado_asistencia(h.estado_asistencia),
            'estado': h.estado or '',
            'reemplazo': reemplazo_nombre,
            # La descripción auto de sacafranco ("Cobertura SACAFRANCO AUTO ...") se muestra vacía.
            'descripcion': '' if _is_auto_sacafranco_desc(h.descripcion) else (h.descripcion or ''),
            'row_color': h.row_color or '',
            'creado_en': h.creado_en.isoformat() if h.creado_en else None,
        })

    return JsonResponse(data, safe=False, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_reporte_asistencia_excel(request):
    if not request.user.has_perm('CoreFisica.export_reporte_asistencia'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    if not request.user.has_perm('CoreFisica.view_reporteasistencia'):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    fecha = request.GET.get('fecha')
    cliente_id = request.GET.get('cliente_id')
    q = (request.GET.get('q') or '').strip()
    zona = _normalize_zona_filter(request.GET.get('zona'))
    headers = [
        'NOMINATIVO', 'CLIENTE', 'PUESTO', 'HORARIO',
        'NOMBRE Y APELLIDOS', 'ASISTENCIA', 'REEMPLAZO', 'ESTADO', 'DESCRIPCIÓN',
    ]
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def render_sheet(ws, turno_val):
        data = _build_reporte_asistencia_data(fecha=fecha, cliente_id=cliente_id, turno=turno_val, zona=zona, q=q)
        header_ctx = _build_header_context(request, fecha, turno_val)
        asistencias, faltos = _build_resumen_asistencia(data)
        grouped = _group_reporte_por_zona_y_provincia(data)
        zona_resumen = []

        _draw_excel_header(ws, header_ctx, border)

        header_row = 6
        data_start_row = header_row + 1

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.border = border
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        current_row = data_start_row
        for zona_group in grouped:
            # Fila de zona
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
            zona_cell = ws.cell(row=current_row, column=1)
            zona_cell.value = _normalize_zona_label(zona_group['zona'])
            zona_cell.font = Font(bold=True, size=11)
            zona_cell.alignment = Alignment(horizontal='left', vertical='center')
            for col_idx in range(1, 10):
                ws.cell(row=current_row, column=col_idx).border = border
            ws.row_dimensions[current_row].height = 24
            current_row += 1

            zona_items = []
            for prov_group in zona_group['provincias']:
                # Fila de provincia
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
                prov_cell = ws.cell(row=current_row, column=1)
                prov_cell.value = str(prov_group['provincia']).upper()
                prov_cell.font = Font(bold=True, size=10)
                prov_cell.alignment = Alignment(horizontal='left', vertical='center')
                for col_idx in range(1, 10):
                    ws.cell(row=current_row, column=col_idx).border = border
                ws.row_dimensions[current_row].height = 22
                current_row += 1

                for item in prov_group['rows']:
                    zona_items.append(item)
                    row_hex = _normalize_hex_color(item.get('row_color'))
                    row_fill = PatternFill(start_color=row_hex, end_color=row_hex, fill_type='solid') if row_hex else None
                    row_vals = [
                        item.get('codigo', ''),
                        item.get('cliente', ''),
                        item.get('puesto', ''),
                        item.get('horario', ''),
                        item.get('nombre_apellidos', ''),
                        'ASISTE' if item.get('estado_asistencia') == 'ASISTIO' else ('FALTO' if item.get('estado_asistencia') == 'FALTO' else ''),
                        item.get('reemplazo', ''),
                        item.get('estado', ''),
                        _descripcion_con_hueca(item),
                    ]
                    for col_idx, value in enumerate(row_vals, start=1):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.value = value
                        cell.border = border
                        if row_fill:
                            cell.fill = row_fill
                        if col_idx == 9:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    ws.row_dimensions[current_row].height = 32
                    current_row += 1

            zona_asistencias, zona_faltos = _build_resumen_asistencia(zona_items)
            zona_resumen.append((zona_group['zona'], zona_asistencias, zona_faltos))
            _write_excel_resumen(ws, current_row, zona_asistencias, zona_faltos, border)
            current_row += 2

        column_widths = {
            1: 11,  # Codigo
            2: 18,  # Cliente
            3: 23,  # Puesto
            4: 12,  # Horario
            5: 40,  # Nombre y Apellidos
            6: 14,  # Asistencia
            7: 38,  # Reemplazo
            8: 12,  # Estado
            9: 28,  # Descripcion
        }
        for col_idx, width in column_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # Resumen global removido; se deja solo el resumen por zona.
        current_row += 1
        for zona_label, zona_asistencias, zona_faltos in zona_resumen:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=5)
            ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=9)

            zcell = ws.cell(row=current_row, column=1)
            zcell.value = _format_zona_label(zona_label)
            zcell.font = Font(bold=True)
            zcell.alignment = Alignment(horizontal='left', vertical='center')

            acell = ws.cell(row=current_row, column=3)
            acell.value = f"Asistencias: {zona_asistencias}"
            acell.alignment = Alignment(horizontal='center', vertical='center')

            fcell = ws.cell(row=current_row, column=6)
            fcell.value = f"Faltas: {zona_faltos}"
            fcell.alignment = Alignment(horizontal='center', vertical='center')

            for col_idx in range(1, 10):
                ws.cell(row=current_row, column=col_idx).border = border
            current_row += 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'DIURNO'
    render_sheet(ws, 'Diurno')

    ws_nocturno = wb.create_sheet('NOCTURNO')
    render_sheet(ws_nocturno, 'Nocturno')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_asistencia.xlsx"'
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_reporte_asistencia_pdf(request):
    if not request.user.has_perm('CoreFisica.export_reporte_asistencia'):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    fecha = request.GET.get('fecha')
    cliente_id = request.GET.get('cliente_id')
    turno = request.GET.get('turno')
    q = (request.GET.get('q') or '').strip()
    zona = _normalize_zona_filter(request.GET.get('zona'))
    data = _build_reporte_asistencia_data(fecha=fecha, cliente_id=cliente_id, turno=turno, zona=zona, q=q)
    header_ctx = _build_header_context(request, fecha, turno)
    asistencias, faltos = _build_resumen_asistencia(data)
    grouped = _group_reporte_por_zona_y_provincia(data)
    zona_resumen = []

    output = BytesIO()
    p = canvas.Canvas(output, pagesize=landscape(letter))
    width, height = landscape(letter)

    x_margin = 0.5 * inch
    y_margin = 0.5 * inch

    headers = [
        'NOMINATIVO', 'CLIENTE', 'PUESTO', 'HORARIO',
        'NOMBRE Y APELLIDOS', 'ASISTENCIA', 'REEMPLAZO', 'ESTADO', 'DESCRIPCIÓN',
    ]

    col_widths = [0.75, 1.15, 1.1, 0.6, 1.75, 0.7, 0.8, 1.5, 1.0]
    col_widths = [w * inch for w in col_widths]

    def ensure_space(y_cursor, needed_height):
        if y_cursor < y_margin + needed_height:
            p.showPage()
            page_width, page_height = landscape(letter)
            new_y = _draw_pdf_header(p, page_width, page_height, x_margin, y_margin, header_ctx)
            new_y = _draw_pdf_table_headers(p, x_margin, new_y, headers, col_widths)
            p.setFont('Helvetica', 6)
            return new_y
        return y_cursor

    def draw_group_row(y_cursor, text, font_size=7):
        y_cursor = ensure_space(y_cursor, 0.3 * inch)
        p.setFont('Helvetica-Bold', font_size)
        p.setFillColor(colors.black)
        p.drawString(x_margin + 6, y_cursor, text)
        p.setFont('Helvetica', 6)
        return y_cursor - 0.2 * inch

    def draw_resumen(y_cursor, asistencias_val, faltos_val):
        y_cursor = ensure_space(y_cursor, 0.9 * inch)
        box_h = 0.55 * inch
        box_gap = 1.2 * inch
        box_w = ((width - (2 * x_margin)) - box_gap) / 2

        x1 = x_margin
        x2 = x1 + box_w + box_gap
        y_bottom = y_cursor - box_h

        p.setStrokeColor(colors.black)

        p.setFillColor(colors.HexColor('#E6B8BE'))
        p.rect(x1, y_bottom, box_w, box_h, stroke=1, fill=1)

        p.setFillColor(colors.white)
        p.rect(x2, y_bottom, box_w, box_h, stroke=1, fill=1)

        p.setFont('Helvetica-Bold', 11)
        p.setFillColor(colors.HexColor('#9C0006'))
        p.drawCentredString(x1 + (box_w / 2), y_bottom + 0.34 * inch, str(asistencias_val))
        p.drawCentredString(x1 + (box_w / 2), y_bottom + 0.12 * inch, 'ASISTENCIAS')

        p.setFillColor(colors.black)
        p.drawCentredString(x2 + (box_w / 2), y_bottom + 0.34 * inch, str(faltos_val))
        p.drawCentredString(x2 + (box_w / 2), y_bottom + 0.12 * inch, 'FALTOS')

        return y_bottom - 0.3 * inch

    y = _draw_pdf_header(p, width, height, x_margin, y_margin, header_ctx)
    y = _draw_pdf_table_headers(p, x_margin, y, headers, col_widths)
    p.setFont('Helvetica', 6)

    for zona_group in grouped:
        y = draw_group_row(y, _normalize_zona_label(zona_group['zona']), 8)
        zona_items = []
        for prov_group in zona_group['provincias']:
            y = draw_group_row(y, str(prov_group['provincia']).upper(), 7)
            for item in prov_group['rows']:
                zona_items.append(item)
                row_vals = [
                    item.get('codigo', ''),
                    item.get('cliente', ''),
                    item.get('puesto', ''),
                    item.get('horario', ''),
                    item.get('nombre_apellidos', ''),
                    'ASISTE' if item.get('estado_asistencia') == 'ASISTIO' else ('FALTO' if item.get('estado_asistencia') == 'FALTO' else ''),
                    item.get('reemplazo', ''),
                    item.get('estado', ''),
                    _descripcion_con_hueca(item)[:240],
                ]

                # Envolver cada celda en varias líneas (no se corta el texto).
                line_h = 0.12 * inch
                cells_lines = [
                    _wrap_text_to_width(str(v) if v is not None else '', col_widths[i] - 6, 'Helvetica', 6)
                    for i, v in enumerate(row_vals)
                ]
                n_lines = max(len(cl) for cl in cells_lines)
                row_h = n_lines * line_h
                y = ensure_space(y, row_h + 0.10 * inch)

                row_hex = _normalize_hex_color(item.get('row_color'))
                if row_hex:
                    x_bg = x_margin
                    bg_bottom = y - (n_lines - 1) * line_h - 0.05 * inch
                    bg_h = row_h + 0.06 * inch
                    p.saveState()
                    p.setFillColor(colors.HexColor(f"#{row_hex}"))
                    for w in col_widths:
                        p.rect(x_bg, bg_bottom, w, bg_h, stroke=0, fill=1)
                        x_bg += w
                    p.restoreState()

                x = x_margin
                for i, lines in enumerate(cells_lines):
                    ly = y
                    for ln in lines:
                        txt_w = pdfmetrics.stringWidth(ln, 'Helvetica', 6)
                        p.drawString(x + max((col_widths[i] - txt_w) / 2, 0), ly, ln)
                        ly -= line_h
                    x += col_widths[i]

                y -= row_h + 0.06 * inch

        zona_asistencias, zona_faltos = _build_resumen_asistencia(zona_items)
        zona_resumen.append((zona_group['zona'], zona_asistencias, zona_faltos))
        y = draw_resumen(y, zona_asistencias, zona_faltos)

    if zona_resumen:
        y = ensure_space(y, 0.4 * inch)
        p.setFont('Helvetica', 7)
        for zona_label, zona_asistencias, zona_faltos in zona_resumen:
            label = _format_zona_label(zona_label)
            p.setFont('Helvetica-Bold', 7)
            p.drawString(x_margin, y, f"{label}:")
            p.setFont('Helvetica', 7)
            p.drawString(x_margin + 1.0 * inch, y, f"Asistencias: {zona_asistencias}")
            p.drawString(x_margin + 3.2 * inch, y, f"Faltas: {zona_faltos}")
            y -= 0.18 * inch
        y -= 0.1 * inch

    p.showPage()
    p.save()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_asistencia.pdf"'
    output.seek(0)
    response.write(output.getvalue())
    return response