"""Importación masiva de clientes desde archivo Excel."""
from datetime import date, datetime

from django.http import JsonResponse
from django.db import transaction
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from ..models import Cliente, Instalacion, Puesto, PuestoHorario, Provincia, Canton
from .importar_puestos_asignaciones import parse_compact_horas_turno_dias


HEADER_MAP = {
    'RUC': 'ruc',
    'RAZON SOCIAL': 'razon_social',
    'RAZÓN SOCIAL': 'razon_social',
    'NOMBRE COMERCIAL': 'nombre_comercial',
    'NOMBRE_COMERCIAL': 'nombre_comercial',
    'NOMBRECOMERCIAL': 'nombre_comercial',
    'CLASIFICACION': 'clasificacion',
    'CLASIFICACIÓN': 'clasificacion',
    'INSTALACION': 'instalacion',
    'INSTALACIÓN': 'instalacion',
    'PROVINCIA': 'provincia',
    'CIUDAD': 'ciudad',

    'FECHA DE INGRESO': 'fecha_ingreso',
    'FECHA INGRESO': 'fecha_ingreso',
    'FECHA_INGRESO': 'fecha_ingreso',
    'FECHA DE INICIO': 'fecha_ingreso',
    'FECHA INICIO': 'fecha_ingreso',
    'FECHA_INICIO': 'fecha_ingreso',
    'FECHA DE INCIO': 'fecha_ingreso',
    'FECHA INCIO': 'fecha_ingreso',
    'FECHA_INCIO': 'fecha_ingreso',
    
    'NOMBRE DE PUESTO': 'puesto_nombre',
    'PUESTO NOMBRE': 'puesto_nombre',
    'PUESTO_NOMBRE': 'puesto_nombre',
    
    'PUESTO': 'puesto',
    'PUESTOS': 'puesto',
    'TIPO DE PUESTO': 'puesto_tipo',
    'TIPO DE PUESTOS': 'puesto_tipo',
    'TIPO PUESTO': 'puesto_tipo',
    'TIPO_PUESTO': 'puesto_tipo',
    'PUESTO TIPO': 'puesto_tipo',
    'PUESTO_TIPO': 'puesto_tipo',
}

# Normalización de clasificación a los valores del modelo
CLASSIF_MAP = {
    'PEQUENO': 'PEQUENO',
    'PEQUEÑO': 'PEQUENO',
    'PEQUENA': 'PEQUENO',
    'PEQUEÑA': 'PEQUENO',
    'MEDIANO': 'MEDIANO',
    'MEDIANA': 'MEDIANO',
    'GRANDE': 'GRANDE',
    'GRAN': 'GRANDE',
    'OFICINA': 'OFICINA',
}


def get_or_create_provincia_token(token):
    if not token:
        return None
    nombre = norm(token)
    if not nombre:
        return None
    from ..utils import buscar_o_crear_provincia
    return buscar_o_crear_provincia(nombre, Provincia)


def get_or_create_canton_token(token, provincia_token=None):
    if not token:
        return None
    nombre = norm(token)
    if not nombre:
        return None
    provincia_obj = get_or_create_provincia_token(provincia_token) if provincia_token else None
    qs = Canton.objects.all()
    if provincia_obj:
        qs = qs.filter(provincia=provincia_obj)
    canton = qs.filter(nombre__iexact=nombre).first()
    if canton:
        return canton
    if provincia_obj:
        return Canton.objects.create(nombre=nombre, provincia=provincia_obj)
    return None


def norm(val):
    if val is None:
        return ''
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val.is_integer():
            return str(int(val))
        return str(val)
    return str(val).strip()


def norm_class(val):
    return CLASSIF_MAP.get(norm(val).upper(), '')


def norm_header_key(val: str) -> str:
    key = norm(val).upper()
    key = key.replace('_', ' ')
    key = ' '.join(key.split())  
    return key


def parse_excel_date(val):
    if not val:
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, (int, float)):
        try:
            return from_excel(val).date()
        except Exception:
            return None
    if isinstance(val, str):
        raw = val.strip()
        if not raw:
            return None
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y'):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
    return None


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def importar_clientes(request):
    if not request.user.has_perm('CoreFisica.import_cliente'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
        
    file = request.FILES.get('file')
    if not file:
        return JsonResponse({'error': 'No se envió archivo'}, status=400)

    try:
        wb = load_workbook(filename=file, read_only=True)
    except Exception as exc:
        return JsonResponse({'error': f'No se pudo abrir el archivo: {exc}'}, status=400)

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return JsonResponse({'error': 'El archivo está vacío'}, status=400)

    header_idx = None
    header_row_num = None
    headers_raw = []
    for ridx, row in enumerate(rows):
        candidate_headers = [norm_header_key(h) for h in row]
        tmp_idx = {HEADER_MAP[h]: i for i, h in enumerate(candidate_headers) if h in HEADER_MAP}
        headers_raw = candidate_headers
        if 'nombre_comercial' in tmp_idx:
            header_idx = tmp_idx
            header_row_num = ridx
            break

    if header_idx is None:
        return JsonResponse({
            'error': 'Falta la columna obligatoria: NOMBRE COMERCIAL',
            'headers_detectados': headers_raw
        }, status=400)

    created_clientes = updated_clientes = 0
    created_inst = updated_inst = 0
    created_puestos = updated_puestos = 0
    errors = []
    # Detalle de lo que se creo (nombres), para mostrarlo en el resumen.
    nuevos_clientes = []
    nuevas_instalaciones = []
    nuevos_puestos = []

    start_row = (header_row_num or 0) + 1

    with transaction.atomic():
        for i, row in enumerate(rows[start_row:], start=start_row + 1):
            def col(key):
                idx = header_idx.get(key)
                return norm(row[idx]) if idx is not None and idx < len(row) else ''

            def col_raw(key):
                idx = header_idx.get(key)
                return row[idx] if idx is not None and idx < len(row) else None

            ruc = col('ruc')
            razon_social = col('razon_social')
            nombre_comercial = col('nombre_comercial') or razon_social
            clasif = norm_class(col('clasificacion'))
            inst_nombre = col('instalacion') or 'SIN NOMBRE'
            provincia = col('provincia')
            ciudad = col('ciudad')
            fecha_ingreso = parse_excel_date(col_raw('fecha_ingreso'))
            
            puesto_nombre = col('puesto_nombre') or col('puesto')
            puesto_tipo = col('puesto_tipo') or None
            # La columna "PUESTO" trae el horario compacto (ej. "10 H D L V") SOLO cuando
            # hay una columna aparte "NOMBRE DE PUESTO"; si no, 'puesto' ES el nombre.
            puesto_horario_txt = col('puesto') if col('puesto_nombre') else ''

            if not nombre_comercial:
                errors.append(f"Fila {i}: sin nombre_comercial")
                continue

            
            # filter().first() (no get_or_create) para tolerar clientes duplicados existentes.
            if ruc:
                cliente = Cliente.objects.filter(ruc=ruc).first()
            else:
                cliente = Cliente.objects.filter(nombre_comercial=nombre_comercial).first()
            created = False
            if not cliente:
                cliente = Cliente.objects.create(
                    ruc=ruc or None,
                    razon_social=razon_social or nombre_comercial,
                    nombre_comercial=nombre_comercial,
                    size=clasif or 'MEDIANO',
                    fecha_ingreso=fecha_ingreso,
                )
                created = True
            if created:
                created_clientes += 1
                nuevos_clientes.append(nombre_comercial)
            else:
                updated = False
                if fecha_ingreso and cliente.fecha_ingreso != fecha_ingreso:
                    cliente.fecha_ingreso = fecha_ingreso
                    updated = True
                if updated:
                    cliente.save(update_fields=['fecha_ingreso'])
                    updated_clientes += 1

            # Instalación por cliente + nombre, resolviendo provincia/cantón.
            # Usamos filter().first() (no get_or_create) para tolerar duplicados ya
            # existentes en la base (si no, get_or_create revienta con MultipleObjectsReturned).
            canton_obj = get_or_create_canton_token(ciudad, provincia)
            instalacion = Instalacion.objects.filter(cliente=cliente, nombre=inst_nombre).first()
            if not instalacion:
                instalacion = Instalacion.objects.create(
                    cliente=cliente, nombre=inst_nombre,
                    **({'canton': canton_obj} if canton_obj else {}),
                )
                created_inst += 1
                nuevas_instalaciones.append(f"{nombre_comercial} - {inst_nombre}")

            # Puesto por instalación + nombre (tolerante a duplicados existentes).
            if puesto_nombre:
                puesto = Puesto.objects.filter(instalacion=instalacion, nombre=puesto_nombre).first()

                # Arma el horario del puesto desde la columna "PUESTO" (ej. "10 H D L V"
                # -> dias/horas/turno). Devuelve True si creó algún PuestoHorario.
                def _armar_horario(p):
                    from ..models import horas_default_turno
                    grupos = parse_compact_horas_turno_dias(puesto_horario_txt) if puesto_horario_txt else []
                    tenia_grupos = bool(grupos)
                    creado_algo = False
                    for grp in grupos:
                        turno_val = grp.get('turno') or 'Diurno'
                        _ing, _sal = horas_default_turno(turno_val)
                        for dia in grp.get('dias', []):
                            # Clave por (puesto, dia, TURNO): un puesto puede tener DIA y NOCHE
                            # el mismo dia (ej. "12 H D L D y 12 H N S"). get_or_create SOLO
                            # agrega los turnos que faltan; NO pisa lo ya configurado a mano.
                            _ph, ph_created = PuestoHorario.objects.get_or_create(
                                puesto=p, dia=dia, turno=turno_val,
                                defaults={'horas': min(max(grp.get('hours', 12), 0), 24),
                                          'hora_ingreso': _ing, 'hora_salida': _sal},
                            )
                            if ph_created:
                                creado_algo = True
                    if creado_algo:
                        try:
                            p.sync_from_horarios()
                            p.save()
                        except Exception:
                            pass
                    elif puesto_horario_txt and not tenia_grupos:
                        errors.append(
                            f"Fila {i}: no se pudo interpretar el horario '{puesto_horario_txt}' "
                            f"del puesto '{puesto_nombre}'"
                        )
                    return creado_algo

                if not puesto:
                    # NUEVO: se crea con nombre + tipo + cantidad=1 y se arma su horario.
                    puesto_defaults = {'cantidad_puestos': 1}
                    if puesto_tipo:
                        puesto_defaults['tipo'] = puesto_tipo
                    puesto = Puesto.objects.create(
                        instalacion=instalacion, nombre=puesto_nombre, **puesto_defaults,
                    )
                    created_puestos += 1
                    nuevos_puestos.append(f"{inst_nombre} - {puesto_nombre}")
                    _armar_horario(puesto)
                else:
                    # EXISTENTE: solo se COMPLETAN los datos VACIOS (no se pisa lo ya
                    # configurado a mano). Tipo si esta vacio, y horario si aun no tiene.
                    toco = False
                    if puesto_tipo and not (puesto.tipo or '').strip():
                        puesto.tipo = puesto_tipo
                        puesto.save(update_fields=['tipo'])
                        toco = True
                    # Siempre intenta agregar los turnos del patron que FALTEN (p.ej. la
                    # NOCHE de un puesto que hoy solo tiene DIA). Solo agrega, no pisa.
                    if puesto_horario_txt:
                        if _armar_horario(puesto):
                            toco = True
                    if toco:
                        updated_puestos += 1

    summary = {
        'clientes_creados': created_clientes,
        'clientes_actualizados': updated_clientes,
        'instalaciones_creadas': created_inst,
        'instalaciones_actualizadas': updated_inst,
        'puestos_creados': created_puestos,
        'puestos_actualizados': updated_puestos,
        'errores': errors,
        'errores_total': len(errors),
        # Detalle de lo creado (nombres) para mostrar en el resumen.
        'nuevos_clientes': nuevos_clientes,
        'nuevas_instalaciones': nuevas_instalaciones,
        'nuevos_puestos': nuevos_puestos,
    }
    return JsonResponse(summary, status=200)
