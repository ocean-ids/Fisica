from django.http import JsonResponse, HttpResponse
from django.utils.dateparse import parse_date
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
import json
import datetime
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.lib import colors
from ..models import Consolidado, Asignacion, ConsolidadoResumen, Persona
from .reporte_asistencia_views import _build_reporte_asistencia_data

ALLOWED_TURNOS = {'Diurno', 'Nocturno'}
ALLOWED_TIPOS = {choice[0] for choice in Consolidado.TIPOS}
TIPO_CONSOLA = 'CONSOLa'
TIPO_GUARDIA = 'GUARDIA'
TIPOS_CENTRO_CONTROL = {'OPERADOR CENTRO CONTROL', 'SUPERVISOR CENTRO CONTROL'}


def _parse_fecha(fecha_param):
    if fecha_param:
        try:
            return datetime.date.fromisoformat(str(fecha_param))
        except ValueError:
            pass
    return timezone.localdate()


def _resolve_row_reference(tipo, persona_ref_id=None, asignacion_ref_id=None):
    if tipo == TIPO_CONSOLA:
        if persona_ref_id:
            return int(persona_ref_id), None
        return None, None
    if tipo == TIPO_GUARDIA:
        if asignacion_ref_id:
            return None, int(asignacion_ref_id)
        return None, None
    return None, None


def _consolidado_key(item: Consolidado):
    ref_id = item.persona_ref_id if item.tipo == TIPO_CONSOLA else item.asignacion_ref_id
    if not ref_id:
        return None
    return item.tipo, ref_id


def _build_consolidado_data(fecha, turno, zona='', q=''):
    fecha_obj = _parse_fecha(fecha)
    turno_val = turno if turno in ALLOWED_TURNOS else None

    consolidados_qs = Consolidado.objects.filter(fecha=fecha_obj)
    if turno_val:
        consolidados_qs = consolidados_qs.filter(turno=turno_val)
    consolidado_map = {}
    for c in consolidados_qs:
        key = _consolidado_key(c)
        if key:
            consolidado_map[key] = c

    data = []
    consola_items = [c for c in consolidados_qs.select_related('persona_ref') if c.tipo == TIPO_CONSOLA and c.persona_ref_id]
    for cons in consola_items:
        p = cons.persona_ref
        if not p:
            continue
        data.append({
            'consolidado_id': cons.id,
            'fecha': fecha_obj.isoformat(),
            'turno': turno_val or '',
            'tipo': TIPO_CONSOLA,
            'persona_ref_id': p.id,
            'asignacion_ref_id': None,
            'nominativo': cons.nominativo or '',
            'proyecto': cons.proyecto or '',
            'puesto': cons.puesto or '',
            'apellidos': p.apellidos,
            'nombres': p.nombres,
            'estado': p.tipo,
            'observacion': cons.observacion or '',
            'zona': 'PERSONAL DE CONSOLA'
        })

    reporte_rows = _build_reporte_asistencia_data(fecha=fecha_obj.isoformat(), turno=turno_val, zona=zona, q=q)
    asig_ids = [r.get('asignacion_id') for r in reporte_rows if r.get('asignacion_id')]
    asig_map = {}
    if asig_ids:
        asig_qs = Asignacion.objects.select_related('instalacion').filter(id__in=asig_ids)
        asig_map = {a.id: a for a in asig_qs}

    for row in reporte_rows:
        asig_id = row.get('asignacion_id')
        if not asig_id:
            continue
        cons = consolidado_map.get((TIPO_GUARDIA, asig_id))
        asig = asig_map.get(asig_id)
        proyecto = getattr(asig.instalacion, 'nombre', '') if asig and asig.instalacion else ''
        nombre_apellidos = (row.get('nombre_apellidos') or '').strip()
        data.append({
            'consolidado_id': cons.id if cons else None,
            'fecha': fecha_obj.isoformat(),
            'turno': turno_val or '',
            'tipo': TIPO_GUARDIA,
            'persona_ref_id': None,
            'asignacion_ref_id': asig_id,
            'nominativo': row.get('codigo') or '',
            'proyecto': proyecto,
            'puesto': row.get('puesto') or '',
            'apellidos': nombre_apellidos,
            'nombres': '',
            'estado': row.get('estado') or '',
            'observacion': cons.observacion if cons else '',
            'zona': (row.get('zona_titulo') or 'SIN ZONA').strip(),
            'provincia': (row.get('provincia') or 'SIN PROVINCIA').strip()
        })

    term = (q or '').strip().lower()
    if term:
        filtered = []
        for item in data:
            haystack = ' '.join([
                str(item.get('nominativo') or ''),
                str(item.get('proyecto') or ''),
                str(item.get('puesto') or ''),
                str(item.get('apellidos') or ''),
                str(item.get('nombres') or ''),
                str(item.get('estado') or ''),
                str(item.get('observacion') or ''),
                str(item.get('zona') or ''),
                str(item.get('provincia') or ''),
                str(item.get('tipo') or ''),
            ]).lower()
            if term in haystack:
                filtered.append(item)
        data = filtered

    return data


def _count_faltas_from_reporte(rows):
    total = 0
    for row in rows:
        if row.get('reemplazo_id'):
            total += 1
            continue
        reemplazo = (row.get('reemplazo') or '').strip()
        if reemplazo and reemplazo != '-':
            total += 1
    return total


def _build_estado_agentes_counts(rows):
    counts = {
        'dobla': 0,
        'franco_trabajados': 0,
        'unidades_eventuales': 0,
        'adelanto_turno': 0,
        'reten': 0,
        'unidades_adicionales': 0,
        'custodio': 0,
    }

    for row in rows:
        estado = (row.get('estado') or '').strip().upper()
        if not estado:
            continue

        if 'DOBL' in estado:
            counts['dobla'] += 1
            continue
        if 'FR/TRABAJADO' in estado or 'FRANCO' in estado:
            counts['franco_trabajados'] += 1
            continue
        if 'EVENTUAL' in estado:
            counts['unidades_eventuales'] += 1
            continue
        if 'ADEL' in estado:
            counts['adelanto_turno'] += 1
            continue
        if 'RETEN' in estado:
            counts['reten'] += 1
            continue
        if 'ADICIONAL' in estado:
            counts['unidades_adicionales'] += 1
            continue
        if 'CUSTODIO' in estado:
            counts['custodio'] += 1
            continue

    counts['total'] = sum(counts.values())
    return counts


def _build_resumen_manual(fecha_obj, turno_val, reporte_rows):
    if not turno_val:
        return None
    faltas_auto = _count_faltas_from_reporte(reporte_rows)
    resumen, created = ConsolidadoResumen.objects.get_or_create(
        fecha=fecha_obj,
        turno=turno_val,
        defaults={'faltas': faltas_auto}
    )
    if not created and resumen.faltas != faltas_auto:
        resumen.faltas = faltas_auto
        resumen.save(update_fields=['faltas'])
    data = {
        'faltas': resumen.faltas,
        'huecas': resumen.huecas,
        'apoyos': resumen.apoyos,
        'capacitacion': resumen.capacitacion,
        'apertura_puesto': resumen.apertura_puesto,
        'servicios_temporales': resumen.servicios_temporales,
        'servicios_adicionales': resumen.servicios_adicionales,
        'aprendiendo_consignas': resumen.aprendiendo_consignas,
    }
    data['total'] = sum(data.values())
    return data


def _group_guardias_por_zona(data):
    guardias = [d for d in data if d.get('tipo') == TIPO_GUARDIA]
    zonas = {}
    for item in guardias:
        zona = item.get('zona') or 'SIN ZONA'
        if zona not in zonas:
            zonas[zona] = []
        zonas[zona].append(item)
    ordered_zonas = sorted(zonas.keys())
    return [(zona, zonas[zona]) for zona in ordered_zonas]


def _serialize_item(item: Consolidado):
    return {
        'id': item.id,
        'fecha': item.fecha.isoformat() if item.fecha else None,
        'turno': item.turno,
        'tipo': item.tipo,
        'persona_ref_id': item.persona_ref_id,
        'asignacion_ref_id': item.asignacion_ref_id,
        'nominativo': item.nominativo or '',
        'proyecto': item.proyecto or '',
        'puesto': item.puesto or '',
        'observacion': item.observacion or ''
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def obtener_consolidado(request):
    if not request.user.has_perm('CoreFisica.view_consolidado'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    fecha = request.GET.get('fecha')
    turno = request.GET.get('turno')
    tipo = request.GET.get('tipo')
    referencia_id = request.GET.get('referencia_id')

    qs = Consolidado.objects.all()
    if fecha:
        fecha_obj = parse_date(fecha)
        if fecha_obj:
            qs = qs.filter(fecha=fecha_obj)
    if turno:
        qs = qs.filter(turno=turno)
    if tipo:
        qs = qs.filter(tipo=tipo)
    if referencia_id:
        try:
            ref_id = int(referencia_id)
            if tipo == TIPO_CONSOLA:
                qs = qs.filter(persona_ref_id=ref_id)
            elif tipo == TIPO_GUARDIA:
                qs = qs.filter(asignacion_ref_id=ref_id)
            else:
                qs = qs.filter(persona_ref_id=ref_id) | qs.filter(asignacion_ref_id=ref_id)
        except (TypeError, ValueError):
            pass

    data = [_serialize_item(item) for item in qs.order_by('fecha', 'turno', 'tipo', 'id')]
    return JsonResponse(data, safe=False)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def obtener_consolidado_armado(request):
    if not request.user.has_perm('CoreFisica.view_consolidado'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    fecha = request.GET.get('fecha')
    turno = request.GET.get('turno')
    zona = (request.GET.get('zona') or '').strip()
    q = (request.GET.get('q') or '').strip()
    data = _build_consolidado_data(fecha, turno, zona=zona, q=q)
    return JsonResponse(data, safe=False)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def obtener_consolidado_resumen(request):
    if not request.user.has_perm('CoreFisica.view_consolidado'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    fecha = request.GET.get('fecha')
    turno = request.GET.get('turno')
    zona = (request.GET.get('zona') or '').strip()
    q = (request.GET.get('q') or '').strip()
    turno_val = turno if turno in ALLOWED_TURNOS else None
    if not turno_val:
        return JsonResponse({'error': 'Turno requerido'}, status=400)
    fecha_obj = _parse_fecha(fecha)
    reporte_rows = _build_reporte_asistencia_data(fecha=fecha_obj.isoformat(), turno=turno_val, zona=zona, q=q)

    manual = _build_resumen_manual(fecha_obj, turno_val, reporte_rows)
    estados = _build_estado_agentes_counts(reporte_rows)
    return JsonResponse({
        'manual': manual,
        'estado_agentes': estados
    })


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def actualizar_consolidado_resumen(request):
    if not request.user.has_perm('CoreFisica.change_consolidado'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON invalido'}, status=400)

    fecha = parse_date(data.get('fecha')) if data.get('fecha') else None
    turno = data.get('turno')
    if not fecha or turno not in ALLOWED_TURNOS:
        return JsonResponse({'error': 'Fecha y turno requeridos'}, status=400)

    resumen, _ = ConsolidadoResumen.objects.get_or_create(fecha=fecha, turno=turno)

    fields = [
        'faltas', 'huecas', 'apoyos', 'capacitacion', 'apertura_puesto',
        'servicios_temporales', 'servicios_adicionales', 'aprendiendo_consignas'
    ]
    for field in fields:
        if field in data:
            try:
                setattr(resumen, field, max(int(data.get(field)), 0))
            except (TypeError, ValueError):
                pass

    resumen.save()

    manual = {
        'faltas': resumen.faltas,
        'huecas': resumen.huecas,
        'apoyos': resumen.apoyos,
        'capacitacion': resumen.capacitacion,
        'apertura_puesto': resumen.apertura_puesto,
        'servicios_temporales': resumen.servicios_temporales,
        'servicios_adicionales': resumen.servicios_adicionales,
        'aprendiendo_consignas': resumen.aprendiendo_consignas,
    }
    manual['total'] = sum(manual.values())
    return JsonResponse({'manual': manual})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def crear_consolidado(request):
    if not request.user.has_perm('CoreFisica.add_consolidado'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON invalido'}, status=400)

    fecha = parse_date(data.get('fecha')) if data.get('fecha') else None
    turno = data.get('turno')
    tipo = data.get('tipo')
    persona_ref_id = data.get('persona_ref_id')
    asignacion_ref_id = data.get('asignacion_ref_id')
    nominativo = (data.get('nominativo') or '').strip() or None
    proyecto = (data.get('proyecto') or '').strip() or None
    puesto = (data.get('puesto') or '').strip() or None
    observacion = (data.get('observacion') or '').strip() or None

    if not fecha or not turno or not tipo:
        return JsonResponse({'error': 'Faltan campos obligatorios'}, status=400)

    if turno not in ALLOWED_TURNOS:
        return JsonResponse({'error': 'Turno invalido'}, status=400)

    if tipo not in ALLOWED_TIPOS:
        return JsonResponse({'error': 'Tipo invalido'}, status=400)

    try:
        resolved_persona_id, resolved_asig_id = _resolve_row_reference(
            tipo,
            persona_ref_id=persona_ref_id,
            asignacion_ref_id=asignacion_ref_id,
        )
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Referencia invalida'}, status=400)

    if tipo == TIPO_CONSOLA and not resolved_persona_id:
        return JsonResponse({'error': 'Falta persona_ref_id para CONSOLa'}, status=400)

    if tipo == TIPO_GUARDIA and not resolved_asig_id:
        return JsonResponse({'error': 'Falta asignacion_ref_id para GUARDIA'}, status=400)

    if resolved_persona_id and not Persona.objects.filter(id=resolved_persona_id).exists():
        return JsonResponse({'error': 'Persona no encontrada'}, status=400)

    if resolved_asig_id and not Asignacion.objects.filter(id=resolved_asig_id).exists():
        return JsonResponse({'error': 'Asignacion no encontrada'}, status=400)

    item = Consolidado.objects.create(
        fecha=fecha,
        turno=turno,
        tipo=tipo,
        persona_ref_id=resolved_persona_id,
        asignacion_ref_id=resolved_asig_id,
        nominativo=nominativo,
        proyecto=proyecto,
        puesto=puesto,
        observacion=observacion
    )

    return JsonResponse(_serialize_item(item), status=201)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def actualizar_consolidado(request, id):
    if not request.user.has_perm('CoreFisica.change_consolidado'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON invalido'}, status=400)

    item = Consolidado.objects.filter(id=id).first()
    if not item:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)

    if 'fecha' in data:
        fecha = parse_date(data.get('fecha')) if data.get('fecha') else None
        if fecha:
            item.fecha = fecha

    if 'turno' in data:
        turno = data.get('turno')
        if turno in ALLOWED_TURNOS:
            item.turno = turno

    if 'tipo' in data:
        tipo = data.get('tipo')
        if tipo in ALLOWED_TIPOS:
            item.tipo = tipo

    has_reference_payload = any(k in data for k in ('persona_ref_id', 'asignacion_ref_id'))
    if has_reference_payload:
        tipo_resuelto = item.tipo
        persona_ref_id = data.get('persona_ref_id', None)
        asignacion_ref_id = data.get('asignacion_ref_id', None)
        try:
            resolved_persona_id, resolved_asig_id = _resolve_row_reference(
                tipo_resuelto,
                persona_ref_id=persona_ref_id,
                asignacion_ref_id=asignacion_ref_id,
            )
            if tipo_resuelto == TIPO_CONSOLA and not resolved_persona_id:
                return JsonResponse({'error': 'Falta persona_ref_id para CONSOLa'}, status=400)
            if tipo_resuelto == TIPO_GUARDIA and not resolved_asig_id:
                return JsonResponse({'error': 'Falta asignacion_ref_id para GUARDIA'}, status=400)

            item.persona_ref_id = resolved_persona_id
            item.asignacion_ref_id = resolved_asig_id
        except (TypeError, ValueError):
            return JsonResponse({'error': 'Referencia invalida'}, status=400)

    if 'nominativo' in data:
        item.nominativo = (data.get('nominativo') or '').strip() or None

    if 'proyecto' in data:
        item.proyecto = (data.get('proyecto') or '').strip() or None

    if 'puesto' in data:
        item.puesto = (data.get('puesto') or '').strip() or None

    if 'observacion' in data:
        item.observacion = (data.get('observacion') or '').strip() or None

    item.save()
    return JsonResponse(_serialize_item(item))


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def eliminar_consolidado(request, id):
    if not request.user.has_perm('CoreFisica.delete_consolidado'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    item = Consolidado.objects.filter(id=id).first()
    if not item:
        return JsonResponse({'error': 'Registro no encontrado'}, status=404)

    item.delete()
    return JsonResponse({'message': 'Registro eliminado'}, status=200)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_consolidado_excel(request):
    if not request.user.has_perm('CoreFisica.export_consolidado'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    fecha = request.GET.get('fecha', None)
    zona = (request.GET.get('zona') or '').strip()
    q = (request.GET.get('q') or '').strip()
    fecha_obj = _parse_fecha(fecha)
    fecha_label = fecha_obj.strftime('%d/%m/%Y')
    operador = ''
    user = getattr(request, 'user', None)
    if user and user.is_authenticated:
        operador = f"{user.first_name} {user.last_name}".strip() or user.get_username()

    headers = ['NOMINATIVO', 'PROYECTO', 'PUESTO', 'APELLIDOS Y NOMBRE', 'ESTADO', 'OBSERVACIONES']
    header_row = 2

    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type='solid', fgColor='D9EEF7')
    section_fill = PatternFill(fill_type='solid', fgColor='8FD1EA')
    total_fill = PatternFill(fill_type='solid', fgColor='FFF200')

    def render_sheet(ws, turno_val):
        turno_label = (turno_val or 'TODOS').upper()
        data = _build_consolidado_data(fecha, turno_val, zona=zona, q=q)
        reporte_rows = _build_reporte_asistencia_data(fecha=fecha_obj.isoformat(), turno=turno_val, zona=zona, q=q)
        manual = _build_resumen_manual(fecha_obj, turno_val, reporte_rows) if turno_val else None
        estados = _build_estado_agentes_counts(reporte_rows)

        ws['A1'] = 'FECHA:'
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)
        ws['B1'] = fecha_label
        ws['D1'] = f"TURNO: {turno_label}"
        ws['F1'] = f"REPORTA: {operador}"

        ws['A1'].font = Font(bold=True)
        ws['D1'].font = Font(bold=True)
        ws['F1'].font = Font(bold=True)
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D1'].alignment = Alignment(horizontal='left', vertical='center')
        ws['F1'].alignment = Alignment(horizontal='left', vertical='center')
        for col in range(1, 7):
            c = ws.cell(row=1, column=col)
            c.border = border

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.border = border
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill

        row_idx = header_row + 1

        consola_rows = [d for d in data if d.get('tipo') == TIPO_CONSOLA]
        if consola_rows:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
            cell = ws.cell(row=row_idx, column=1)
            cell.value = 'PERSONAL DE CONSOLA Y OFICINAS'
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill
            for col in range(1, 7):
                c = ws.cell(row=row_idx, column=col)
                c.border = border
            row_idx += 1

            for item in consola_rows:
                nombre = f"{item.get('apellidos', '')} {item.get('nombres', '')}".strip()
                values = [
                    item.get('nominativo', ''),
                    item.get('proyecto', ''),
                    item.get('puesto', ''),
                    nombre,
                    item.get('estado', ''),
                    item.get('observacion', ''),
                ]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = border
                    align = 'left' if col_idx in (4, 6) else 'center'
                    cell.alignment = Alignment(horizontal=align, vertical='center')
                row_idx += 1

        for zona_label, items in _group_guardias_por_zona(data):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
            cell = ws.cell(row=row_idx, column=1)
            cell.value = str(zona_label).upper()
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill
            for col in range(1, 7):
                c = ws.cell(row=row_idx, column=col)
                c.border = border
            row_idx += 1

            for item in items:
                nombre = f"{item.get('apellidos', '')} {item.get('nombres', '')}".strip()
                values = [
                    item.get('nominativo', ''),
                    item.get('proyecto', ''),
                    item.get('puesto', ''),
                    nombre,
                    item.get('estado', ''),
                    item.get('observacion', ''),
                ]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = border
                    align = 'left' if col_idx in (4, 6) else 'center'
                    cell.alignment = Alignment(horizontal=align, vertical='center')
                row_idx += 1

        def write_summary_row(row_num, label, value, is_total=False):
            ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=4)
            label_cell = ws.cell(row=row_num, column=2)
            value_cell = ws.cell(row=row_num, column=5)
            label_cell.value = label
            value_cell.value = value
            label_cell.alignment = Alignment(horizontal='center', vertical='center')
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
            if is_total:
                value_cell.fill = total_fill
            for col in range(2, 6):
                c = ws.cell(row=row_num, column=col)
                c.border = border
            return row_num + 1

        if manual:
            row_idx += 1
            row_idx = write_summary_row(row_idx, 'FALTOS', manual['faltas'])
            row_idx = write_summary_row(row_idx, 'HUECAS', manual['huecas'])
            row_idx = write_summary_row(row_idx, 'APOYOS', manual['apoyos'])
            row_idx = write_summary_row(row_idx, 'CAPACITACION', manual['capacitacion'])
            row_idx = write_summary_row(row_idx, 'APERTURA DE PUESTO', manual['apertura_puesto'])
            row_idx = write_summary_row(row_idx, 'SERVICIOS TEMPORALES', manual['servicios_temporales'])
            row_idx = write_summary_row(row_idx, 'SERVICIOS ADICIONALES', manual['servicios_adicionales'])
            row_idx = write_summary_row(row_idx, 'APRENDIENDO CONSIGNAS', manual['aprendiendo_consignas'])
            row_idx = write_summary_row(row_idx, 'TOTAL=', manual['total'], is_total=True)

        if estados:
            row_idx += 1
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=5)
            header_cell = ws.cell(row=row_idx, column=2)
            header_cell.value = 'ESTADO DE AGENTES'
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.fill = section_fill
            for col in range(2, 6):
                c = ws.cell(row=row_idx, column=col)
                c.border = border
            row_idx += 1
            row_idx = write_summary_row(row_idx, 'DOBLA', estados['dobla'])
            row_idx = write_summary_row(row_idx, 'FRANCO TRABAJADOS', estados['franco_trabajados'])
            row_idx = write_summary_row(row_idx, 'UNIDADES EVENTUALES', estados['unidades_eventuales'])
            row_idx = write_summary_row(row_idx, 'ADELANTO DE TURNO', estados['adelanto_turno'])
            row_idx = write_summary_row(row_idx, 'RETEN', estados['reten'])
            row_idx = write_summary_row(row_idx, 'UNIDADES ADICIONALES', estados['unidades_adicionales'])
            row_idx = write_summary_row(row_idx, 'CUSTODIO', estados['custodio'])
            row_idx = write_summary_row(row_idx, 'TOTAL=', estados['total'], is_total=True)

        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 34
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 34

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'DIURNO'
    render_sheet(ws, 'Diurno')

    ws_nocturno = wb.create_sheet('NOCTURNO')
    render_sheet(ws_nocturno, 'Nocturno')

    meses = [
        'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
        'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE'
    ]
    fecha_archivo = timezone.localdate()
    nombre_archivo = f"CONSOLIDADO - {meses[fecha_archivo.month - 1]} {fecha_archivo.year}.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_consolidado_pdf(request):
    if not request.user.has_perm('CoreFisica.export_consolidado'):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    fecha = request.GET.get('fecha', None)
    turno = request.GET.get('turno')
    zona = (request.GET.get('zona') or '').strip()
    q = (request.GET.get('q') or '').strip()
    data = _build_consolidado_data(fecha, turno, zona=zona, q=q)
    turno_val = turno if turno in ALLOWED_TURNOS else None
    reporte_rows = _build_reporte_asistencia_data(fecha=_parse_fecha(fecha).isoformat(), turno=turno_val, zona=zona, q=q)
    manual = _build_resumen_manual(_parse_fecha(fecha), turno_val, reporte_rows) if turno_val else None
    estados = _build_estado_agentes_counts(reporte_rows)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="consolidado.pdf"'

    p = canvas.Canvas(response, pagesize=landscape(letter))
    width, height = landscape(letter)
    x_margin = 0.5 * inch
    y_margin = 0.5 * inch

    fecha_label = _parse_fecha(fecha).strftime('%d/%m/%Y')
    turno_label = (turno or 'TODOS').upper()
    operador = ''
    user = getattr(request, 'user', None)
    if user and user.is_authenticated:
        operador = f"{user.first_name} {user.last_name}".strip() or user.get_username()

    p.setFont('Helvetica-Bold', 11)
    p.drawString(x_margin, height - y_margin, f"FECHA: {fecha_label}")
    p.drawString(x_margin + 3.5 * inch, height - y_margin, f"TURNO: {turno_label}")
    p.drawRightString(width - x_margin, height - y_margin, f"REPORTA: {operador}")

    headers = ['NOMINATIVO', 'PROYECTO', 'PUESTO', 'APELLIDOS Y NOMBRE', 'ESTADO', 'OBSERVACIONES']
    col_widths = [0.9, 1.8, 1.3, 2.3, 0.9, 2.7]
    col_widths = [w * inch for w in col_widths]

    y = height - y_margin - 0.4 * inch
    p.setFont('Helvetica-Bold', 8)
    x = x_margin
    for i, header in enumerate(headers):
        p.drawString(x + 2, y, header)
        x += col_widths[i]

    y -= 0.2 * inch
    p.setFont('Helvetica', 7)

    def ensure_space(y_cursor):
        if y_cursor < y_margin + 0.4 * inch:
            p.showPage()
            p.setFont('Helvetica-Bold', 8)
            x = x_margin
            for i, header in enumerate(headers):
                p.drawString(x + 2, height - y_margin, header)
                x += col_widths[i]
            return height - y_margin - 0.25 * inch
        return y_cursor

    def draw_section(label, y_cursor, font_size=8):
        y_cursor = ensure_space(y_cursor)
        p.setFont('Helvetica-Bold', font_size)
        p.setFillColor(colors.black)
        p.drawString(x_margin + 2, y_cursor, label)
        p.setFont('Helvetica', 7)
        return y_cursor - 0.2 * inch

    consola_rows = [d for d in data if d.get('tipo') == TIPO_CONSOLA]
    if consola_rows:
        y = draw_section('PERSONAL DE CONSOLA Y OFICINAS', y)
        for item in consola_rows:
            y = ensure_space(y)
            nombre = f"{item.get('apellidos', '')} {item.get('nombres', '')}".strip()
            row_vals = [
                item.get('nominativo', ''),
                item.get('proyecto', ''),
                item.get('puesto', ''),
                nombre,
                item.get('estado', ''),
                item.get('observacion', ''),
            ]
            x = x_margin
            for i, val in enumerate(row_vals):
                text = str(val or '')
                max_width = col_widths[i] - 6
                while pdfmetrics.stringWidth(text, 'Helvetica', 7) > max_width and len(text) > 0:
                    text = text[:-1]
                p.drawString(x + 2, y, text)
                x += col_widths[i]
            y -= 0.18 * inch

    for zona_label, items in _group_guardias_por_zona(data):
        y = draw_section(str(zona_label).upper(), y)
        for item in items:
            y = ensure_space(y)
            nombre = f"{item.get('apellidos', '')} {item.get('nombres', '')}".strip()
            row_vals = [
                item.get('nominativo', ''),
                item.get('proyecto', ''),
                item.get('puesto', ''),
                nombre,
                item.get('estado', ''),
                item.get('observacion', ''),
            ]
            x = x_margin
            for i, val in enumerate(row_vals):
                text = str(val or '')
                max_width = col_widths[i] - 6
                while pdfmetrics.stringWidth(text, 'Helvetica', 7) > max_width and len(text) > 0:
                    text = text[:-1]
                p.drawString(x + 2, y, text)
                x += col_widths[i]
            y -= 0.18 * inch

    if manual:
        y -= 0.25 * inch
        p.setFont('Helvetica-Bold', 8)
        p.drawString(x_margin, y, 'FALTAS')
        p.drawRightString(x_margin + 2.8 * inch, y, str(manual['faltas']))
        y -= 0.18 * inch
        p.drawString(x_margin, y, 'HUECAS')
        p.drawRightString(x_margin + 2.8 * inch, y, str(manual['huecas']))
        y -= 0.18 * inch
        p.drawString(x_margin, y, 'APOYOS')
        p.drawRightString(x_margin + 2.8 * inch, y, str(manual['apoyos']))
        y -= 0.18 * inch
        p.drawString(x_margin, y, 'CAPACITACION')
        p.drawRightString(x_margin + 2.8 * inch, y, str(manual['capacitacion']))
        y -= 0.18 * inch
        p.drawString(x_margin, y, 'APERTURA DE PUESTO')
        p.drawRightString(x_margin + 2.8 * inch, y, str(manual['apertura_puesto']))
        y -= 0.18 * inch
        p.drawString(x_margin, y, 'SERVICIOS TEMPORALES')
        p.drawRightString(x_margin + 2.8 * inch, y, str(manual['servicios_temporales']))
        y -= 0.18 * inch
        p.drawString(x_margin, y, 'SERVICIOS ADICIONALES')
        p.drawRightString(x_margin + 2.8 * inch, y, str(manual['servicios_adicionales']))
        y -= 0.18 * inch
        p.drawString(x_margin, y, 'APRENDIENDO CONSIGNAS')
        p.drawRightString(x_margin + 2.8 * inch, y, str(manual['aprendiendo_consignas']))
        y -= 0.18 * inch
        p.drawString(x_margin, y, 'TOTAL=')
        p.drawRightString(x_margin + 2.8 * inch, y, str(manual['total']))

    if estados:
        y -= 0.3 * inch
        p.setFont('Helvetica-Bold', 8)
        p.drawString(x_margin, y, 'ESTADO DE AGENTES')
        y -= 0.18 * inch
        p.drawString(x_margin, y, 'DOBLA')
        p.drawRightString(x_margin + 2.8 * inch, y, str(estados['dobla']))
        y -= 0.18 * inch
        p.drawString(x_margin, y, 'FRANCO TRABAJADOS')
        p.drawRightString(x_margin + 2.8 * inch, y, str(estados['franco_trabajados']))
        y -= 0.18 * inch
        p.drawString(x_margin, y, 'UNIDADES EVENTUALES')
        p.drawRightString(x_margin + 2.8 * inch, y, str(estados['unidades_eventuales']))
        y -= 0.18 * inch
        p.drawString(x_margin, y, 'ADELANTO DE TURNO')
        p.drawRightString(x_margin + 2.8 * inch, y, str(estados['adelanto_turno']))
        y -= 0.18 * inch
        p.drawString(x_margin, y, 'RETEN')
        p.drawRightString(x_margin + 2.8 * inch, y, str(estados['reten']))
        y -= 0.18 * inch
        p.drawString(x_margin, y, 'UNIDADES ADICIONALES')
        p.drawRightString(x_margin + 2.8 * inch, y, str(estados['unidades_adicionales']))
        y -= 0.18 * inch
        p.drawString(x_margin, y, 'CUSTODIO')
        p.drawRightString(x_margin + 2.8 * inch, y, str(estados['custodio']))
        y -= 0.18 * inch
        p.drawString(x_margin, y, 'TOTAL=')
        p.drawRightString(x_margin + 2.8 * inch, y, str(estados['total']))

    p.showPage()
    p.save()
    return response
