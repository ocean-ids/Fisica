from django.shortcuts import render
from django.contrib.auth import authenticate, login, logout
from django.http import JsonResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from django.contrib.auth.models import User
from rest_framework.response import Response
from rest_framework import status
import json
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from ..models import Cliente, Instalacion, Puesto, Persona, Horario, Asignacion
from reportlab.pdfgen import canvas
from io import BytesIO
from reportlab.platypus.tables import Table
cm = 8.54
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet , ParagraphStyle
from reportlab.lib.enums import TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.units import cm
from openpyxl.styles import Alignment, Font



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Encabezados
    ws.append(['ID', 'Nombre', 'Correo'])

    # Datos de ejemplo (podrías sacar de la BD)
    datos = [
        [1, 'Juan Pérez', 'juan@example.com'],
        [2, 'Ana Gómez', 'ana@example.com'],
        [3, 'Luis Torres', 'luis@example.com'],
    ]

    for fila in datos:
        ws.append(fila)

    # Preparar respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte.xlsx'
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte.pdf"'

    p = canvas.Canvas(response)
    p.drawString(100, 800, "Reporte de usuarios")
    
    datos = [
        [1, 'Juan Pérez', 'juan@example.com'],
        [2, 'Ana Gómez', 'ana@example.com'],
        [3, 'Luis Torres', 'luis@example.com'],
    ]

    y = 760
    for fila in datos:
        texto = f"{fila[0]} - {fila[1]} - {fila[2]}"
        p.drawString(100, y, texto)
        y -= 20

    p.showPage()
    p.save()
    return response



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generar_excel_horario(request):
    asignaciones = Asignacion.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Horario"

    # Encabezado
    encabezado = [
        "H. INGRESO", "H. SALIDA", "", "CLIENTE", "PUESTO NOMBRE", "HORARIO", "#", "CEDULA", "APELLIDOS Y NOMBRES"
    ] + list(range(1, 31))

    ws.append(encabezado)

    # Estilo encabezado
    header_font = Font(bold=True)
    for col in range(1, len(encabezado) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Datos
    for asignacion in asignaciones:
        fila = [
            asignacion.horario.hora_ingreso.strftime("%H:%M") if asignacion.horario and asignacion.horario.hora_ingreso else "",
            asignacion.horario.hora_salida.strftime("%H:%M") if asignacion.horario and asignacion.horario.hora_salida else "",
            asignacion.instalacion.codigo if asignacion.instalacion and asignacion.instalacion.codigo else "",
            asignacion.cliente.razon_social if asignacion.cliente and asignacion.cliente.razon_social else "",
            asignacion.puesto.nombre if asignacion.puesto and asignacion.puesto.nombre else "",
            asignacion.horario.denominativo if asignacion.horario and asignacion.horario.denominativo else "",
            asignacion.id,
            asignacion.persona.cedula if asignacion.persona and asignacion.persona.cedula else "",
            f"{asignacion.persona.apellidos} {asignacion.persona.nombres}" if asignacion.persona else "",
        ]

        # Días 1 al 30
        for dia in range(1, 31):
            valor = getattr(asignacion, f'dia_{dia}', '')
            fila.append(valor)

        ws.append(fila)

    # Ajustes de estilo y formato por columna
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.row == 1:
                continue
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min((max_length + 2), 50)
        ws.column_dimensions[column].width = adjusted_width

    # Configurar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Horario.xlsx'
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generar_pdf_Horario(request):
    cm = 8.54
    asignaciones = Asignacion.objects.all()
    print(asignaciones)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=Horario.pdf'
    
    doc = SimpleDocTemplate(
        response,
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=7 * cm,
        bottomMargin=1 * cm,
        pagesize=landscape(A4)
    )

    # Encabezado de la tabla
    encabezado = [
        "H. INGRESO", "H. SALIDA", "", "CLIENTE", "PUESTO NOMBRE", "HORARIO", "#", "CEDULA", "APELLIDOS Y NOMBRES"
    ] + list(range(1, 31))  # Días del 1 al 30

    data = [encabezado]

    styles = getSampleStyleSheet()
    style_nombre = ParagraphStyle(
    name="NombreWrap",
    parent=styles["Normal"],
    fontSize=6.5,         # Más pequeño para que quepa mejor
    leading=7.5,          # Espaciado entre líneas
    alignment=TA_LEFT,
    wordWrap='CJK',       # Permite saltos de línea
    )
    # Anchos de columna
    colWidths = [
        5 * cm,   # H. INGRESO
        5 * cm,   # H. SALIDA
        3 * cm,   # INSTALACIÓN
        7 * cm,   # CLIENTE
        8 * cm,   # PUESTO NOMBRE
        7 * cm,   # HORARIO
        2 * cm,   # #
        6 * cm,   # CEDULA
        13 * cm,  # APELLIDOS Y NOMBRES
    ] + [1.4 * cm] * 30  # Días

    for asignacion in asignaciones:
        try:
            fila = [
                asignacion.horario.hora_ingreso.strftime("%H:%M") if asignacion.horario and asignacion.horario.hora_ingreso else "",
                asignacion.horario.hora_salida.strftime("%H:%M") if asignacion.horario and asignacion.horario.hora_salida else "",
                asignacion.instalacion.codigo if asignacion.instalacion and asignacion.instalacion.codigo else "",
                asignacion.cliente.razon_social if asignacion.cliente and asignacion.cliente.razon_social else "",
                asignacion.puesto.nombre if asignacion.puesto and asignacion.puesto.nombre else "",
                asignacion.horario.denominativo if asignacion.horario and asignacion.horario.denominativo else "",
                asignacion.id,
                asignacion.persona.cedula if asignacion.persona and asignacion.persona.cedula else "",
                Paragraph(f"{asignacion.persona.apellidos} {asignacion.persona.nombres}", style_nombre)
                if asignacion.persona else "",
            ]

            # Agregar los valores de los días 1 al 30
            for dia in range(1, 31):
                valor = getattr(asignacion, f'dia_{dia}', '')
                fila.append(valor)

            data.append(fila)
        except Exception as e:
            print(f"Error en asignación ID {getattr(asignacion, 'id', 'sin ID')}: {e}")

    table = Table(data, colWidths=colWidths, rowHeights=18)

    style = TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
    ])
    table.setStyle(style)

    elements = [table]
    doc.build(elements)

    return response
