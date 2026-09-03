"""Cobertura de FIJO: un fijo con token D/N + nominativo (ej. DK37) ese dia cubre OTRA
instalacion -> sale en el reporte bajo ese nominativo, no en su puesto propio."""
import datetime
from django.test import TestCase
from django.utils import timezone
from rest_framework.test import APIRequestFactory

from CoreFisica.models import (
    Provincia, Canton, Cliente, Instalacion, Puesto, Persona, Asignacion, AsignacionSemanal,
)
from CoreFisica.views.reporte_asistencia_views import _build_reporte_asistencia_data
from CoreFisica.views.importar_puestos_asignaciones import (
    _cal_valor_ok, importar_formato_reporte, _split_horario_rango, _rep_detectar_columnas,
)


class CoberturaFijoTests(TestCase):
    def test_cal_valor_ok_acepta_cobertura_en_fijo(self):
        # Fijo ahora acepta tokens de COBERTURA (D/N + nominativo), ademas de D/N/F/T/V.
        self.assertTrue(_cal_valor_ok('DK37', es_saca=False))
        self.assertTrue(_cal_valor_ok('NK35', es_saca=False))
        self.assertTrue(_cal_valor_ok('D', es_saca=False))
        self.assertTrue(_cal_valor_ok('F', es_saca=False))
        # Base (DB/NB) NO es cobertura -> el fijo no lo acepta.
        self.assertFalse(_cal_valor_ok('DB', es_saca=False))
        # Basura -> no reconocido.
        self.assertFalse(_cal_valor_ok('XYZ', es_saca=False))

    def _rows(self, res):
        return res[0] if isinstance(res, tuple) else res

    def test_fijo_con_token_cobertura_sale_bajo_nominativo(self):
        prov, _ = Provincia.objects.get_or_create(nombre='GUAYAS')
        can, _ = Canton.objects.get_or_create(nombre='GUAYAQUIL', provincia=prov)
        cliA = Cliente.objects.create(razon_social='CLIA', nombre_comercial='CLIENTE A')
        cliB = Cliente.objects.create(razon_social='CLIB', nombre_comercial='CELCO GYE')
        instA = Instalacion.objects.create(cliente=cliA, canton=can, nombre='PUESTO PROPIO', codigo='K35')
        instB = Instalacion.objects.create(cliente=cliB, canton=can, nombre='CELCO GUAYAQUIL', codigo='K37')
        puestoA = Puesto.objects.create(instalacion=instA, nombre='PUESTO PROPIO', tipo='CONTROL')
        Puesto.objects.create(instalacion=instB, nombre='CELCO GUAYAQUIL', tipo='CONTROL')
        persona = Persona.objects.create(cedula='0900000001', nombres='LUIS', apellidos='TEST', tipo='FIJOS', is_active=True)

        hoy = timezone.localdate()
        fecha = datetime.date(hoy.year, hoy.month, 15)
        month_start = fecha.replace(day=1)
        if fecha.month == 12:
            end = datetime.date(fecha.year, 12, 31)
        else:
            end = datetime.date(fecha.year, fecha.month + 1, 1) - datetime.timedelta(days=1)

        asig = Asignacion.objects.create(
            persona=persona, cliente=cliA, instalacion=instA, puesto=puestoA,
            mes=fecha.month, anio=fecha.year, estado='ACTIVO',
            recurring=True, start_date=month_start, end_date=end, fecha=None,
        )
        # Calendario: el dia 15 con token de cobertura DK37 (cubre K37 en Diurno).
        ws = month_start + datetime.timedelta(days=((15 - 1) // 7) * 7)
        day_field = ['mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun'][fecha.weekday()]
        sem = AsignacionSemanal.objects.create(asignacion=asig, puesto=puestoA, week_start=ws)
        setattr(sem, day_field, 'DK37')
        sem.save()

        rows = self._rows(_build_reporte_asistencia_data(fecha=fecha.isoformat(), turno='Diurno'))
        fila = [r for r in rows if isinstance(r, dict) and r.get('asignacion_id') == asig.id]
        self.assertEqual(len(fila), 1, 'el fijo debe aparecer en Diurno ese dia')
        # Debe salir bajo el nominativo CUBIERTO (K37), no en su puesto propio (K35).
        self.assertEqual(fila[0].get('codigo'), 'K37')
        self.assertEqual(fila[0].get('cliente'), 'CELCO GYE')
        self.assertEqual(fila[0].get('instalacion_nombre'), 'CELCO GUAYAQUIL')

        # En un dia NORMAL (D) sale en su puesto propio (K35).
        otro = datetime.date(hoy.year, hoy.month, 16)
        ws2 = month_start + datetime.timedelta(days=((16 - 1) // 7) * 7)
        df2 = ['mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun'][otro.weekday()]
        sem2, _ = AsignacionSemanal.objects.get_or_create(asignacion=asig, week_start=ws2, defaults={'puesto': puestoA})
        setattr(sem2, df2, 'D')
        sem2.save()
        rows2 = self._rows(_build_reporte_asistencia_data(fecha=otro.isoformat(), turno='Diurno'))
        fila2 = [r for r in rows2 if isinstance(r, dict) and r.get('asignacion_id') == asig.id]
        self.assertEqual(len(fila2), 1)
        self.assertEqual(fila2[0].get('codigo'), 'K35', 'dia normal -> su puesto propio')


class HuecaImportParcialTests(TestCase):
    """Una fila SIN cedula ni nombre (puesto sin guardia) se importa como HUECA
    (persona=None) AUNQUE el cronograma del mes venga PARCIAL: no se exige el mes
    completo para huecas; los dias en blanco simplemente = no hay hueca ese dia."""

    def _build_wb(self, mes, anio, dias_marcados):
        from openpyxl import Workbook
        days_in_month = ((datetime.date(anio, mes + 1, 1) if mes < 12 else datetime.date(anio, 12, 31))
                         - datetime.timedelta(days=1)).day if mes < 12 else 31
        wb = Workbook()
        ws = wb.active
        ws.title = 'HOJA'
        # Encabezado que reconoce _rep_detectar_columnas: nominativo = CLIENTE-1.
        # cols: 0=nominativo 1=CLIENTE 2=PUESTO 3=TIPO 4=CEDULA 5=NOMBRES 6=HING 7=HSAL 8..=dias
        header = [None, 'CLIENTE', 'PUESTO', 'TIPO', 'CEDULA', 'APELLIDOS Y NOMBRES',
                  'H INGRESO', 'H SALIDA'] + list(range(1, days_in_month + 1))
        ws.append(header)
        cal = ['D' if (d in dias_marcados) else '' for d in range(1, days_in_month + 1)]
        # Fila HUECA: SIN cedula y SIN nombre.
        fila = ['K35', 'CELCO GYE', 'CELCO BODEGA', '24H', '', '', '07:00', '19:00'] + cal
        ws.append(fila)
        return wb

    def test_hueca_con_mes_parcial_se_importa(self):
        prov, _ = Provincia.objects.get_or_create(nombre='GUAYAS')
        can, _ = Canton.objects.get_or_create(nombre='GUAYAQUIL', provincia=prov)
        cli = Cliente.objects.create(razon_social='CELCO SA', nombre_comercial='CELCO GYE')
        inst = Instalacion.objects.create(cliente=cli, canton=can, nombre='CELCO GUAYAQUIL', codigo='K35')
        puesto = Puesto.objects.create(instalacion=inst, nombre='CELCO BODEGA', tipo='CONTROL')

        hoy = timezone.localdate()
        mes, anio = hoy.month, hoy.year
        dias_marcados = [21, 22, 23, 24, 25]  # SOLO 5 dias -> mes parcial

        wb = self._build_wb(mes, anio, dias_marcados)
        req = APIRequestFactory().get(f'/x?mes={mes}&anio={anio}&meses=0&meses_sync=0')
        importar_formato_reporte(req, wb, None)

        huecas = Asignacion.objects.filter(persona__isnull=True, instalacion=inst, puesto=puesto,
                                           mes=mes, anio=anio)
        self.assertEqual(huecas.count(), 1, 'la hueca con mes parcial debe importarse')
        self.assertTrue(huecas.first().es_hueca, 'debe quedar marcada como HUECA (no "No Cubierto")')

        # Y sale como HUECA en el reporte un dia marcado (persona vacia -> "HUECA").
        fecha = datetime.date(anio, mes, 22)
        rows = _build_reporte_asistencia_data(fecha=fecha.isoformat(), turno='Diurno')
        rows = rows[0] if isinstance(rows, tuple) else rows
        fila = [r for r in rows if isinstance(r, dict) and r.get('asignacion_id') == huecas.first().id]
        self.assertEqual(len(fila), 1, 'la hueca debe salir en el reporte el dia marcado')
        self.assertEqual((fila[0].get('nombre_apellidos') or '').upper(), 'HUECA')


class FormatoDescargableImportableTests(TestCase):
    """El importador acepta el formato NATIVO del descargable: encabezados
    NOMBRE PUESTO / PERSONA / RESUMEN y una sola columna HORARIO en rango
    'HH:MM - HH:MM' (que se parte en ingreso/salida)."""

    def test_split_horario_rango(self):
        self.assertEqual(_split_horario_rango('07:00 - 19:00'), ('07:00', '19:00'))
        self.assertEqual(_split_horario_rango('06:30 – 21:00'), ('06:30', '21:00'))  # guion largo
        self.assertEqual(_split_horario_rango('07:00'), ('07:00', None))
        self.assertEqual(_split_horario_rango(None), (None, None))

    def test_detecta_ambos_formatos(self):
        # Formato PARA IMPORTAR (clasico)
        clasico = [['NOMINATIVO', 'CLIENTE', 'PUESTO', 'TIPO', 'CEDULA',
                    'APELLIDOS Y NOMBRES', 'H INGRESO', 'H SALIDA', 1, 2, 3]]
        ri, col = _rep_detectar_columnas(clasico)
        self.assertEqual(ri, 0)
        self.assertIsNotNone(col.get('ing'))
        # Formato DESCARGABLE (nativo)
        desc = [['HORARIO', 'NOMINATIVO', 'CLIENTE', 'NOMBRE PUESTO', 'RESUMEN',
                 'CEDULA', 'PERSONA', 1, 2, 3]]
        ri2, col2 = _rep_detectar_columnas(desc)
        self.assertEqual(ri2, 0)
        self.assertIsNotNone(col2.get('hor'))      # columna HORARIO detectada
        self.assertIsNotNone(col2.get('pue'))      # NOMBRE PUESTO -> pue
        self.assertIsNotNone(col2.get('nombre'))   # PERSONA -> nombre

    def test_importa_formato_descargable(self):
        from openpyxl import Workbook
        prov, _ = Provincia.objects.get_or_create(nombre='GUAYAS')
        can, _ = Canton.objects.get_or_create(nombre='GUAYAQUIL', provincia=prov)
        cli = Cliente.objects.create(razon_social='CELCO SA', nombre_comercial='CELCO GYE')
        inst = Instalacion.objects.create(cliente=cli, canton=can, nombre='CELCO GUAYAQUIL', codigo='K35')
        Puesto.objects.create(instalacion=inst, nombre='CELCO BODEGA', tipo='CONTROL')
        persona = Persona.objects.create(cedula='0900000123', nombres='JUAN', apellidos='PEREZ',
                                         tipo='FIJOS', is_active=True)

        hoy = timezone.localdate()
        mes, anio = hoy.month, hoy.year
        dim = ((datetime.date(anio, mes + 1, 1) if mes < 12 else datetime.date(anio, 12, 31))
               - datetime.timedelta(days=1)).day if mes < 12 else 31

        wb = Workbook(); ws = wb.active; ws.title = 'HOJA'
        # Encabezado NATIVO del descargable
        ws.append(['HORARIO', 'NOMINATIVO', 'CLIENTE', 'NOMBRE PUESTO', 'RESUMEN',
                   'CEDULA', 'PERSONA'] + list(range(1, dim + 1)))
        ws.append(['07:00 - 19:00', 'K35', 'CELCO GYE', 'CELCO BODEGA', '1 24HLD',
                   '0900000123', 'PEREZ JUAN'] + ['D'] * dim)

        req = APIRequestFactory().get(f'/x?mes={mes}&anio={anio}&meses=0&meses_sync=0')
        importar_formato_reporte(req, wb, None)

        asig = Asignacion.objects.filter(persona=persona, mes=mes, anio=anio).first()
        self.assertIsNotNone(asig, 'la asignacion del formato descargable debe importarse')
        self.assertEqual(asig.instalacion_id, inst.id)
        # El HORARIO en rango se partio en ingreso/salida
        self.assertIsNotNone(asig.horario)
        self.assertEqual(str(asig.horario.hora_ingreso)[:5], '07:00')
        self.assertEqual(str(asig.horario.hora_salida)[:5], '19:00')

    def test_fila_con_puesto_vacio_no_se_importa_y_avisa(self):
        """Una fila con persona pero PUESTO en blanco NO se importa (no hereda el
        puesto de la fila de arriba) y genera una alerta 'PUESTO vacia'."""
        from openpyxl import Workbook
        prov, _ = Provincia.objects.get_or_create(nombre='GUAYAS')
        can, _ = Canton.objects.get_or_create(nombre='GUAYAQUIL', provincia=prov)
        cli = Cliente.objects.create(razon_social='CELCO SA', nombre_comercial='CELCO GYE')
        inst = Instalacion.objects.create(cliente=cli, canton=can, nombre='CELCO GUAYAQUIL', codigo='K35')
        Puesto.objects.create(instalacion=inst, nombre='CELCO BODEGA', tipo='CONTROL')
        p_ok = Persona.objects.create(cedula='0900000201', nombres='ANA', apellidos='LOPEZ',
                                      tipo='FIJOS', is_active=True)
        p_sin = Persona.objects.create(cedula='0900000202', nombres='BETO', apellidos='RUIZ',
                                       tipo='FIJOS', is_active=True)

        hoy = timezone.localdate()
        mes, anio = hoy.month, hoy.year
        dim = ((datetime.date(anio, mes + 1, 1) if mes < 12 else datetime.date(anio, 12, 31))
               - datetime.timedelta(days=1)).day if mes < 12 else 31

        wb = Workbook(); ws = wb.active; ws.title = 'HOJA'
        ws.append(['NOMINATIVO', 'CLIENTE', 'PUESTO', 'TIPO', 'CEDULA',
                   'APELLIDOS Y NOMBRES', 'H INGRESO', 'H SALIDA'] + list(range(1, dim + 1)))
        # Fila 1: valida (con puesto)
        ws.append(['K35', 'CELCO GYE', 'CELCO BODEGA', '1 24HLD', '0900000201',
                   'LOPEZ ANA', '07:00', '19:00'] + ['D'] * dim)
        # Fila 2: PUESTO en blanco (aunque la de arriba SI tiene puesto -> NO se hereda)
        ws.append(['K35', 'CELCO GYE', '', '', '0900000202',
                   'RUIZ BETO', '', ''] + ['D'] * dim)

        req = APIRequestFactory().get(f'/x?mes={mes}&anio={anio}&meses=0&meses_sync=0')
        resp = importar_formato_reporte(req, wb, None)
        data = resp.data if hasattr(resp, 'data') else resp

        # La de arriba SI se importa
        self.assertIsNotNone(Asignacion.objects.filter(persona=p_ok, mes=mes, anio=anio).first())
        # La de puesto vacio NO se importa (y NO hereda 'CELCO BODEGA')
        self.assertIsNone(Asignacion.objects.filter(persona=p_sin, mes=mes, anio=anio).first(),
                          'la fila sin puesto no debe importarse ni heredar el puesto de arriba')
        # Y sale la alerta
        errores = data.get('errores', []) if isinstance(data, dict) else []
        self.assertTrue(any('PUESTO vacia' in e and '0900000202' in e for e in errores),
                        f'debe avisar PUESTO vacia para la fila sin puesto. Alertas: {errores}')

    def test_cedula_no_coincide_con_nombre_no_se_importa(self):
        """Si la cedula del Excel pertenece a OTRA persona (el nombre no coincide),
        la fila NO se importa y sale alerta 'revisar cedula'."""
        from openpyxl import Workbook
        prov, _ = Provincia.objects.get_or_create(nombre='GUAYAS')
        can, _ = Canton.objects.get_or_create(nombre='GUAYAQUIL', provincia=prov)
        cli = Cliente.objects.create(razon_social='CELCO SA', nombre_comercial='CELCO GYE')
        inst = Instalacion.objects.create(cliente=cli, canton=can, nombre='CELCO GUAYAQUIL', codigo='K35')
        Puesto.objects.create(instalacion=inst, nombre='CELCO BODEGA', tipo='CONTROL')
        # La cedula 0900000301 pertenece a PEREZ JUAN...
        p = Persona.objects.create(cedula='0900000301', nombres='JUAN', apellidos='PEREZ',
                                   tipo='FIJOS', is_active=True)

        hoy = timezone.localdate()
        mes, anio = hoy.month, hoy.year
        dim = ((datetime.date(anio, mes + 1, 1) if mes < 12 else datetime.date(anio, 12, 31))
               - datetime.timedelta(days=1)).day if mes < 12 else 31

        wb = Workbook(); ws = wb.active; ws.title = 'HOJA'
        ws.append(['NOMINATIVO', 'CLIENTE', 'PUESTO', 'TIPO', 'CEDULA',
                   'APELLIDOS Y NOMBRES', 'H INGRESO', 'H SALIDA'] + list(range(1, dim + 1)))
        # ...pero en el Excel esa cedula viene con el nombre de OTRA persona.
        ws.append(['K35', 'CELCO GYE', 'CELCO BODEGA', '1 24HLD', '0900000301',
                   'GOMEZ LUIS CARLOS', '07:00', '19:00'] + ['D'] * dim)

        req = APIRequestFactory().get(f'/x?mes={mes}&anio={anio}&meses=0&meses_sync=0')
        resp = importar_formato_reporte(req, wb, None)
        data = resp.data if hasattr(resp, 'data') else resp

        # NO se importa a PEREZ JUAN
        self.assertIsNone(Asignacion.objects.filter(persona=p, mes=mes, anio=anio).first(),
                          'no debe importar cuando la cedula no corresponde al nombre')
        errores = data.get('errores', []) if isinstance(data, dict) else []
        self.assertTrue(any('revisar cedula' in e.lower() and '0900000301' in e for e in errores),
                        f'debe avisar que la cedula no coincide. Alertas: {errores}')
