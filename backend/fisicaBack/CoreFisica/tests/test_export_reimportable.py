"""Round-trip: el export re-importable se puede volver a importar (actualiza sin duplicar)."""
import io, json
from django.test import TestCase
from django.contrib.auth.models import User, Permission
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook, load_workbook

from CoreFisica.models import (
    Provincia, Canton, Cliente, Instalacion, Persona, Asignacion,
    AsignacionSemanal, SacafrancoFila,
)
from CoreFisica.views.importar_puestos_asignaciones import es_formato_reporte


def _login(c, u, p):
    return c.post('/api/login/', data=json.dumps({'username': u, 'password': p}),
                  content_type='application/json').json().get('access')


def _xlsx():
    wb = Workbook(); ws = wb.active
    ws.append(['NOMINATIVO', 'CLIENTE', 'PUESTO', 'TIPO', 'CEDULA',
               'APELLIDOS Y NOMBRES', 'H INGRESO', 'H SALIDA'] + list(range(1, 32)))
    ws.append(['N-001', 'CX', 'PUESTO 1', '1 12HD', '0912345678',
               'PEREZ LOPEZ JUAN', '08:00', '20:00'] + ['D'] * 31)
    ws.append(['', '', 'SACAFRANCO', '', '0987654321',
               'GOMEZ RUIZ MARIA', '', ''] + ['F'] * 31)
    b = io.BytesIO(); wb.save(b); return b.getvalue()


class ExportReimportableTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='er', password='ErPass123!', email='er@e.com')
        for cod in ('import_puestos_asignaciones', 'export_asignacion', 'view_asignacion'):
            self.user.user_permissions.add(Permission.objects.get(codename=cod))
        self.access = _login(self.client, 'er', 'ErPass123!')
        prov, _ = Provincia.objects.get_or_create(nombre='GUAYAS')
        can, _ = Canton.objects.get_or_create(nombre='GUAYAQUIL', provincia=prov)
        cli = Cliente.objects.create(razon_social='CX', nombre_comercial='CX')
        Instalacion.objects.create(cliente=cli, canton=can, nombre='INST', codigo='N-001')

    def _auth(self): return {'HTTP_AUTHORIZATION': f'Bearer {self.access}'}

    def _importar(self, contenido, nombre='r.xlsx', con_mes=True):
        f = SimpleUploadedFile(nombre, contenido, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # con_mes=False simula la importacion desde Puestos (sin ?mes=&anio=):
        # el mes/anio deben detectarse desde la propia hoja.
        url = '/api/v1/importar-puestos-asignaciones/'
        if con_mes:
            url += '?mes=7&anio=2026'
        return self.client.post(url, {'file': f}, **self._auth())

    def test_roundtrip_con_export_reimportable(self):
        # 1) import inicial
        self.assertEqual(self._importar(_xlsx()).status_code, 200)
        p = Persona.objects.get(cedula='0912345678')
        asig_count = Asignacion.objects.filter(persona=p).count()
        saca_count = SacafrancoFila.objects.filter(persona__cedula='0987654321').count()

        # 2) DESCARGAR el reporte de Asignaciones (el "Descargar" de siempre)
        exp = self.client.get('/api/v1/reporte-asignaciones/?mes=07&anio=2026', **self._auth())
        self.assertEqual(exp.status_code, 200, exp.content[:200])

        # 3) el descargable DEBE ser re-importable (trae hoja DATOS en formato reporte)
        wb = load_workbook(io.BytesIO(exp.content), data_only=True)
        reconocible = es_formato_reporte(wb)
        print(f"\n[DESCARGABLE Asignaciones] ¿es re-importable? -> {reconocible}")
        self.assertTrue(reconocible, 'El descargable de Asignaciones DEBE ser reconocible por el importador')

        # 4) re-importar el descargable SIN pasar mes/anio (como desde Puestos):
        #    el mes/anio debe detectarse desde la hoja (fila "JULIO 2026").
        r2 = self._importar(exp.content, nombre='reexport.xlsx', con_mes=False)
        cuerpo = r2.json()
        print(f"[RE-IMPORT sin mes] HTTP {r2.status_code} -> {json.dumps(cuerpo, ensure_ascii=False)[:250]}")
        self.assertEqual(r2.status_code, 200, r2.content)
        self.assertEqual(cuerpo.get('errores', []), [], 'no debe haber errores de deteccion de mes')

        # 5) NO duplica
        self.assertEqual(Asignacion.objects.filter(persona=p).count(), asig_count, 'no debe duplicar asignaciones')
        self.assertEqual(SacafrancoFila.objects.filter(persona__cedula='0987654321').count(), saca_count, 'no debe duplicar sacafranco')
        # 6) calendario preservado (D en julio 1 = miercoles)
        asig_jul = Asignacion.objects.get(persona=p, mes=7, anio=2026)
        sem = AsignacionSemanal.objects.filter(asignacion=asig_jul).first()
        self.assertEqual((sem.wed or '').upper(), 'D')
